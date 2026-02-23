import os
import tempfile
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from pathlib import Path
from core.loader import SetupLoader
from core.models import CourseMetadata
from core.setup_validator import SetupValidator
from core.filled_validator import FilledMarksValidator
from core.exceptions import ValidationError, SystemError

from core.constants import (
    header_fill, table_border, table_center_alignment,
    DIRECT_RATIO, INDIRECT_RATIO,
    ABSENT_SYMBOL, WORKBOOK_PASSWORD,
    MIN_MARK_VALUE, LIKERT_MAX, LIKERT_MIN,
    COURSE_CODE, SECTION, SEMESTER, ACADEMIC_YEAR,
    MARGIN_LEFT,MARGIN_RIGHT,MARGIN_TOP,MARGIN_BOTTOM, MARGIN_HEADER, MARGIN_FOOTER,
    DIRECT, INDIRECT
)

class COCalculator:
    def __init__(self, metadata: CourseMetadata, setup_path: str, filled_path: str):
        self.setup_path = setup_path
        self.metadata = metadata
        self.filled_path = filled_path
        self.startrow = 0
        
        loader = SetupLoader(setup_path)
        config = loader.load_config()
        students = loader.load_students()
        qmap = loader.load_question_map()

        # Capture the original order of components from the config sheet
        self.original_component_order = list(config.dataframe["Component"].unique())

        validator = SetupValidator(
            metadata=self.metadata,
            config=config,
            students=students,
            question_map=qmap,
            only_CA=False,
        )
        self.validated = validator.validate()

        self.students_df = pd.DataFrame(
            [(str(s.reg_no).strip(), s.name) for s in self.validated.students],
            columns=["RegNo", "Student_Name"]
        )
        filled_validator = FilledMarksValidator(
            validated_setup=self.validated,
            filled_path=self.filled_path,
        )
        filled_validator.validate()
        self.filled_excel = pd.ExcelFile(self.filled_path)

        self.sheet_map = {
            str(name).strip().lower(): name
            for name in self.filled_excel.sheet_names
        }
        
        bold = Font(bold=True)
        self.bold_font = bold
        self.center_style = table_center_alignment()
        self.thin_border = table_border()
        self.grey_fill = header_fill()

    def _resolve_sheet_name(self, requested_name: str) -> str:
        actual_name = self.sheet_map.get(requested_name.strip().lower())
        if actual_name is None:
            raise ValidationError(
                f"Sheet '{requested_name}' not found."
            )
        return str(actual_name)
    
    def _apply_header_footer(self, ws):
        # Extract data from self.metadata (adjust keys based on your CourseMetadata model)
        course_code = getattr(self.metadata, COURSE_CODE, 'N/A')
        section = getattr(self.metadata, SECTION, 'N/A')
        semester = getattr(self.metadata, SEMESTER, 'N/A')
        academic_year = getattr(self.metadata, ACADEMIC_YEAR, 'N/A')

        # Format: Left: Course & Section | Right: Sem & Year        
        # Apply to the worksheet
        ws.oddHeader.left.text = f"{course_code} - {section}"
        ws.oddHeader.right.text = f"{semester} | {academic_year}"
        
        # Optional: Add page numbers to the footer
        ws.oddFooter.center.text = "Page &P of &N"
    def _set_page_layout(self, ws, report_type):
        ws.sheet_properties.pageSetup = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToPage = True
        ws.page_setup.scale = None
        # Set Minimum Margins (0.25 inches is standard for 'Narrow')
        # Units are in inches
        ws.page_margins.left = MARGIN_LEFT
        ws.page_margins.right = MARGIN_RIGHT
        ws.page_margins.top = MARGIN_TOP    # Slightly more for the header
        ws.page_margins.bottom = MARGIN_BOTTOM # Slightly more for the footer
        ws.page_margins.header = MARGIN_HEADER
        ws.page_margins.footer = MARGIN_FOOTER
        if report_type== DIRECT:
            # Direct reports usually have many columns, so A3 Landscape is better
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
        else:
            # Indirect reports are usually narrower, so A4 Portrait fits well
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
        # Optional: Scale to fit one page width
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # Allow it to span multiple pages vertically
        
    def _setup_report_template(self, ws, title_text, headers):
        # --- 1. Title ---
        total_cols = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = self.bold_font # Ensure these are class attributes or passed in
        title_cell.alignment = self.center_style
        
        for col in range(1, total_cols + 1):
            ws.cell(1, col).border = self.thin_border
        ws.append([""]) # Spacer row after title

        # --- 2. Metadata (Rows 2-7) ---
        metadata_items = list(vars(self.metadata).items())[:6]
        for key, value in metadata_items:
            ws.append([key.replace("_", " ").title(), value])
        
        for row in ws.iter_rows(min_row=ws.max_row-len(metadata_items)+1, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = self.thin_border
                cell.alignment = self.center_style
                cell.font = self.bold_font if cell.column == 1 else Font(bold=False)

        # --- 3. Spacer & Headers ---
        for _ in range(2): 
                ws.append([""]) # Empty rows for spacing
        header_row = ws.max_row + 1
        for c, h in enumerate(headers, 1):
            cell = ws.cell(header_row, c, h)
            cell.font = self.bold_font
            cell.fill = self.grey_fill
            cell.alignment = self.center_style
            cell.border = self.thin_border
        self.startrow = ws.max_row
    
    def _collect_direct(self) -> pd.DataFrame:
        records = []
        for comp_name, comp in self.validated.components.items():
            if not comp.direct: continue
            #df = pd.read_excel(self.filled_excel, sheet_name=comp_name, header=None)
            actual_name = self._resolve_sheet_name(comp_name)
            df = pd.read_excel(self.filled_excel, sheet_name=actual_name, header=None)
            student_block = df.iloc[3:].reset_index(drop=True)
            regnos = student_block.iloc[:, 0].astype(str).str.strip()

            for q_index, question in enumerate(comp.questions):
                col_index = q_index + 2
                raw_series = student_block.iloc[:, col_index]
                absent_mask = (
                    raw_series
                    .astype(str)
                    .str.strip()
                    .str.lower()
                    .eq(ABSENT_SYMBOL.strip().lower())
                )
                numeric_series = pd.to_numeric(raw_series, errors='coerce')
                raw_alloc_clean = numeric_series.mask(absent_mask)

                num_cos = len(question.co_list)

                for co in question.co_list:
                    if not comp.co_split:
                        # If they were absent (NaN), NaN / num_cos is still NaN. 
                        # If they got 0, 0 / num_cos is 0. This is perfect.
                        raw_alloc = raw_alloc_clean / num_cos
                        max_alloc = question.max_marks / num_cos
                    else:
                        raw_alloc = raw_alloc_clean
                        max_alloc = question.max_marks

                    records.append(pd.DataFrame({
                        "RegNo": regnos,
                        "CO": int(co),
                        "Component": comp_name,
                        "Raw": raw_alloc,
                        "Max": max_alloc,
                        "Weight": comp.weight,
                    }))


        if not records: return pd.DataFrame()
        
        df_direct = pd.concat(records, ignore_index=True)
        return df_direct

    def _collect_indirect(self) -> pd.DataFrame:
        if not self.validated.indirect_tools: 
            return pd.DataFrame()
            
        records = []
        for tool in self.validated.indirect_tools:
            # Load the specific survey sheet
            indirect_extension = INDIRECT.strip().lower()
            safe_sheet_name = f"{tool.name}_{indirect_extension}"[:31]
            #df = pd.read_excel(self.filled_excel, sheet_name=safe_sheet_name)
            actual_name = self._resolve_sheet_name(safe_sheet_name)
            df = pd.read_excel(self.filled_excel, sheet_name=actual_name)
            
            # Ensure RegNo is a clean string
            df["RegNo"] = df["RegNo"].astype(str).str.strip()
            
            # Unpivot the CO columns into rows
            melted = df.melt(id_vars=["RegNo"], var_name="CO", value_name="Likert")
            
            # Filter only the columns that are COs
            melted = melted[melted["CO"].astype(str).str.contains("CO", case=False)]
            
            # Add the Tool name to every row in this specific sheet's data
            melted["Tool"] = tool.name
            absent_mask = (
                melted["Likert"]
                .astype(str)
                .str.strip()
                .str.lower()
                .eq(ABSENT_SYMBOL.strip().lower())
            )
            
            # Process Likert scores
            melted["Likert"] = pd.to_numeric(melted["Likert"], errors='coerce')
            melted["Likert"] = melted["Likert"].mask(absent_mask)
            
            # Calculate percentage for this specific tool entry
            #melted["IndirectPercent"] = ((melted["Likert"] - 1) / 4 * 100).clip(lower=0)
            scale_range = LIKERT_MAX - LIKERT_MIN
            if scale_range <= 0:
                raise SystemError("Invalid Likert scale configuration.")
                
            # (Score - Min) / (Max - Min) * 100
            # If Likert is NaN (Absent), IndirectPercent becomes NaN automatically
            melted["IndirectPercent"] = (
                (melted["Likert"] - LIKERT_MIN) / scale_range * 100
            ).clip(lower=0)
            
            records.append(melted)
            
        if not records: 
            return pd.DataFrame()
        
        # Combine all records into one large DataFrame
        # By removing .groupby(), the row count will be = (Students * COs * Number of Tools)
        df_final = pd.concat(records, ignore_index=True)
        
        # Clean up CO column to ensure consistent format (e.g., '1')
        df_final["CO"] = (
            df_final["CO"]
            .astype(str)
            .str.upper()
            .str.replace("CO", "", regex=False)
            .astype(int)
        )
        
        return df_final

    def compute(self, output_path: str):
        try:
            df_direct = self._collect_direct()
            df_indirect = self._collect_indirect()

            # --- DATA SANITIZATION ---
            # Strip spaces from RegNo and Components
            for df in [df_direct, df_indirect]:
                if not df.empty:
                    if "RegNo" in df.columns:
                        df["RegNo"] = df["RegNo"].astype(str).str.strip()
                    if "CO" in df.columns:
                        # Strip "CO" if present, then convert to int
                        df["CO"] = (
                            df["CO"]
                            .astype(str)
                            .str.upper()
                            .str.replace("CO", "", regex=False)
                            .str.strip()
                            .astype(int)
                        )

            if df_direct.empty:
                raise ValidationError("No direct marks found.")
            
            df_direct = df_direct.drop_duplicates(subset=["RegNo", "CO", "Component"], keep="last")
            df_indirect = df_indirect.drop_duplicates(subset=["RegNo", "CO", "Tool"], keep="last")

            # -------------------------
            # DIRECT CALCULATION
            # -------------------------
            df_work = df_direct.groupby(["RegNo", "CO", "Component"], as_index=False).agg({
                "Raw": lambda x: x.sum(min_count=1),
                "Max": "sum",
                "Weight": "first"
            })

           
            df_work["ComponentPercent"] = np.where(
                (df_work["Raw"].notna()) & (df_work["Max"] > 0),
                (df_work["Raw"] / df_work["Max"]) * 100,
                0
            )
 
            df_work["WeightedScore"] = df_work["ComponentPercent"] * df_work["Weight"]
            df_work["ScaledComponent"] = df_work["WeightedScore"] / 100

            # Compute expected total direct weight per CO (fixed denominator)
            expected_weight_per_co = (
                df_work
                .drop_duplicates(["CO", "Component"])
                .groupby("CO")["Weight"]
                .sum()
                .to_dict()
            )

            direct_res = (
                df_work
                .groupby(["RegNo", "CO"], as_index=False)
                .agg({
                    "WeightedScore": "sum"
                })
            )

            direct_res["ExpectedWeight"] = direct_res["CO"].map(expected_weight_per_co)

            direct_res["DirectPercent"] = np.where(
                direct_res["ExpectedWeight"] > 0,
                direct_res["WeightedScore"] / direct_res["ExpectedWeight"],
                0
            )

            direct_res = direct_res[["RegNo", "CO", "DirectPercent"]]

            # -------------------------
            # INDIRECT CALCULATION
            # -------------------------
            if df_indirect.empty:
                indirect_res = pd.DataFrame(columns=["RegNo", "CO", "IndirectPercent"])
            else:
                tool_weight_map = {
                    tool.name: tool.weight
                    for tool in self.validated.indirect_tools
                }
                df_indirect["ToolWeight"] = df_indirect["Tool"].map(tool_weight_map)
                df_indirect["ScaledTool"] = (
                    df_indirect["IndirectPercent"] * df_indirect["ToolWeight"] / 100
                )

                indirect_res = (
                    df_indirect
                    .groupby(["RegNo", "CO"], as_index=False)
                    .agg({"ScaledTool": "sum"})
                    .rename(columns={"ScaledTool": "IndirectPercent"})
                )

            # -------------------------
            # MERGE DIRECT + INDIRECT
            # -------------------------
            final_df = direct_res.merge(
                indirect_res,
                on=["RegNo", "CO"],
                how="left"
            )

            final_df["IndirectPercent"] = pd.to_numeric(
                final_df["IndirectPercent"],
                errors="coerce"
            ).fillna(MIN_MARK_VALUE)

            final_df["FinalPercent"] = (
                final_df["DirectPercent"] * DIRECT_RATIO +
                final_df["IndirectPercent"] * INDIRECT_RATIO
            )

            # Store for report usage
            self.final_df = final_df
            self.df_direct = df_direct
            self.df_indirect = df_indirect
            self.df_direct_breakdown = df_work
            self.df_indirect_breakdown = df_indirect

            self.final_df["DirectPercent"] = self.final_df["DirectPercent"].clip(0, 100)
            self.final_df["IndirectPercent"] = self.final_df["IndirectPercent"].clip(0, 100)
            self.final_df["FinalPercent"] = self.final_df["FinalPercent"].clip(0, 100)

            self._generate_report(output_path)

        except Exception as e:
            raise SystemError(f"Computation failed: {str(e)}")

    def _generate_report(self, output_path: str):

        final_df = self.final_df
        df_direct = self.df_direct_breakdown
        df_indirect = self.df_indirect_breakdown
        df_direct["RegNo"] = df_direct["RegNo"].astype(str)
        df_indirect["RegNo"] = df_indirect["RegNo"].astype(str)
        final_df["RegNo"] = final_df["RegNo"].astype(str)

        # Pre-build lookup dictionaries for fast access
        direct_lookup = (
            df_direct
            .set_index(["RegNo", "CO", "Component"])
            .to_dict("index")
        )

        indirect_lookup = (
            df_indirect
            .set_index(["RegNo", "CO", "Tool"])
            .to_dict("index")
        )

        final_lookup = (
            final_df
            .set_index(["RegNo", "CO"])
            .to_dict("index")
        )
        direct_by_co = {
            co: group
            for co, group in df_direct.groupby("CO")
        }

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        all_cos = sorted(final_df["CO"].unique())
        students_df = self.students_df.copy()
        students_df["RegNo"] = students_df["RegNo"].astype(str)

        for co_label in all_cos:

            co_str = f"CO{co_label}"

            # ======================================================
            # DIRECT SHEET
            # ======================================================
            ws_d = wb.create_sheet(f"{co_str}_Direct")

            co_direct = direct_by_co.get(co_label)

            if co_direct is not None:
                components_set = set(co_direct["Component"].unique())
            else:
                components_set = set()

            present_components = [
                comp for comp in self.original_component_order
                if comp in components_set
            ]


            headers = ["RegNo", "Student Name"]

            if co_direct is not None:
                weight_map = (
                    co_direct
                    .drop_duplicates("Component")
                    .set_index("Component")["Weight"]
                    .to_dict()
                )
            else:
                weight_map = {}

            if co_direct is not None:
                for comp in present_components:
                    max_map = (
                        co_direct
                        .drop_duplicates("Component")
                        .set_index("Component")["Max"]
                        .to_dict()
                    )

                    max_val = max_map.get(comp, 0)
                    weight_val = weight_map.get(comp, 0)
                    headers += [
                        f"{comp} ({max_val})",
                        f"{comp} wtd. ({weight_val})"
                    ]
            else:
                pass            

            max_total_direct = sum(weight_map.get(comp, 0) for comp in present_components)

            headers += [
                f"Total ({max_total_direct})",
                "100 %",
                f"{int(DIRECT_RATIO*100)}%"
            ]

            self._setup_report_template(
                ws_d,
                f"{co_str} Direct Attainment",
                headers
            )

            start_row = self.startrow + 1

            for r_idx, row in zip(range(start_row, start_row + len(students_df)), students_df.itertuples(index=False)):

                regno = str(row.RegNo)
                name = row.Student_Name

                ws_d.cell(r_idx, 1, regno)
                ws_d.cell(r_idx, 2, name)

                curr_col = 3
                total_scaled = 0
                all_absent = True

                for comp in present_components:

                    key = (regno, co_label, comp)
                    row_obj = direct_lookup.get(key)

                    if row_obj and not pd.isna(row_obj["Raw"]):
                        raw = float(row_obj["Raw"])
                        scaled = float(row_obj["ScaledComponent"])

                        ws_d.cell(r_idx, curr_col, raw)
                        ws_d.cell(r_idx, curr_col + 1, round(scaled, 2))

                        total_scaled += scaled
                        all_absent = False
                    else:
                        ws_d.cell(r_idx, curr_col, ABSENT_SYMBOL.upper())
                        ws_d.cell(r_idx, curr_col + 1, ABSENT_SYMBOL.upper())

                    curr_col += 2

                final_obj = final_lookup.get((regno, co_label))

                if final_obj:
                    direct_percent = float(final_obj["DirectPercent"])
                else:
                    direct_percent = MIN_MARK_VALUE

                direct_contribution = direct_percent * DIRECT_RATIO


                ws_d.cell(r_idx, curr_col, 0 if all_absent else round(total_scaled, 2))
                ws_d.cell(r_idx, curr_col + 1, round(direct_percent, 2))
                ws_d.cell(r_idx, curr_col + 2, round(direct_contribution, 2))
            ws_d.freeze_panes = f"C{start_row}"

            # ======================================================
            # INDIRECT SHEET
            # ======================================================
            ws_i = wb.create_sheet(f"{co_str}_Indirect")

            headers = ["RegNo", "Student Name"]

            for tool in self.validated.indirect_tools:
                headers += [
                    f"{tool.name} (1-5)",
                    f"{tool.name} wtd. ({tool.weight})"
                ]

            headers += [
                "100%",
                f"{int(INDIRECT_RATIO*100)}%"
            ]

            self._setup_report_template(
                ws_i,
                f"{co_str} Indirect Attainment",
                headers
            )

            start_row = self.startrow + 1

            for r_idx, row in zip(
                    range(start_row, start_row + len(students_df)),
                    students_df.itertuples(index=False)
            ):
                regno = str(row.RegNo)
                name = row.Student_Name

                ws_i.cell(r_idx, 1, regno)
                ws_i.cell(r_idx, 2, name)

                curr_col = 3
                all_absent = True

                for tool in self.validated.indirect_tools:

                    key = (regno, co_label, tool.name)
                    row_obj = indirect_lookup.get(key)

                    if row_obj and not pd.isna(row_obj["Likert"]):
                        likert = float(row_obj["Likert"])
                        scaled = float(row_obj["ScaledTool"])

                        ws_i.cell(r_idx, curr_col, likert)
                        ws_i.cell(r_idx, curr_col + 1, round(scaled, 2))
                        all_absent = False
                    else:
                        ws_i.cell(r_idx, curr_col, ABSENT_SYMBOL.upper())
                        ws_i.cell(r_idx, curr_col + 1, ABSENT_SYMBOL.upper())


                    curr_col += 2

                final_obj = final_lookup.get((regno, co_label))
                if final_obj:
                    indirect_percent = float(final_obj["IndirectPercent"])
                else:
                    indirect_percent = MIN_MARK_VALUE

                indirect_contribution = indirect_percent * INDIRECT_RATIO


                ws_i.cell(r_idx, curr_col, round(indirect_percent, 2))
                ws_i.cell(r_idx, curr_col + 1, round(indirect_contribution, 2))
            ws_i.freeze_panes = f"C{start_row}"

            for ws in [ws_d, ws_i]:
                # ----- Formatting -----
                for row in ws.iter_rows(min_row=start_row):
                    for cell in row:
                        cell.border = self.thin_border
                        if cell.column == 2:
                            cell.alignment = Alignment(wrap_text=True, vertical="center")
                        else:
                            cell.alignment = self.center_style

                # Auto-fit
                for col_idx in range(1, ws.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    for row in range(2, ws.max_row + 1):
                        val = ws.cell(row=row, column=col_idx).value
                        if val is not None:
                            max_length = max(max_length, len(str(val)))
                    ws.column_dimensions[column_letter].width = max_length + 1

                ws.protection.set_password(WORKBOOK_PASSWORD)
                ws.protection.enable()
                ws.protection.sort = True
                ws.protection.autoFilter = True
                ws.protection.selectLockedCells = False
                ws.protection.selectUnlockedCells = False
                self._apply_header_footer(ws)
                if ws.title.endswith("_Direct"):
                    self._set_page_layout(ws, report_type=DIRECT)
                else:
                    self._set_page_layout(ws, report_type=INDIRECT)

        # -------- SAVE --------
        output_path_obj = Path(output_path)
        output_path_obj.parent.mkdir(parents=True, exist_ok=True)

        temp_file = None

        try:
            with tempfile.NamedTemporaryFile(
                delete=False,
                dir=output_path_obj.parent,
                suffix=".xlsx"
            ) as tmp:
                temp_file = Path(tmp.name)

            wb.save(temp_file)
            os.replace(temp_file, output_path_obj)

        finally:
            if temp_file and temp_file.exists():
                try:
                    temp_file.unlink()
                except Exception:
                    pass
            wb.close()
            del wb
