import hashlib
import os
import tempfile

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Protection, Font, Alignment, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from core.constants import WORKBOOK_PASSWORD
from core.constants import LIKERT_MIN, LIKERT_MAX, ABSENT_SYMBOL
from core.constants import DEFAULT_NUMBER_FORMAT
from core.models import ValidatedSetup, CourseMetadata
from core.exceptions import ValidationError


class MarksWorkbookGenerator:

    def __init__(
        self,
        metadata: CourseMetadata,
        validated_setup: ValidatedSetup,
    ):
        self.metadata = metadata
        self.setup = validated_setup

    # --------------------------------------------------
    # Utilities
    # --------------------------------------------------

    @staticmethod
    def _autofit_columns(ws):
        for column_cells in ws.columns:
            max_length = 0
            column_index = column_cells[0].column

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[
                get_column_letter(column_index)
            ].width = max_length + 2

    # ... inside your MarksWorkbookGenerator class ...

    def _apply_header_footer(self, ws):
        # Extract data from self.metadata (adjust keys based on your CourseMetadata model)
        course_code = getattr(self.metadata, 'course_code', 'N/A')
        section = getattr(self.metadata, 'section', 'N/A')
        semester = getattr(self.metadata, 'semester', 'N/A')
        academic_year = getattr(self.metadata, 'academic_year', 'N/A')

        # Format: Left: Course & Section | Right: Sem & Year        
        # Apply to the worksheet
        ws.oddHeader.left.text = f"Course: {course_code} - Section: {section}"
        ws.oddHeader.right.text = f"{semester} | {academic_year}"
        
        # Optional: Add page numbers to the footer
        ws.oddFooter.center.text = "Page &P of &N"
    def _set_page_layout(self, ws):
        # Set Orientation and Paper Size
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Set Minimum Margins (0.25 inches is standard for 'Narrow')
        # Units are in inches
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5    # Slightly more for the header
        ws.page_margins.bottom = 0.5 # Slightly more for the footer
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

    def _embed_system_hash(self, wb, validated_setup):
        hasher = hashlib.sha256()

        # ---- Components + Questions ----
        for comp_name, comp in sorted(validated_setup.components.items()):
            hasher.update(comp_name.encode())

            for q in comp.questions:
                hasher.update(q.identifier.encode())
                hasher.update(str(q.max_marks).encode())
                hasher.update(",".join(q.co_list).encode())

        # ---- Students ----
        for student in validated_setup.students:
            hasher.update(student.reg_no.encode())

        # Final digest
        fingerprint = hasher.hexdigest()

        # Create hidden sheet
        sheet = wb.create_sheet("__SYSTEM_HASH__")
        sheet.sheet_state = "hidden"
        sheet.append([fingerprint])

    
    # --------------------------------------------------
    # Main
    # --------------------------------------------------

    def generate(self, output_path: str, progress_callback=None):

        if os.path.exists(output_path):
            try:
                os.rename(output_path, output_path)
            except OSError:
                raise ValidationError(
                    f"File '{os.path.basename(output_path)}' is open in Excel."
                )

        temp_dir = os.path.dirname(output_path)
        fd, temp_path = tempfile.mkstemp(
            suffix=".xlsx",
            dir=temp_dir,
        )
        os.close(fd)

        try:
            wb = Workbook()
            if wb.active:
                wb.remove(wb.active)
            bold = Font(bold=True)
            center = Alignment(
                horizontal="center",
                vertical="center"
            )
            thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            grey = PatternFill(
                start_color="F2F2F2",
                end_color="F2F2F2",
                fill_type="solid"
            )
            # --------------------------------------------------
            # 1. Course Info Sheet
            # --------------------------------------------------

            ws_info = wb.create_sheet("Course_Info")
            self._apply_header_footer(ws_info)
            self._set_page_layout(ws_info)

            ws_info.append(["Field", "Value"])

            for key, value in vars(self.metadata).items():
                ws_info.append([key, value])
            for cell in ws_info[1][:2]:
                cell.font = bold
                cell.fill = grey
            for row in ws_info.iter_rows(
                    min_row=1,
                    max_row=ws_info.max_row,
                    min_col=1,
                    max_col=2
                ):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = center

            self._autofit_columns(ws_info)
            ws_info.protection.set_password(WORKBOOK_PASSWORD)
            ws_info.protection.sheet = True
            ws_info.protection.enable()

            # --------------------------------------------------
            # 2. Component Sheets
            # --------------------------------------------------

            students = self.setup.students
            student_rows = [
                (student.reg_no, student.name)
                for student in students
            ]

            components = self.setup.components

            total_components = len(components)

            for idx, (component, comp_info) in enumerate(components.items()):

                if progress_callback:
                    progress_callback(
                        f"Generating {component} "
                        f"({idx+1}/{total_components})"
                    )

                if not comp_info.direct:
                    continue

                ws = wb.create_sheet(component)
                self._apply_header_footer(ws)
                self._set_page_layout(ws)
                questions = comp_info.questions
                question_ids = [q.identifier for q in questions]

                headers = (
                    ["RegNo", "Student_Name"]
                    + question_ids
                    + ["Total"]
                )

                ws.append(headers)

                # CO Mapping Row
                ws.append(
                    ["CO", ""]
                    + [
                        ",".join(q.co_list)
                        for q in questions
                    ]
                    + [""]
                )

                # Max Marks Row
                ws.append(
                    ["Max", ""]
                    + [q.max_marks for q in questions]
                    + [""]
                )

                # Student Rows
                empty_marks = [""] * (len(question_ids) + 1)

                for regno, name in student_rows:
                    ws.append([regno, name] + empty_marks)

                # --------------------------------------------------
                # Styling (Optimized)
                # --------------------------------------------------

                ws.freeze_panes = "C4"

                max_row = ws.max_row
                max_col = ws.max_column
                end_col = 2 + len(question_ids)

                # Lock header rows only
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=3,
                    max_col=max_col
                ):
                    for cell in row:
                        cell.font = bold
                        cell.fill = grey
                        cell.alignment = center
                        cell.border = thin_border
                        cell.protection = Protection(locked=True)

                # Lock RegNo & Name columns
                for row in ws.iter_rows(
                    min_row=4,
                    max_row=max_row,
                    min_col=1,
                    max_col=2
                ):
                    for cell in row:
                        cell.fill = grey
                        cell.border = thin_border
                        cell.protection = Protection(locked=True)

                validators = {}
                for col_idx in range(3, end_col + 1):
                    q_index = col_idx - 3
                    max_mark = questions[q_index].max_marks
                    col_letter = get_column_letter(col_idx)
                    
                    dv = DataValidation(
                        type="custom",
                        formula1=f'=OR(AND(ISNUMBER({col_letter}4),{col_letter}4>=0,{col_letter}4<={max_mark}),UPPER({col_letter}4)="A")',
                        allow_blank=True
                    )
                    dv.error = f"Enter 0 to {max_mark} or A for absent."
                    dv.errorTitle = "Invalid Mark"
                    dv.prompt = f"Allowed: 0 to {max_mark} or A"
                    dv.promptTitle = "Mark Entry Rule"
                    dv.showErrorMessage = True
                    dv.showInputMessage = True
                    
                    ws.add_data_validation(dv)
                    validators[col_idx] = dv

                # Unlock mark entry area
                for row in ws.iter_rows(
                    min_row=4,
                    max_row=max_row,
                    min_col=3,
                    max_col=end_col
                ):
                    for cell in row:
                        cell.protection = Protection(locked=False)
                        cell.alignment = center
                        cell.border = thin_border
                        cell.number_format = DEFAULT_NUMBER_FORMAT
                        validators[cell.column].add(cell)

                # Totals
                total_col_letter = get_column_letter(max_col)
                first_q_letter = get_column_letter(3)
                last_q_letter = get_column_letter(end_col)

                for r in range(3, max_row + 1):
                    ws[
                        f"{total_col_letter}{r}"
                    ] = (
                        f"=SUM({first_q_letter}{r}:"
                        f"{last_q_letter}{r})"
                    )
                    ws[
                        f"{total_col_letter}{r}"
                    ].alignment = center
                    ws[
                        f"{total_col_letter}{r}"
                    ].border = thin_border
                    ws[
                        f"{total_col_letter}{r}" 
                    ].number_format = DEFAULT_NUMBER_FORMAT
                    ws[
                        f"{total_col_letter}{r}"
                    ].fill = grey

                ws.protection.set_password(WORKBOOK_PASSWORD)
                ws.protection.sheet = True
                ws.protection.enable()

                self._autofit_columns(ws)
            # --------------------------------------------------
            # 3. Indirect Tool Sheets
            # --------------------------------------------------
            total_cos = self.metadata.total_cos
            headers = ["RegNo", "Student_Name"] + [
                    f"CO{i}" for i in range(1, total_cos + 1)
                ]
            for tool in self.setup.indirect_tools:
                sheet_name = f"{tool.name}_INDIRECT"
                ws = wb.create_sheet(title=sheet_name[:31])
                self._apply_header_footer(ws)
                self._set_page_layout(ws)
                ws.append(headers)
                # Add students
                empty_likert = [""] * total_cos
                for regno, name in student_rows:
                    ws.append([regno, name] + empty_likert)
                
                ws.freeze_panes = "C2"
                max_row = ws.max_row
                max_col = ws.max_column
                end_col = 2 + total_cos

                for cell in ws[1][:max_col]:
                    cell.font = bold
                    cell.fill = grey
                    cell.alignment = center
                    cell.border = thin_border
                    cell.protection = Protection(locked=True)
                
                # Lock RegNo & Name columns
                for row in ws.iter_rows(
                    min_row=2,
                    max_row=max_row,
                    min_col=1,
                    max_col=2
                ):
                    for cell in row:
                        cell.fill = grey
                        cell.border = thin_border
                        cell.protection = Protection(locked=True)

                for col_idx in range(3, end_col + 1):
                    col_letter = get_column_letter(col_idx)

                    # We change '4' to '2' to match your starting row
                    dv = DataValidation(
                        type="custom",
                        #formula1=f'=OR(AND(ISNUMBER({col_letter}2),{col_letter}2>=1,{col_letter}2<=5),UPPER({col_letter}2)="A")',
                        formula1 = (
                            f'=OR('
                            f'AND(ISNUMBER({col_letter}2),'
                            f'{col_letter}2>={LIKERT_MIN},'
                            f'{col_letter}2<={LIKERT_MAX}),'
                            f'UPPER({col_letter}2)="{ABSENT_SYMBOL}")'
                        ),
                        allow_blank=True
                    )

                    dv.error = "Enter 1-5 or A for absent."
                    dv.errorTitle = "Invalid Survey Response"
                    dv.prompt = f"Allowed: 1 to 5 or A"
                    dv.promptTitle = "Survey Entry Rule"
                    dv.showErrorMessage = True
                    dv.showInputMessage = True
                    
                    ws.add_data_validation(dv)
                    
                    # Apply from row 2 down to the last student row
                    dv.add(f"{col_letter}2:{col_letter}{max_row}")

                # Unlock mark entry area
                for row in ws.iter_rows(
                    min_row=2,
                    max_row=max_row,
                    min_col=3,
                    max_col=end_col
                ):
                    for cell in row:
                        cell.protection = Protection(locked=False)
                        cell.alignment = center
                        cell.border = thin_border

                self._autofit_columns(ws)
                ws.protection.set_password(WORKBOOK_PASSWORD)
                ws.protection.sheet = True
                ws.protection.enable()
            # --------------------------------------------------
            # 4. System Hash Embedding
            # --------------------------------------------------
            self._embed_system_hash(wb, self.setup)
            
            wb.save(temp_path)

            if os.path.exists(output_path):
                os.remove(output_path)

            os.replace(temp_path, output_path)

        except Exception:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise
