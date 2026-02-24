import hashlib
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Border,
    Protection,
    Font,
    Alignment,
    PatternFill,
    Side,
)
from openpyxl.worksheet.datavalidation import DataValidation

from core.constants import (
    WORKBOOK_PASSWORD,
    LIKERT_MIN,
    LIKERT_MAX,
    ABSENT_SYMBOL,
    DEFAULT_NUMBER_FORMAT,
)
from core.models import ValidatedSetup, CourseMetadata


class MarksTemplateWorkbookRenderer:
    """
    Excel behavior layer.
    Applies styling, validation, formulas, protection.
    No file I/O.
    """

    def __init__(self, metadata: CourseMetadata, setup: ValidatedSetup):
        self.metadata = metadata
        self.setup = setup

        self.bold = Font(bold=True)
        self.center = Alignment(horizontal="center", vertical="center")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.grey = PatternFill(
            start_color="F2F2F2",
            end_color="F2F2F2",
            fill_type="solid",
        )

    # --------------------------------------------------

    def render(self, structure: dict) -> Workbook:
        wb = Workbook()

        default = wb.active
        if default is not None:
            wb.remove(default)

        self._render_course_info(wb, structure["course_info"])
        self._render_components(wb, structure["components"])
        self._render_indirect(wb, structure["indirect"])
        self._embed_system_hash(wb)

        return wb

    # --------------------------------------------------

    def _apply_layout(self, ws):
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        ws.oddHeader.left.text = (
            f"Course: {self.metadata.course_code} - "
            f"Section: {self.metadata.section}"
        )
        ws.oddHeader.right.text = (
            f"{self.metadata.semester} | "
            f"{self.metadata.academic_year}"
        )
        ws.oddFooter.center.text = "Page &P of &N"

    # --------------------------------------------------

    def _autofit(self, ws):
        for column_cells in ws.columns:
            max_len = 0
            col_index = column_cells[0].column
            for cell in column_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[
                get_column_letter(col_index)
            ].width = max_len + 2

    # --------------------------------------------------

    def _render_course_info(self, wb, data):
        ws = wb.create_sheet("Course_Info")
        self._apply_layout(ws)

        ws.append(data["headers"])

        for row in data["rows"]:
            ws.append(row)

        for cell in ws[1][:2]:
            cell.font = self.bold
            cell.fill = self.grey

        ws.protection.set_password(WORKBOOK_PASSWORD)
        ws.protection.sheet = True
        ws.protection.enable()

        self._autofit(ws)

    # --------------------------------------------------

    def _render_components(self, wb, components):
        for name, comp in components.items():

            ws = wb.create_sheet(name)
            self._apply_layout(ws)

            ws.append(comp["headers"])
            ws.append(comp["co_row"])
            ws.append(comp["max_row"])

            for row in comp["students"]:
                ws.append(row)

            ws.freeze_panes = "C4"

            self._style_component_sheet(ws, comp)

            ws.protection.set_password(WORKBOOK_PASSWORD)
            ws.protection.sheet = True
            ws.protection.enable()

            self._autofit(ws)

    # --------------------------------------------------

    def _style_component_sheet(self, ws, comp):
        questions = comp["questions"]
        max_row = ws.max_row
        max_col = ws.max_column
        end_col = 2 + len(questions)

        # Header rows
        for row in ws.iter_rows(min_row=1, max_row=3, max_col=max_col):
            for cell in row:
                cell.font = self.bold
                cell.fill = self.grey
                cell.alignment = self.center
                cell.border = self.thin_border
                cell.protection = Protection(locked=True)

        # Lock RegNo & Name
        for row in ws.iter_rows(
            min_row=4, max_row=max_row, min_col=1, max_col=2
        ):
            for cell in row:
                cell.fill = self.grey
                cell.border = self.thin_border
                cell.protection = Protection(locked=True)

        # Data validation
        validators = {}

        for col_idx in range(3, end_col + 1):
            q_index = col_idx - 3
            max_mark = questions[q_index].max_marks
            col_letter = get_column_letter(col_idx)

            dv = DataValidation(
                type="custom",
                formula1=(
                    f'=OR(AND(ISNUMBER({col_letter}4),'
                    f'{col_letter}4>=0,'
                    f'{col_letter}4<={max_mark}),'
                    f'UPPER({col_letter}4)="A")'
                ),
                allow_blank=True,
            )

            ws.add_data_validation(dv)
            validators[col_idx] = dv

        # Unlock marks
        for row in ws.iter_rows(
            min_row=4, max_row=max_row,
            min_col=3, max_col=end_col
        ):
            for cell in row:
                cell.protection = Protection(locked=False)
                cell.alignment = self.center
                cell.border = self.thin_border
                cell.number_format = DEFAULT_NUMBER_FORMAT
                validators[cell.column].add(cell)

        # Totals
        total_col = get_column_letter(max_col)
        first_q = get_column_letter(3)
        last_q = get_column_letter(end_col)

        for r in range(3, max_row + 1):
            ws[f"{total_col}{r}"] = (
                f"=SUM({first_q}{r}:{last_q}{r})"
            )
            ws[f"{total_col}{r}"].alignment = self.center
            ws[f"{total_col}{r}"].border = self.thin_border
            ws[f"{total_col}{r}"].number_format = DEFAULT_NUMBER_FORMAT
            ws[f"{total_col}{r}"].fill = self.grey

    # --------------------------------------------------

    def _render_indirect(self, wb, indirect):
        for tool_name, data in indirect.items():

            ws = wb.create_sheet(f"{tool_name}_INDIRECT"[:31])
            self._apply_layout(ws)

            ws.append(data["headers"])

            for row in data["students"]:
                ws.append(row)

            ws.freeze_panes = "C2"

            max_row = ws.max_row
            max_col = ws.max_column
            end_col = max_col

            for cell in ws[1][:max_col]:
                cell.font = self.bold
                cell.fill = self.grey
                cell.alignment = self.center
                cell.border = self.thin_border
                cell.protection = Protection(locked=True)

            for row in ws.iter_rows(
                min_row=2, max_row=max_row,
                min_col=1, max_col=2
            ):
                for cell in row:
                    cell.fill = self.grey
                    cell.border = self.thin_border
                    cell.protection = Protection(locked=True)

            for col_idx in range(3, end_col + 1):
                col_letter = get_column_letter(col_idx)

                dv = DataValidation(
                    type="custom",
                    formula1=(
                        f'=OR(AND(ISNUMBER({col_letter}2),'
                        f'{col_letter}2>={LIKERT_MIN},'
                        f'{col_letter}2<={LIKERT_MAX}),'
                        f'UPPER({col_letter}2)="{ABSENT_SYMBOL}")'
                    ),
                    allow_blank=True,
                )

                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}{max_row}")

                for row in ws.iter_rows(
                    min_row=2, max_row=max_row,
                    min_col=col_idx, max_col=col_idx
                ):
                    for cell in row:
                        cell.protection = Protection(locked=False)
                        cell.alignment = self.center
                        cell.border = self.thin_border

            ws.protection.set_password(WORKBOOK_PASSWORD)
            ws.protection.sheet = True
            ws.protection.enable()

            self._autofit(ws)

    # --------------------------------------------------

    def _embed_system_hash(self, wb):
        hasher = hashlib.sha256()

        for comp_name, comp in sorted(self.setup.components.items()):
            hasher.update(comp_name.encode())
            for q in comp.questions:
                hasher.update(q.identifier.encode())
                hasher.update(str(q.max_marks).encode())
                hasher.update(",".join(q.co_list).encode())

        for student in self.setup.students:
            hasher.update(student.reg_no.encode())

        fingerprint = hasher.hexdigest()

        sheet = wb.create_sheet("__SYSTEM_HASH__")
        sheet.sheet_state = "hidden"
        sheet.append([fingerprint])