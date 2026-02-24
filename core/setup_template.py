from openpyxl import Workbook
from openpyxl.styles import Font
from core.constants import MAX_CO_LIMIT
from openpyxl.worksheet.datavalidation import DataValidation


class SetupTemplateBuilder:
    """
    Builds the initial Course Setup Excel template.
    Pure builder.
    Does not save to disk.
    """

    def build(self) -> Workbook:
        wb = Workbook()

        # =====================================================
        # COURSE METADATA SHEET
        # =====================================================
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
        ws_meta = wb.create_sheet("Course_Metadata")
        ws_meta.title = "Course_Metadata"
        ws_meta.append(["Field", "Value"])
        ws_meta["A1"].font = Font(bold=True)
        ws_meta["B1"].font = Font(bold=True)

        rows = [
            ("Course_Code", "ECE101"),
            ("Course_Name", "SAMPLE COURSE"),
            ("Section", "A, B"),
            ("Semester", "III"),
            ("Academic_Year", "2026-27"),
            ("Faculty_Name", "ABCECE"),
            ("Total_Outcomes", 5),
        ]

        r = 2
        for k, v in rows:
            ws_meta.cell(r, 1, k)
            ws_meta.cell(r, 2, v)
            r += 1
        ws_meta.column_dimensions["A"].width = 22
        ws_meta.column_dimensions["B"].width = 30


        # =====================================================
        # ASSESSMENT CONFIG SHEET
        # =====================================================
        ws_config = wb.create_sheet("Assessment_Config")

        headers = [
            "Component",
            "Weight (%)",
            "CIA",
            "CO_Wise_Marks_Breakup",
            "Direct",
        ]

        ws_config.append(headers)

        for col in range(1, len(headers) + 1):
            ws_config.cell(row=1, column=col).font = Font(bold=True)

        # Example rows (editable)
        ws_config.append(["S1", "80", "yes", "yes", "yes"])
        ws_config.append(["S2", "20", "yes", "yes", "yes"])
        ws_config.append(["Survey", "100", "no", "no", "no"])
        dv_yesno = DataValidation(type="list", formula1='"yes,no"', allow_blank=False)
        ws_config.add_data_validation(dv_yesno)
        dv_yesno.add("C2:C200")
        dv_yesno.add("D2:D200")
        dv_yesno.add("E2:E200")


        # =====================================================
        # STUDENTS SHEET
        # =====================================================
        ws_students = wb.create_sheet("Students")

        ws_students.append(["RegNo", "Student_Name"])

        ws_students["A1"].font = Font(bold=True)
        ws_students["B1"].font = Font(bold=True)

        # =====================================================
        # QUESTION MAP SHEET
        # =====================================================
        ws_qmap = wb.create_sheet("Question_Map")

        q_headers = [
            "Component",
            "Q_No/Rubric_Parameter",
            "Max_Marks",
            "CO",
        ]

        ws_qmap.append(q_headers)

        for col in range(1, len(q_headers) + 1):
            ws_qmap.cell(row=1, column=col).font = Font(bold=True)

        ws_qmap.append(["S1", "Q1", 20, "1"])
        ws_qmap.append(["S1", "Q2", 20, "2"])
        ws_qmap.append(["S1", "Q3", 20, "3"])
        ws_qmap.append(["S1", "Q4", 20, "4"])
        ws_qmap.append(["S2", "Q1", 10, "4"])
        ws_qmap.append(["S2", "Q2", 10, "5"])
        ws_qmap.append(["Survey", "R1", 100, "1,2,3,4,5"])

        return wb