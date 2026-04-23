"""Shared Excel sheet layout/style helpers for openpyxl workflows."""

from __future__ import annotations

from copy import copy
from typing import Any, Callable, Sequence

from common.constants import (
    ALLOW_FILTER,
    ALLOW_SELECT_LOCKED,
    ALLOW_SELECT_UNLOCKED,
    ALLOW_SORT,
    ID_COURSE_SETUP,
)
from common.exceptions import ConfigurationError
from common.registry import BLUEPRINT_REGISTRY
from common.utils import normalize
from common.workbook_integrity.workbook_secret import ensure_workbook_secret_policy, get_workbook_password

_STYLE_REGISTRY_HEADER = "header"
_STYLE_REGISTRY_BODY = "body"
_BORDER_THIN = "thin"
_BORDER_COLOR_BLACK = "000000"
_COLUMN_MIN_WIDTH = 8
_COLUMN_MAX_WIDTH = 60
_COLUMN_WIDTH_PADDING = 2
_HEADER_ACTIVE_COLUMN = "A"
_AUTOSIZE_SAMPLE_ROWS = 300
XLSX_PAPER_SIZE_A4 = 9
XLSX_PAGE_MIN_MARGIN_IN = 0.25
XLSX_AUTOFIT_SAMPLE_ROWS = 30
XLSX_AUTOFIT_PADDING = 2
XLSX_AUTOFIT_MIN_WIDTH = _COLUMN_MIN_WIDTH
XLSX_AUTOFIT_MAX_WIDTH = _COLUMN_MAX_WIDTH
_STYLE_PROVIDER_TEMPLATE = tuple[dict[str, Any], dict[str, Any]]
_STYLE_PROVIDER_FN = Callable[[], _STYLE_PROVIDER_TEMPLATE]
_XLSXWRITER_PROTECT_OPTIONS = {
    "sort": ALLOW_SORT,
    "autofilter": ALLOW_FILTER,
    "select_locked_cells": ALLOW_SELECT_LOCKED,
    "select_unlocked_cells": ALLOW_SELECT_UNLOCKED,
}

def _course_setup_v2_style_provider() -> _STYLE_PROVIDER_TEMPLATE:
    """Fallback style provider for COURSE_SETUP_V2 when blueprint lookup is unavailable."""
    return (
        {
            "bold": True,
            "bg_color": "#D9EAD3",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        },
        {
            "locked": False,
            "border": 1,
        },
    )


_TEMPLATE_STYLE_PROVIDERS: dict[str, _STYLE_PROVIDER_FN] = {
    ID_COURSE_SETUP: _course_setup_v2_style_provider,
}


def _resolve_effective_template_id(template_id: str | None) -> str:
    """Resolve effective template id.
    
    Args:
        template_id: Parameter value (str | None).
    
    Returns:
        str: Return value.
    
    Raises:
        None.
    """
    effective = (template_id or ID_COURSE_SETUP).strip()
    if not effective:
        raise ConfigurationError("Template id is required for style resolution.")
    return effective


def _style_registry_from_blueprint(template_id: str) -> _STYLE_PROVIDER_TEMPLATE | None:
    """Style registry from blueprint.
    
    Args:
        template_id: Parameter value (str).
    
    Returns:
        _STYLE_PROVIDER_TEMPLATE | None: Return value.
    
    Raises:
        None.
    """
    blueprint = BLUEPRINT_REGISTRY.get(template_id)
    if blueprint is None:
        return None
    if str(getattr(blueprint, "type_id", "")).strip() != str(template_id).strip():
        raise ConfigurationError(
            f"Blueprint type_id '{getattr(blueprint, 'type_id', None)}' does not match template_id '{template_id}'."
        )
    return (
        dict(blueprint.style_registry.get(_STYLE_REGISTRY_HEADER, {})),
        dict(blueprint.style_registry.get(_STYLE_REGISTRY_BODY, {})),
    )


def style_registry_for_template(template_id: str | None = None) -> _STYLE_PROVIDER_TEMPLATE:
    """Style registry for template.
    
    Args:
        template_id: Parameter value (str | None).
    
    Returns:
        _STYLE_PROVIDER_TEMPLATE: Return value.
    
    Raises:
        None.
    """
    effective_template_id = _resolve_effective_template_id(template_id)
    from_blueprint = _style_registry_from_blueprint(effective_template_id)
    if from_blueprint is not None:
        return from_blueprint
    provider = _TEMPLATE_STYLE_PROVIDERS.get(effective_template_id)
    if callable(provider):
        header_style, body_style = provider()
        return dict(header_style), dict(body_style)
    raise ConfigurationError(f"No style provider found for template_id='{effective_template_id}'.")


def thin_border() -> Any:
    """Thin border.
    
    Args:
        None.
    
    Returns:
        Any: Return value.
    
    Raises:
        None.
    """
    from openpyxl.styles import Border, Side

    thin = Side(style=_BORDER_THIN, color=_BORDER_COLOR_BLACK)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def color_without_hash(color: str) -> str:
    """Color without hash.
    
    Args:
        color: Parameter value (str).
    
    Returns:
        str: Return value.
    
    Raises:
        None.
    """
    return color[1:] if color.startswith("#") else color


def excel_col_name(col_index_1_based: int) -> str:
    """Excel col name.
    
    Args:
        col_index_1_based: Parameter value (int).
    
    Returns:
        str: Return value.
    
    Raises:
        None.
    """
    index = col_index_1_based
    label = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        label = chr(65 + rem) + label
    return label


def autosize_columns(ws: Any, max_col: int) -> None:
    """Autosize columns.
    
    Args:
        ws: Parameter value (Any).
        max_col: Parameter value (int).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    sampled_max_row = min(int(ws.max_row), _AUTOSIZE_SAMPLE_ROWS)
    if sampled_max_row <= 0:
        for col in range(1, max_col + 1):
            col_label = excel_col_name(col)
            ws.column_dimensions[col_label].width = _COLUMN_MIN_WIDTH
        return
    for col in range(1, max_col + 1):
        col_label = excel_col_name(col)
        max_len = 0
        for values in ws.iter_cols(
            min_col=col,
            max_col=col,
            min_row=1,
            max_row=sampled_max_row,
            values_only=True,
        ):
            for value in values:
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            break
        ws.column_dimensions[col_label].width = min(
            _COLUMN_MAX_WIDTH,
            max(_COLUMN_MIN_WIDTH, max_len + _COLUMN_WIDTH_PADDING),
        )


def compute_sampled_column_widths(
    sample_rows: Sequence[Sequence[Any]],
    last_col: int,
    *,
    min_width: int = _COLUMN_MIN_WIDTH,
    max_width: int = _COLUMN_MAX_WIDTH,
    padding: int = _COLUMN_WIDTH_PADDING,
) -> dict[int, int]:
    """Compute sampled column widths.
    
    Args:
        sample_rows: Parameter value (Sequence[Sequence[Any]]).
        last_col: Parameter value (int).
        min_width: Parameter value (int).
        max_width: Parameter value (int).
        padding: Parameter value (int).
    
    Returns:
        dict[int, int]: Return value.
    
    Raises:
        None.
    """
    widths: dict[int, int] = {}
    for col_index in range(last_col + 1):
        max_len = 0
        for row in sample_rows:
            if col_index >= len(row):
                continue
            value = row[col_index]
            if value is None:
                continue
            max_len = max(max_len, len(str(value).strip()))
        widths[col_index] = min(
            max_width,
            max(min_width, max_len + padding),
        )
    return widths


def set_header_selected_cell(ws: Any, header_row: int) -> None:
    """Set header selected cell.
    
    Args:
        ws: Parameter value (Any).
        header_row: Parameter value (int).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    active = f"{_HEADER_ACTIVE_COLUMN}{header_row}"
    ws.sheet_view.selection[0].activeCell = active
    ws.sheet_view.selection[0].sqref = active


def apply_sheet_layout_and_protection(
    *,
    ws: Any,
    header_row: int,
    header_count: int,
    paper_size: Any,
    orientation: Any,
) -> None:
    """Apply sheet layout and protection.
    
    Args:
        ws: Parameter value (Any).
        header_row: Parameter value (int).
        header_count: Parameter value (int).
        paper_size: Parameter value (Any).
        orientation: Parameter value (Any).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)
    set_header_selected_cell(ws, header_row)
    autosize_columns(ws, header_count)
    ws.page_setup.paperSize = paper_size
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    protect_openpyxl_sheet(ws)


def build_xlsxwriter_header_format(workbook: Any, header_style: dict[str, Any]) -> Any:
    """Build xlsxwriter header format.
    
    Args:
        workbook: Parameter value (Any).
        header_style: Parameter value (dict[str, Any]).
    
    Returns:
        Any: Return value.
    
    Raises:
        None.
    """
    return workbook.add_format(
        {
            "bold": bool(header_style.get("bold", True)),
            "bg_color": str(header_style.get("bg_color", "")),
            "border": int(header_style.get("border", 1)),
            "align": str(header_style.get("align", "center")),
            "valign": str(header_style.get("valign", "vcenter")),
        }
    )


def build_xlsxwriter_body_format(workbook: Any, body_style: dict[str, Any]) -> Any:
    """Build xlsxwriter body format.
    
    Args:
        workbook: Parameter value (Any).
        body_style: Parameter value (dict[str, Any]).
    
    Returns:
        Any: Return value.
    
    Raises:
        None.
    """
    return workbook.add_format(
        {
            "locked": bool(body_style.get("locked", False)),
            "border": int(body_style.get("border", 1)),
        }
    )


def build_template_xlsxwriter_formats(
    workbook: Any,
    *,
    template_id: str | None = None,
    cache_attr: str | None = None,
    include_column_wrap: bool = False,
    normalize_header_valign_to_center: bool = False,
) -> dict[str, Any]:
    """Build template xlsxwriter formats.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str | None).
        cache_attr: Parameter value (str | None).
        include_column_wrap: Parameter value (bool).
        normalize_header_valign_to_center: Parameter value (bool).
    
    Returns:
        dict[str, Any]: Return value.
    
    Raises:
        None.
    """
    if cache_attr:
        cached = getattr(workbook, cache_attr, None)
        if isinstance(cached, dict):
            return cached

    header_style, body_style = style_registry_for_template(template_id)
    border_enabled = int(body_style.get("border", 1)) > 0
    header_border_enabled = int(header_style.get("border", 1)) > 0
    header_bg = color_without_hash(str(header_style.get("bg_color", "")))

    header_default_valign = "center" if normalize_header_valign_to_center else "vcenter"
    header_valign = str(header_style.get("valign", header_default_valign))
    if normalize_header_valign_to_center and header_valign == "vcenter":
        header_valign = "center"

    border_value = 1 if border_enabled else 0
    header_border_value = 1 if header_border_enabled else 0
    formats = {
        "header": workbook.add_format(
            {
                "bold": bool(header_style.get("bold", True)),
                "border": header_border_value,
                "align": str(header_style.get("align", "center")),
                "valign": header_valign,
                "text_wrap": True,
                "fg_color": header_bg,
                "pattern": 1,
            }
        ),
        "body": workbook.add_format(
            {
                "border": border_value,
                "valign": "vcenter",
            }
        ),
        "body_center": workbook.add_format(
            {
                "border": border_value,
                "align": "center",
                "valign": "vcenter",
            }
        ),
        "body_wrap": workbook.add_format(
            {
                "border": border_value,
                "align": "left",
                "valign": "vcenter",
                "text_wrap": True,
            }
        ),
    }
    if include_column_wrap:
        formats["column_wrap"] = workbook.add_format(
            {
                "align": "left",
                "valign": "vcenter",
                "text_wrap": True,
            }
        )

    if cache_attr:
        setattr(workbook, cache_attr, formats)
    return formats


def build_marks_template_xlsxwriter_formats(
    workbook: Any,
    *,
    template_id: str | None = None,
    cache_attr: str | None = None,
    include_column_wrap: bool = False,
    normalize_header_valign_to_center: bool = False,
) -> dict[str, Any]:
    """Build marks template xlsxwriter formats.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str | None).
        cache_attr: Parameter value (str | None).
        include_column_wrap: Parameter value (bool).
        normalize_header_valign_to_center: Parameter value (bool).
    
    Returns:
        dict[str, Any]: Return value.
    
    Raises:
        None.
    """
    if cache_attr:
        cached = getattr(workbook, cache_attr, None)
        if isinstance(cached, dict):
            return cached

    formats = dict(
        build_template_xlsxwriter_formats(
            workbook,
            template_id=template_id,
            include_column_wrap=include_column_wrap,
            normalize_header_valign_to_center=normalize_header_valign_to_center,
        )
    )
    header_style, body_style = style_registry_for_template(template_id)
    border_value = 1 if int(body_style.get("border", 1)) > 0 else 0
    header_border_value = 1 if int(header_style.get("border", 1)) > 0 else 0
    header_bg = color_without_hash(str(header_style.get("bg_color", "")))

    header_default_valign = "center" if normalize_header_valign_to_center else "vcenter"
    header_valign = str(header_style.get("valign", header_default_valign))
    if normalize_header_valign_to_center and header_valign == "vcenter":
        header_valign = "center"

    formats["num"] = workbook.add_format(
        {
            "border": border_value,
            "num_format": "0.00",
        }
    )
    formats["header_num"] = workbook.add_format(
        {
            "bold": bool(header_style.get("bold", True)),
            "border": header_border_value,
            "align": str(header_style.get("align", "center")),
            "valign": header_valign,
            "num_format": "0.00",
            "fg_color": header_bg,
            "pattern": 1,
        }
    )
    formats["unlocked_body"] = workbook.add_format(
        {
            "border": border_value,
            "locked": False,
        }
    )
    formats["column_wrap_unbordered"] = workbook.add_format(
        {
            "align": "left",
            "valign": "vcenter",
            "text_wrap": True,
        }
    )

    if cache_attr:
        setattr(workbook, cache_attr, formats)
    return formats


def apply_xlsxwriter_layout(
    ws: Any,
    *,
    header_row_index: int,
    paper_size: int,
    landscape: bool,
    selection_col: int = 0,
) -> None:
    """Apply xlsxwriter layout.
    
    Args:
        ws: Parameter value (Any).
        header_row_index: Parameter value (int).
        paper_size: Parameter value (int).
        landscape: Parameter value (bool).
        selection_col: Parameter value (int).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    apply_xlsxwriter_page_setup(
        ws,
        paper_size=paper_size,
        landscape=landscape,
        fit_width=1,
        fit_height=0,
    )
    protect_xlsxwriter_sheet(ws)
    ws.set_selection(header_row_index, selection_col, header_row_index, selection_col)


def apply_xlsxwriter_page_setup(
    ws: Any,
    *,
    paper_size: int,
    landscape: bool,
    fit_width: int = 1,
    fit_height: int = 0,
    margins: tuple[float, float, float, float] | None = None,
) -> None:
    """Apply xlsxwriter page setup.
    
    Args:
        ws: Parameter value (Any).
        paper_size: Parameter value (int).
        landscape: Parameter value (bool).
        fit_width: Parameter value (int).
        fit_height: Parameter value (int).
        margins: Parameter value (tuple[float, float, float, float] | None).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    if landscape:
        ws.set_landscape()
    else:
        ws.set_portrait()
    ws.set_paper(paper_size)
    ws.fit_to_pages(fit_width, fit_height)
    if margins is not None:
        left, right, top, bottom = margins
        ws.set_margins(left, right, top, bottom)


def apply_xlsxwriter_column_widths(
    ws: Any,
    widths: dict[int, int],
    *,
    default_width: int = _COLUMN_MIN_WIDTH,
    wrap_columns: Sequence[int] = (),
    wrap_format: Any | None = None,
) -> None:
    """Apply xlsxwriter column widths.
    
    Args:
        ws: Parameter value (Any).
        widths: Parameter value (dict[int, int]).
        default_width: Parameter value (int).
        wrap_columns: Parameter value (Sequence[int]).
        wrap_format: Parameter value (Any | None).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    max_col = max(widths.keys(), default=-1)
    wrap_set = set(wrap_columns)
    for col in range(0, max_col + 1):
        width = widths.get(col, default_width)
        if col in wrap_set and wrap_format is not None:
            ws.set_column(col, col, width, wrap_format)
        else:
            ws.set_column(col, col, width)


def apply_xlsxwriter_sheet_frame(
    ws: Any,
    *,
    repeat_last_row: int,
    freeze_row: int,
    freeze_col: int,
    select_row: int,
    select_col: int,
    paper_size: int,
    landscape: bool,
    margins: tuple[float, float, float, float] | None = None,
    protect: bool = True,
) -> None:
    """Apply xlsxwriter sheet frame.
    
    Args:
        ws: Parameter value (Any).
        repeat_last_row: Parameter value (int).
        freeze_row: Parameter value (int).
        freeze_col: Parameter value (int).
        select_row: Parameter value (int).
        select_col: Parameter value (int).
        paper_size: Parameter value (int).
        landscape: Parameter value (bool).
        margins: Parameter value (tuple[float, float, float, float] | None).
        protect: Parameter value (bool).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    apply_xlsxwriter_viewport(
        ws,
        repeat_last_row=repeat_last_row,
        freeze_row=freeze_row,
        freeze_col=freeze_col,
        select_row=select_row,
        select_col=select_col,
    )
    apply_xlsxwriter_page_setup(
        ws,
        paper_size=paper_size,
        landscape=landscape,
        fit_width=1,
        fit_height=0,
        margins=margins,
    )
    if protect:
        protect_xlsxwriter_sheet(ws)


def apply_xlsxwriter_viewport(
    ws: Any,
    *,
    repeat_last_row: int | None = None,
    freeze_row: int | None = None,
    freeze_col: int = 0,
    select_row: int | None = None,
    select_col: int = 0,
) -> None:
    """Apply xlsxwriter viewport.
    
    Args:
        ws: Parameter value (Any).
        repeat_last_row: Parameter value (int | None).
        freeze_row: Parameter value (int | None).
        freeze_col: Parameter value (int).
        select_row: Parameter value (int | None).
        select_col: Parameter value (int).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    if isinstance(repeat_last_row, int):
        ws.repeat_rows(0, repeat_last_row)
    if isinstance(freeze_row, int):
        ws.freeze_panes(freeze_row, freeze_col)
    if isinstance(select_row, int):
        ws.set_selection(select_row, select_col, select_row, select_col)


def set_two_column_metadata_widths(
    ws: Any,
    *,
    col1_title: str,
    col2_title: str,
    rows: Sequence[tuple[Any, Any]],
    col1_index: int,
    col2_index: int,
    default_width: int = 8,
) -> None:
    """Set two column metadata widths.
    
    Args:
        ws: Parameter value (Any).
        col1_title: Parameter value (str).
        col2_title: Parameter value (str).
        rows: Parameter value (Sequence[tuple[Any, Any]]).
        col1_index: Parameter value (int).
        col2_index: Parameter value (int).
        default_width: Parameter value (int).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    sample_rows: list[list[Any]] = [["", col1_title, col2_title]]
    sample_rows.extend(["", field, value] for field, value in rows)
    widths = compute_sampled_column_widths(sample_rows, 2)
    ws.set_column(col1_index, col1_index, widths.get(1, default_width))
    ws.set_column(col2_index, col2_index, widths.get(2, default_width))


def write_two_column_metadata_rows(
    ws: Any,
    *,
    rows: Sequence[tuple[Any, Any]],
    label_col_index: int,
    value_col_index: int,
    label_format: Any,
    value_format: Any,
    start_row_index: int = 0,
    centered_row_labels: Sequence[str] = (),
    centered_label_format: Any | None = None,
    centered_value_format: Any | None = None,
) -> int:
    """Write two-column metadata rows with optional centered-label rows.

    Args:
        ws: Parameter value (Any).
        rows: Parameter value (Sequence[tuple[Any, Any]]).
        label_col_index: Parameter value (int).
        value_col_index: Parameter value (int).
        label_format: Parameter value (Any).
        value_format: Parameter value (Any).
        start_row_index: Parameter value (int).
        centered_row_labels: Parameter value (Sequence[str]).
        centered_label_format: Parameter value (Any | None).
        centered_value_format: Parameter value (Any | None).

    Returns:
        int: Return value.

    Raises:
        None.
    """
    centered_set = {normalize(label) for label in centered_row_labels}
    center_label_fmt = centered_label_format if centered_label_format is not None else label_format
    center_value_fmt = centered_value_format if centered_value_format is not None else value_format
    row_index = int(start_row_index)
    for label, value in rows:
        if normalize(label) in centered_set:
            ws.write(row_index, label_col_index, label, center_label_fmt)
            ws.write(row_index, value_col_index, value, center_value_fmt)
        else:
            ws.write(row_index, label_col_index, label, label_format)
            ws.write(row_index, value_col_index, value, value_format)
        row_index += 1
    return row_index


def write_sheet_footer_xlsxwriter(
    ws: Any,
    *,
    footer_text: str,
    row_index: int,
    col_index: int = 0,
    cell_format: Any | None = None,
) -> None:
    """Write a footer text cell on an xlsxwriter worksheet.

    Args:
        ws: Parameter value (Any).
        footer_text: Parameter value (str).
        row_index: Parameter value (int).
        col_index: Parameter value (int).
        cell_format: Parameter value (Any | None).

    Returns:
        None.

    Raises:
        None.
    """
    if not str(footer_text).strip():
        return
    target_row = max(0, int(row_index))
    target_col = max(0, int(col_index))
    if cell_format is None:
        ws.write(target_row, target_col, footer_text)
    else:
        ws.write(target_row, target_col, footer_text, cell_format)


def protect_openpyxl_sheet(ws: Any) -> None:
    """Protect openpyxl sheet.
    
    Args:
        ws: Parameter value (Any).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    ensure_workbook_secret_policy()
    ws.protection.sheet = True
    ws.protection.password = get_workbook_password()
    ws.protection.sort = ALLOW_SORT
    ws.protection.autoFilter = ALLOW_FILTER
    ws.protection.selectLockedCells = ALLOW_SELECT_LOCKED
    ws.protection.selectUnlockedCells = ALLOW_SELECT_UNLOCKED


def protect_xlsxwriter_sheet(ws: Any) -> None:
    """Protect xlsxwriter sheet.
    
    Args:
        ws: Parameter value (Any).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    ensure_workbook_secret_policy()
    ws.protect(get_workbook_password(), dict(_XLSXWRITER_PROTECT_OPTIONS))


def copy_openpyxl_sheet(source_sheet: Any, target_sheet: Any) -> None:
    """Copy openpyxl sheet.
    
    Args:
        source_sheet: Parameter value (Any).
        target_sheet: Parameter value (Any).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    source = source_sheet
    target = target_sheet
    for row in source.iter_rows(min_row=1, max_row=source.max_row, min_col=1, max_col=source.max_column):
        for cell in row:
            target_cell = target.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            copy_openpyxl_cell_style(cell, target_cell, include_extended_style=True)
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
    for idx, dimension in source.row_dimensions.items():
        target.row_dimensions[idx].height = dimension.height
        target.row_dimensions[idx].hidden = dimension.hidden
    target.freeze_panes = source.freeze_panes
    target.page_setup.orientation = source.page_setup.orientation
    target.page_setup.paperSize = source.page_setup.paperSize
    target.page_setup.fitToWidth = source.page_setup.fitToWidth
    target.page_setup.fitToHeight = source.page_setup.fitToHeight
    target.print_title_rows = source.print_title_rows
    target.print_title_cols = source.print_title_cols
    target.auto_filter.ref = source.auto_filter.ref


def copy_openpyxl_cell_style(
    source_cell: Any,
    target_cell: Any,
    *,
    include_extended_style: bool = False,
) -> None:
    """Copy openpyxl cell style.
    
    Args:
        source_cell: Parameter value (Any).
        target_cell: Parameter value (Any).
        include_extended_style: Parameter value (bool).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if not include_extended_style:
        return
    if source_cell.protection is not None:
        target_cell.protection = copy(source_cell.protection)
    if source_cell.alignment is not None:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.fill is not None:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.font is not None:
        target_cell.font = copy(source_cell.font)
    if source_cell.border is not None:
        target_cell.border = copy(source_cell.border)


def find_header_row_by_value(
    sheet: Any,
    *,
    header_value: str,
    header_col: int = 1,
    max_scan_rows: int = 300,
) -> int:
    """Find header row by value.
    
    Args:
        sheet: Parameter value (Any).
        header_value: Parameter value (str).
        header_col: Parameter value (int).
        max_scan_rows: Parameter value (int).
    
    Returns:
        int: Return value.
    
    Raises:
        None.
    """
    upper = min(max_scan_rows, int(sheet.max_row))
    wanted = normalize(header_value)
    for row in range(1, upper + 1):
        if normalize(sheet.cell(row=row, column=header_col).value) == wanted:
            return row
    return 0

