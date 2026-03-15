"""Course coordinator module for collecting Final CO report Excel files."""

from __future__ import annotations

import logging
import json
import re
from datetime import datetime
from dataclasses import dataclass
from html import escape
from pathlib import Path
from typing import Any, Iterable

from PySide6.QtCore import Qt, QSize, QUrl, Signal
from PySide6.QtGui import (
    QColor,
    QDesktopServices,
    QDropEvent,
    QDragEnterEvent,
    QDragLeaveEvent,
    QDragMoveEvent,
    QFont,
    QKeySequence,
    QPalette,
    QPainter,
    QPen,
    QShortcut,
)
from PySide6.QtWidgets import (
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QPlainTextEdit,
    QPushButton,
    QSizePolicy,
    QStyle,
    QTabWidget,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)

from common.constants import (
    APP_NAME,
    CO_REPORT_ABSENT_TOKEN,
    CO_REPORT_HEADER_REG_NO,
    CO_REPORT_HEADER_SERIAL,
    CO_REPORT_HEADER_STUDENT_NAME,
    CO_REPORT_HEADER_TOTAL_RATIO_TEMPLATE,
    CO_REPORT_MAX_DECIMAL_PLACES,
    CO_REPORT_NOT_APPLICABLE_TOKEN,
    CO_REPORT_DIRECT_SHEET_SUFFIX,
    CO_REPORT_INDIRECT_SHEET_SUFFIX,
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SHEET,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    DIRECT_RATIO,
    INDIRECT_RATIO,
    INSTRUCTOR_CARD_MARGIN,
    INSTRUCTOR_CARD_SPACING,
    INSTRUCTOR_ACTIVE_TITLE_FONT_SIZE,
    INSTRUCTOR_INFO_TAB_FIXED_HEIGHT,
    INSTRUCTOR_INFO_TAB_LAYOUT_MARGINS,
    INSTRUCTOR_INFO_TAB_LAYOUT_SPACING,
    ID_COURSE_SETUP,
    SYSTEM_HASH_SHEET,
    SYSTEM_HASH_TEMPLATE_HASH_HEADER,
    SYSTEM_HASH_TEMPLATE_ID_HEADER,
    SYSTEM_REPORT_INTEGRITY_HASH_HEADER,
    SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER,
    SYSTEM_REPORT_INTEGRITY_SHEET,
    UI_FONT_FAMILY,
)
from common.excel_sheet_layout import (
    apply_sheet_layout_and_protection as _apply_sheet_layout_and_protection,
    color_without_hash as _color_without_hash,
    style_registry_for_setup as _style_registry_for_setup,
    thin_border as _thin_border,
)
from common.exceptions import JobCancelledError
from common.jobs import CancellationToken, generate_job_id
from common.qt_jobs import run_in_background
from common.texts import t
from common.toast import show_toast
from common.utils import (
    coerce_excel_number,
    emit_user_status,
    log_process_message,
    normalize,
    remember_dialog_dir,
    remember_dialog_dir_safe,
    resolve_dialog_start_path,
)
from common.ui_logging import (
    UILogHandler,
    build_i18n_log_message,
    format_log_line_at,
    parse_i18n_log_message,
    resolve_i18n_log_message,
)
from common.workbook_signing import verify_payload_signature

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}

_logger = logging.getLogger(__name__)
_COURSE_METADATA_COURSE_NAME_KEY = "course_name"
_STYLE_CACHE_ATTR = "_focus_coordinator_style_cache"
_STYLE_KEY_BORDER = "border"
_STYLE_KEY_BG_COLOR = "bg_color"
_STYLE_KEY_ALIGN = "align"
_STYLE_KEY_VALIGN = "valign"
_STYLE_KEY_BOLD = "bold"
_ALIGN_CENTER = "center"
_ALIGN_VCENTER = "vcenter"
_PATTERN_SOLID = "solid"
_CO_REPORT_NAME_TOKEN_RE = re.compile(r"(?:[_\-\s]*co[_\-\s]*report)+$", re.IGNORECASE)


@dataclass(slots=True)
class CoordinatorWorkflowState:
    busy: bool = False
    active_job_id: str | None = None

    def set_busy(self, value: bool, *, job_id: str | None = None) -> None:
        self.busy = value
        self.active_job_id = job_id if value else None


@dataclass(slots=True, frozen=True)
class _FinalReportSignature:
    template_id: str
    course_code: str
    total_outcomes: int
    section: str
    direct_sheet_count: int
    indirect_sheet_count: int


@dataclass(slots=True, frozen=True)
class _CoAttainmentRow:
    reg_no: str
    student_name: str
    direct_score: float | str
    indirect_score: float | str


def _path_key(path: Path) -> str:
    return str(path.resolve()).casefold()


def _build_co_attainment_default_name(source_path: Path, *, section: str = "") -> str:
    stem = source_path.stem.strip()
    cleaned = _CO_REPORT_NAME_TOKEN_RE.sub("", stem).rstrip("_- ").strip()
    section_token = section.strip()
    if section_token:
        parts = [part for part in cleaned.split("_") if normalize(part) != normalize(section_token)]
        cleaned = "_".join(parts).strip("_- ")
    base = cleaned if cleaned else stem
    return f"{base}_CO_Attainment.xlsx"


def _is_supported_excel_file(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES


def _filter_excel_paths(paths: Iterable[str]) -> list[Path]:
    collected: list[Path] = []
    seen: set[str] = set()
    for value in paths:
        path = Path(value)
        if not _is_supported_excel_file(path):
            continue
        key = _path_key(path)
        if key in seen:
            continue
        seen.add(key)
        collected.append(path.resolve())
    return collected


def _has_valid_final_co_report(path: Path) -> bool:
    return _extract_final_report_signature(path) is not None


def _extract_final_report_signature(path: Path) -> _FinalReportSignature | None:
    # openpyxl does not support legacy .xls files.
    if path.suffix.lower() == ".xls":
        return None
    try:
        from openpyxl import load_workbook
    except Exception:
        _logger.exception("openpyxl is unavailable while validating '%s'.", path)
        return None
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            if SYSTEM_HASH_SHEET not in workbook.sheetnames:
                return None
            if SYSTEM_REPORT_INTEGRITY_SHEET not in workbook.sheetnames:
                return None
            if COURSE_METADATA_SHEET not in workbook.sheetnames:
                return None

            hash_sheet = workbook[SYSTEM_HASH_SHEET]
            if normalize(hash_sheet["A1"].value) != normalize(SYSTEM_HASH_TEMPLATE_ID_HEADER):
                return None
            if normalize(hash_sheet["B1"].value) != normalize(SYSTEM_HASH_TEMPLATE_HASH_HEADER):
                return None

            template_id_raw = hash_sheet["A2"].value
            template_hash_raw = hash_sheet["B2"].value
            template_id = str(template_id_raw).strip() if template_id_raw is not None else ""
            template_hash = str(template_hash_raw).strip() if template_hash_raw is not None else ""
            if not template_id or not template_hash:
                return None
            if not verify_payload_signature(template_id, template_hash):
                return None
            if normalize(template_id) != normalize(ID_COURSE_SETUP):
                return None

            integrity_sheet = workbook[SYSTEM_REPORT_INTEGRITY_SHEET]
            if normalize(integrity_sheet["A1"].value) != normalize(SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER):
                return None
            if normalize(integrity_sheet["B1"].value) != normalize(SYSTEM_REPORT_INTEGRITY_HASH_HEADER):
                return None

            manifest_text_raw = integrity_sheet["A2"].value
            manifest_hash_raw = integrity_sheet["B2"].value
            manifest_text = str(manifest_text_raw).strip() if manifest_text_raw is not None else ""
            manifest_hash = str(manifest_hash_raw).strip() if manifest_hash_raw is not None else ""
            if not manifest_text or not manifest_hash:
                return None
            if not verify_payload_signature(manifest_text, manifest_hash):
                return None

            manifest = json.loads(manifest_text)
            if not isinstance(manifest, dict):
                return None
            sheet_order = manifest.get("sheet_order")
            if not isinstance(sheet_order, list):
                return None
            sheet_names = [str(name).strip() for name in sheet_order if isinstance(name, str)]
            if not sheet_names:
                return None

            direct_sheet_count = sum(
                1 for name in sheet_names if name.endswith(CO_REPORT_DIRECT_SHEET_SUFFIX)
            )
            indirect_sheet_count = sum(
                1 for name in sheet_names if name.endswith(CO_REPORT_INDIRECT_SHEET_SUFFIX)
            )
            if direct_sheet_count <= 0 or indirect_sheet_count <= 0:
                return None
            if direct_sheet_count != indirect_sheet_count:
                return None

            metadata_sheet = workbook[COURSE_METADATA_SHEET]
            metadata: dict[str, str] = {}
            row = 2
            while True:
                key = metadata_sheet.cell(row=row, column=1).value
                value = metadata_sheet.cell(row=row, column=2).value
                if normalize(key) == "" and normalize(value) == "":
                    break
                key_text = str(key).strip() if key is not None else ""
                value_text = str(value).strip() if value is not None else ""
                if key_text:
                    metadata[normalize(key_text)] = value_text
                row += 1

            course_code = metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "").strip()
            total_outcomes_text = metadata.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY), "").strip()
            section = metadata.get(normalize(COURSE_METADATA_SECTION_KEY), "").strip()
            if not course_code or not total_outcomes_text or not section:
                return None
            try:
                total_outcomes = int(float(total_outcomes_text))
            except (TypeError, ValueError):
                return None
            if total_outcomes <= 0:
                return None

            return _FinalReportSignature(
                template_id=template_id,
                course_code=course_code,
                total_outcomes=total_outcomes,
                section=section,
                direct_sheet_count=direct_sheet_count,
                indirect_sheet_count=indirect_sheet_count,
            )
        finally:
            workbook.close()
    except Exception:
        _logger.exception("Failed to validate final CO report workbook '%s'.", path)
        return None


def _analyze_dropped_files(
    dropped_files: list[str],
    *,
    existing_keys: set[str],
    existing_paths: list[str],
    token: CancellationToken,
) -> dict[str, object]:
    accepted = _filter_excel_paths(dropped_files)
    seen = set(existing_keys)
    added: list[str] = []
    duplicates = 0
    invalid_final_report: list[str] = []
    existing_resolved = [Path(path).resolve() for path in existing_paths if path]
    baseline_signature: _FinalReportSignature | None = None
    seen_sections: set[str] = set()

    for path in existing_resolved:
        token.raise_if_cancelled()
        signature = _extract_final_report_signature(path)
        if signature is None:
            continue
        if baseline_signature is None:
            baseline_signature = signature
        section_key = normalize(signature.section)
        if section_key:
            seen_sections.add(section_key)

    for path in accepted:
        token.raise_if_cancelled()
        key = _path_key(path)
        if key in seen:
            duplicates += 1
            continue
        signature = _extract_final_report_signature(path)
        if signature is None:
            invalid_final_report.append(str(path))
            continue
        if baseline_signature is None:
            baseline_signature = signature
        else:
            is_mismatch = (
                signature.template_id != baseline_signature.template_id
                or signature.course_code != baseline_signature.course_code
                or signature.total_outcomes != baseline_signature.total_outcomes
                or signature.direct_sheet_count != baseline_signature.direct_sheet_count
                or signature.indirect_sheet_count != baseline_signature.indirect_sheet_count
            )
            if is_mismatch:
                invalid_final_report.append(str(path))
                continue
            if normalize(signature.section) in seen_sections:
                invalid_final_report.append(str(path))
                continue
        seen_sections.add(normalize(signature.section))
        seen.add(key)
        added.append(str(path))

    ignored = (len(dropped_files) - len(accepted)) + duplicates + len(invalid_final_report)
    return {
        "added": added,
        "duplicates": duplicates,
        "invalid_final_report": invalid_final_report,
        "ignored": ignored,
    }


def _ratio_percent_token(ratio: float) -> str:
    percent = ratio * 100.0
    if abs(percent - round(percent)) <= 1e-9:
        return f"{int(round(percent))}"
    return f"{percent:g}"


def _ratio_total_header(ratio: float) -> str:
    return CO_REPORT_HEADER_TOTAL_RATIO_TEMPLATE.format(ratio=_ratio_percent_token(ratio))


def _coerce_numeric_score(value: Any) -> float | str | None:
    if normalize(value) == normalize(CO_REPORT_NOT_APPLICABLE_TOKEN):
        return CO_REPORT_ABSENT_TOKEN
    parsed = coerce_excel_number(value)
    if isinstance(parsed, bool):
        return None
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


def _metadata_rows_for_output(metadata: dict[str, str], co_index: int) -> list[tuple[str, str]]:
    return [
        ("Course Code", metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")),
        ("Course Name", metadata.get(normalize(_COURSE_METADATA_COURSE_NAME_KEY), "")),
        ("Semester", metadata.get(normalize(COURSE_METADATA_SEMESTER_KEY), "")),
        ("Academic Year", metadata.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")),
        ("CO Number", f"CO{co_index}"),
    ]


def _extract_course_metadata_fields(sheet: Any) -> dict[str, str]:
    metadata: dict[str, str] = {}
    row = 2
    while True:
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        key_text = str(key).strip() if key is not None else ""
        if key_text:
            coerced = coerce_excel_number(value)
            metadata[normalize(key_text)] = str(coerced).strip() if coerced is not None else ""
        row += 1
    return metadata


def _extract_co_scores_by_reg(sheet: Any, *, ratio: float) -> dict[str, tuple[str, str, float | str]]:
    required_headers = {
        normalize(CO_REPORT_HEADER_SERIAL),
        normalize(CO_REPORT_HEADER_REG_NO),
        normalize(CO_REPORT_HEADER_STUDENT_NAME),
        normalize(_ratio_total_header(ratio)),
    }
    header_row = 0
    column_map: dict[str, int] = {}
    for row_idx in range(1, int(sheet.max_row) + 1):
        row_map: dict[str, int] = {}
        for col_idx in range(1, int(sheet.max_column) + 1):
            key = normalize(sheet.cell(row=row_idx, column=col_idx).value)
            if key and key not in row_map:
                row_map[key] = col_idx
        if required_headers.issubset(row_map.keys()):
            header_row = row_idx
            column_map = row_map
            break
    if header_row <= 0:
        raise ValueError(f"Required headers are missing in sheet '{sheet.title}'.")

    reg_col = column_map[normalize(CO_REPORT_HEADER_REG_NO)]
    name_col = column_map[normalize(CO_REPORT_HEADER_STUDENT_NAME)]
    score_col = column_map[normalize(_ratio_total_header(ratio))]

    rows: dict[str, tuple[str, str, float | str]] = {}
    for row_idx in range(header_row + 1, int(sheet.max_row) + 1):
        reg_raw = sheet.cell(row=row_idx, column=reg_col).value
        reg_value = coerce_excel_number(reg_raw)
        reg_no = str(reg_value).strip() if reg_value is not None else ""
        if not reg_no:
            continue
        reg_key = normalize(reg_no)
        if reg_key in rows:
            continue

        score = _coerce_numeric_score(sheet.cell(row=row_idx, column=score_col).value)
        if score is None:
            continue
        student_raw = sheet.cell(row=row_idx, column=name_col).value
        student_name = str(student_raw).strip() if student_raw is not None else ""
        normalized_score = (
            round(score, CO_REPORT_MAX_DECIMAL_PLACES) if isinstance(score, (int, float)) else score
        )
        rows[reg_key] = (reg_no, student_name, normalized_score)
    return rows


def _collect_co_rows_from_workbook(workbook: Any, *, co_index: int) -> list[_CoAttainmentRow]:
    direct_name = f"CO{co_index}{CO_REPORT_DIRECT_SHEET_SUFFIX}"
    indirect_name = f"CO{co_index}{CO_REPORT_INDIRECT_SHEET_SUFFIX}"
    if direct_name not in workbook.sheetnames or indirect_name not in workbook.sheetnames:
        raise ValueError(f"Missing CO sheets for CO{co_index}.")

    direct_rows = _extract_co_scores_by_reg(workbook[direct_name], ratio=DIRECT_RATIO)
    indirect_rows = _extract_co_scores_by_reg(workbook[indirect_name], ratio=INDIRECT_RATIO)
    rows: list[_CoAttainmentRow] = []
    for reg_key, (reg_no, direct_name_value, direct_score) in direct_rows.items():
        indirect_entry = indirect_rows.get(reg_key)
        if indirect_entry is None:
            continue
        _, indirect_name_value, indirect_score = indirect_entry
        student_name = direct_name_value or indirect_name_value
        direct_is_absent = isinstance(direct_score, str) and normalize(direct_score) == normalize(CO_REPORT_ABSENT_TOKEN)
        indirect_is_absent = isinstance(indirect_score, str) and normalize(indirect_score) == normalize(CO_REPORT_ABSENT_TOKEN)
        if direct_is_absent or indirect_is_absent:
            direct_score = CO_REPORT_ABSENT_TOKEN
            indirect_score = CO_REPORT_ABSENT_TOKEN
        rows.append(
            _CoAttainmentRow(
                reg_no=reg_no,
                student_name=student_name,
                direct_score=direct_score,
                indirect_score=indirect_score,
            )
        )
    return rows


def _write_co_attainment_sheet(
    workbook: Any,
    *,
    co_index: int,
    metadata: dict[str, str],
    rows: list[_CoAttainmentRow],
) -> None:
    sheet = workbook.create_sheet(f"CO{co_index}")
    style_cache = _style_cache_for_sheet(sheet)
    metadata_rows = _metadata_rows_for_output(metadata, co_index)
    for row_idx, (label, value) in enumerate(metadata_rows, start=1):
        label_cell = sheet.cell(row=row_idx, column=2, value=label)
        value_cell = sheet.cell(row=row_idx, column=3, value=value)
        label_cell.border = style_cache["body_border"]
        value_cell.border = style_cache["body_border"]
        label_cell.alignment = style_cache["body_alignment"]
        value_cell.alignment = style_cache["body_alignment"]

    header_row = len(metadata_rows) + 2
    headers = [
        "#",
        "Regno",
        "Student name",
        f"Direct ({_ratio_percent_token(DIRECT_RATIO)}%)",
        f"Indirect ({_ratio_percent_token(INDIRECT_RATIO)}%)",
        "Total (100%)",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=header)
        cell.border = style_cache["header_border"]
        cell.fill = style_cache["header_fill"]
        cell.font = style_cache["header_font"]
        cell.alignment = style_cache["header_alignment"]

    for serial, row in enumerate(rows, start=1):
        row_idx = header_row + serial
        if isinstance(row.direct_score, (int, float)) and isinstance(row.indirect_score, (int, float)):
            total: float | str = round(row.direct_score + row.indirect_score, CO_REPORT_MAX_DECIMAL_PLACES)
        else:
            total = CO_REPORT_ABSENT_TOKEN
        values = [serial, row.reg_no, row.student_name, row.direct_score, row.indirect_score, total]
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = style_cache["body_border"]
            cell.alignment = (
                style_cache["body_alignment_center"] if col_idx >= 4 else style_cache["body_alignment"]
            )

    _apply_sheet_layout_and_protection(
        ws=sheet,
        header_row=header_row,
        header_count=len(headers),
        paper_size=sheet.PAPERSIZE_A4,
        orientation=sheet.ORIENTATION_LANDSCAPE,
    )


def _style_cache_for_sheet(ws: Any) -> dict[str, Any]:
    from openpyxl.styles import Alignment, Border, Font, PatternFill

    cached = getattr(ws, _STYLE_CACHE_ATTR, None)
    if isinstance(cached, dict):
        return cached
    header_style, body_style = _style_registry_for_setup()
    border_enabled = int(body_style.get(_STYLE_KEY_BORDER, 1)) > 0
    body_border = _thin_border() if border_enabled else Border()
    header_border_enabled = int(header_style.get(_STYLE_KEY_BORDER, 1)) > 0
    header_border = _thin_border() if header_border_enabled else Border()
    header_bg = _color_without_hash(str(header_style.get(_STYLE_KEY_BG_COLOR, "")))
    header_fill = (
        PatternFill(fill_type=_PATTERN_SOLID, fgColor=header_bg)
        if header_bg
        else PatternFill(fill_type=None)
    )
    header_alignment = Alignment(
        horizontal=str(header_style.get(_STYLE_KEY_ALIGN, _ALIGN_CENTER)),
        vertical=str(header_style.get(_STYLE_KEY_VALIGN, _ALIGN_CENTER)).replace(_ALIGN_VCENTER, _ALIGN_CENTER),
        wrap_text=True,
    )
    body_align = str(body_style.get(_STYLE_KEY_ALIGN, "")).strip()
    body_valign = str(body_style.get(_STYLE_KEY_VALIGN, _ALIGN_CENTER)).replace(_ALIGN_VCENTER, _ALIGN_CENTER)
    body_alignment = Alignment(
        horizontal=body_align if body_align else None,
        vertical=body_valign,
    )
    body_alignment_center = Alignment(horizontal=_ALIGN_CENTER, vertical=body_valign)
    style_cache = {
        "header_border": header_border,
        "body_border": body_border,
        "header_fill": header_fill,
        "header_font": Font(bold=bool(header_style.get(_STYLE_KEY_BOLD, True))),
        "header_alignment": header_alignment,
        "body_alignment": body_alignment,
        "body_alignment_center": body_alignment_center,
    }
    setattr(ws, _STYLE_CACHE_ATTR, style_cache)
    return style_cache


def _generate_co_attainment_workbook(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
) -> Path:
    if not source_paths:
        raise ValueError("No source files provided for CO attainment calculation.")

    first_signature = _extract_final_report_signature(source_paths[0])
    if first_signature is None:
        raise ValueError(f"Invalid final CO report file: {source_paths[0]}")
    if normalize(first_signature.template_id) != normalize(ID_COURSE_SETUP):
        raise ValueError(
            f"Unsupported template id '{first_signature.template_id}'. Expected '{ID_COURSE_SETUP}'."
        )
    return _generate_co_attainment_workbook_course_setup_v1(
        source_paths,
        output_path,
        token=token,
        total_outcomes=first_signature.total_outcomes,
    )


def _generate_co_attainment_workbook_course_setup_v1(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    total_outcomes: int,
) -> Path:
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:  # pragma: no cover - guarded by runtime dependency availability
        raise RuntimeError("openpyxl is required for CO attainment calculation.") from exc

    if total_outcomes <= 0:
        raise ValueError("No CO outcomes available in the uploaded reports.")

    metadata: dict[str, str] = {}
    per_co: dict[int, dict[str, _CoAttainmentRow]] = {co: {} for co in range(1, total_outcomes + 1)}

    for source in source_paths:
        token.raise_if_cancelled()
        workbook = load_workbook(filename=source, data_only=True, read_only=True)
        try:
            if not metadata and COURSE_METADATA_SHEET in workbook.sheetnames:
                metadata = _extract_course_metadata_fields(workbook[COURSE_METADATA_SHEET])
            for co_index in range(1, total_outcomes + 1):
                rows = _collect_co_rows_from_workbook(workbook, co_index=co_index)
                sink = per_co[co_index]
                for row in rows:
                    reg_key = normalize(row.reg_no)
                    if not reg_key or reg_key in sink:
                        continue
                    sink[reg_key] = row
        finally:
            workbook.close()

    output_workbook = Workbook()
    try:
        if output_workbook.active is not None:
            output_workbook.remove(output_workbook.active)
        for co_index in range(1, total_outcomes + 1):
            ordered_rows = list(per_co[co_index].values())
            _write_co_attainment_sheet(
                output_workbook,
                co_index=co_index,
                metadata=metadata,
                rows=ordered_rows,
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_workbook.save(output_path)
    finally:
        output_workbook.close()
    return output_path


class _ExcelDropList(QListWidget):
    files_dropped = Signal(list)
    drag_state_changed = Signal(bool)
    browse_requested = Signal()

    def __init__(self) -> None:
        super().__init__()
        self._placeholder_text = ""
        self.setAcceptDrops(True)
        self.setDragEnabled(False)
        self.setDropIndicatorShown(False)
        self.setSpacing(2)
        self.setAlternatingRowColors(False)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setFrameShape(QFrame.Shape.NoFrame)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

    def set_placeholder_text(self, text: str) -> None:
        self._placeholder_text = text
        self.viewport().update()

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        if self.count() != 0 or not self._placeholder_text:
            return
        painter = QPainter(self.viewport())
        painter.setPen(QColor("gray"))
        painter.setFont(QFont(UI_FONT_FAMILY, 10))
        painter.drawText(
            self.viewport().rect().adjusted(16, 16, -16, -16),
            Qt.AlignmentFlag.AlignCenter | Qt.TextFlag.TextWordWrap,
            self._placeholder_text,
        )
        painter.end()

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasUrls():
            self.drag_state_changed.emit(True)
            event.acceptProposedAction()
            return
        event.ignore()

    def dragMoveEvent(self, event: QDragMoveEvent) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            return
        event.ignore()

    def dragLeaveEvent(self, event: QDragLeaveEvent) -> None:
        self.drag_state_changed.emit(False)
        super().dragLeaveEvent(event)

    def dropEvent(self, event: QDropEvent) -> None:
        urls = event.mimeData().urls()
        dropped = [url.toLocalFile() for url in urls if url.isLocalFile()]
        self.drag_state_changed.emit(False)
        if dropped:
            self.files_dropped.emit(dropped)
            event.acceptProposedAction()
            return
        event.ignore()

    def mouseDoubleClickEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self.browse_requested.emit()
            event.accept()
            return
        super().mouseDoubleClickEvent(event)


class _DropZoneFrame(QFrame):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("coordinatorDropZone")

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        palette = self.palette()
        active = bool(self.property("dragActive"))
        bg_color = palette.color(QPalette.ColorRole.AlternateBase)
        if active:
            bg_color.setAlpha(220)
        border_color = palette.color(QPalette.ColorRole.Highlight) if active else palette.color(QPalette.ColorRole.Mid)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(bg_color)
        painter.drawRoundedRect(self.rect().adjusted(1, 1, -2, -2), 12, 12)
        pen = QPen(border_color, 2, Qt.PenStyle.DashLine)
        pen.setDashPattern([4, 3])
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(self.rect().adjusted(6, 6, -6, -6), 10, 10)
        painter.end()


class _ElidedFileNameLabel(QLabel):
    def __init__(self, text: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._full_text = text
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.setMinimumWidth(0)
        self._apply_elided_text()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._apply_elided_text()

    def _apply_elided_text(self) -> None:
        width = self.contentsRect().width()
        if width <= 0:
            return
        super().setText(self.fontMetrics().elidedText(self._full_text, Qt.TextElideMode.ElideMiddle, width))


class _CoordinatorFileItemWidget(QWidget):
    removed = Signal(str)

    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.file_path = file_path

        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 4, 12, 4)
        layout.setSpacing(12)

        file_name = Path(file_path).name
        name_label = _ElidedFileNameLabel(file_name)
        name_label.setFont(QFont(UI_FONT_FAMILY, 10))
        name_label.setToolTip(file_path)
        layout.addWidget(name_label, 1)

        self.remove_btn = QPushButton()
        self.remove_btn.setObjectName("coordinatorFileRemoveButton")
        self.remove_btn.setFixedSize(24, 24)
        self.remove_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        icon = self.style().standardIcon(QStyle.StandardPixmap.SP_TrashIcon)
        if not icon.isNull():
            self.remove_btn.setIcon(icon)
            self.remove_btn.setIconSize(QSize(16, 16))
        else:
            self.remove_btn.setText(t("coordinator.file.remove_fallback"))
        self.remove_btn.setStyleSheet(
            """
            QPushButton {
                background-color: transparent;
                border: none;
                padding: 0px;
                margin: 0px;
                min-width: 24px;
                min-height: 24px;
                max-width: 24px;
                max-height: 24px;
            }
            QPushButton:hover {
                background-color: rgba(231, 76, 60, 0.15);
                border-radius: 4px;
            }
            """
        )
        self.remove_btn.clicked.connect(lambda: self.removed.emit(self.file_path))
        layout.addWidget(self.remove_btn, 0)


class CoordinatorModule(QWidget):
    status_changed = Signal(str)
    OUTPUT_LINK_OPEN_FILE_KEY = "instructor.links.open_file"
    OUTPUT_LINK_OPEN_FOLDER_KEY = "instructor.links.open_folder"
    OUTPUT_LINK_NOT_AVAILABLE_KEY = "instructor.links.not_available"
    OUTPUT_LINK_OPEN_FAILED_KEY = "instructor.links.open_failed"

    def __init__(self) -> None:
        super().__init__()
        self._files: list[Path] = []
        self._downloaded_outputs: list[Path] = []
        self._logger = _logger
        self.state = CoordinatorWorkflowState()
        self._cancel_token: CancellationToken | None = None
        self._active_jobs: list[object] = []
        self._pending_drop_batches: list[list[str]] = []
        self._ui_log_handler: UILogHandler | None = None
        self._user_log_entries: list[dict[str, object]] = []
        self._build_ui()
        self._setup_ui_logging()
        self.retranslate_ui()
        self._refresh_ui()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(
            INSTRUCTOR_CARD_MARGIN,
            INSTRUCTOR_CARD_MARGIN,
            INSTRUCTOR_CARD_MARGIN,
            INSTRUCTOR_CARD_MARGIN,
        )
        root.setSpacing(max(10, INSTRUCTOR_CARD_SPACING))

        header_card = QFrame()
        header_card.setObjectName("coordinatorHeaderCard")
        header_layout = QVBoxLayout(header_card)
        header_layout.setContentsMargins(16, 14, 16, 14)
        header_layout.setSpacing(6)
        self.title_label = QLabel()
        self.title_label.setObjectName("coordinatorTitle")
        self.title_label.setFont(QFont(UI_FONT_FAMILY, INSTRUCTOR_ACTIVE_TITLE_FONT_SIZE, QFont.Weight.Bold))
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        header_layout.addWidget(self.title_label)
        self.hint_label = QLabel()
        self.hint_label.setObjectName("coordinatorHint")
        self.hint_label.setFont(QFont(UI_FONT_FAMILY, 10))
        self.hint_label.setWordWrap(True)
        self.hint_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        header_layout.addWidget(self.hint_label)
        root.addWidget(header_card)

        self.drop_zone = _DropZoneFrame()
        self.drop_zone.setProperty("dragActive", False)
        zone_layout = QVBoxLayout(self.drop_zone)
        zone_layout.setContentsMargins(14, 14, 14, 14)
        zone_layout.setSpacing(0)
        self.drop_list = _ExcelDropList()
        self.drop_list.setObjectName("coordinatorDropList")
        self.drop_list.setMinimumHeight(220)
        self.drop_list.files_dropped.connect(self._on_files_dropped)
        self.drop_list.drag_state_changed.connect(self._set_drop_active)
        self.drop_list.browse_requested.connect(self._browse_files)
        zone_layout.addWidget(self.drop_list)
        root.addWidget(self.drop_zone, 1)

        controls_row = QHBoxLayout()
        controls_row.setContentsMargins(6, 0, 6, 0)
        controls_row.setSpacing(10)
        self.summary_label = QLabel()
        self.summary_label.setObjectName("coordinatorSummary")
        self.summary_label.setFont(QFont(UI_FONT_FAMILY, 9))
        controls_row.addWidget(self.summary_label)
        controls_row.addStretch(1)
        self.clear_button = QPushButton()
        self.clear_button.setObjectName("coordinatorClearButton")
        self.clear_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.clear_button.setMinimumWidth(150)
        self.clear_button.setDefault(True)
        self.clear_button.clicked.connect(self._clear_all)
        controls_row.addWidget(self.clear_button)
        self.calculate_button = QPushButton()
        self.calculate_button.setObjectName("coordinatorCalculateButton")
        self.calculate_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.calculate_button.setMinimumWidth(150)
        self.calculate_button.setAutoDefault(False)
        self.calculate_button.setDefault(False)
        self.calculate_button.clicked.connect(self._on_calculate_clicked)
        controls_row.addWidget(self.calculate_button)
        root.addLayout(controls_row)

        self.info_tabs = QTabWidget()
        self.info_tabs.setObjectName("instructorInfoTabs")
        self.info_tabs.setFixedHeight(INSTRUCTOR_INFO_TAB_FIXED_HEIGHT)
        self.info_tabs.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.info_tabs.tabBar().setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.info_tabs.currentChanged.connect(self._on_info_tab_changed)

        log_tab = QWidget()
        log_tab_layout = QVBoxLayout(log_tab)
        log_tab_layout.setContentsMargins(*INSTRUCTOR_INFO_TAB_LAYOUT_MARGINS)
        log_tab_layout.setSpacing(INSTRUCTOR_INFO_TAB_LAYOUT_SPACING)

        self.user_log_view = QPlainTextEdit()
        self.user_log_view.setReadOnly(True)
        self.user_log_view.setObjectName("userLogView")
        self.user_log_view.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth)
        self.user_log_view.setFrameShape(QFrame.Shape.NoFrame)
        self.user_log_view.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        log_tab_layout.addWidget(self.user_log_view)

        links_tab = QWidget()
        links_tab_layout = QVBoxLayout(links_tab)
        links_tab_layout.setContentsMargins(*INSTRUCTOR_INFO_TAB_LAYOUT_MARGINS)
        links_tab_layout.setSpacing(INSTRUCTOR_INFO_TAB_LAYOUT_SPACING)

        self.generated_outputs_view = QTextBrowser()
        self.generated_outputs_view.setObjectName("generatedOutputsView")
        self.generated_outputs_view.setOpenExternalLinks(False)
        self.generated_outputs_view.setOpenLinks(False)
        self.generated_outputs_view.setFrameShape(QFrame.Shape.NoFrame)
        self.generated_outputs_view.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.generated_outputs_view.anchorClicked.connect(
            lambda url: self._on_output_link_activated(url.toString())
        )
        links_tab_layout.addWidget(self.generated_outputs_view)

        self.info_tabs.addTab(log_tab, t("instructor.log.title"))
        self.info_tabs.addTab(links_tab, t("instructor.links.title"))
        root.addWidget(self.info_tabs)

        self.shortcut_add_file = QShortcut(QKeySequence("Ctrl+O"), self)
        self.shortcut_add_file.activated.connect(self._browse_files)
        self.shortcut_save_output = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_save_output.activated.connect(self._on_save_shortcut_activated)

        panel_style = """
        QFrame#coordinatorHeaderCard { border: 1px solid palette(mid); border-radius: 12px; background-color: palette(base); }
        QLabel#coordinatorTitle { letter-spacing: 0.3px; }
        QLabel#coordinatorSummary { padding: 5px 10px; border: 1px solid palette(mid); border-radius: 10px; background-color: palette(alternate-base); }
        QFrame#coordinatorDropZone { border: none; background: transparent; }
        QListWidget#coordinatorDropList { border: none; padding: 10px; background: transparent; }
        QListWidget#coordinatorDropList::item { margin: 2px 0; }
        QPushButton#coordinatorClearButton, QPushButton#coordinatorCalculateButton { padding: 6px 12px; min-width: 150px; min-height: 30px; border-radius: 6px; border: none; }
        QPushButton#coordinatorClearButton:disabled, QPushButton#coordinatorCalculateButton:disabled { border: 1px solid palette(mid); }
        QPushButton#coordinatorCalculateButton:enabled { background-color: palette(highlight); color: palette(highlighted-text); border: none; font-weight: 600; }
        QPushButton#coordinatorCalculateButton:disabled { border: 1px solid palette(mid); }
        QTabWidget#instructorInfoTabs::pane { border: none; background: palette(base); }
        QTabWidget#instructorInfoTabs QTabBar::tab:first { margin-left: 8px; }
        QTabWidget#instructorInfoTabs QPlainTextEdit, QTabWidget#instructorInfoTabs QTextBrowser { border: 1px solid palette(mid); border-radius: 8px; background: palette(base); padding: 8px; }
        """

        self.setStyleSheet(panel_style)

    def retranslate_ui(self) -> None:
        self._rerender_user_log()
        self.title_label.setText(t("coordinator.title"))
        self.hint_label.setText(t("coordinator.drop_hint"))
        self.drop_list.set_placeholder_text(t("coordinator.list_placeholder"))
        self.clear_button.setText(t("coordinator.clear_all"))
        self.calculate_button.setText(t("coordinator.calculate"))
        self.info_tabs.setTabText(0, t("instructor.log.title"))
        self.info_tabs.setTabText(1, t("instructor.links.title"))
        self._refresh_output_links()
        self._refresh_summary()

    def _publish_status(self, message: str) -> None:
        self._append_user_log(message)
        emit_user_status(self.status_changed, message, logger=self._logger)

    def _publish_status_key(self, text_key: str, **kwargs: Any) -> None:
        localized = t(text_key, **kwargs)
        payload = build_i18n_log_message(text_key, kwargs=kwargs, fallback=localized)
        self._append_user_log(payload)
        emit_user_status(self.status_changed, payload, logger=self._logger)

    def _set_busy(self, busy: bool, *, job_id: str | None = None) -> None:
        self.state.set_busy(busy, job_id=job_id)
        self._refresh_ui()

    def _refresh_ui(self) -> None:
        has_files = bool(self._files)
        self.clear_button.setEnabled(has_files and not self.state.busy)
        self.calculate_button.setEnabled(has_files and not self.state.busy)
        self.drop_list.setEnabled(not self.state.busy)
        for row in range(self.drop_list.count()):
            item = self.drop_list.item(row)
            widget = self.drop_list.itemWidget(item)
            if isinstance(widget, _CoordinatorFileItemWidget):
                widget.remove_btn.setEnabled(not self.state.busy)
        self._refresh_output_links()
        self._refresh_summary()

    def _on_calculate_clicked(self) -> None:
        if self.state.busy or not self._files:
            return

        signature = _extract_final_report_signature(self._files[0])
        default_name = _build_co_attainment_default_name(
            self._files[0],
            section=signature.section if signature is not None else "",
        )
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            t("coordinator.calculate"),
            resolve_dialog_start_path(APP_NAME, default_name),
            t("coordinator.dialog.filter"),
        )
        if not save_path:
            return

        process_name = "calculating coordinator co attainment"
        token = CancellationToken()
        job_id = generate_job_id()
        self._cancel_token = token
        self._set_busy(True, job_id=job_id)
        self._publish_status_key("coordinator.status.processing_started")

        def _finalize(job: object) -> None:
            if job in self._active_jobs:
                self._active_jobs.remove(job)
            self._cancel_token = None
            self._set_busy(False)
            self._drain_next_batch()

        def _on_finished(result: object) -> None:
            try:
                output_path = Path(str(result)) if result else Path(save_path)
                if all(_path_key(path) != _path_key(output_path) for path in self._downloaded_outputs):
                    self._downloaded_outputs.append(output_path)
                self._remember_dialog_dir_safe(str(output_path))
                self._publish_status_key("coordinator.status.calculate_completed")
                log_process_message(
                    process_name,
                    logger=self._logger,
                    success_message=f"{process_name} completed successfully. output={output_path}",
                    user_success_message=build_i18n_log_message(
                        "coordinator.status.calculate_completed",
                        fallback=t("coordinator.status.calculate_completed"),
                    ),
                    job_id=job_id,
                    step_id="coordinator_calculate_attainment",
                )
                show_toast(
                    self,
                    t("coordinator.status.calculate_completed"),
                    title=t("coordinator.title"),
                    level="info",
                )
            finally:
                _finalize(job)

        def _on_failed(exc: Exception) -> None:
            try:
                if isinstance(exc, JobCancelledError):
                    self._publish_status_key("coordinator.status.operation_cancelled")
                    self._logger.info(
                        "%s cancelled by user/system request.",
                        process_name,
                        extra={
                            "user_message": build_i18n_log_message(
                                "coordinator.status.operation_cancelled",
                                fallback=t("coordinator.status.operation_cancelled"),
                            ),
                            "job_id": job_id,
                            "step_id": "coordinator_calculate_attainment",
                        },
                    )
                    return
                log_process_message(
                    process_name,
                    logger=self._logger,
                    error=exc,
                    user_error_message=build_i18n_log_message(
                        "coordinator.status.processing_failed",
                        fallback=t("coordinator.status.processing_failed"),
                    ),
                    job_id=job_id,
                    step_id="coordinator_calculate_attainment",
                )
                show_toast(
                    self,
                    t("coordinator.status.processing_failed"),
                    title=t("coordinator.title"),
                    level="error",
                )
            finally:
                _finalize(job)

        job = run_in_background(
            _generate_co_attainment_workbook,
            list(self._files),
            Path(save_path),
            token=token,
            on_finished=_on_finished,
            on_failed=_on_failed,
        )
        self._active_jobs.append(job)

    def _on_save_shortcut_activated(self) -> None:
        if self.state.busy:
            return
        if self.calculate_button.isEnabled():
            self._on_calculate_clicked()

    def _drain_next_batch(self) -> None:
        if self.state.busy or not self._pending_drop_batches:
            return
        next_batch = self._pending_drop_batches.pop(0)
        self._process_files_async(next_batch)

    def _process_files_async(self, dropped_files: list[str]) -> None:
        if not dropped_files:
            return
        if self.state.busy:
            self._pending_drop_batches.append(dropped_files)
            self._publish_status_key("coordinator.status.queued", count=len(dropped_files))
            return

        process_name = "collecting coordinator files"
        token = CancellationToken()
        job_id = generate_job_id()
        existing_keys = {_path_key(path) for path in self._files}
        existing_paths = [str(path) for path in self._files]
        self._cancel_token = token
        self._set_busy(True, job_id=job_id)
        self._publish_status_key("coordinator.status.processing_started")

        def _finalize(job: object) -> None:
            if job in self._active_jobs:
                self._active_jobs.remove(job)
            self._cancel_token = None
            self._set_busy(False)
            self._drain_next_batch()

        def _on_finished(result: object) -> None:
            try:
                if not isinstance(result, dict):
                    raise RuntimeError("Coordinator processing returned unexpected result type.")
                added_paths = [Path(value) for value in result.get("added", [])]
                duplicates = int(result.get("duplicates", 0))
                invalid_paths = [Path(value) for value in result.get("invalid_final_report", [])]
                ignored = int(result.get("ignored", 0))

                for path in added_paths:
                    self._files.append(path)
                    item = QListWidgetItem()
                    item.setToolTip(str(path))
                    item.setData(Qt.ItemDataRole.UserRole, str(path))
                    self.drop_list.addItem(item)
                    row_widget = _CoordinatorFileItemWidget(str(path), parent=self.drop_list)
                    row_widget.removed.connect(self._remove_file_by_path)
                    item.setSizeHint(row_widget.sizeHint())
                    self.drop_list.setItemWidget(item, row_widget)

                if added_paths:
                    self._publish_status_key(
                        "coordinator.status.added",
                        added=len(added_paths),
                        total=len(self._files),
                    )
                if duplicates:
                    show_toast(
                        self,
                        t("coordinator.duplicate.body", count=duplicates),
                        title=t("coordinator.duplicate.title"),
                        level="info",
                    )
                if invalid_paths:
                    file_names = "\n".join(path.name for path in invalid_paths)
                    show_toast(
                        self,
                        t(
                            "coordinator.invalid_final_report.body",
                            count=len(invalid_paths),
                            files=file_names,
                        ),
                        title=t("coordinator.invalid_final_report.title"),
                        level="warning",
                    )
                if ignored:
                    self._publish_status_key("coordinator.status.ignored", count=ignored)

                log_process_message(
                    process_name,
                    logger=self._logger,
                    success_message=(
                        f"{process_name} completed successfully. "
                        f"added={len(added_paths)}, duplicates={duplicates}, "
                        f"invalid={len(invalid_paths)}, ignored={ignored}"
                    ),
                    user_success_message=build_i18n_log_message(
                        "coordinator.status.processing_completed",
                        fallback=t("coordinator.status.processing_completed"),
                    ),
                    job_id=job_id,
                    step_id="coordinator_collect_files",
                )
            finally:
                _finalize(job)

        def _on_failed(exc: Exception) -> None:
            try:
                if isinstance(exc, JobCancelledError):
                    self._publish_status_key("coordinator.status.operation_cancelled")
                    self._logger.info(
                        "%s cancelled by user/system request.",
                        process_name,
                        extra={
                            "user_message": build_i18n_log_message(
                                "coordinator.status.operation_cancelled",
                                fallback=t("coordinator.status.operation_cancelled"),
                            ),
                            "job_id": job_id,
                            "step_id": "coordinator_collect_files",
                        },
                    )
                    return
                log_process_message(
                    process_name,
                    logger=self._logger,
                    error=exc,
                    user_error_message=build_i18n_log_message(
                        "coordinator.status.processing_failed",
                        fallback=t("coordinator.status.processing_failed"),
                    ),
                    job_id=job_id,
                    step_id="coordinator_collect_files",
                )
                show_toast(
                    self,
                    t("coordinator.status.processing_failed"),
                    title=t("coordinator.title"),
                    level="error",
                )
            finally:
                _finalize(job)

        job = run_in_background(
            _analyze_dropped_files,
            dropped_files,
            existing_keys=existing_keys,
            existing_paths=existing_paths,
            token=token,
            on_finished=_on_finished,
            on_failed=_on_failed,
        )
        self._active_jobs.append(job)

    def _on_files_dropped(self, dropped_files: list[str]) -> None:
        first_path = next((value for value in dropped_files if value), "")
        if first_path:
            self._remember_dialog_dir_safe(first_path)
        self._process_files_async(dropped_files)

    def _browse_files(self) -> None:
        if self.state.busy:
            return
        selected_files, _ = QFileDialog.getOpenFileNames(
            self,
            t("coordinator.dialog.title"),
            resolve_dialog_start_path(APP_NAME),
            t("coordinator.dialog.filter"),
        )
        if selected_files:
            self._remember_dialog_dir_safe(selected_files[0])
            self._process_files_async(selected_files)

    def _remember_dialog_dir_safe(self, selected_path: str) -> None:
        try:
            remember_dialog_dir(selected_path, app_name=APP_NAME)
        except OSError:
            remember_dialog_dir_safe(
                selected_path,
                app_name=APP_NAME,
                logger=self._logger,
            )

    def _setup_ui_logging(self) -> None:
        if self._ui_log_handler is not None:
            return
        self._ui_log_handler = UILogHandler(self._append_user_log)
        self._logger.addHandler(self._ui_log_handler)
        self._append_user_log(
            build_i18n_log_message(
                "instructor.log.ready",
                fallback=t("instructor.log.ready"),
            )
        )

    def _append_user_log(self, message: str) -> None:
        parsed = parse_i18n_log_message(message)
        localized = resolve_i18n_log_message(message)
        timestamp = datetime.now()
        if parsed is None:
            self._user_log_entries.append({"timestamp": timestamp, "message": localized})
        else:
            key, kwargs, fallback = parsed
            self._user_log_entries.append(
                {
                    "timestamp": timestamp,
                    "message": localized,
                    "text_key": key,
                    "kwargs": kwargs,
                    "fallback": fallback,
                }
            )
        line = format_log_line_at(localized, timestamp=timestamp)
        if line is None:
            return
        self.user_log_view.appendPlainText(line)

    def _rerender_user_log(self) -> None:
        self.user_log_view.clear()
        for entry in self._user_log_entries:
            timestamp = entry.get("timestamp")
            text_key = entry.get("text_key")
            fallback = entry.get("fallback")
            kwargs = entry.get("kwargs")
            message = entry.get("message")
            if isinstance(text_key, str):
                safe_kwargs = kwargs if isinstance(kwargs, dict) else {}
                try:
                    resolved = t(text_key, **safe_kwargs)
                except Exception:
                    resolved = fallback if isinstance(fallback, str) else str(message or "")
            else:
                resolved = str(message or "")
            ts = timestamp if isinstance(timestamp, datetime) else None
            line = format_log_line_at(resolved, timestamp=ts)
            if line is None:
                continue
            self.user_log_view.appendPlainText(line)

    def _output_link_markup(self, label: str, path: str | None) -> str:
        if not path:
            return f"<b>{escape(label)}</b>: {t(self.OUTPUT_LINK_NOT_AVAILABLE_KEY)}"
        href_path = Path(path).as_posix()
        file_link = f'<a href="file::{href_path}">{t(self.OUTPUT_LINK_OPEN_FILE_KEY)}</a>'
        folder_link = f'<a href="folder::{href_path}">{t(self.OUTPUT_LINK_OPEN_FOLDER_KEY)}</a>'
        name = escape(Path(path).name)
        full_path = escape(str(Path(path)))
        return (
            f"<b>{escape(label)}</b>: {name}<br>"
            f"<span>{full_path}</span><br>"
            f"{file_link} | {folder_link}"
        )

    def _output_links_html(self) -> str:
        rows: list[str] = []
        for path in self._files:
            rows.append(
                f"<div style='margin-bottom:10px'>{self._output_link_markup(t('coordinator.links.uploaded_report'), str(path))}</div>"
            )
        if not rows:
            rows.append(
                f"<div style='margin-bottom:10px'>{self._output_link_markup(t('coordinator.links.uploaded_report'), None)}</div>"
            )

        if self._downloaded_outputs:
            for path in self._downloaded_outputs:
                rows.append(
                    f"<div style='margin-bottom:10px'>{self._output_link_markup(t('coordinator.links.downloaded_output'), str(path))}</div>"
                )
        else:
            rows.append(
                f"<div style='margin-bottom:10px'>{self._output_link_markup(t('coordinator.links.downloaded_output'), None)}</div>"
            )
        return "".join(rows)

    def _refresh_output_links(self) -> None:
        self.generated_outputs_view.setHtml(self._output_links_html())

    def _on_output_link_activated(self, href: str) -> None:
        mode, _, raw_path = href.partition("::")
        path = raw_path.strip()
        if not path:
            return
        target = Path(path).parent if mode == "folder" else Path(path)
        opened = QDesktopServices.openUrl(QUrl.fromLocalFile(str(target)))
        if opened:
            return
        show_toast(
            self,
            t(self.OUTPUT_LINK_OPEN_FAILED_KEY),
            title=t("instructor.msg.error_title"),
            level="error",
        )

    def _clear_info_text_selection(self) -> None:
        for view in (self.user_log_view, self.generated_outputs_view):
            cursor = view.textCursor()
            if cursor.hasSelection():
                cursor.clearSelection()
                view.setTextCursor(cursor)

    def _on_info_tab_changed(self, _index: int) -> None:
        self._clear_info_text_selection()

    def _refresh_summary(self) -> None:
        self.summary_label.setText(t("coordinator.summary", count=len(self._files)))

    def _set_drop_active(self, active: bool) -> None:
        self.drop_zone.setProperty("dragActive", active)
        self.drop_zone.update()

    def set_shared_activity_log_mode(self, enabled: bool) -> None:
        self.info_tabs.setVisible(not enabled)

    def get_shared_outputs_html(self) -> str:
        return self._output_links_html()

    def _remove_file_by_path(self, file_path: str) -> None:
        if self.state.busy:
            return
        target_key = _path_key(Path(file_path))
        before_count = len(self._files)
        self._files = [path for path in self._files if _path_key(path) != target_key]
        if len(self._files) == before_count:
            return

        for row in range(self.drop_list.count()):
            item = self.drop_list.item(row)
            path_value = str(item.data(Qt.ItemDataRole.UserRole) or "")
            if _path_key(Path(path_value)) == target_key:
                self.drop_list.takeItem(row)
                break

        self._refresh_ui()
        self._publish_status_key("coordinator.status.removed", count=1)
        log_process_message(
            "removing selected coordinator files",
            logger=self._logger,
            success_message="removing selected coordinator files completed successfully. removed=1",
            user_success_message=build_i18n_log_message(
                "coordinator.status.removed",
                kwargs={"count": 1},
                fallback=t("coordinator.status.removed", count=1),
            ),
        )

    def _clear_all(self) -> None:
        if self.state.busy:
            return
        if not self._files:
            return
        total = len(self._files)
        self._files.clear()
        self.drop_list.clear()
        self._refresh_ui()
        self._publish_status_key("coordinator.status.cleared", count=total)
        log_process_message(
            "clearing coordinator files",
            logger=self._logger,
            success_message=f"clearing coordinator files completed successfully. removed={total}",
            user_success_message=build_i18n_log_message(
                "coordinator.status.cleared",
                kwargs={"count": total},
                fallback=t("coordinator.status.cleared", count=total),
            ),
        )

    def closeEvent(self, event) -> None:
        if self._cancel_token is not None:
            self._cancel_token.cancel()
            self._cancel_token = None
        self._active_jobs.clear()
        if self._ui_log_handler is not None:
            self._logger.removeHandler(self._ui_log_handler)
            self._ui_log_handler = None
        super().closeEvent(event)
