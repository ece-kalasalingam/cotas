"""Shared workbook operations for instructor workflow steps."""

from __future__ import annotations

import os
import re
import shutil
import tempfile
from pathlib import Path

from common.constants import (
    CO_REPORT_TEMPLATE_NAME_SUFFIX,
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_SHEET,
    FILE_EXTENSION_XLSX,
    FILENAME_JOIN_SEPARATOR,
    MARKS_TEMPLATE_NAME_SUFFIX,
    WORKBOOK_TEMP_SUFFIX,
)
from common.texts import t
from common.utils import coerce_excel_number, normalize


def sanitize_filename_token(value: object) -> str:
    token = str(value).strip()
    token = re.sub(r'[<>:"/\\|?*]+', "_", token)
    token = re.sub(r"\s+", "", token)
    token = token.strip(" ._")
    return token


def build_marks_template_default_name(course_details_path: str | None) -> str:
    return _build_workbook_default_name_from_metadata(
        workbook_path=course_details_path,
        suffix_label=MARKS_TEMPLATE_NAME_SUFFIX,
        fallback=t("instructor.dialog.step3.default_name"),
    )


def build_final_report_default_name(filled_marks_path: str | None) -> str:
    return _build_workbook_default_name_from_metadata(
        workbook_path=filled_marks_path,
        suffix_label=CO_REPORT_TEMPLATE_NAME_SUFFIX,
        fallback=t("instructor.dialog.step4.default_name"),
    )


def _build_workbook_default_name_from_metadata(
    *,
    workbook_path: str | None,
    suffix_label: str,
    fallback: str,
) -> str:
    if not workbook_path:
        return fallback

    try:
        import openpyxl
    except ModuleNotFoundError:
        return fallback

    workbook = None
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
        if COURSE_METADATA_SHEET not in workbook.sheetnames:
            return fallback
        sheet = workbook[COURSE_METADATA_SHEET]
        fields: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = normalize(row[0] if len(row) > 0 else None)
            if not key:
                continue
            value = row[1] if len(row) > 1 else None
            coerced = coerce_excel_number(value)
            fields[key] = str(coerced).strip() if coerced is not None else ""

        parts = [
            sanitize_filename_token(fields.get(COURSE_METADATA_COURSE_CODE_KEY, "")),
            sanitize_filename_token(fields.get(COURSE_METADATA_SEMESTER_KEY, "")),
            sanitize_filename_token(fields.get(COURSE_METADATA_SECTION_KEY, "")),
            sanitize_filename_token(fields.get(COURSE_METADATA_ACADEMIC_YEAR_KEY, "")),
            suffix_label,
        ]
        if any(not part for part in parts[:4]):
            return fallback
        return f"{FILENAME_JOIN_SEPARATOR.join(parts)}{FILE_EXTENSION_XLSX}"
    except Exception:
        return fallback
    finally:
        if workbook is not None:
            workbook.close()


def atomic_copy_file(source_path: str | Path, output_path: str | Path, *, logger: object | None = None) -> Path:
    source = Path(source_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_name = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            delete=False,
            dir=str(output.parent),
            prefix=f"{output.name}.",
            suffix=WORKBOOK_TEMP_SUFFIX,
        ) as temp_file:
            temp_name = temp_file.name
        shutil.copyfile(str(source), temp_name)
        os.replace(temp_name, output)
    except Exception:
        if temp_name:
            try:
                Path(temp_name).unlink(missing_ok=True)
            except OSError:
                if logger is not None:
                    warning = getattr(logger, "warning", None)
                    if callable(warning):
                        warning("Failed to cleanup temp report file: %s", temp_name)
        raise
    return output
