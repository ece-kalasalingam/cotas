"""Generator for the Course Details workbook template."""

from __future__ import annotations

import logging
import os
import hashlib
import re
from pathlib import Path
from typing import Any, Sequence
from uuid import uuid4

from common.constants import (
    ALLOW_FILTER,
    ALLOW_SELECT_LOCKED,
    ALLOW_SELECT_UNLOCKED,
    ALLOW_SORT,
    ASSESSMENT_CONFIG_HEADERS,
    ASSESSMENT_CONFIG_SHEET,
    ASSESSMENT_VALIDATION_YES_NO_OPTIONS,
    COURSE_METADATA_HEADERS,
    COURSE_METADATA_SHEET,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    HEADER_PATTERNFILL_COLOR,
    ID_COURSE_SETUP,
    LIKERT_MAX,
    LIKERT_MIN,
    MARKS_ENTRY_CO_MARKS_LABEL_PREFIX,
    MARKS_ENTRY_CO_PREFIX,
    MARKS_ENTRY_QUESTION_PREFIX,
    MARKS_ENTRY_ROW_HEADERS,
    MARKS_ENTRY_TOTAL_LABEL,
    MARKS_ENTRY_VALIDATION_ERROR_MESSAGE,
    MARKS_ENTRY_VALIDATION_ERROR_TITLE,
    MARKS_ENTRY_VALIDATION_FORMULA,
    MAX_EXCEL_SHEETNAME_LENGTH,
    MIN_MARK_VALUE,
    QUESTION_MAP_HEADERS,
    QUESTION_MAP_SHEET,
    STUDENTS_HEADERS,
    STUDENTS_SHEET,
    SYSTEM_HASH_SHEET,
    SYSTEM_HASH_TEMPLATE_HASH_HEADER,
    SYSTEM_HASH_TEMPLATE_HASH_KEY,
    SYSTEM_HASH_TEMPLATE_ID_HEADER,
    SYSTEM_HASH_TEMPLATE_ID_KEY,
    WEIGHT_TOTAL_EXPECTED,
    WEIGHT_TOTAL_ROUND_DIGITS,
    WORKBOOK_PASSWORD,
)
from common.exceptions import AppSystemError, ValidationError
from common.registry import BLUEPRINT_REGISTRY
from common.sample_setup_data import SAMPLE_SETUP_DATA
from common.sheet_schema import ValidationRule, WorkbookBlueprint
from common.texts import t
from common.utils import coerce_excel_number, normalize

_logger = logging.getLogger(__name__)
_YES_NO_TOKENS = {normalize(option) for option in ASSESSMENT_VALIDATION_YES_NO_OPTIONS}
_YES_TOKEN = normalize(ASSESSMENT_VALIDATION_YES_NO_OPTIONS[0])


def generate_course_details_template(
    output_path: str | Path, template_id: str = ID_COURSE_SETUP
) -> Path:
    """Generate and save the course details template workbook with atomic replace."""
    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise ValidationError(t("instructor.validation.xlsxwriter_missing")) from exc

    blueprint = _get_blueprint(template_id)
    sample_data = SAMPLE_SETUP_DATA
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    temp_output = output.with_name(f"{output.name}.{uuid4().hex}.tmp")
    workbook = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    workbook_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal workbook_closed
        if not workbook_closed:
            try:
                workbook.close()
                workbook_closed = True
            except Exception:
                pass
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp template file: %s", temp_output)

    try:
        header_format = _build_header_format(workbook, blueprint.style_registry.get("header", {}))
        body_format = _build_body_format(workbook, blueprint.style_registry.get("body", {}))

        for sheet_schema in blueprint.sheets:
            if len(sheet_schema.header_matrix) != 1:
                raise ValidationError(
                    t(
                        "instructor.validation.sheet_single_header_row",
                        sheet_name=sheet_schema.name,
                    )
                )

            worksheet = generate_worksheet(
                workbook=workbook,
                sheet_name=sheet_schema.name,
                headers=sheet_schema.header_matrix[0],
                data=sample_data.get(sheet_schema.name, []),
                header_format=header_format,
                body_format=body_format,
            )

            for validation in sheet_schema.validations:
                _apply_validation(worksheet, validation)

            if sheet_schema.is_protected:
                _protect_sheet(worksheet)

        _add_system_hash_sheet(workbook, template_id)

        workbook.close()
        workbook_closed = True
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate course details template. template_id=%s output=%s",
            template_id,
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc

    return output


def generate_marks_template_from_course_details(
    course_details_path: str | Path,
    output_path: str | Path,
) -> Path:
    """Generate marks-entry workbook from a validated course-details workbook."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise ValidationError(t("instructor.validation.openpyxl_missing")) from exc

    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise ValidationError(t("instructor.validation.xlsxwriter_missing")) from exc

    source_file = Path(course_details_path)
    if not source_file.exists():
        raise ValidationError(t("instructor.validation.workbook_not_found", workbook=source_file))

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output.with_name(f"{output.name}.{uuid4().hex}.tmp")

    try:
        workbook = openpyxl.load_workbook(source_file, data_only=True)
    except Exception as exc:
        raise ValidationError(t("instructor.validation.workbook_open_failed", workbook=source_file)) from exc

    target = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    target_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal target_closed
        if not target_closed:
            try:
                target.close()
                target_closed = True
            except Exception:
                pass
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp marks template file: %s", temp_output)

    try:
        context = _extract_marks_template_context(workbook)
        _write_marks_template_workbook(target, context)
        _copy_system_hash_sheet(workbook, target)
        target.close()
        target_closed = True
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate marks template. source=%s output=%s",
            source_file,
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc
    finally:
        workbook.close()

    return output


def _extract_marks_template_context(workbook: Any) -> dict[str, Any]:
    metadata_rows = _iter_data_rows(workbook[COURSE_METADATA_SHEET], len(COURSE_METADATA_HEADERS))
    assessment_rows = _iter_data_rows(
        workbook[ASSESSMENT_CONFIG_SHEET], len(ASSESSMENT_CONFIG_HEADERS)
    )
    question_rows = _iter_data_rows(workbook[QUESTION_MAP_SHEET], len(QUESTION_MAP_HEADERS))
    student_rows = _iter_data_rows(workbook[STUDENTS_SHEET], len(STUDENTS_HEADERS))

    total_outcomes = _extract_total_outcomes(metadata_rows)
    students = _extract_students(student_rows)
    components = _extract_components(assessment_rows)
    questions_by_component = _extract_questions(question_rows)

    return {
        "metadata_rows": metadata_rows,
        "assessment_rows": assessment_rows,
        "total_outcomes": total_outcomes,
        "students": students,
        "components": components,
        "questions_by_component": questions_by_component,
    }


def _extract_total_outcomes(metadata_rows: Sequence[Sequence[Any]]) -> int:
    for row in metadata_rows:
        if len(row) < 2:
            continue
        if normalize(row[0]) == "total_outcomes":
            value = coerce_excel_number(row[1])
            if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
                break
            return value
    raise ValidationError(t("instructor.validation.course_metadata_total_outcomes_invalid"))


def _extract_students(student_rows: Sequence[Sequence[Any]]) -> list[tuple[str, str]]:
    students: list[tuple[str, str]] = []
    for row in student_rows:
        reg_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if reg_no and name:
            students.append((reg_no, name))
    if not students:
        raise ValidationError(t("instructor.validation.students_row_required_one"))
    return students


def _extract_components(assessment_rows: Sequence[Sequence[Any]]) -> list[dict[str, Any]]:
    components: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in assessment_rows:
        if len(row) < 5:
            continue
        component_name = str(row[0]).strip()
        component_key = normalize(component_name)
        if not component_key or component_key in seen:
            continue
        seen.add(component_key)
        components.append(
            {
                "key": component_key,
                "display_name": component_name,
                "co_wise_breakup": _parse_yes_no(
                    row[3],
                    ASSESSMENT_CONFIG_SHEET,
                    0,
                    ASSESSMENT_CONFIG_HEADERS[3],
                ),
                "is_direct": _parse_yes_no(
                    row[4],
                    ASSESSMENT_CONFIG_SHEET,
                    0,
                    ASSESSMENT_CONFIG_HEADERS[4],
                ),
            }
        )
    if not components:
        raise ValidationError(t("instructor.validation.assessment_component_required_one"))
    return components


def _extract_questions(question_rows: Sequence[Sequence[Any]]) -> dict[str, list[dict[str, Any]]]:
    questions_by_component: dict[str, list[dict[str, Any]]] = {}
    for row in question_rows:
        if len(row) < 4:
            continue
        component_key = normalize(row[0])
        if not component_key:
            continue
        max_marks = coerce_excel_number(row[2])
        if isinstance(max_marks, bool) or not isinstance(max_marks, (int, float)):
            continue
        co_values = _co_tokens(row[3])
        if not co_values:
            continue
        questions_by_component.setdefault(component_key, []).append(
            {
                "max_marks": float(max_marks),
                "co_values": co_values,
            }
        )
    return questions_by_component


def _write_marks_template_workbook(workbook: Any, context: dict[str, Any]) -> None:
    header_fmt = workbook.add_format({"bold": True, "border": 1, "align": "center", "valign": "vcenter"})
    body_fmt = workbook.add_format({"border": 1})
    num_fmt = workbook.add_format({"border": 1, "num_format": "0.00"})

    _write_two_column_copy_sheet(
        workbook=workbook,
        title=COURSE_METADATA_SHEET,
        header=COURSE_METADATA_HEADERS,
        rows=context["metadata_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
    )
    _write_multi_column_copy_sheet(
        workbook=workbook,
        title=ASSESSMENT_CONFIG_SHEET,
        header=ASSESSMENT_CONFIG_HEADERS,
        rows=context["assessment_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
        num_fmt=num_fmt,
    )

    used_sheet_names = {normalize(COURSE_METADATA_SHEET), normalize(ASSESSMENT_CONFIG_SHEET)}
    for component in context["components"]:
        component_name = component["display_name"]
        sheet_name = _safe_sheet_name(component_name, used_sheet_names)
        questions = context["questions_by_component"].get(component["key"], [])
        if component["is_direct"]:
            if not questions:
                raise ValidationError(t("instructor.validation.question_map_row_required_one"))
            if component["co_wise_breakup"]:
                _write_direct_co_wise_sheet(
                    workbook,
                    sheet_name,
                    context["students"],
                    questions,
                    header_fmt,
                    body_fmt,
                    num_fmt,
                )
            else:
                _write_direct_non_co_wise_sheet(
                    workbook,
                    sheet_name,
                    context["students"],
                    questions,
                    header_fmt,
                    body_fmt,
                    num_fmt,
                )
        else:
            _write_indirect_sheet(
                workbook,
                sheet_name,
                context["students"],
                context["total_outcomes"],
                header_fmt,
                body_fmt,
            )


def _copy_system_hash_sheet(source_workbook: Any, target_workbook: Any) -> None:
    if SYSTEM_HASH_SHEET not in source_workbook.sheetnames:
        return

    source = source_workbook[SYSTEM_HASH_SHEET]
    target = target_workbook.add_worksheet(SYSTEM_HASH_SHEET)
    target.write_row(0, 0, [source["A1"].value, source["B1"].value])
    target.write_row(1, 0, [source["A2"].value, source["B2"].value])
    target.hide()


def _write_two_column_copy_sheet(
    workbook: Any,
    title: str,
    header: tuple[str, str],
    rows: Sequence[Sequence[Any]],
    header_fmt: Any,
    body_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(title)
    ws.write_row(0, 0, list(header), header_fmt)
    ws.set_column(0, 0, 24)
    ws.set_column(1, 1, 24)
    for row_index, row in enumerate(rows, start=1):
        ws.write(row_index, 0, row[0] if len(row) > 0 else "", body_fmt)
        ws.write(row_index, 1, row[1] if len(row) > 1 else "", body_fmt)
    ws.freeze_panes(1, 0)


def _write_multi_column_copy_sheet(
    workbook: Any,
    title: str,
    header: Sequence[str],
    rows: Sequence[Sequence[Any]],
    header_fmt: Any,
    body_fmt: Any,
    num_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(title)
    ws.write_row(0, 0, list(header), header_fmt)
    for col in range(len(header)):
        ws.set_column(col, col, 22)
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row[: len(header)]):
            cell_fmt = num_fmt if col_index == 1 and isinstance(value, (int, float)) else body_fmt
            ws.write(row_index, col_index, value, cell_fmt)
    ws.freeze_panes(1, 0)


def _write_direct_co_wise_sheet(
    workbook: Any,
    sheet_name: str,
    students: Sequence[tuple[str, str]],
    questions: Sequence[dict[str, Any]],
    header_fmt: Any,
    body_fmt: Any,
    num_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(sheet_name)
    question_count = len(questions)
    total_col = 3 + question_count
    ws.write_row(0, 0, list(MARKS_ENTRY_ROW_HEADERS), header_fmt)
    for idx in range(question_count):
        ws.write(0, 3 + idx, f"{MARKS_ENTRY_QUESTION_PREFIX}{idx + 1}", header_fmt)
    ws.write(0, total_col, MARKS_ENTRY_TOTAL_LABEL, header_fmt)

    ws.write_row(1, 0, ["", "", ""], header_fmt)
    for idx, question in enumerate(questions):
        co_number = question["co_values"][0]
        ws.write(1, 3 + idx, f"{MARKS_ENTRY_CO_PREFIX}{co_number}", header_fmt)
    ws.write(1, total_col, "", header_fmt)

    ws.write_row(2, 0, ["", "", ""], header_fmt)
    component_total = 0.0
    for idx, question in enumerate(questions):
        max_marks = float(question["max_marks"])
        component_total += max_marks
        ws.write_number(2, 3 + idx, max_marks, num_fmt)
    ws.write_number(2, total_col, component_total, num_fmt)

    for row_offset, (reg_no, student_name) in enumerate(students, start=3):
        ws.write_number(row_offset, 0, row_offset - 2, body_fmt)
        ws.write(row_offset, 1, reg_no, body_fmt)
        ws.write(row_offset, 2, student_name, body_fmt)
        for col in range(3, total_col):
            ws.write_blank(row_offset, col, None, body_fmt)
        first_mark_col = _excel_col_name(3)
        last_mark_col = _excel_col_name(total_col - 1)
        ws.write_formula(
            row_offset,
            total_col,
            f"=SUM({first_mark_col}{row_offset + 1}:{last_mark_col}{row_offset + 1})",
            num_fmt,
        )

    if students and question_count > 0:
        first_row = 3
        last_row = 2 + len(students)
        validation_formula = MARKS_ENTRY_VALIDATION_FORMULA
        ws.data_validation(
            first_row,
            3,
            last_row,
            total_col - 1,
            {
                "validate": "custom",
                "value": validation_formula,
                "error_title": MARKS_ENTRY_VALIDATION_ERROR_TITLE,
                "error_message": MARKS_ENTRY_VALIDATION_ERROR_MESSAGE,
                "ignore_blank": True,
            },
        )

    _set_common_student_columns(ws, total_col)
    ws.freeze_panes(3, 3)


def _write_direct_non_co_wise_sheet(
    workbook: Any,
    sheet_name: str,
    students: Sequence[tuple[str, str]],
    questions: Sequence[dict[str, Any]],
    header_fmt: Any,
    body_fmt: Any,
    num_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(sheet_name)
    covered_cos = sorted({co for q in questions for co in q["co_values"]})
    co_count = max(1, len(covered_cos))
    total_max = sum(float(question["max_marks"]) for question in questions)
    per_co = total_max / co_count if co_count else total_max

    ws.write_row(0, 0, list(MARKS_ENTRY_ROW_HEADERS) + [MARKS_ENTRY_TOTAL_LABEL], header_fmt)
    for idx, co_number in enumerate(covered_cos):
        ws.write(0, 4 + idx, f"{MARKS_ENTRY_CO_MARKS_LABEL_PREFIX}{co_number}", header_fmt)

    ws.write_row(1, 0, ["", "", "", "COs"], header_fmt)
    for idx, co_number in enumerate(covered_cos):
        ws.write(1, 4 + idx, f"{MARKS_ENTRY_CO_PREFIX}{co_number}", header_fmt)

    ws.write_row(2, 0, ["", "", "", ""], header_fmt)
    ws.write_number(2, 3, total_max, num_fmt)
    for idx in range(len(covered_cos)):
        ws.write_number(2, 4 + idx, per_co, num_fmt)

    for row_offset, (reg_no, student_name) in enumerate(students, start=3):
        ws.write_number(row_offset, 0, row_offset - 3, body_fmt)
        ws.write(row_offset, 1, reg_no, body_fmt)
        ws.write(row_offset, 2, student_name, body_fmt)
        ws.write_blank(row_offset, 3, None, body_fmt)
        for idx in range(len(covered_cos)):
            co_col = 4 + idx
            col_name_total = _excel_col_name(3)
            divisor = len(covered_cos) if covered_cos else 1
            ws.write_formula(
                row_offset,
                co_col,
                (
                    f'=IF(OR(${col_name_total}{row_offset + 1}="A",${col_name_total}{row_offset + 1}="a"),'
                    f'"A",IF(${col_name_total}{row_offset + 1}="","",${col_name_total}{row_offset + 1}/{divisor}))'
                ),
                num_fmt,
            )

    if students:
        first_row = 3
        last_row = 2 + len(students)
        validation_formula = MARKS_ENTRY_VALIDATION_FORMULA
        ws.data_validation(
            first_row,
            3,
            last_row,
            3,
            {
                "validate": "custom",
                "value": validation_formula,
                "error_title": MARKS_ENTRY_VALIDATION_ERROR_TITLE,
                "error_message": MARKS_ENTRY_VALIDATION_ERROR_MESSAGE,
                "ignore_blank": True,
            },
        )

    _set_common_student_columns(ws, 3 + len(covered_cos))
    ws.freeze_panes(3, 3)


def _write_indirect_sheet(
    workbook: Any,
    sheet_name: str,
    students: Sequence[tuple[str, str]],
    total_outcomes: int,
    header_fmt: Any,
    body_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(sheet_name)
    headers = list(MARKS_ENTRY_ROW_HEADERS) + [
        f"{MARKS_ENTRY_CO_PREFIX}{i}" for i in range(1, total_outcomes + 1)
    ]
    ws.write_row(0, 0, headers, header_fmt)

    for row_offset, (reg_no, student_name) in enumerate(students, start=1):
        ws.write_number(row_offset, 0, row_offset, body_fmt)
        ws.write(row_offset, 1, reg_no, body_fmt)
        ws.write(row_offset, 2, student_name, body_fmt)
        for col in range(3, 3 + total_outcomes):
            ws.write_blank(row_offset, col, None, body_fmt)

    if students and total_outcomes > 0:
        first_row = 1
        last_row = len(students)
        ws.data_validation(
            first_row,
            3,
            last_row,
            2 + total_outcomes,
            {
                "validate": "custom",
                "value": (
                    f'=OR(D2="A",D2="a",AND(ISNUMBER(D2),D2>={MIN_MARK_VALUE},'
                    f'D2>={LIKERT_MIN},D2<={LIKERT_MAX}))'
                ),
                "error_title": "Invalid marks",
                "error_message": "Enter A/a or a numeric Likert value within allowed range.",
                "ignore_blank": True,
            },
        )

    _set_common_student_columns(ws, 2 + total_outcomes)
    ws.freeze_panes(1, 3)


def _set_common_student_columns(ws: Any, last_col: int) -> None:
    ws.set_column(0, 0, 10)
    ws.set_column(1, 1, 18)
    ws.set_column(2, 2, 28)
    if last_col >= 3:
        ws.set_column(3, last_col, 12)


def _excel_col_name(col_index: int) -> str:
    index = col_index + 1
    label = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        label = chr(65 + rem) + label
    return label


def _safe_sheet_name(name: str, used_sheet_names: set[str]) -> str:
    token = re.sub(r"[:\\/?*\[\]]", "_", name).strip() or "Component"
    token = token[:MAX_EXCEL_SHEETNAME_LENGTH]
    base_key = normalize(token)
    if base_key not in used_sheet_names:
        used_sheet_names.add(base_key)
        return token

    counter = 2
    while True:
        suffix = f"_{counter}"
        trimmed = token[: max(1, MAX_EXCEL_SHEETNAME_LENGTH - len(suffix))]
        candidate = f"{trimmed}{suffix}"
        key = normalize(candidate)
        if key not in used_sheet_names:
            used_sheet_names.add(key)
            return candidate
        counter += 1


def _get_blueprint(template_id: str) -> WorkbookBlueprint:
    blueprint = BLUEPRINT_REGISTRY.get(template_id)
    if blueprint is None:
        available = ", ".join(sorted(BLUEPRINT_REGISTRY))
        raise ValidationError(
            t(
                "instructor.validation.unknown_template",
                template_id=template_id,
                available=available,
            )
        )
    return blueprint


def _build_header_format(workbook: Any, header_style: dict[str, Any]) -> Any:
    bg_color = str(header_style.get("bg_color", HEADER_PATTERNFILL_COLOR))
    return workbook.add_format(
        {
            "bold": bool(header_style.get("bold", True)),
            "bg_color": bg_color,
            "border": int(header_style.get("border", 1)),
            "align": str(header_style.get("align", "center")),
            "valign": str(header_style.get("valign", "vcenter")),
        }
    )


def _build_body_format(workbook: Any, body_style: dict[str, Any]) -> Any:
    return workbook.add_format(
        {
            "locked": bool(body_style.get("locked", False)),
            "border": int(body_style.get("border", 1)),
        }
    )


def generate_worksheet(
    workbook: Any,
    sheet_name: str,
    headers: Sequence[str],
    data: Sequence[Sequence[Any]],
    header_format: Any,
    body_format: Any,
) -> Any:
    """Create a worksheet with strict validation and efficient row writes."""
    if not sheet_name or not isinstance(sheet_name, str):
        raise ValidationError(t("instructor.validation.invalid_sheet_name"))

    if not headers:
        raise ValidationError(t("instructor.validation.headers_empty", sheet_name=sheet_name))

    if len(set(headers)) != len(headers):
        raise ValidationError(t("instructor.validation.headers_unique", sheet_name=sheet_name))

    column_count = len(headers)
    for row_index, row in enumerate(data, start=1):
        if len(row) != column_count:
            raise ValidationError(
                t(
                    "instructor.validation.row_length_mismatch",
                    row=row_index,
                    sheet_name=sheet_name,
                    expected=column_count,
                    found=len(row),
                )
            )

    worksheet = workbook.add_worksheet(sheet_name)
    col_widths: dict[int, int] = {}
    write_row = worksheet.write_row
    write_row(0, 0, headers, header_format)
    for col_index, value in enumerate(headers):
        col_widths[col_index] = max(12, len(str(value)) + 2)

    for col_index, width in col_widths.items():
        worksheet.set_column(col_index, col_index, width)

    for row_offset, row_values in enumerate(data, start=1):
        write_row(row_offset, 0, row_values, body_format)

    worksheet.freeze_panes(1, 0)
    return worksheet


def _apply_validation(worksheet: Any, rule: ValidationRule) -> None:
    options = dict(rule.options)
    validation_type = options.pop("validate", None)
    if not validation_type:
        return

    options["validate"] = validation_type
    if "ignore_blank" not in options:
        options["ignore_blank"] = True

    worksheet.data_validation(
        rule.first_row,
        rule.first_col,
        rule.last_row,
        rule.last_col,
        options,
    )


def _protect_sheet(worksheet: Any) -> None:
    worksheet.protect(
        WORKBOOK_PASSWORD,
        {
            "sort": ALLOW_SORT,
            "autofilter": ALLOW_FILTER,
            "select_locked_cells": ALLOW_SELECT_LOCKED,
            "select_unlocked_cells": ALLOW_SELECT_UNLOCKED,
        },
    )


def _add_system_hash_sheet(workbook: Any, template_id: str) -> None:
    worksheet = workbook.add_worksheet(SYSTEM_HASH_SHEET)
    template_hash = _compute_template_hash(template_id)

    worksheet.write_row(
        0,
        0,
        [SYSTEM_HASH_TEMPLATE_ID_HEADER, SYSTEM_HASH_TEMPLATE_HASH_HEADER],
    )
    worksheet.write_row(1, 0, [template_id, template_hash])
    worksheet.hide()


def _compute_template_hash(template_id: str) -> str:
    payload = f"{template_id}|{WORKBOOK_PASSWORD}".encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def validate_course_details_workbook(workbook_path: str | Path) -> str:
    """Validate uploaded course details workbook and return template id."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise ValidationError(t("instructor.validation.openpyxl_missing")) from exc

    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise ValidationError(t("instructor.validation.workbook_not_found", workbook=workbook_file))

    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=True)
    except Exception as exc:
        raise ValidationError(t("instructor.validation.workbook_open_failed", workbook=workbook_file)) from exc

    try:
        template_id = _extract_and_validate_template_id(workbook)
        blueprint = _get_blueprint(template_id)
        _validate_workbook_schema(workbook, blueprint)
        _validate_template_specific_rules(workbook, template_id)
    finally:
        workbook.close()

    return template_id


def _extract_and_validate_template_id(workbook: Any) -> str:
    if SYSTEM_HASH_SHEET not in workbook.sheetnames:
        raise ValidationError(
            t("instructor.validation.system_sheet_missing", sheet=SYSTEM_HASH_SHEET)
        )

    hash_sheet = workbook[SYSTEM_HASH_SHEET]
    if normalize(hash_sheet["A1"].value) != normalize(SYSTEM_HASH_TEMPLATE_ID_KEY):
        raise ValidationError(t("instructor.validation.system_hash_missing_template_id_header"))
    if normalize(hash_sheet["B1"].value) != normalize(SYSTEM_HASH_TEMPLATE_HASH_KEY):
        raise ValidationError(t("instructor.validation.system_hash_missing_template_hash_header"))

    template_id = str(hash_sheet["A2"].value).strip() if hash_sheet["A2"].value is not None else ""
    template_hash = (
        str(hash_sheet["B2"].value).strip() if hash_sheet["B2"].value is not None else ""
    )
    if not template_id:
        raise ValidationError(t("instructor.validation.system_hash_template_id_missing"))
    expected_hash = _compute_template_hash(template_id)
    if template_hash != expected_hash:
        raise ValidationError(t("instructor.validation.system_hash_mismatch"))
    return template_id


def _validate_workbook_schema(workbook: Any, blueprint: WorkbookBlueprint) -> None:
    expected_sheet_names = [schema.name for schema in blueprint.sheets] + [SYSTEM_HASH_SHEET]
    actual_sheet_names = list(workbook.sheetnames)
    if actual_sheet_names != expected_sheet_names:
        raise ValidationError(
            t(
                "instructor.validation.workbook_sheet_mismatch",
                template_id=blueprint.type_id,
                expected=expected_sheet_names,
                found=actual_sheet_names,
            )
        )

    for sheet_schema in blueprint.sheets:
        if len(sheet_schema.header_matrix) != 1:
            raise ValidationError(
                t(
                    "instructor.validation.sheet_single_header_row",
                    sheet_name=sheet_schema.name,
                )
            )

        expected_headers = [normalize(h) for h in sheet_schema.header_matrix[0]]
        worksheet = workbook[sheet_schema.name]
        actual_headers = [
            normalize(worksheet.cell(row=1, column=col_index + 1).value)
            for col_index in range(len(expected_headers))
        ]
        if actual_headers != expected_headers:
            raise ValidationError(
                t(
                    "instructor.validation.header_mismatch",
                    sheet_name=sheet_schema.name,
                    expected=sheet_schema.header_matrix[0],
                )
            )


def _iter_data_rows(worksheet: Any, expected_col_count: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=2, max_col=expected_col_count, values_only=True):
        values = list(row)
        if any(normalize(value) != "" for value in values):
            rows.append(values)
    return rows


def _header_index_map(worksheet: Any, headers: Sequence[str]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col_index, expected_header in enumerate(headers, start=1):
        header_value = worksheet.cell(row=1, column=col_index).value
        index[normalize(expected_header)] = col_index - 1
        if normalize(header_value) != normalize(expected_header):
            raise ValidationError(
                t(
                    "instructor.validation.unexpected_header",
                    sheet_name=worksheet.title,
                    col=col_index,
                )
            )
    return index


def _validate_template_specific_rules(workbook: Any, template_id: str) -> None:
    if template_id == ID_COURSE_SETUP:
        _validate_course_setup_v1(workbook)
        return
    raise ValidationError(t("instructor.validation.validator_missing", template_id=template_id))


def _validate_course_setup_v1(workbook: Any) -> None:
    metadata_sheet = workbook[COURSE_METADATA_SHEET]
    assessment_sheet = workbook[ASSESSMENT_CONFIG_SHEET]
    question_map_sheet = workbook[QUESTION_MAP_SHEET]
    students_sheet = workbook[STUDENTS_SHEET]

    total_outcomes = _validate_course_metadata(metadata_sheet)
    component_config = _validate_assessment_config(assessment_sheet)
    _validate_question_map(question_map_sheet, component_config, total_outcomes)
    _validate_students(students_sheet)


def _validate_course_metadata(worksheet: Any) -> int:
    expected_headers = list(COURSE_METADATA_HEADERS)
    header_map = _header_index_map(worksheet, expected_headers)
    rows = _iter_data_rows(worksheet, len(expected_headers))

    expected_field_rows = SAMPLE_SETUP_DATA.get(COURSE_METADATA_SHEET, [])
    expected_field_types: dict[str, type] = {}
    for field_name, sample_value in expected_field_rows:
        key = normalize(field_name)
        expected_field_types[key] = int if isinstance(sample_value, int) else str

    actual_values: dict[str, Any] = {}
    for row_number, row in enumerate(rows, start=2):
        field_raw = row[header_map["field"]]
        value_raw = row[header_map["value"]]
        field_key = normalize(field_raw)
        if not field_key:
            raise ValidationError(t("instructor.validation.course_metadata_field_empty", row=row_number))
        if field_key in actual_values:
            raise ValidationError(
                t(
                    "instructor.validation.course_metadata_duplicate_field",
                    row=row_number,
                    field=field_raw,
                )
            )
        if field_key not in expected_field_types:
            raise ValidationError(
                t(
                    "instructor.validation.course_metadata_unknown_field",
                    row=row_number,
                    field=field_raw,
                )
            )
        if normalize(value_raw) == "":
            raise ValidationError(
                t(
                    "instructor.validation.course_metadata_value_required",
                    row=row_number,
                    field=field_raw,
                )
            )
        actual_values[field_key] = coerce_excel_number(value_raw)

    missing_fields = [name for name in expected_field_types if name not in actual_values]
    if missing_fields:
        raise ValidationError(
            t(
                "instructor.validation.course_metadata_missing_fields",
                fields=", ".join(missing_fields),
            )
        )

    for field_key, expected_type in expected_field_types.items():
        value = actual_values[field_key]
        if expected_type is int:
            if isinstance(value, bool) or not isinstance(value, int):
                raise ValidationError(
                    t("instructor.validation.course_metadata_field_must_be_int", field=field_key)
                )
        else:
            if not isinstance(value, str) or normalize(value) == "":
                raise ValidationError(
                    t(
                        "instructor.validation.course_metadata_field_must_be_non_empty_str",
                        field=field_key,
                    )
                )

    total_outcomes = actual_values.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY))
    if isinstance(total_outcomes, bool) or not isinstance(total_outcomes, int) or total_outcomes <= 0:
        raise ValidationError(t("instructor.validation.course_metadata_total_outcomes_invalid"))
    return total_outcomes


def _parse_yes_no(value: Any, sheet_name: str, row_number: int, field_name: str) -> bool:
    token = normalize(value)
    if token not in _YES_NO_TOKENS:
        raise ValidationError(
            t(
                "instructor.validation.yes_no_required",
                sheet_name=sheet_name,
                row=row_number,
                field=field_name,
            )
        )
    return token == _YES_TOKEN


def _validate_assessment_config(worksheet: Any) -> dict[str, dict[str, Any]]:
    expected_headers = list(ASSESSMENT_CONFIG_HEADERS)
    header_map = _header_index_map(worksheet, expected_headers)
    rows = _iter_data_rows(worksheet, len(expected_headers))

    if not rows:
        raise ValidationError(t("instructor.validation.assessment_component_required_one"))

    component_config: dict[str, dict[str, Any]] = {}
    direct_weight_total = 0.0
    indirect_weight_total = 0.0
    direct_count = 0
    indirect_count = 0

    for row_number, row in enumerate(rows, start=2):
        component_raw = row[header_map["component"]]
        component_key = normalize(component_raw)
        if not component_key:
            raise ValidationError(
                t("instructor.validation.assessment_component_required", row=row_number)
            )
        if component_key in component_config:
            raise ValidationError(
                t(
                    "instructor.validation.assessment_component_duplicate",
                    row=row_number,
                    component=component_raw,
                )
            )

        weight_value = coerce_excel_number(row[header_map["weight (%)"]])
        if (
            isinstance(weight_value, bool)
            or not isinstance(weight_value, (int, float))
        ):
            raise ValidationError(
                t("instructor.validation.assessment_weight_numeric", row=row_number)
            )

        is_direct = _parse_yes_no(
            row[header_map["direct"]],
            ASSESSMENT_CONFIG_SHEET,
            row_number,
            ASSESSMENT_CONFIG_HEADERS[4],
        )
        co_wise_breakup = _parse_yes_no(
            row[header_map["co_wise_marks_breakup"]],
            ASSESSMENT_CONFIG_SHEET,
            row_number,
            ASSESSMENT_CONFIG_HEADERS[3],
        )
        _parse_yes_no(
            row[header_map["cia"]],
            ASSESSMENT_CONFIG_SHEET,
            row_number,
            ASSESSMENT_CONFIG_HEADERS[2],
        )

        if is_direct:
            direct_weight_total += float(weight_value)
            direct_count += 1
        else:
            indirect_weight_total += float(weight_value)
            indirect_count += 1

        component_config[component_key] = {
            "display_name": str(component_raw).strip(),
            "co_wise_breakup": co_wise_breakup,
        }

    if direct_count == 0:
        raise ValidationError(t("instructor.validation.assessment_direct_missing"))
    if indirect_count == 0:
        raise ValidationError(t("instructor.validation.assessment_indirect_missing"))
    if round(direct_weight_total, WEIGHT_TOTAL_ROUND_DIGITS) != WEIGHT_TOTAL_EXPECTED:
        raise ValidationError(
            t("instructor.validation.assessment_direct_total_invalid", found=direct_weight_total)
        )
    if round(indirect_weight_total, WEIGHT_TOTAL_ROUND_DIGITS) != WEIGHT_TOTAL_EXPECTED:
        raise ValidationError(
            t("instructor.validation.assessment_indirect_total_invalid", found=indirect_weight_total)
        )

    return component_config


def _co_tokens(value: Any) -> list[int]:
    if value is None:
        return []
    if isinstance(value, bool):
        return []
    if isinstance(value, int):
        return [value]
    if isinstance(value, float):
        return [int(value)] if value.is_integer() else []

    token = str(value).strip()
    if not token:
        return []

    numbers: list[int] = []
    for item in token.split(","):
        part = item.strip()
        if not part:
            return []
        match = re.fullmatch(r"(?:co)?\s*(\d+)", part, flags=re.IGNORECASE)
        if not match:
            return []
        numbers.append(int(match.group(1)))
    return numbers


def _validate_question_map(
    worksheet: Any,
    component_config: dict[str, dict[str, Any]],
    total_outcomes: int,
) -> None:
    expected_headers = list(QUESTION_MAP_HEADERS)
    header_map = _header_index_map(worksheet, expected_headers)
    rows = _iter_data_rows(worksheet, len(expected_headers))

    if not rows:
        raise ValidationError(t("instructor.validation.question_map_row_required_one"))

    seen_co_wise_questions: set[tuple[str, str]] = set()
    for row_number, row in enumerate(rows, start=2):
        component_raw = row[header_map["component"]]
        component_key = normalize(component_raw)
        if not component_key:
            raise ValidationError(t("instructor.validation.question_component_required", row=row_number))
        if component_key not in component_config:
            raise ValidationError(
                t(
                    "instructor.validation.question_component_unknown",
                    row=row_number,
                    component=component_raw,
                )
            )

        question_raw = row[header_map["q_no/rubric_parameter"]]
        question_key = normalize(question_raw)
        if not question_key:
            raise ValidationError(
                t("instructor.validation.question_label_required", row=row_number)
            )

        max_marks = coerce_excel_number(row[header_map["max_marks"]])
        if isinstance(max_marks, bool) or not isinstance(max_marks, (int, float)):
            raise ValidationError(t("instructor.validation.question_max_marks_numeric", row=row_number))
        if float(max_marks) <= 0:
            raise ValidationError(
                t("instructor.validation.question_max_marks_positive", row=row_number)
            )

        co_values = _co_tokens(row[header_map["co"]])
        if not co_values:
            raise ValidationError(t("instructor.validation.question_co_required", row=row_number))
        if len(set(co_values)) != len(co_values):
            raise ValidationError(t("instructor.validation.question_co_no_repeat", row=row_number))
        if any(co_number <= 0 or co_number > total_outcomes for co_number in co_values):
            raise ValidationError(
                t(
                    "instructor.validation.question_co_out_of_range",
                    row=row_number,
                    total_outcomes=total_outcomes,
                )
            )

        is_co_wise = bool(component_config[component_key]["co_wise_breakup"])
        if is_co_wise:
            if len(co_values) != 1:
                raise ValidationError(
                    t(
                        "instructor.validation.question_co_wise_requires_one",
                        row=row_number,
                        component=component_raw,
                    )
                )
            question_id = (component_key, question_key)
            if question_id in seen_co_wise_questions:
                raise ValidationError(
                    t(
                        "instructor.validation.question_duplicate_for_component",
                        row=row_number,
                        question=question_raw,
                        component=component_raw,
                    )
                )
            seen_co_wise_questions.add(question_id)


def _validate_students(worksheet: Any) -> None:
    expected_headers = list(STUDENTS_HEADERS)
    header_map = _header_index_map(worksheet, expected_headers)
    rows = _iter_data_rows(worksheet, len(expected_headers))

    if not rows:
        raise ValidationError(t("instructor.validation.students_row_required_one"))

    seen_reg_numbers: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        reg_no_raw = row[header_map["reg_no"]]
        student_name_raw = row[header_map["student_name"]]

        reg_no = str(reg_no_raw).strip() if reg_no_raw is not None else ""
        student_name = str(student_name_raw).strip() if student_name_raw is not None else ""

        if not reg_no or not student_name:
            raise ValidationError(
                t("instructor.validation.students_reg_and_name_required", row=row_number)
            )

        reg_key = normalize(reg_no)
        if reg_key in seen_reg_numbers:
            raise ValidationError(
                t("instructor.validation.students_duplicate_reg_no", row=row_number, reg_no=reg_no)
            )
        seen_reg_numbers.add(reg_key)
