"""Generator for the Course Details workbook template."""

from __future__ import annotations

import logging
import os
import json
import re
from pathlib import Path
from typing import Any, Sequence
from uuid import uuid4

from common.constants import (
    ALLOW_FILTER,
    ALLOW_SELECT_LOCKED,
    ALLOW_SELECT_UNLOCKED,
    ALLOW_SORT,
    ASSESSMENT_CONFIG_HEADERS,
    ASSESSMENT_CONFIG_SHEET,
    ASSESSMENT_VALIDATION_YES_NO_OPTIONS,
    COURSE_METADATA_HEADERS,
    COURSE_METADATA_SHEET,
    HEADER_PATTERNFILL_COLOR,
    ID_COURSE_SETUP,
    LIKERT_MAX,
    LIKERT_MIN,
    MARKS_ENTRY_CO_MARKS_LABEL_PREFIX,
    MARKS_ENTRY_CO_PREFIX,
    MARKS_ENTRY_QUESTION_PREFIX,
    MARKS_ENTRY_ROW_HEADERS,
    MARKS_ENTRY_TOTAL_LABEL,
    MARKS_ENTRY_VALIDATION_ERROR_TITLE,
    MAX_EXCEL_SHEETNAME_LENGTH,
    MIN_MARK_VALUE,
    QUESTION_MAP_HEADERS,
    QUESTION_MAP_SHEET,
    STUDENTS_HEADERS,
    STUDENTS_SHEET,
    SYSTEM_HASH_SHEET,
    SYSTEM_HASH_TEMPLATE_HASH_HEADER,
    SYSTEM_HASH_TEMPLATE_HASH_KEY,
    SYSTEM_HASH_TEMPLATE_ID_HEADER,
    SYSTEM_HASH_TEMPLATE_ID_KEY,
    SYSTEM_LAYOUT_MANIFEST_HASH_HEADER,
    SYSTEM_LAYOUT_MANIFEST_HEADER,
    SYSTEM_LAYOUT_SHEET,
    WORKBOOK_PASSWORD,
)
from common.exceptions import AppSystemError, JobCancelledError, ValidationError
from common.jobs import CancellationToken
from common.registry import BLUEPRINT_REGISTRY
from common.sample_setup_data import SAMPLE_SETUP_DATA
from common.sheet_schema import ValidationRule, WorkbookBlueprint
from common.texts import t
from common.utils import coerce_excel_number, normalize
from common.workbook_signing import sign_payload, verify_payload_signature
from modules.instructor.template_versions import course_setup_v1

_logger = logging.getLogger(__name__)
_YES_NO_TOKENS = {normalize(option) for option in ASSESSMENT_VALIDATION_YES_NO_OPTIONS}
_YES_TOKEN = normalize(ASSESSMENT_VALIDATION_YES_NO_OPTIONS[0])
_AUTO_FIT_SAMPLE_ROWS = 6
_AUTO_FIT_PADDING = 2
_AUTO_FIT_MIN_WIDTH = 8
_AUTO_FIT_MAX_WIDTH = 60
_PAGE_MIN_MARGIN_IN = 0.25
_COMPONENT_NAME_LABEL = "Component name"
_CO_LABEL = "CO"
_MAX_LABEL = "Max."


def _ve(message: str, *, code: str, **context: object) -> ValidationError:
    return ValidationError(message, code=code, context=context)


def generate_course_details_template(
    output_path: str | Path,
    template_id: str = ID_COURSE_SETUP,
    *,
    cancel_token: CancellationToken | None = None,
) -> Path:
    """Generate and save the course details template workbook with atomic replace."""
    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise _ve(
            t("instructor.validation.xlsxwriter_missing"),
            code="XLSXWRITER_MISSING",
        ) from exc

    blueprint = _get_blueprint(template_id)
    sample_data = SAMPLE_SETUP_DATA
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    temp_output = output.with_name(f"{output.name}.{uuid4().hex}.tmp")
    workbook = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    workbook_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal workbook_closed
        if not workbook_closed:
            try:
                workbook.close()
                workbook_closed = True
            except Exception:
                _logger.debug("Suppressing workbook close error during cleanup.", exc_info=True)
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp template file: %s", temp_output)

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        header_format = _build_header_format(workbook, blueprint.style_registry.get("header", {}))
        body_format = _build_body_format(workbook, blueprint.style_registry.get("body", {}))

        for sheet_schema in blueprint.sheets:
            if cancel_token is not None:
                cancel_token.raise_if_cancelled()
            if len(sheet_schema.header_matrix) != 1:
                raise ValidationError(
                    t(
                        "instructor.validation.sheet_single_header_row",
                        sheet_name=sheet_schema.name,
                    )
                )

            worksheet = generate_worksheet(
                workbook=workbook,
                sheet_name=sheet_schema.name,
                headers=sheet_schema.header_matrix[0],
                data=sample_data.get(sheet_schema.name, []),
                header_format=header_format,
                body_format=body_format,
            )

            for validation in sheet_schema.validations:
                if cancel_token is not None:
                    cancel_token.raise_if_cancelled()
                _apply_validation(worksheet, validation)

            if sheet_schema.is_protected:
                _protect_sheet(worksheet)

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        _add_system_hash_sheet(workbook, template_id)

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        workbook.close()
        workbook_closed = True
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except JobCancelledError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate course details template. template_id=%s output=%s",
            template_id,
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc

    return output


def generate_marks_template_from_course_details(
    course_details_path: str | Path,
    output_path: str | Path,
    *,
    cancel_token: CancellationToken | None = None,
) -> Path:
    """Generate marks-entry workbook from a validated course-details workbook."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise _ve(
            t("instructor.validation.openpyxl_missing"),
            code="OPENPYXL_MISSING",
        ) from exc

    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise _ve(
            t("instructor.validation.xlsxwriter_missing"),
            code="XLSXWRITER_MISSING",
        ) from exc

    source_file = Path(course_details_path)
    if not source_file.exists():
        raise _ve(
            t("instructor.validation.workbook_not_found", workbook=source_file),
            code="WORKBOOK_NOT_FOUND",
            workbook=str(source_file),
        )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output.with_name(f"{output.name}.{uuid4().hex}.tmp")

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        workbook = openpyxl.load_workbook(source_file, data_only=True)
    except JobCancelledError:
        raise
    except Exception as exc:
        raise _ve(
            t("instructor.validation.workbook_open_failed", workbook=source_file),
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(source_file),
        ) from exc

    target = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    target_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal target_closed
        if not target_closed:
            try:
                target.close()
                target_closed = True
            except Exception:
                _logger.debug("Suppressing target close error during cleanup.", exc_info=True)
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp marks template file: %s", temp_output)

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        template_id = _extract_and_validate_template_id(workbook)
        context = _extract_marks_template_context_by_template(workbook, template_id)
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        layout_manifest = _write_marks_template_workbook_by_template(
            target,
            context,
            template_id=template_id,
            cancel_token=cancel_token,
        )
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        _copy_system_hash_sheet(workbook, target)
        _add_system_layout_sheet(target, layout_manifest)
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        target.close()
        target_closed = True
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except JobCancelledError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate marks template. source=%s output=%s",
            source_file,
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc
    finally:
        workbook.close()

    return output


def _extract_marks_template_context(workbook: Any) -> dict[str, Any]:
    metadata_rows = _iter_data_rows(workbook[COURSE_METADATA_SHEET], len(COURSE_METADATA_HEADERS))
    assessment_rows = _iter_data_rows(workbook[ASSESSMENT_CONFIG_SHEET], len(ASSESSMENT_CONFIG_HEADERS))
    question_rows = _iter_data_rows(workbook[QUESTION_MAP_SHEET], len(QUESTION_MAP_HEADERS))
    student_rows = _iter_data_rows(workbook[STUDENTS_SHEET], len(STUDENTS_HEADERS))

    total_outcomes = _extract_total_outcomes(metadata_rows)
    students = _extract_students(student_rows)
    components = _extract_components(assessment_rows)
    questions_by_component = _extract_questions(question_rows)

    return {
        "metadata_rows": metadata_rows,
        "assessment_rows": assessment_rows,
        "total_outcomes": total_outcomes,
        "students": students,
        "components": components,
        "questions_by_component": questions_by_component,
    }


def _extract_marks_template_context_by_template(workbook: Any, template_id: str) -> dict[str, Any]:
    extractor = _template_context_extractors().get(template_id)
    if extractor is None:
        raise ValidationError(t("instructor.validation.validator_missing", template_id=template_id))
    return extractor(workbook)


def _write_marks_template_workbook_by_template(
    workbook: Any,
    context: dict[str, Any],
    *,
    template_id: str,
    cancel_token: CancellationToken | None = None,
) -> dict[str, Any]:
    writer = _template_marks_writers().get(template_id)
    if writer is None:
        raise ValidationError(t("instructor.validation.validator_missing", template_id=template_id))
    return writer(workbook, context, template_id=template_id, cancel_token=cancel_token)


def _extract_total_outcomes(metadata_rows: Sequence[Sequence[Any]]) -> int:
    for row in metadata_rows:
        if len(row) < 2:
            continue
        if normalize(row[0]) == "total_outcomes":
            value = coerce_excel_number(row[1])
            if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
                break
            return value
    raise ValidationError(t("instructor.validation.course_metadata_total_outcomes_invalid"))


def _extract_students(student_rows: Sequence[Sequence[Any]]) -> list[tuple[str, str]]:
    students: list[tuple[str, str]] = []
    for row in student_rows:
        reg_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if reg_no and name:
            students.append((reg_no, name))
    if not students:
        raise ValidationError(t("instructor.validation.students_row_required_one"))
    return students


def _extract_components(assessment_rows: Sequence[Sequence[Any]]) -> list[dict[str, Any]]:
    components: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in assessment_rows:
        if len(row) < 5:
            continue
        component_name = str(row[0]).strip()
        component_key = normalize(component_name)
        if not component_key or component_key in seen:
            continue
        seen.add(component_key)
        components.append(
            {
                "key": component_key,
                "display_name": component_name,
                "co_wise_breakup": _parse_yes_no(
                    row[3],
                    ASSESSMENT_CONFIG_SHEET,
                    0,
                    ASSESSMENT_CONFIG_HEADERS[3],
                ),
                "is_direct": _parse_yes_no(
                    row[4],
                    ASSESSMENT_CONFIG_SHEET,
                    0,
                    ASSESSMENT_CONFIG_HEADERS[4],
                ),
            }
        )
    if not components:
        raise ValidationError(t("instructor.validation.assessment_component_required_one"))
    return components


def _extract_questions(question_rows: Sequence[Sequence[Any]]) -> dict[str, list[dict[str, Any]]]:
    questions_by_component: dict[str, list[dict[str, Any]]] = {}
    for row in question_rows:
        if len(row) < 4:
            continue
        component_key = normalize(row[0])
        if not component_key:
            continue
        max_marks = coerce_excel_number(row[2])
        if isinstance(max_marks, bool) or not isinstance(max_marks, (int, float)):
            continue
        co_values = _co_tokens(row[3])
        if not co_values:
            continue
        questions_by_component.setdefault(component_key, []).append(
            {
                "max_marks": float(max_marks),
                "co_values": co_values,
            }
        )
    return questions_by_component


def _write_marks_template_workbook(
    workbook: Any,
    context: dict[str, Any],
    *,
    template_id: str = ID_COURSE_SETUP,
    cancel_token: CancellationToken | None = None,
) -> dict[str, Any]:
    setup_blueprint = _get_blueprint(template_id)
    header_fmt = _build_header_format(workbook, setup_blueprint.style_registry.get("header", {}))
    body_fmt = workbook.add_format({"border": 1})
    wrapped_body_fmt = workbook.add_format({"border": 1, "text_wrap": True})
    wrapped_column_fmt = workbook.add_format({"text_wrap": True})
    num_fmt = workbook.add_format({"border": 1, "num_format": "0.00"})
    header_num_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": str(setup_blueprint.style_registry.get("header", {}).get("bg_color", HEADER_PATTERNFILL_COLOR)),
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "num_format": "0.00",
        }
    )
    unlocked_body_fmt = workbook.add_format({"border": 1, "locked": False})

    layout_sheets, component_plans = _precompute_marks_layout_manifest(
        context=context,
        cancel_token=cancel_token,
    )

    students = context["students"]
    _write_two_column_copy_sheet(
        workbook=workbook,
        title=COURSE_METADATA_SHEET,
        header=COURSE_METADATA_HEADERS,
        rows=context["metadata_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
    )
    _write_multi_column_copy_sheet(
        workbook=workbook,
        title=ASSESSMENT_CONFIG_SHEET,
        header=ASSESSMENT_CONFIG_HEADERS,
        rows=context["assessment_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
        num_fmt=num_fmt,
    )

    for plan in component_plans:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        if plan["kind"] == "direct_co_wise":
            _write_direct_co_wise_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                plan["questions"],
                header_fmt,
                body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
                num_fmt,
                header_num_fmt,
                unlocked_body_fmt,
            )
        elif plan["kind"] == "direct_non_co_wise":
            _write_direct_non_co_wise_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                plan["questions"],
                header_fmt,
                body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
                num_fmt,
                header_num_fmt,
                unlocked_body_fmt,
            )
        else:
            _write_indirect_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                context["total_outcomes"],
                header_fmt,
                body_fmt,
                unlocked_body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
            )

    return {
        "schema_version": 1,
        "sheet_order": [entry["name"] for entry in layout_sheets] + [SYSTEM_HASH_SHEET, SYSTEM_LAYOUT_SHEET],
        "sheets": layout_sheets,
    }


def _precompute_marks_layout_manifest(
    *,
    context: dict[str, Any],
    cancel_token: CancellationToken | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    students = context["students"]
    metadata_rows = context["metadata_rows"]
    assessment_rows = context["assessment_rows"]
    questions_by_component = context["questions_by_component"]
    total_outcomes = context["total_outcomes"]
    student_identity_hash = _student_identity_hash(students)

    layout_sheets: list[dict[str, Any]] = [
        _build_two_column_copy_sheet_spec(
            title=COURSE_METADATA_SHEET,
            header=COURSE_METADATA_HEADERS,
            rows=metadata_rows,
        ),
        _build_multi_column_copy_sheet_spec(
            title=ASSESSMENT_CONFIG_SHEET,
            header=ASSESSMENT_CONFIG_HEADERS,
            rows=assessment_rows,
        ),
    ]
    component_plans: list[dict[str, Any]] = []

    used_sheet_names = {normalize(COURSE_METADATA_SHEET), normalize(ASSESSMENT_CONFIG_SHEET)}
    for component in context["components"]:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        component_name = component["display_name"]
        sheet_name = _safe_sheet_name(component_name, used_sheet_names)
        questions = questions_by_component.get(component["key"], [])
        if component["is_direct"]:
            if not questions:
                raise ValidationError(t("instructor.validation.question_map_row_required_one"))
            if component["co_wise_breakup"]:
                layout_sheets.append(
                    _build_direct_co_wise_sheet_spec(
                        sheet_name=sheet_name,
                        metadata_rows=metadata_rows,
                        component_name=component_name,
                        students=students,
                        questions=questions,
                        student_identity_hash=student_identity_hash,
                    )
                )
                component_plans.append(
                    {
                        "kind": "direct_co_wise",
                        "sheet_name": sheet_name,
                        "component_name": component_name,
                        "questions": questions,
                    }
                )
            else:
                layout_sheets.append(
                    _build_direct_non_co_wise_sheet_spec(
                        sheet_name=sheet_name,
                        metadata_rows=metadata_rows,
                        component_name=component_name,
                        students=students,
                        questions=questions,
                        student_identity_hash=student_identity_hash,
                    )
                )
                component_plans.append(
                    {
                        "kind": "direct_non_co_wise",
                        "sheet_name": sheet_name,
                        "component_name": component_name,
                        "questions": questions,
                    }
                )
        else:
            layout_sheets.append(
                _build_indirect_sheet_spec(
                    sheet_name=sheet_name,
                    metadata_rows=metadata_rows,
                    component_name=component_name,
                    students=students,
                    total_outcomes=total_outcomes,
                    student_identity_hash=student_identity_hash,
                )
            )
            component_plans.append(
                {
                    "kind": "indirect",
                    "sheet_name": sheet_name,
                    "component_name": component_name,
                    "questions": [],
                }
            )

    return layout_sheets, component_plans


def _copy_system_hash_sheet(source_workbook: Any, target_workbook: Any) -> None:
    if SYSTEM_HASH_SHEET not in source_workbook.sheetnames:
        return

    source = source_workbook[SYSTEM_HASH_SHEET]
    target = target_workbook.add_worksheet(SYSTEM_HASH_SHEET)
    target.write_row(0, 0, [source["A1"].value, source["B1"].value])
    target.write_row(1, 0, [source["A2"].value, source["B2"].value])
    target.hide()


def _write_two_column_copy_sheet(
    workbook: Any,
    title: str,
    header: tuple[str, str],
    rows: Sequence[Sequence[Any]],
    header_fmt: Any,
    body_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(title)
    ws.write_row(0, 0, list(header), header_fmt)
    ws.set_column(0, 0, 24)
    ws.set_column(1, 1, 24)
    for row_index, row in enumerate(rows, start=1):
        ws.write(row_index, 0, row[0] if len(row) > 0 else "", body_fmt)
        ws.write(row_index, 1, row[1] if len(row) > 1 else "", body_fmt)
    ws.repeat_rows(0, 0)
    ws.freeze_panes(1, 0)
    ws.set_selection(1, 0, 1, 0)


def _build_two_column_copy_sheet_spec(
    *,
    title: str,
    header: tuple[str, str],
    rows: Sequence[Sequence[Any]],
) -> dict[str, Any]:
    anchors = []
    for row_index, row in enumerate(rows, start=2):
        anchors.append([f"A{row_index}", row[0] if len(row) > 0 else ""])
        anchors.append([f"B{row_index}", row[1] if len(row) > 1 else ""])
    return {
        "name": title,
        "header_row": 1,
        "headers": list(header),
        "anchors": anchors,
        "formula_anchors": [],
    }


def _write_multi_column_copy_sheet(
    workbook: Any,
    title: str,
    header: Sequence[str],
    rows: Sequence[Sequence[Any]],
    header_fmt: Any,
    body_fmt: Any,
    num_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(title)
    ws.write_row(0, 0, list(header), header_fmt)
    for col in range(len(header)):
        ws.set_column(col, col, 22)
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row[: len(header)]):
            cell_fmt = num_fmt if col_index == 1 and isinstance(value, (int, float)) else body_fmt
            ws.write(row_index, col_index, value, cell_fmt)
    ws.repeat_rows(0, 0)
    ws.freeze_panes(1, 0)
    ws.set_selection(1, 0, 1, 0)


def _build_multi_column_copy_sheet_spec(
    *,
    title: str,
    header: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> dict[str, Any]:
    anchors = []
    for row_index, row in enumerate(rows, start=2):
        for col_index, _header in enumerate(header):
            anchors.append(
                [
                    f"{_excel_col_name(col_index)}{row_index}",
                    row[col_index] if col_index < len(row) else "",
                ]
            )
    return {
        "name": title,
        "header_row": 1,
        "headers": list(header),
        "anchors": anchors,
        "formula_anchors": [],
    }


def _build_direct_co_wise_sheet_spec(
    *,
    sheet_name: str,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    students: Sequence[tuple[str, str]],
    questions: Sequence[dict[str, Any]],
    student_identity_hash: str,
) -> dict[str, Any]:
    header_start_row = len(metadata_rows) + 2
    header_row = header_start_row + 1
    question_count = len(questions)
    total_col = 3 + question_count
    row_headers = list(MARKS_ENTRY_ROW_HEADERS)
    question_headers = [f"{MARKS_ENTRY_QUESTION_PREFIX}{idx + 1}" for idx in range(question_count)]
    max_marks_values = [float(question["max_marks"]) for question in questions]
    sheet_headers = row_headers + question_headers + [MARKS_ENTRY_TOTAL_LABEL]

    anchors = _component_metadata_anchor_cells(metadata_rows)
    component_row = len(metadata_rows) + 1
    anchors.extend(
        [
            [f"B{component_row}", _COMPONENT_NAME_LABEL],
            [f"C{component_row}", component_name],
            [f"C{header_row + 1}", _CO_LABEL],
            [f"C{header_row + 2}", _MAX_LABEL],
            [f"{_excel_col_name(total_col)}{header_row}", MARKS_ENTRY_TOTAL_LABEL],
        ]
    )

    formula_anchors: list[list[str]] = []
    if students and question_count > 0:
        first_data_row = header_start_row + 3
        first_mark_col = _excel_col_name(3)
        last_mark_col = _excel_col_name(total_col - 1)
        first_row_formula = f"=SUM({first_mark_col}{first_data_row + 1}:{last_mark_col}{first_data_row + 1})"
        formula_anchors.append([f"{_excel_col_name(total_col)}{first_data_row + 1}", first_row_formula])

    return {
        "name": sheet_name,
        "kind": "direct_co_wise",
        "header_row": header_row,
        "headers": sheet_headers,
        "anchors": anchors,
        "formula_anchors": formula_anchors,
        "student_count": len(students),
        "student_identity_hash": student_identity_hash,
        "mark_structure": {
            "mark_maxima": max_marks_values,
        },
    }


def _build_direct_non_co_wise_sheet_spec(
    *,
    sheet_name: str,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    students: Sequence[tuple[str, str]],
    questions: Sequence[dict[str, Any]],
    student_identity_hash: str,
) -> dict[str, Any]:
    header_start_row = len(metadata_rows) + 2
    header_row = header_start_row + 1
    covered_cos = sorted({co for q in questions for co in q["co_values"]})
    co_mark_headers = [f"{MARKS_ENTRY_CO_MARKS_LABEL_PREFIX}{co}" for co in covered_cos]
    total_max = sum(float(question["max_marks"]) for question in questions)
    max_marks_per_co = _split_equal_with_residual(total_max, max(1, len(covered_cos)))
    mark_maxima = [total_max] + [float(value) for value in max_marks_per_co]
    sheet_headers = list(MARKS_ENTRY_ROW_HEADERS) + [MARKS_ENTRY_TOTAL_LABEL] + co_mark_headers

    anchors = _component_metadata_anchor_cells(metadata_rows)
    component_row = len(metadata_rows) + 1
    anchors.extend(
        [
            [f"B{component_row}", _COMPONENT_NAME_LABEL],
            [f"C{component_row}", component_name],
            [f"C{header_row + 1}", _CO_LABEL],
            [f"C{header_row + 2}", _MAX_LABEL],
            [f"D{header_row}", MARKS_ENTRY_TOTAL_LABEL],
        ]
    )

    formula_anchors: list[list[str]] = []
    if students and covered_cos:
        first_data_row = header_start_row + 3
        first_row = first_data_row + 1
        divisor = len(covered_cos)
        col_name_total = _excel_col_name(3)
        first_co_col_name = _excel_col_name(4) if divisor > 1 else ""
        for idx in range(len(covered_cos)):
            co_col = 4 + idx
            if idx == len(covered_cos) - 1 and len(covered_cos) > 1:
                prev_co_col_name = _excel_col_name(co_col - 1)
                formula = (
                    f'=IF(OR(${col_name_total}{first_row}="A",${col_name_total}{first_row}="a"),'
                    f'"A",IF(${col_name_total}{first_row}="","",${col_name_total}{first_row}-SUM('
                    f"{first_co_col_name}{first_row}:{prev_co_col_name}{first_row})))"
                )
            else:
                formula = (
                    f'=IF(OR(${col_name_total}{first_row}="A",${col_name_total}{first_row}="a"),'
                    f'"A",IF(${col_name_total}{first_row}="","",ROUND(${col_name_total}{first_row}/{divisor},2)))'
                )
            formula_anchors.append([f"{_excel_col_name(co_col)}{first_row}", formula])

    return {
        "name": sheet_name,
        "kind": "direct_non_co_wise",
        "header_row": header_row,
        "headers": sheet_headers,
        "anchors": anchors,
        "formula_anchors": formula_anchors,
        "student_count": len(students),
        "student_identity_hash": student_identity_hash,
        "mark_structure": {
            "mark_maxima": mark_maxima,
        },
    }


def _build_indirect_sheet_spec(
    *,
    sheet_name: str,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    students: Sequence[tuple[str, str]],
    total_outcomes: int,
    student_identity_hash: str,
) -> dict[str, Any]:
    header_start_row = len(metadata_rows) + 2
    header_row = header_start_row + 1
    headers = list(MARKS_ENTRY_ROW_HEADERS) + [
        f"{MARKS_ENTRY_CO_PREFIX}{i}" for i in range(1, total_outcomes + 1)
    ]
    anchors = _component_metadata_anchor_cells(metadata_rows)
    component_row = len(metadata_rows) + 1
    anchors.extend(
        [
            [f"B{component_row}", _COMPONENT_NAME_LABEL],
            [f"C{component_row}", component_name],
        ]
    )
    return {
        "name": sheet_name,
        "kind": "indirect",
        "header_row": header_row,
        "headers": headers,
        "anchors": anchors,
        "formula_anchors": [],
        "student_count": len(students),
        "student_identity_hash": student_identity_hash,
        "mark_structure": {
            "likert_range": [LIKERT_MIN, LIKERT_MAX],
        },
    }


def _write_direct_co_wise_sheet(
    workbook: Any,
    sheet_name: str,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    students: Sequence[tuple[str, str]],
    questions: Sequence[dict[str, Any]],
    header_fmt: Any,
    body_fmt: Any,
    wrapped_body_fmt: Any,
    wrapped_column_fmt: Any,
    num_fmt: Any,
    header_num_fmt: Any,
    unlocked_body_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(sheet_name)
    header_start_row = _write_component_course_metadata(ws, metadata_rows, component_name, body_fmt)
    question_count = len(questions)
    total_col = 3 + question_count
    row_headers = list(MARKS_ENTRY_ROW_HEADERS)
    question_headers = [f"{MARKS_ENTRY_QUESTION_PREFIX}{idx + 1}" for idx in range(question_count)]
    co_labels = [f"{MARKS_ENTRY_CO_PREFIX}{question['co_values'][0]}" for question in questions]
    max_marks_values = [float(question["max_marks"]) for question in questions]
    sheet_headers = row_headers + question_headers + [MARKS_ENTRY_TOTAL_LABEL]

    ws.write_row(header_start_row, 0, row_headers, header_fmt)
    for idx, question_header in enumerate(question_headers):
        ws.write(
            header_start_row,
            3 + idx,
            question_header,
            header_fmt,
        )
    ws.write(header_start_row, total_col, MARKS_ENTRY_TOTAL_LABEL, header_fmt)

    ws.write_row(header_start_row + 1, 0, ["", "", _CO_LABEL], header_fmt)
    for idx, co_label in enumerate(co_labels):
        ws.write(header_start_row + 1, 3 + idx, co_label, header_fmt)
    ws.write(header_start_row + 1, total_col, "", header_fmt)

    ws.write_row(header_start_row + 2, 0, ["", "", _MAX_LABEL], header_fmt)
    component_total = sum(max_marks_values)
    for idx, max_marks in enumerate(max_marks_values):
        ws.write_number(header_start_row + 2, 3 + idx, max_marks, header_num_fmt)
    ws.write_number(header_start_row + 2, total_col, component_total, header_num_fmt)

    first_data_row = header_start_row + 3
    first_mark_col = _excel_col_name(3)
    last_mark_col = _excel_col_name(total_col - 1)
    for row_offset, (reg_no, student_name) in enumerate(students, start=first_data_row):
        ws.write_number(row_offset, 0, row_offset - (first_data_row - 1), body_fmt)
        ws.write(row_offset, 1, reg_no, body_fmt)
        ws.write(row_offset, 2, student_name, wrapped_body_fmt)
        for col in range(3, total_col):
            ws.write_blank(row_offset, col, None, unlocked_body_fmt)
        ws.write_formula(
            row_offset,
            total_col,
            f"=SUM({first_mark_col}{row_offset + 1}:{last_mark_col}{row_offset + 1})",
            num_fmt,
        )

    if students and question_count > 0:
        first_row = first_data_row
        last_row = first_data_row + len(students) - 1
        max_marks_row = header_start_row + 2
        for idx, max_marks_value in enumerate(max_marks_values):
            col_index = 3 + idx
            validation_formula = _build_marks_validation_formula_for_column(
                col_index=col_index,
                first_data_row=first_data_row,
                max_marks_row=max_marks_row,
            )
            ws.data_validation(
                first_row,
                col_index,
                last_row,
                col_index,
                {
                    "validate": "custom",
                    "value": validation_formula,
                    "error_title": MARKS_ENTRY_VALIDATION_ERROR_TITLE,
                    "error_message": _build_marks_validation_error_message(max_marks_value),
                    "ignore_blank": True,
                },
            )

    sample_rows: list[list[Any]] = _component_metadata_sample_rows(metadata_rows, component_name) + [
        sheet_headers,
        ["", "", _CO_LABEL] + co_labels + [""],
        ["", "", _MAX_LABEL] + max_marks_values + [component_total],
    ]
    preview_students = students[: max(0, _AUTO_FIT_SAMPLE_ROWS - len(sample_rows))]
    for row_offset, (reg_no, student_name) in enumerate(preview_students, start=first_data_row):
        sample_rows.append(
            [row_offset - (first_data_row - 1), reg_no, student_name] + [""] * question_count + [""]
        )
    _set_common_student_columns(ws, total_col, sample_rows, wrapped_column_fmt)
    ws.repeat_rows(0, header_start_row + 2)
    ws.freeze_panes(header_start_row + 3, 3)
    ws.set_selection(first_data_row, 3, first_data_row, 3)
    _protect_sheet(ws)


def _write_direct_non_co_wise_sheet(
    workbook: Any,
    sheet_name: str,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    students: Sequence[tuple[str, str]],
    questions: Sequence[dict[str, Any]],
    header_fmt: Any,
    body_fmt: Any,
    wrapped_body_fmt: Any,
    wrapped_column_fmt: Any,
    num_fmt: Any,
    header_num_fmt: Any,
    unlocked_body_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(sheet_name)
    header_start_row = _write_component_course_metadata(ws, metadata_rows, component_name, body_fmt)
    covered_cos = sorted({co for q in questions for co in q["co_values"]})
    co_count = max(1, len(covered_cos))
    total_max = sum(float(question["max_marks"]) for question in questions)
    max_marks_per_co = _split_equal_with_residual(total_max, co_count)
    row_headers = list(MARKS_ENTRY_ROW_HEADERS)
    co_mark_headers = [f"{MARKS_ENTRY_CO_MARKS_LABEL_PREFIX}{co}" for co in covered_cos]
    co_prefix_labels = [f"{MARKS_ENTRY_CO_PREFIX}{co}" for co in covered_cos]
    sheet_headers = row_headers + [MARKS_ENTRY_TOTAL_LABEL] + co_mark_headers

    ws.write_row(header_start_row, 0, row_headers + [MARKS_ENTRY_TOTAL_LABEL], header_fmt)
    for idx, co_header in enumerate(co_mark_headers):
        ws.write(
            header_start_row,
            4 + idx,
            co_header,
            header_fmt,
        )

    ws.write_row(header_start_row + 1, 0, ["", "", _CO_LABEL, ""], header_fmt)
    for idx, co_prefix in enumerate(co_prefix_labels):
        ws.write(header_start_row + 1, 4 + idx, co_prefix, header_fmt)

    ws.write_row(header_start_row + 2, 0, ["", "", _MAX_LABEL, ""], header_fmt)
    ws.write_number(header_start_row + 2, 3, total_max, header_num_fmt)
    for idx, value in enumerate(max_marks_per_co):
        ws.write_number(header_start_row + 2, 4 + idx, value, header_num_fmt)

    first_data_row = header_start_row + 3
    co_total = len(covered_cos)
    col_name_total = _excel_col_name(3)
    divisor = co_total if co_total else 1
    first_co_col_name = _excel_col_name(4) if co_total > 1 else ""
    for row_offset, (reg_no, student_name) in enumerate(students, start=first_data_row):
        ws.write_number(row_offset, 0, row_offset - first_data_row, body_fmt)
        ws.write(row_offset, 1, reg_no, body_fmt)
        ws.write(row_offset, 2, student_name, wrapped_body_fmt)
        ws.write_blank(row_offset, 3, None, unlocked_body_fmt)
        for idx in range(co_total):
            co_col = 4 + idx
            if idx == co_total - 1 and co_total > 1:
                prev_co_col_name = _excel_col_name(co_col - 1)
                formula = (
                    f'=IF(OR(${col_name_total}{row_offset + 1}="A",${col_name_total}{row_offset + 1}="a"),'
                    f'"A",IF(${col_name_total}{row_offset + 1}="","",'
                    f'${col_name_total}{row_offset + 1}-SUM({first_co_col_name}{row_offset + 1}:'
                    f'{prev_co_col_name}{row_offset + 1})))'
                )
            else:
                formula = (
                    f'=IF(OR(${col_name_total}{row_offset + 1}="A",${col_name_total}{row_offset + 1}="a"),'
                    f'"A",IF(${col_name_total}{row_offset + 1}="","",ROUND(${col_name_total}{row_offset + 1}/'
                    f"{divisor},2)))"
                )
            ws.write_formula(
                row_offset,
                co_col,
                formula,
                num_fmt,
            )

    if students:
        first_row = first_data_row
        last_row = first_data_row + len(students) - 1
        validation_formula = _build_marks_validation_formula_for_column(
            col_index=3,
            first_data_row=first_data_row,
            max_marks_row=header_start_row + 2,
        )
        ws.data_validation(
            first_row,
            3,
            last_row,
            3,
            {
                "validate": "custom",
                "value": validation_formula,
                "error_title": MARKS_ENTRY_VALIDATION_ERROR_TITLE,
                "error_message": _build_marks_validation_error_message(total_max),
                "ignore_blank": True,
            },
        )

    sample_rows: list[list[Any]] = _component_metadata_sample_rows(metadata_rows, component_name) + [
        sheet_headers,
        ["", "", _CO_LABEL, ""] + co_prefix_labels,
        ["", "", _MAX_LABEL, total_max] + max_marks_per_co,
    ]
    preview_students = students[: max(0, _AUTO_FIT_SAMPLE_ROWS - len(sample_rows))]
    for row_offset, (reg_no, student_name) in enumerate(preview_students, start=first_data_row):
        sample_rows.append([row_offset - first_data_row, reg_no, student_name, ""] + [""] * len(covered_cos))
    _set_common_student_columns(ws, 3 + len(covered_cos), sample_rows, wrapped_column_fmt)
    ws.repeat_rows(0, header_start_row + 2)
    ws.freeze_panes(header_start_row + 3, 3)
    ws.set_selection(first_data_row, 3, first_data_row, 3)
    _protect_sheet(ws)


def _write_indirect_sheet(
    workbook: Any,
    sheet_name: str,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    students: Sequence[tuple[str, str]],
    total_outcomes: int,
    header_fmt: Any,
    body_fmt: Any,
    unlocked_body_fmt: Any,
    wrapped_body_fmt: Any,
    wrapped_column_fmt: Any,
) -> None:
    ws = workbook.add_worksheet(sheet_name)
    header_start_row = _write_component_course_metadata(ws, metadata_rows, component_name, body_fmt)
    headers = list(MARKS_ENTRY_ROW_HEADERS) + [
        f"{MARKS_ENTRY_CO_PREFIX}{i}" for i in range(1, total_outcomes + 1)
    ]
    ws.write_row(header_start_row, 0, headers, header_fmt)

    first_data_row = header_start_row + 1
    for row_offset, (reg_no, student_name) in enumerate(students, start=first_data_row):
        ws.write_number(row_offset, 0, row_offset - header_start_row, body_fmt)
        ws.write(row_offset, 1, reg_no, body_fmt)
        ws.write(row_offset, 2, student_name, wrapped_body_fmt)
        for col in range(3, 3 + total_outcomes):
            ws.write_blank(row_offset, col, None, unlocked_body_fmt)

    if students and total_outcomes > 0:
        first_row = first_data_row
        last_row = first_data_row + len(students) - 1
        ws.data_validation(
            first_row,
            3,
            last_row,
            2 + total_outcomes,
            {
                "validate": "custom",
                "value": (
                    f'=OR(D{first_data_row + 1}="A",D{first_data_row + 1}="a",'
                    f'AND(ISNUMBER(D{first_data_row + 1}),D{first_data_row + 1}>={MIN_MARK_VALUE},'
                    f'D{first_data_row + 1}>={LIKERT_MIN},D{first_data_row + 1}<={LIKERT_MAX}))'
                ),
                "error_title": MARKS_ENTRY_VALIDATION_ERROR_TITLE,
                "error_message": (
                    f"Enter A/a or a numeric Likert value between {LIKERT_MIN} and {LIKERT_MAX}."
                ),
                "ignore_blank": True,
            },
        )

    sample_rows: list[list[Any]] = _component_metadata_sample_rows(metadata_rows, component_name) + [headers]
    preview_students = students[: max(0, _AUTO_FIT_SAMPLE_ROWS - len(sample_rows))]
    for row_index, (reg_no, student_name) in enumerate(preview_students, start=1):
        sample_rows.append([row_index, reg_no, student_name] + [""] * total_outcomes)
    _set_common_student_columns(ws, 2 + total_outcomes, sample_rows, wrapped_column_fmt)
    ws.repeat_rows(0, header_start_row)
    ws.freeze_panes(header_start_row + 1, 3)
    ws.set_selection(first_data_row, 3, first_data_row, 3)
    _protect_sheet(ws)


def _write_component_course_metadata(
    ws: Any,
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
    body_fmt: Any,
) -> int:
    for row_index, row in enumerate(metadata_rows):
        ws.write(row_index, 1, row[0] if len(row) > 0 else "", body_fmt)
        ws.write(row_index, 2, row[1] if len(row) > 1 else "", body_fmt)
    component_row = len(metadata_rows)
    ws.write(component_row, 1, _COMPONENT_NAME_LABEL, body_fmt)
    ws.write(component_row, 2, component_name, body_fmt)
    return len(metadata_rows) + 2


def _component_metadata_sample_rows(
    metadata_rows: Sequence[Sequence[Any]],
    component_name: str,
) -> list[list[Any]]:
    sample_rows: list[list[Any]] = []
    for row in metadata_rows:
        sample_rows.append(["", row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else ""])
    sample_rows.append(["", _COMPONENT_NAME_LABEL, component_name])
    return sample_rows


def _component_metadata_anchor_cells(metadata_rows: Sequence[Sequence[Any]]) -> list[list[Any]]:
    anchors: list[list[Any]] = []
    for row_index, row in enumerate(metadata_rows, start=1):
        anchors.append([f"B{row_index}", row[0] if len(row) > 0 else ""])
        anchors.append([f"C{row_index}", row[1] if len(row) > 1 else ""])
    return anchors


def _student_identity_hash(students: Sequence[tuple[str, str]]) -> str:
    # Stable signature of ordered student identities copied from course details.
    payload = "\n".join(f"{reg_no.strip()}|{student_name.strip()}" for reg_no, student_name in students)
    return sign_payload(payload)


def _build_marks_validation_formula_for_column(
    col_index: int,
    first_data_row: int,
    max_marks_row: int,
) -> str:
    col_name = _excel_col_name(col_index)
    excel_data_row = first_data_row + 1
    excel_max_row = max_marks_row + 1
    return (
        f'=OR({col_name}{excel_data_row}="A",{col_name}{excel_data_row}="a",'
        f"AND(ISNUMBER({col_name}{excel_data_row}),{col_name}{excel_data_row}>={MIN_MARK_VALUE},"
        f"{col_name}{excel_data_row}<={col_name}${excel_max_row}))"
    )


def _build_marks_validation_error_message(max_marks_value: Any) -> str:
    coerced_max = coerce_excel_number(max_marks_value)
    if isinstance(coerced_max, bool) or not isinstance(coerced_max, (int, float)):
        max_value_text = str(max_marks_value).strip()
    else:
        max_value_text = f"{coerced_max:g}"
    return (
        f"Enter A/a or a numeric mark between {MIN_MARK_VALUE:g} and {max_value_text}."
    )


def _set_common_student_columns(
    ws: Any,
    last_col: int,
    sample_rows: Sequence[Sequence[Any]],
    wrapped_c_column_format: Any,
) -> None:
    ws.set_paper(9)  # A4
    ws.set_landscape()
    ws.set_margins(_PAGE_MIN_MARGIN_IN, _PAGE_MIN_MARGIN_IN, _PAGE_MIN_MARGIN_IN, _PAGE_MIN_MARGIN_IN)
    ws.fit_to_pages(1, 0)

    widths = _compute_sampled_column_widths(sample_rows, last_col)

    for col in range(0, last_col + 1):
        width = widths.get(col, _AUTO_FIT_MIN_WIDTH)
        if col == 2:
            ws.set_column(col, col, width, wrapped_c_column_format)
        else:
            ws.set_column(col, col, width)


def _compute_sampled_column_widths(
    sample_rows: Sequence[Sequence[Any]],
    last_col: int,
) -> dict[int, int]:
    widths: dict[int, int] = {}
    for col_index in range(last_col + 1):
        max_len = 0
        for row in sample_rows:
            if col_index >= len(row):
                continue
            value = row[col_index]
            if value is None:
                continue
            max_len = max(max_len, len(str(value).strip()))
        widths[col_index] = min(
            _AUTO_FIT_MAX_WIDTH,
            max(_AUTO_FIT_MIN_WIDTH, max_len + _AUTO_FIT_PADDING),
        )
    return widths


def _excel_col_name(col_index: int) -> str:
    index = col_index + 1
    label = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        label = chr(65 + rem) + label
    return label


def _split_equal_with_residual(total: float, parts: int) -> list[float]:
    if parts <= 0:
        return []
    if parts == 1:
        return [round(total, 2)]
    base = round(total / parts, 2)
    values = [base] * parts
    values[-1] = round(total - sum(values[:-1]), 2)
    return values


def _safe_sheet_name(name: str, used_sheet_names: set[str]) -> str:
    token = re.sub(r"[:\\/?*\[\]]", "_", name).strip() or "Component"
    token = token[:MAX_EXCEL_SHEETNAME_LENGTH]
    base_key = normalize(token)
    if base_key not in used_sheet_names:
        used_sheet_names.add(base_key)
        return token

    counter = 2
    while True:
        suffix = f"_{counter}"
        trimmed = token[: max(1, MAX_EXCEL_SHEETNAME_LENGTH - len(suffix))]
        candidate = f"{trimmed}{suffix}"
        key = normalize(candidate)
        if key not in used_sheet_names:
            used_sheet_names.add(key)
            return candidate
        counter += 1


def _get_blueprint(template_id: str) -> WorkbookBlueprint:
    blueprint = BLUEPRINT_REGISTRY.get(template_id)
    if blueprint is None:
        available = ", ".join(sorted(BLUEPRINT_REGISTRY))
        raise _ve(
            t(
                "instructor.validation.unknown_template",
                template_id=template_id,
                available=available,
            ),
            code="UNKNOWN_TEMPLATE",
            template_id=template_id,
            available=available,
        )
    return blueprint


def _build_header_format(workbook: Any, header_style: dict[str, Any]) -> Any:
    bg_color = str(header_style.get("bg_color", HEADER_PATTERNFILL_COLOR))
    return workbook.add_format(
        {
            "bold": bool(header_style.get("bold", True)),
            "bg_color": bg_color,
            "border": int(header_style.get("border", 1)),
            "align": str(header_style.get("align", "center")),
            "valign": str(header_style.get("valign", "vcenter")),
        }
    )


def _build_body_format(workbook: Any, body_style: dict[str, Any]) -> Any:
    return workbook.add_format(
        {
            "locked": bool(body_style.get("locked", False)),
            "border": int(body_style.get("border", 1)),
        }
    )


def generate_worksheet(
    workbook: Any,
    sheet_name: str,
    headers: Sequence[str],
    data: Sequence[Sequence[Any]],
    header_format: Any,
    body_format: Any,
) -> Any:
    """Create a worksheet with strict validation and efficient row writes."""
    if not sheet_name or not isinstance(sheet_name, str):
        raise ValidationError(t("instructor.validation.invalid_sheet_name"))

    if not headers:
        raise ValidationError(t("instructor.validation.headers_empty", sheet_name=sheet_name))

    if len(set(headers)) != len(headers):
        raise ValidationError(t("instructor.validation.headers_unique", sheet_name=sheet_name))

    column_count = len(headers)
    for row_index, row in enumerate(data, start=1):
        if len(row) != column_count:
            raise ValidationError(
                t(
                    "instructor.validation.row_length_mismatch",
                    row=row_index,
                    sheet_name=sheet_name,
                    expected=column_count,
                    found=len(row),
                )
            )

    worksheet = workbook.add_worksheet(sheet_name)
    col_widths: dict[int, int] = {}
    write_row = worksheet.write_row
    write_row(0, 0, headers, header_format)
    for col_index, value in enumerate(headers):
        col_widths[col_index] = max(12, len(str(value)) + 2)

    for col_index, width in col_widths.items():
        worksheet.set_column(col_index, col_index, width)

    for row_offset, row_values in enumerate(data, start=1):
        write_row(row_offset, 0, row_values, body_format)

    worksheet.freeze_panes(1, 0)
    return worksheet


def _apply_validation(worksheet: Any, rule: ValidationRule) -> None:
    options = dict(rule.options)
    validation_type = options.pop("validate", None)
    if not validation_type:
        return

    options["validate"] = validation_type
    if "ignore_blank" not in options:
        options["ignore_blank"] = True

    worksheet.data_validation(
        rule.first_row,
        rule.first_col,
        rule.last_row,
        rule.last_col,
        options,
    )


def _protect_sheet(worksheet: Any) -> None:
    # Keep locked-cell selection disabled and unlocked-cell selection enabled so
    # keyboard navigation (Tab) jumps between mark-entry cells.
    worksheet.protect(
        WORKBOOK_PASSWORD,
        {
            "sort": ALLOW_SORT,
            "autofilter": ALLOW_FILTER,
            "select_locked_cells": ALLOW_SELECT_LOCKED,
            "select_unlocked_cells": ALLOW_SELECT_UNLOCKED,
        },
    )


def _add_system_hash_sheet(workbook: Any, template_id: str) -> None:
    worksheet = workbook.add_worksheet(SYSTEM_HASH_SHEET)
    template_hash = _compute_template_hash(template_id)

    worksheet.write_row(
        0,
        0,
        [SYSTEM_HASH_TEMPLATE_ID_HEADER, SYSTEM_HASH_TEMPLATE_HASH_HEADER],
    )
    worksheet.write_row(1, 0, [template_id, template_hash])
    worksheet.hide()


def _add_system_layout_sheet(workbook: Any, layout_manifest: dict[str, Any]) -> None:
    worksheet = workbook.add_worksheet(SYSTEM_LAYOUT_SHEET)
    manifest_text = _serialize_layout_manifest(layout_manifest)
    manifest_hash = _compute_layout_manifest_hash(manifest_text)
    worksheet.write_row(
        0,
        0,
        [SYSTEM_LAYOUT_MANIFEST_HEADER, SYSTEM_LAYOUT_MANIFEST_HASH_HEADER],
    )
    worksheet.write_row(1, 0, [manifest_text, manifest_hash])
    worksheet.hide()


def _compute_template_hash(template_id: str) -> str:
    return sign_payload(template_id)


def _serialize_layout_manifest(layout_manifest: dict[str, Any]) -> str:
    return json.dumps(layout_manifest, sort_keys=True, separators=(",", ":"), ensure_ascii=True)


def _compute_layout_manifest_hash(manifest_text: str) -> str:
    return sign_payload(manifest_text)


def validate_course_details_workbook(workbook_path: str | Path) -> str:
    """Validate uploaded course details workbook and return template id."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise _ve(
            t("instructor.validation.openpyxl_missing"),
            code="OPENPYXL_MISSING",
        ) from exc

    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise _ve(
            t("instructor.validation.workbook_not_found", workbook=workbook_file),
            code="WORKBOOK_NOT_FOUND",
            workbook=str(workbook_file),
        )

    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=True)
    except Exception as exc:
        raise _ve(
            t("instructor.validation.workbook_open_failed", workbook=workbook_file),
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(workbook_file),
        ) from exc

    try:
        template_id = _extract_and_validate_template_id(workbook)
        blueprint = _get_blueprint(template_id)
        _validate_workbook_schema(workbook, blueprint)
        _validate_template_specific_rules(workbook, template_id)
    finally:
        workbook.close()

    return template_id


def _extract_and_validate_template_id(workbook: Any) -> str:
    if SYSTEM_HASH_SHEET not in workbook.sheetnames:
        raise ValidationError(
            t("instructor.validation.system_sheet_missing", sheet=SYSTEM_HASH_SHEET)
        )

    hash_sheet = workbook[SYSTEM_HASH_SHEET]
    if normalize(hash_sheet["A1"].value) != normalize(SYSTEM_HASH_TEMPLATE_ID_KEY):
        raise ValidationError(t("instructor.validation.system_hash_missing_template_id_header"))
    if normalize(hash_sheet["B1"].value) != normalize(SYSTEM_HASH_TEMPLATE_HASH_KEY):
        raise ValidationError(t("instructor.validation.system_hash_missing_template_hash_header"))

    template_id = str(hash_sheet["A2"].value).strip() if hash_sheet["A2"].value is not None else ""
    template_hash = (
        str(hash_sheet["B2"].value).strip() if hash_sheet["B2"].value is not None else ""
    )
    if not template_id:
        raise ValidationError(t("instructor.validation.system_hash_template_id_missing"))
    if not verify_payload_signature(template_id, template_hash):
        raise ValidationError(t("instructor.validation.system_hash_mismatch"))
    return template_id


def _validate_workbook_schema(workbook: Any, blueprint: WorkbookBlueprint) -> None:
    expected_sheet_names = [schema.name for schema in blueprint.sheets] + [SYSTEM_HASH_SHEET]
    actual_sheet_names = list(workbook.sheetnames)
    if actual_sheet_names != expected_sheet_names:
        raise ValidationError(
            t(
                "instructor.validation.workbook_sheet_mismatch",
                template_id=blueprint.type_id,
                expected=expected_sheet_names,
                found=actual_sheet_names,
            )
        )

    for sheet_schema in blueprint.sheets:
        if len(sheet_schema.header_matrix) != 1:
            raise ValidationError(
                t(
                    "instructor.validation.sheet_single_header_row",
                    sheet_name=sheet_schema.name,
                )
            )

        expected_headers = [normalize(h) for h in sheet_schema.header_matrix[0]]
        worksheet = workbook[sheet_schema.name]
        actual_headers = [
            normalize(worksheet.cell(row=1, column=col_index + 1).value)
            for col_index in range(len(expected_headers))
        ]
        if actual_headers != expected_headers:
            raise ValidationError(
                t(
                    "instructor.validation.header_mismatch",
                    sheet_name=sheet_schema.name,
                    expected=sheet_schema.header_matrix[0],
                )
            )


def _iter_data_rows(worksheet: Any, expected_col_count: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=2, max_col=expected_col_count, values_only=True):
        values = list(row)
        if any(normalize(value) != "" for value in values):
            rows.append(values)
    return rows


def _validate_template_specific_rules(workbook: Any, template_id: str) -> None:
    validator = _template_rule_validators().get(template_id)
    if validator is None:
        raise ValidationError(t("instructor.validation.validator_missing", template_id=template_id))
    validator(workbook)


def _template_context_extractors() -> dict[str, Any]:
    # Register per-template context builders here. New template versions can be
    # implemented in separate modules and wired into this dispatch map.
    return {
        ID_COURSE_SETUP: course_setup_v1.extract_marks_template_context,
    }


def _template_marks_writers() -> dict[str, Any]:
    # Register per-template marks workbook writers here.
    return {
        ID_COURSE_SETUP: course_setup_v1.write_marks_template_workbook,
    }


def _template_rule_validators() -> dict[str, Any]:
    # Register per-template course-details validators here.
    return {
        ID_COURSE_SETUP: course_setup_v1.validate_course_details_rules,
    }


def _parse_yes_no(value: Any, sheet_name: str, row_number: int, field_name: str) -> bool:
    token = normalize(value)
    if token not in _YES_NO_TOKENS:
        raise ValidationError(
            t(
                "instructor.validation.yes_no_required",
                sheet_name=sheet_name,
                row=row_number,
                field=field_name,
            )
        )
    return token == _YES_TOKEN


def _co_tokens(value: Any) -> list[int]:
    if value is None:
        return []
    if isinstance(value, bool):
        return []
    if isinstance(value, int):
        return [value]
    if isinstance(value, float):
        return [int(value)] if value.is_integer() else []

    token = str(value).strip()
    if not token:
        return []

    numbers: list[int] = []
    for item in token.split(","):
        part = item.strip()
        if not part:
            return []
        match = re.fullmatch(r"(?:co)?\s*(\d+)", part, flags=re.IGNORECASE)
        if not match:
            return []
        numbers.append(int(match.group(1)))
    return numbers


