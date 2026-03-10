"""Validation logic for uploaded filled-marks workbooks."""

from __future__ import annotations

import json
from pathlib import Path

from common.constants import (
    ID_COURSE_SETUP,
    SYSTEM_HASH_SHEET,
    SYSTEM_HASH_TEMPLATE_HASH_KEY,
    SYSTEM_HASH_TEMPLATE_ID_KEY,
    SYSTEM_LAYOUT_MANIFEST_HASH_KEY,
    SYSTEM_LAYOUT_MANIFEST_KEY,
    SYSTEM_LAYOUT_SHEET,
)
from common.exceptions import ValidationError
from common.texts import t
from common.utils import normalize
from common.workbook_signing import verify_payload_signature
from domain.template_versions import course_setup_v1


def filled_marks_manifest_validators() -> dict[str, object]:
    return {
        ID_COURSE_SETUP: course_setup_v1.validate_filled_marks_manifest_schema,
    }


def validate_filled_marks_manifest_schema_by_template(
    workbook: object,
    manifest: object,
    *,
    template_id: str,
) -> None:
    validator = filled_marks_manifest_validators().get(template_id)
    if validator is None:
        raise ValidationError(
            t("instructor.validation.step3.template_validator_missing", template_id=template_id)
        )
    validator(workbook, manifest)


def validate_uploaded_filled_marks_workbook(workbook_path: str | Path) -> None:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise ValidationError(
            t("instructor.validation.openpyxl_missing"),
            code="OPENPYXL_MISSING",
        ) from exc

    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise ValidationError(
            t("instructor.validation.workbook_not_found", workbook=workbook_file),
            code="WORKBOOK_NOT_FOUND",
            context={"workbook": str(workbook_file)},
        )

    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=False)
    except Exception as exc:
        raise ValidationError(
            t("instructor.validation.workbook_open_failed", workbook=workbook_file),
            code="WORKBOOK_OPEN_FAILED",
            context={"workbook": str(workbook_file)},
        ) from exc

    try:
        if SYSTEM_HASH_SHEET not in workbook.sheetnames:
            raise ValidationError(t("instructor.validation.system_sheet_missing", sheet=SYSTEM_HASH_SHEET))

        hash_sheet = workbook[SYSTEM_HASH_SHEET]
        if normalize(hash_sheet["A1"].value) != normalize(SYSTEM_HASH_TEMPLATE_ID_KEY):
            raise ValidationError(t("instructor.validation.system_hash_missing_template_id_header"))
        if normalize(hash_sheet["B1"].value) != normalize(SYSTEM_HASH_TEMPLATE_HASH_KEY):
            raise ValidationError(t("instructor.validation.system_hash_missing_template_hash_header"))

        template_id = str(hash_sheet["A2"].value).strip() if hash_sheet["A2"].value is not None else ""
        template_hash = str(hash_sheet["B2"].value).strip() if hash_sheet["B2"].value is not None else ""
        if not template_id:
            raise ValidationError(t("instructor.validation.system_hash_template_id_missing"))
        if not verify_payload_signature(template_id, template_hash):
            raise ValidationError(t("instructor.validation.system_hash_mismatch"))

        if SYSTEM_LAYOUT_SHEET not in workbook.sheetnames:
            raise ValidationError(
                t("instructor.validation.step3.layout_sheet_missing", sheet=SYSTEM_LAYOUT_SHEET)
            )
        layout_sheet = workbook[SYSTEM_LAYOUT_SHEET]
        if normalize(layout_sheet["A1"].value) != normalize(SYSTEM_LAYOUT_MANIFEST_KEY):
            raise ValidationError(
                t(
                    "instructor.validation.step3.layout_header_mismatch",
                    column="A1",
                    expected=SYSTEM_LAYOUT_MANIFEST_KEY,
                )
            )
        if normalize(layout_sheet["B1"].value) != normalize(SYSTEM_LAYOUT_MANIFEST_HASH_KEY):
            raise ValidationError(
                t(
                    "instructor.validation.step3.layout_header_mismatch",
                    column="B1",
                    expected=SYSTEM_LAYOUT_MANIFEST_HASH_KEY,
                )
            )

        manifest_text = str(layout_sheet["A2"].value).strip() if layout_sheet["A2"].value is not None else ""
        manifest_hash = str(layout_sheet["B2"].value).strip() if layout_sheet["B2"].value is not None else ""
        if not manifest_text or not manifest_hash:
            raise ValidationError(t("instructor.validation.step3.layout_manifest_missing"))
        if not verify_payload_signature(manifest_text, manifest_hash):
            raise ValidationError(t("instructor.validation.step3.layout_hash_mismatch"))

        try:
            manifest = json.loads(manifest_text)
        except Exception as exc:
            raise ValidationError(t("instructor.validation.step3.layout_manifest_json_invalid")) from exc
        validate_filled_marks_manifest_schema_by_template(workbook, manifest, template_id=template_id)
    finally:
        workbook.close()
