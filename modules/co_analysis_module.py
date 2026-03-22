"""CO Analysis module."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path, PureWindowsPath
from typing import Any

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QListWidgetItem,
    QScrollArea,
    QVBoxLayout,
    QWidget,
)

from common.constants import (
    APP_NAME,
    CO_ATTAINMENT_LEVEL_DEFAULT,
    CO_ATTAINMENT_PERCENT_DEFAULT,
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_HEADERS,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_SHEET,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    INSTRUCTOR_INFO_TAB_FIXED_HEIGHT,
    LEVEL_1_THRESHOLD,
    LEVEL_2_THRESHOLD,
    LEVEL_3_THRESHOLD,
    MODULE_LEFT_PANE_CONTENT_MARGINS,
    MODULE_LEFT_PANE_LAYOUT_SPACING,
    MODULE_LEFT_PANE_WIDTH_OFFSET,
)
from common.drag_drop_file_widget import ManagedDropFileWidget
from common.module_messages import rerender_user_log as _rerender_user_log_impl
from common.module_runtime import ModuleRuntime
from common.module_ui_engine import ModuleUIEngine, ModuleUIEngineConfig
from common.output_panel import OutputPanelData
from common.removable_file_item_widget import (
    ElidedFileNameLabel as _SharedElidedFileNameLabel,
)
from common.removable_file_item_widget import (
    RemovableFileItemWidget as _SharedRemovableFileItemWidget,
)
from common.texts import t
from common.toast import show_toast
from common.ui_logging import (
    UILogHandler,
    build_i18n_log_message,
    format_log_line_at,
    parse_i18n_log_message,
    resolve_i18n_log_message,
)
from common.ui_stylings import GLOBAL_QPUSHBUTTON_MIN_WIDTH
from common.utils import emit_user_status, log_process_message, normalize, resolve_dialog_start_path
from domain.coordinator_engine import _path_key
from modules.coordinator.workflow_controller import CoordinatorWorkflowController
from modules.instructor.validators.step2_filled_marks_validator import (
    validate_uploaded_filled_marks_workbook,
)

_logger = logging.getLogger(__name__)
_LEFT_PANE_WIDTH = GLOBAL_QPUSHBUTTON_MIN_WIDTH + MODULE_LEFT_PANE_WIDTH_OFFSET
_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
_COURSE_METADATA_DUPLICATE_FIELDS = (
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
)


@dataclass(slots=True)
class COAnalysisWorkflowState:
    busy: bool = False

    def set_busy(self, value: bool) -> None:
        self.busy = value


def _messages_namespace() -> dict[str, object]:
    return {
        "emit_user_status": emit_user_status,
        "t": t,
        "build_i18n_log_message": build_i18n_log_message,
        "parse_i18n_log_message": parse_i18n_log_message,
        "resolve_i18n_log_message": resolve_i18n_log_message,
        "format_log_line_at": format_log_line_at,
        "UILogHandler": UILogHandler,
    }


class _LogSink:
    def appendPlainText(self, _text: str) -> None:  # noqa: N802 - Qt-style name
        return

    def clear(self) -> None:
        return


class _ElidedFileNameLabel(_SharedElidedFileNameLabel):
    pass


class _COAnalysisFileItemWidget(_SharedRemovableFileItemWidget):
    def __init__(self, file_path: str, parent: QWidget | None = None) -> None:
        super().__init__(
            file_path,
            remove_fallback_text=t("coordinator.file.remove_fallback"),
            open_file_tooltip=t("instructor.links.open_file"),
            open_folder_tooltip=t("instructor.links.open_folder"),
            remove_tooltip=t("coordinator.file.remove_tooltip"),
            parent=parent,
        )


def _extract_course_metadata_signature(path: Path) -> tuple[str, ...] | None:
    try:
        import openpyxl
    except Exception:
        return None
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if COURSE_METADATA_SHEET not in workbook.sheetnames:
            return None
        sheet = workbook[COURSE_METADATA_SHEET]
        field_header = normalize(sheet.cell(row=1, column=1).value)
        value_header = normalize(sheet.cell(row=1, column=2).value)
        expected_field = normalize(COURSE_METADATA_HEADERS[0])
        expected_value = normalize(COURSE_METADATA_HEADERS[1])
        if field_header != expected_field or value_header != expected_value:
            return None
        metadata: dict[str, str] = {}
        row = 2
        while True:
            field_raw = sheet.cell(row=row, column=1).value
            value_raw = sheet.cell(row=row, column=2).value
            if normalize(field_raw) == "" and normalize(value_raw) == "":
                break
            field_key = normalize(field_raw)
            if field_key:
                value_text = str(value_raw).strip() if value_raw is not None else ""
                metadata[field_key] = value_text
            row += 1
        signature = tuple(
            metadata.get(normalize(field_name), "").strip()
            for field_name in _COURSE_METADATA_DUPLICATE_FIELDS
        )
        if any(part for part in signature):
            return signature
        return None
    finally:
        workbook.close()


class COAnalysisModule(QWidget):
    status_changed = Signal(str)
    _THRESHOLD_VALIDATION_KEY = "coordinator.thresholds.invalid_rule"
    _CO_ATTAINMENT_TARGET_VALIDATION_KEY = "coordinator.co_attainment.invalid_percent"

    def __init__(self) -> None:
        super().__init__()
        self._files: list[Path] = []
        self._logger = _logger
        self.state = COAnalysisWorkflowState()
        self._ui_log_handler: UILogHandler | None = None
        self._user_log_entries: list[dict[str, object]] = []
        self._threshold_violation_active = False
        self._workflow_controller = CoordinatorWorkflowController(self)
        self._runtime = ModuleRuntime(
            module=self,
            app_name=APP_NAME,
            logger=self._logger,
            async_runner=_NoopAsyncRunner(),
            messages_namespace_factory=_messages_namespace,
        )
        self._build_ui()
        self._setup_ui_logging()
        self.retranslate_ui()
        self._refresh_ui()

    def _build_ui(self) -> None:
        self._ui_engine = ModuleUIEngine(
            self,
            config=ModuleUIEngineConfig(
                top_object_name="coAnalysisTopRegion",
                footer_height=INSTRUCTOR_INFO_TAB_FIXED_HEIGHT,
            ),
        )
        top_pane = QWidget()
        top_layout = QHBoxLayout(top_pane)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(0)
        self._ui_engine.set_top_widget(top_pane)

        left_card = QFrame()
        left_card.setObjectName("coordinatorLeftCard")
        left_card.setFrameShape(QFrame.Shape.StyledPanel)
        left_card.setFrameShadow(QFrame.Shadow.Raised)
        left_card_layout = QVBoxLayout(left_card)
        left_card_layout.setContentsMargins(0, 0, 0, 0)
        left_card_layout.setSpacing(0)

        left_scroll = QScrollArea()
        left_scroll.setObjectName("coordinatorLeftScroll")
        left_scroll.setFrameShape(QFrame.Shape.NoFrame)
        left_scroll.setWidgetResizable(True)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        left_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        left_scroll.viewport().setObjectName("coordinatorLeftScrollViewport")
        left_content = QWidget()
        left_layout = QVBoxLayout(left_content)
        left_layout.setContentsMargins(*MODULE_LEFT_PANE_CONTENT_MARGINS)
        left_layout.setSpacing(MODULE_LEFT_PANE_LAYOUT_SPACING)
        left_scroll.setWidget(left_content)
        left_card_layout.addWidget(left_scroll, 1)
        left_card.setFixedWidth(_LEFT_PANE_WIDTH)
        top_layout.addWidget(left_card, 0)

        self.title_label = QLabel()
        self.title_label.setObjectName("coordinatorTitle")
        self.title_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        left_layout.addWidget(self.title_label)

        self.hint_label = QLabel()
        self.hint_label.setObjectName("coordinatorHint")
        self.hint_label.setWordWrap(True)
        self.hint_label.setAlignment(Qt.AlignmentFlag.AlignJustify | Qt.AlignmentFlag.AlignTop)
        left_layout.addWidget(self.hint_label)

        thresholds_layout = QVBoxLayout()
        thresholds_layout.setContentsMargins(0, 0, 0, 0)
        thresholds_layout.setSpacing(MODULE_LEFT_PANE_LAYOUT_SPACING)
        self.threshold_title_label = QLabel()
        self.threshold_title_label.setObjectName("coordinatorThresholdTitle")
        thresholds_layout.addWidget(self.threshold_title_label)
        self.threshold_description_label = QLabel()
        self.threshold_description_label.setWordWrap(True)
        self.threshold_description_label.setAlignment(
            Qt.AlignmentFlag.AlignJustify | Qt.AlignmentFlag.AlignTop
        )
        thresholds_layout.addWidget(self.threshold_description_label)

        threshold_rows = QGridLayout()
        threshold_rows.setColumnStretch(0, 0)
        threshold_rows.setColumnStretch(1, 1)

        self.threshold_l1_label = QLabel()
        self.threshold_l1_label.setObjectName("coordinatorThresholdL1Label")
        self.threshold_l1_input = QDoubleSpinBox()
        self.threshold_l1_input.setRange(0.0, 100.0)
        self.threshold_l1_input.setDecimals(2)
        self.threshold_l1_input.setSingleStep(0.5)
        self.threshold_l1_input.setValue(float(LEVEL_1_THRESHOLD))
        threshold_rows.addWidget(self.threshold_l1_label, 0, 0)
        threshold_rows.addWidget(
            self.threshold_l1_input,
            0,
            1,
            alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
        )

        self.threshold_l2_label = QLabel()
        self.threshold_l2_label.setObjectName("coordinatorThresholdInputLabel")
        self.threshold_l2_input = QDoubleSpinBox()
        self.threshold_l2_input.setRange(0.0, 100.0)
        self.threshold_l2_input.setDecimals(2)
        self.threshold_l2_input.setSingleStep(0.5)
        self.threshold_l2_input.setValue(float(LEVEL_2_THRESHOLD))
        threshold_rows.addWidget(self.threshold_l2_label, 1, 0)
        threshold_rows.addWidget(
            self.threshold_l2_input,
            1,
            1,
            alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
        )

        self.threshold_l3_label = QLabel()
        self.threshold_l3_label.setObjectName("coordinatorThresholdInputLabel")
        self.threshold_l3_input = QDoubleSpinBox()
        self.threshold_l3_input.setRange(0.0, 100.0)
        self.threshold_l3_input.setDecimals(2)
        self.threshold_l3_input.setSingleStep(0.5)
        self.threshold_l3_input.setValue(float(LEVEL_3_THRESHOLD))
        threshold_rows.addWidget(self.threshold_l3_label, 2, 0)
        threshold_rows.addWidget(
            self.threshold_l3_input,
            2,
            1,
            alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
        )
        self.threshold_l1_input.valueChanged.connect(self._on_threshold_value_changed)
        self.threshold_l2_input.valueChanged.connect(self._on_threshold_value_changed)
        self.threshold_l3_input.valueChanged.connect(self._on_threshold_value_changed)
        self.threshold_l1_input.editingFinished.connect(self._on_threshold_editing_finished)
        self.threshold_l2_input.editingFinished.connect(self._on_threshold_editing_finished)
        self.threshold_l3_input.editingFinished.connect(self._on_threshold_editing_finished)

        thresholds_layout.addLayout(threshold_rows)

        self.co_attainment_description_label = QLabel()
        self.co_attainment_description_label.setWordWrap(True)
        self.co_attainment_description_label.setAlignment(
            Qt.AlignmentFlag.AlignJustify | Qt.AlignmentFlag.AlignTop
        )
        thresholds_layout.addWidget(self.co_attainment_description_label)

        co_attainment_rows = QGridLayout()
        co_attainment_rows.setColumnStretch(0, 0)
        co_attainment_rows.setColumnStretch(1, 1)

        self.co_attainment_percent_label = QLabel()
        self.co_attainment_percent_label.setObjectName("coordinatorThresholdInputLabel")
        self.co_attainment_percent_input = QDoubleSpinBox()
        self.co_attainment_percent_input.setRange(0.0, 100.0)
        self.co_attainment_percent_input.setDecimals(2)
        self.co_attainment_percent_input.setSingleStep(0.5)
        self.co_attainment_percent_input.setValue(float(CO_ATTAINMENT_PERCENT_DEFAULT))
        co_attainment_rows.addWidget(self.co_attainment_percent_label, 0, 0)
        co_attainment_rows.addWidget(
            self.co_attainment_percent_input,
            0,
            1,
            alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
        )

        self.co_attainment_level_label = QLabel()
        self.co_attainment_level_label.setObjectName("coordinatorThresholdInputLabel")
        self.co_attainment_level_input = QComboBox()
        self.co_attainment_level_input.addItem("L1", 1)
        self.co_attainment_level_input.addItem("L2", 2)
        self.co_attainment_level_input.addItem("L3", 3)
        default_level_index = max(
            0,
            min(self.co_attainment_level_input.count() - 1, CO_ATTAINMENT_LEVEL_DEFAULT - 1),
        )
        self.co_attainment_level_input.setCurrentIndex(default_level_index)
        co_attainment_rows.addWidget(self.co_attainment_level_label, 1, 0)
        co_attainment_rows.addWidget(
            self.co_attainment_level_input,
            1,
            1,
            alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
        )

        self.co_attainment_percent_input.valueChanged.connect(self._on_threshold_value_changed)
        self.co_attainment_percent_input.editingFinished.connect(self._on_threshold_editing_finished)
        self.co_attainment_level_input.currentIndexChanged.connect(
            lambda _idx: self._on_threshold_value_changed(0.0)
        )
        self.co_attainment_level_input.activated.connect(lambda _idx: self._on_threshold_editing_finished())
        thresholds_layout.addLayout(co_attainment_rows)
        left_layout.addLayout(thresholds_layout)
        left_layout.addStretch(1)

        right_pane = QWidget()
        right_pane.setObjectName("coordinatorActiveCard")
        right_layout = QVBoxLayout(right_pane)
        right_scroll = QScrollArea()
        right_scroll.setObjectName("coordinatorRightScroll")
        right_scroll.setFrameShape(QFrame.Shape.NoFrame)
        right_scroll.setWidgetResizable(True)
        right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        right_scroll.viewport().setObjectName("coordinatorRightScrollViewport")
        right_scroll.setWidget(right_pane)
        top_layout.addWidget(right_scroll, 1)

        self.drop_widget = ManagedDropFileWidget(
            drop_mode="multiple",
            remove_fallback_text=t("coordinator.file.remove_fallback"),
            open_file_tooltip=t("instructor.links.open_file"),
            open_folder_tooltip=t("instructor.links.open_folder"),
            remove_tooltip=t("coordinator.file.remove_tooltip"),
        )
        self.drop_widget.set_summary_text_builder(
            lambda _count: t("coordinator.summary", count=len(self._files))
        )
        self.drop_widget.drop_list.set_placeholder_text(t("common.dropzone.placeholder"))
        self.drop_widget.drop_list.setObjectName("coordinatorDropList")
        self.drop_widget.files_dropped.connect(self._on_files_dropped)
        self.drop_widget.files_rejected.connect(self._on_drop_files_rejected)
        self.drop_widget.browse_requested.connect(self._on_drop_browse_requested)
        self.drop_widget.browse_requested.connect(self._browse_files)
        self.drop_widget.files_changed.connect(self._on_drop_files_changed)
        self.drop_widget.clear_button.clicked.connect(self._clear_all)
        self.drop_widget.submit_requested.connect(self._on_submit_requested)
        self.drop_widget.set_clear_button_text(t("coordinator.clear_all"))
        right_layout.addWidget(self.drop_widget, 1)

        self.drop_zone = self.drop_widget.drop_zone
        self.drop_list = self.drop_widget.drop_list
        self.clear_button = self.drop_widget.clear_button
        self.calculate_button = self.drop_widget.submit_button
        self.calculate_button.setObjectName("primaryAction")
        self.calculate_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.calculate_button.setAutoDefault(False)
        self.calculate_button.setDefault(False)

        self.user_log_view = _LogSink()
        self._ui_engine.set_footer_visible(False)

        self.shortcut_add_file = QShortcut(QKeySequence(QKeySequence.StandardKey.Open), self)
        self.shortcut_add_file.activated.connect(self._browse_files)

    def retranslate_ui(self) -> None:
        self.title_label.setText(t("coordinator.title"))
        self.hint_label.setText(t("coordinator.drop_hint"))
        self.drop_widget.drop_list.set_placeholder_text(t("common.dropzone.placeholder"))
        self.drop_widget.set_clear_button_text(t("coordinator.clear_all"))
        self.drop_widget.set_summary_text_builder(
            lambda _count: t("coordinator.summary", count=len(self._files))
        )
        self.drop_widget.set_submit_button_text(t("coordinator.calculate"))
        self.threshold_title_label.setText(t("coordinator.thresholds.title"))
        self.threshold_description_label.setText(t("coordinator.thresholds.description"))
        self.threshold_l1_label.setText(t("coordinator.thresholds.l1.label"))
        self.threshold_l2_label.setText(t("coordinator.thresholds.l2.label"))
        self.threshold_l3_label.setText(t("coordinator.thresholds.l3.label"))
        self.co_attainment_description_label.setText(t("coordinator.co_attainment.description"))
        self.co_attainment_percent_label.setText(t("coordinator.co_attainment.percent.label"))
        self.co_attainment_level_label.setText(t("coordinator.co_attainment.level.label"))
        self._refresh_summary()
        self._rerender_user_log()

    def _publish_status(self, message: str) -> None:
        self._runtime.publish_status(message)

    def _publish_status_key(self, text_key: str, **kwargs: Any) -> None:
        self._runtime.publish_status_key(text_key, **kwargs)

    def _refresh_ui(self) -> None:
        has_files = bool(self._files)
        can_submit = has_files and self._has_valid_attainment_thresholds() and self._has_valid_co_attainment_target()
        self.drop_widget.setEnabled(True)
        self.drop_widget.set_submit_allowed(can_submit)
        self.threshold_l1_input.setEnabled(True)
        self.threshold_l2_input.setEnabled(True)
        self.threshold_l3_input.setEnabled(True)
        self.co_attainment_percent_input.setEnabled(True)
        self.co_attainment_level_input.setEnabled(True)
        self.drop_list.setEnabled(True)
        self.clear_button.setEnabled(has_files)
        self.calculate_button.setEnabled(can_submit)
        for row in range(self.drop_list.count()):
            item = self.drop_list.item(row)
            widget = self.drop_list.itemWidget(item)
            if isinstance(widget, _COAnalysisFileItemWidget):
                widget.remove_btn.setEnabled(True)
        self._refresh_summary()

    def _read_attainment_thresholds(self) -> tuple[float, float, float]:
        return self._workflow_controller.read_attainment_thresholds()

    def _has_valid_attainment_thresholds(self) -> bool:
        return self._workflow_controller.has_valid_attainment_thresholds()

    def _read_co_attainment_target(self) -> tuple[float, int]:
        return self._workflow_controller.read_co_attainment_target()

    def _has_valid_co_attainment_target(self) -> bool:
        return self._workflow_controller.has_valid_co_attainment_target()

    def _notify_threshold_violation(self, *, force: bool) -> None:
        self._workflow_controller.notify_threshold_violation(force=force)

    def _show_threshold_validation_toast(self, *, message_key: str) -> None:
        show_toast(
            self,
            t(message_key),
            title=t("coordinator.title"),
            level="error",
        )

    def _on_threshold_value_changed(self, _value: float) -> None:
        self._workflow_controller.on_threshold_value_changed()
        self._refresh_ui()

    def _on_threshold_editing_finished(self) -> None:
        self._workflow_controller.on_threshold_editing_finished()
        self._refresh_ui()

    def _on_submit_requested(self) -> None:
        # UI-only action requested by product; no processing logic is attached.
        return

    def _on_drop_browse_requested(self) -> None:
        self._publish_status_key("instructor.status.step2_drop_browse_requested")

    def _on_files_dropped(self, dropped_files: list[str]) -> None:
        self._publish_status_key("instructor.status.step2_drop_files_dropped", count=len(dropped_files))
        if dropped_files:
            self._remember_dialog_dir_safe(dropped_files[0])
        self._collect_valid_files(dropped_files)

    def _on_drop_files_rejected(self, files: list[str]) -> None:
        self._publish_status_key("instructor.status.step2_drop_files_rejected", count=len(files))

    def _on_drop_files_changed(self, files: list[str]) -> None:
        self._publish_status_key("instructor.status.step2_drop_files_changed", count=len(files))

    def _browse_files(self) -> None:
        selected_files, _ = QFileDialog.getOpenFileNames(
            self,
            t("instructor.dialog.step2.upload.title"),
            resolve_dialog_start_path(APP_NAME),
            t("instructor.dialog.filter.excel_open"),
        )
        if not selected_files:
            return
        self._remember_dialog_dir_safe(selected_files[0])
        self._collect_valid_files(selected_files)

    def _remember_dialog_dir_safe(self, selected_path: str) -> None:
        self._runtime.remember_dialog_dir_safe(selected_path)

    def _setup_ui_logging(self) -> None:
        self._runtime.setup_ui_logging()

    def _append_user_log(self, message: str) -> None:
        self._runtime.append_user_log(message)

    def _rerender_user_log(self) -> None:
        _rerender_user_log_impl(self, ns=_messages_namespace())

    def _refresh_summary(self) -> None:
        count = len(self._files)
        self.drop_widget.set_summary_text_builder(lambda _count: t("coordinator.summary", count=count))
        # ManagedDropFileWidget enables this label using its own internal list length.
        # CO Analysis maintains file state externally, so force the visual state from module state.
        self.drop_widget.summary_label.setEnabled(count > 0)

    def _collect_valid_files(self, candidate_paths: list[str]) -> None:
        if not candidate_paths:
            return
        self._publish_status_key("coordinator.status.processing_started")
        existing_keys = {_path_key(path) for path in self._files}
        seen_metadata_signatures = {
            signature
            for signature in (_extract_course_metadata_signature(path) for path in self._files)
            if signature is not None
        }
        added_paths: list[Path] = []
        duplicates = 0
        invalid = 0
        validated_candidates: list[tuple[Path, str, tuple[str, ...] | None]] = []
        for raw_path in candidate_paths:
            path = Path(raw_path)
            suffix = path.suffix.lower()
            if suffix not in _SUPPORTED_EXTENSIONS or not path.exists():
                invalid += 1
                continue
            key = _path_key(path)
            try:
                validate_uploaded_filled_marks_workbook(path)
            except Exception as exc:
                invalid += 1
                self._logger.info("Rejected invalid marks template workbook '%s': %s", path, exc)
                continue
            metadata_signature = _extract_course_metadata_signature(path)
            validated_candidates.append((path, key, metadata_signature))

        path_counts: dict[str, int] = {}
        metadata_counts: dict[tuple[str, ...], int] = {}
        for _path, key, metadata_signature in validated_candidates:
            path_counts[key] = path_counts.get(key, 0) + 1
            if metadata_signature is not None:
                metadata_counts[metadata_signature] = metadata_counts.get(metadata_signature, 0) + 1

        for path, key, metadata_signature in validated_candidates:
            is_duplicate = False
            if key in existing_keys:
                is_duplicate = True
            elif metadata_signature is not None and metadata_signature in seen_metadata_signatures:
                is_duplicate = True
            elif path_counts.get(key, 0) > 1:
                is_duplicate = True
            elif metadata_signature is not None and metadata_counts.get(metadata_signature, 0) > 1:
                is_duplicate = True

            if is_duplicate:
                duplicates += 1
                continue
            added_paths.append(path)

        self._add_uploaded_paths(added_paths)
        if added_paths:
            self._publish_status_key(
                "coordinator.status.added",
                added=len(added_paths),
                total=len(self._files),
            )
        if duplicates or invalid:
            show_toast(
                self,
                t(
                    "instructor.toast.step2_upload_reject_summary",
                    invalid=invalid,
                    duplicates=duplicates,
                ),
                title=t("instructor.step2.title"),
                level="warning",
            )
        ignored = duplicates + invalid
        if ignored:
            self._publish_status_key("coordinator.status.ignored", count=ignored)
        log_process_message(
            "collecting co analysis files",
            logger=self._logger,
            success_message=(
                "collecting co analysis files completed successfully. "
                f"added={len(added_paths)}, duplicates={duplicates}, invalid={invalid}"
            ),
            user_success_message=build_i18n_log_message(
                "coordinator.status.processing_completed",
                fallback=t("coordinator.status.processing_completed"),
            ),
        )
        self._refresh_ui()

    def _new_file_item_widget(
        self,
        file_path: str,
        *,
        parent: QWidget | None = None,
    ) -> _COAnalysisFileItemWidget:
        return _COAnalysisFileItemWidget(file_path, parent=parent)

    def _add_uploaded_paths(self, added_paths: list[Path]) -> None:
        for path in added_paths:
            self._files.append(path)
            path_text = str(path)
            if len(path_text) >= 2 and path_text[1] == ":":
                path_text = str(PureWindowsPath(path_text))
            item = QListWidgetItem()
            item.setToolTip(path_text)
            item.setData(Qt.ItemDataRole.UserRole, path_text)
            self.drop_list.addItem(item)
            row_widget = self._new_file_item_widget(path_text, parent=self.drop_list)
            row_widget.removed.connect(self._remove_file_by_path)
            item.setSizeHint(row_widget.sizeHint())
            self.drop_list.setItemWidget(item, row_widget)

    def _remove_file_by_path(self, file_path: str) -> None:
        target_key = _path_key(Path(file_path))
        before_count = len(self._files)
        self._files = [path for path in self._files if _path_key(path) != target_key]
        if len(self._files) == before_count:
            return
        for row in range(self.drop_list.count()):
            item = self.drop_list.item(row)
            path_value = str(item.data(Qt.ItemDataRole.UserRole) or "")
            if _path_key(Path(path_value)) == target_key:
                self.drop_list.takeItem(row)
                break
        self._refresh_ui()
        self._publish_status_key("coordinator.status.removed", count=1)
        log_process_message(
            "removing selected co analysis files",
            logger=self._logger,
            success_message="removing selected co analysis files completed successfully. removed=1",
            user_success_message=build_i18n_log_message(
                "coordinator.status.removed",
                kwargs={"count": 1},
                fallback=t("coordinator.status.removed", count=1),
            ),
        )

    def _clear_all(self) -> None:
        if not self._files:
            return
        total = len(self._files)
        self._files.clear()
        self.drop_list.clear()
        self._refresh_ui()
        self._publish_status_key("coordinator.status.cleared", count=total)
        log_process_message(
            "clearing co analysis files",
            logger=self._logger,
            success_message=f"clearing co analysis files completed successfully. removed={total}",
            user_success_message=build_i18n_log_message(
                "coordinator.status.cleared",
                kwargs={"count": total},
                fallback=t("coordinator.status.cleared", count=total),
            ),
        )

    def set_shared_activity_log_mode(self, enabled: bool) -> None:
        self._ui_engine.set_footer_visible(False)

    def get_shared_outputs_data(self) -> OutputPanelData:
        return OutputPanelData(items=tuple())

    def closeEvent(self, event) -> None:
        if self._ui_log_handler is not None:
            self._logger.removeHandler(self._ui_log_handler)
            self._ui_log_handler = None
        super().closeEvent(event)
class _NoopAsyncRunner:
    def start(self, **_kwargs: object) -> None:
        return

