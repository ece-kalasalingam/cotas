"""Course Instructor CO module UI (list-based step navigation)."""

from __future__ import annotations

import logging
import os
import re
import shutil
import tempfile
from html import escape
from hashlib import sha256
from datetime import datetime
from pathlib import Path

from PySide6.QtCore import Qt, QUrl, Signal
from PySide6.QtGui import QDesktopServices, QFont
from PySide6.QtWidgets import (
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QPlainTextEdit,
    QPushButton,
    QTabWidget,
    QTextBrowser,
    QVBoxLayout,
    QWidget,
)

from common.constants import (
    APP_NAME,
    COURSE_METADATA_SHEET,
    ID_COURSE_SETUP,
    MARKS_ENTRY_ROW_HEADERS,
    SYSTEM_HASH_SHEET,
    SYSTEM_HASH_TEMPLATE_HASH_KEY,
    SYSTEM_HASH_TEMPLATE_ID_KEY,
    WORKBOOK_PASSWORD,
    INSTRUCTOR_ACTIVE_TITLE_FONT_SIZE,
    INSTRUCTOR_CARD_MARGIN,
    INSTRUCTOR_CARD_SPACING,
    INSTRUCTOR_PANEL_STYLESHEET,
    INSTRUCTOR_RAIL_MAX_WIDTH,
    INSTRUCTOR_RAIL_TITLE_FONT_SIZE,
    INSTRUCTOR_STEP2_ACTION_MARGIN,
    INSTRUCTOR_STEP2_ACTION_SPACING,
    INSTRUCTOR_STEP_LIST_SPACING,
    INSTRUCTOR_TOP_LAYOUT_MARGINS,
    INSTRUCTOR_INFO_TAB_FIXED_HEIGHT,
    INSTRUCTOR_INFO_TAB_LAYOUT_MARGINS,
    INSTRUCTOR_INFO_TAB_LAYOUT_SPACING,
    UI_FONT_FAMILY,
)
from common.exceptions import AppSystemError, JobCancelledError, ValidationError
from common.jobs import CancellationToken
from common.qt_jobs import run_in_background
from common.toast import show_toast
from common.texts import t
from common.utils import (
    coerce_excel_number,
    emit_user_status,
    log_process_message,
    normalize,
    remember_dialog_dir,
    remember_dialog_dir_safe,
    resolve_dialog_start_path,
)
from domain import InstructorWorkflowState
from modules.instructor import (
    generate_course_details_template,
    generate_marks_template_from_course_details,
    validate_course_details_workbook,
)
from services import InstructorWorkflowService

_logger = logging.getLogger(__name__)


class _UILogHandler(logging.Handler):
    """Forward logger messages to the in-panel user log view."""

    def __init__(self, sink):
        super().__init__(level=logging.INFO)
        self._sink = sink

    def emit(self, record: logging.LogRecord) -> None:
        try:
            user_message = getattr(record, "user_message", None)
            message = f"{record.levelname}: {user_message or record.getMessage()}"
            self._sink(message)
        except Exception:
            self.handleError(record)


def _publish_status_compat(target: object, message: str) -> None:
    """Publish status for both full widget instances and lightweight test doubles."""
    publish = getattr(target, "_publish_status", None)
    if callable(publish):
        publish(message)
        return
    emit_user_status(getattr(target, "status_changed", None), message, logger=_logger)


def _set_busy_compat(target: object, busy: bool, *, job_id: str | None = None) -> None:
    setter = getattr(target, "_set_busy", None)
    if callable(setter):
        setter(busy, job_id=job_id)


def _start_async_operation_compat(
    target: object,
    *,
    token: CancellationToken,
    job_id: str | None,
    work,
    on_success,
    on_failure,
) -> None:
    starter = getattr(target, "_start_async_operation", None)
    if callable(starter):
        starter(
            token=token,
            job_id=job_id,
            work=work,
            on_success=on_success,
            on_failure=on_failure,
        )
        return

    setattr(target, "_cancel_token", token)
    _set_busy_compat(target, True, job_id=job_id)
    job_ref: dict[str, object] = {}

    def _finalize() -> None:
        active_jobs = getattr(target, "_active_jobs", None)
        tracked_job = job_ref.get("job")
        if isinstance(active_jobs, list) and tracked_job in active_jobs:
            active_jobs.remove(tracked_job)
        setattr(target, "_cancel_token", None)
        _set_busy_compat(target, False)
        refresh = getattr(target, "_refresh_ui", None)
        if callable(refresh):
            refresh()

    def _on_finished(result: object) -> None:
        try:
            on_success(result)
        finally:
            _finalize()

    def _on_failed(exc: Exception) -> None:
        try:
            on_failure(exc)
        finally:
            _finalize()

    job = run_in_background(work, on_finished=_on_finished, on_failed=_on_failed)
    job_ref["job"] = job
    active_jobs = getattr(target, "_active_jobs", None)
    if isinstance(active_jobs, list):
        active_jobs.append(job)


def _localized_log_messages(process_key: str) -> tuple[str, str]:
    user_process_name = t(process_key)
    return (
        t("instructor.log.completed_process", process=user_process_name),
        t("instructor.log.error_while_process", process=user_process_name),
    )


def _sanitize_filename_token(value: object) -> str:
    token = str(value).strip()
    token = re.sub(r'[<>:"/\\|?*]+', "_", token)
    token = re.sub(r"\s+", "", token)
    token = token.strip(" ._")
    return token


def _build_marks_template_default_name(course_details_path: str | None) -> str:
    fallback = t("instructor.dialog.step3.default_name")
    if not course_details_path:
        return fallback

    try:
        import openpyxl
    except ModuleNotFoundError:
        return fallback

    workbook = None
    try:
        workbook = openpyxl.load_workbook(course_details_path, data_only=True)
        if COURSE_METADATA_SHEET not in workbook.sheetnames:
            return fallback
        sheet = workbook[COURSE_METADATA_SHEET]
        fields: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = normalize(row[0] if len(row) > 0 else None)
            if not key:
                continue
            value = row[1] if len(row) > 1 else None
            coerced = coerce_excel_number(value)
            fields[key] = str(coerced).strip() if coerced is not None else ""

        parts = [
            _sanitize_filename_token(fields.get("course_code", "")),
            _sanitize_filename_token(fields.get("semester", "")),
            _sanitize_filename_token(fields.get("section", "")),
            _sanitize_filename_token(fields.get("academic_year", "")),
            "Marks",
        ]
        if any(not part for part in parts[:4]):
            return fallback
        return f"{'_'.join(parts)}.xlsx"
    except Exception:
        return fallback
    finally:
        if workbook is not None:
            workbook.close()


def _atomic_copy_file(source_path: str | Path, output_path: str | Path) -> Path:
    source = Path(source_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_name = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            delete=False,
            dir=str(output.parent),
            prefix=f"{output.name}.",
            suffix=".tmp",
        ) as temp_file:
            temp_name = temp_file.name
        shutil.copyfile(str(source), temp_name)
        os.replace(temp_name, output)
    except Exception:
        if temp_name:
            try:
                Path(temp_name).unlink(missing_ok=True)
            except OSError:
                _logger.warning("Failed to cleanup temp report file: %s", temp_name)
        raise
    return output


def _validate_uploaded_filled_marks_workbook(workbook_path: str | Path) -> None:
    """Validate a filled marks workbook before enabling final-report generation."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise ValidationError(
            t("instructor.validation.openpyxl_missing"),
            code="OPENPYXL_MISSING",
        ) from exc

    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise ValidationError(
            t("instructor.validation.workbook_not_found", workbook=workbook_file),
            code="WORKBOOK_NOT_FOUND",
            context={"workbook": str(workbook_file)},
        )

    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=True)
    except Exception as exc:
        raise ValidationError(
            t("instructor.validation.workbook_open_failed", workbook=workbook_file),
            code="WORKBOOK_OPEN_FAILED",
            context={"workbook": str(workbook_file)},
        ) from exc

    try:
        if SYSTEM_HASH_SHEET not in workbook.sheetnames:
            raise ValidationError(t("instructor.validation.system_sheet_missing", sheet=SYSTEM_HASH_SHEET))

        hash_sheet = workbook[SYSTEM_HASH_SHEET]
        if normalize(hash_sheet["A1"].value) != normalize(SYSTEM_HASH_TEMPLATE_ID_KEY):
            raise ValidationError(t("instructor.validation.system_hash_missing_template_id_header"))
        if normalize(hash_sheet["B1"].value) != normalize(SYSTEM_HASH_TEMPLATE_HASH_KEY):
            raise ValidationError(t("instructor.validation.system_hash_missing_template_hash_header"))

        template_id = str(hash_sheet["A2"].value).strip() if hash_sheet["A2"].value is not None else ""
        template_hash = str(hash_sheet["B2"].value).strip() if hash_sheet["B2"].value is not None else ""
        if not template_id:
            raise ValidationError(t("instructor.validation.system_hash_template_id_missing"))
        expected_hash = sha256(f"{template_id}|{WORKBOOK_PASSWORD}".encode("utf-8")).hexdigest()
        if template_hash != expected_hash:
            raise ValidationError(t("instructor.validation.system_hash_mismatch"))

        marks_sheet_names = [name for name in workbook.sheetnames if name != SYSTEM_HASH_SHEET]
        if not marks_sheet_names:
            raise ValidationError(t("instructor.validation.filled_marks_workbook_invalid"))

        expected_headers = [normalize(header) for header in MARKS_ENTRY_ROW_HEADERS]
        for sheet_name in marks_sheet_names:
            worksheet = workbook[sheet_name]
            actual_headers = [
                normalize(worksheet.cell(row=1, column=col_index + 1).value)
                for col_index in range(len(expected_headers))
            ]
            if actual_headers != expected_headers:
                raise ValidationError(t("instructor.validation.filled_marks_workbook_invalid"))
    finally:
        workbook.close()


class InstructorModule(QWidget):
    """Simple wizard-like UI for CO score workflow."""

    status_changed = Signal(str)
    WORKFLOW_STEPS = (1, 2, 3)
    RAIL_LINK_TITLE_KEY = "instructor.links.title"
    RAIL_LINKS = (
        ("instructor.links.course_details_generated", "step1_path"),
        ("instructor.links.course_details_uploaded", "step2_course_details_path"),
        ("instructor.links.marks_template_generated", "step2_path"),
        ("instructor.links.marks_template_uploaded", "step3_path"),
    )
    RAIL_LINK_OPEN_FILE_KEY = "instructor.links.open_file"
    RAIL_LINK_OPEN_FOLDER_KEY = "instructor.links.open_folder"
    RAIL_LINK_NOT_AVAILABLE_KEY = "instructor.links.not_available"
    RAIL_LINK_OPEN_FAILED_KEY = "instructor.links.open_failed"

    STEP_TITLE_KEYS = {
        1: "instructor.step1.title",
        2: "instructor.step2.title",
        3: "instructor.step3.title",
    }

    STEP_DESC_KEYS = {
        1: "instructor.step1.desc",
        2: "instructor.step2.desc",
        3: "instructor.step3.desc",
    }

    PATH_ATTRS = {
        1: "step1_path",
        2: "step2_path",
        3: "step4_path",
    }

    DONE_ATTRS = {
        1: "step1_done",
        2: "step2_done",
        3: "step4_done",
    }

    OUTDATED_ATTRS = {
        3: "step4_outdated",
    }

    ACTION_DEFAULT_KEYS = {
        1: "instructor.action.step1.default",
        2: "instructor.action.step2.default",
        3: "instructor.action.step5.default",
    }

    ACTION_REDO_KEYS = {
        1: "instructor.action.step1.redo",
        2: "instructor.action.step2.redo",
        3: "instructor.action.step5.redo",
    }

    def __init__(self):
        super().__init__()
        self.current_step = 1

        self.step1_path: str | None = None
        self.step2_path: str | None = None
        self.step2_course_details_path: str | None = None
        self.step3_path: str | None = None
        self.step4_path: str | None = None

        self.step1_done = False
        self.step2_done = False
        self.step3_done = False
        self.step4_done = False
        self.step2_upload_ready = False

        self.step3_outdated = False  # Filled marks
        self.step4_outdated = False  # Final report

        self.state = InstructorWorkflowState()
        self._workflow_service = InstructorWorkflowService()
        self._cancel_token: CancellationToken | None = None
        self._active_jobs: list[object] = []
        self._is_closing = False
        self._step2_marks_default_name = t("instructor.dialog.step3.default_name")

        self._ui_log_handler: _UILogHandler | None = None
        self._step_items: list[QListWidgetItem] = []
        self._build_ui()
        self._setup_ui_logging()
        self._refresh_ui()

    def _build_ui(self) -> None:
        root = QHBoxLayout(self)

        left = QFrame()
        left.setObjectName("stepRail")
        left_layout = QVBoxLayout(left)
        

        rail_title = QLabel(t("instructor.workflow_title"))
        rail_title.setFont(
            QFont(UI_FONT_FAMILY, INSTRUCTOR_RAIL_TITLE_FONT_SIZE, QFont.Weight.Bold)
        )
        left_layout.addWidget(rail_title)

        self.progress_label = QLabel()
        self.progress_label.setObjectName("progressText")
        left_layout.addWidget(self.progress_label)

        self.step_list = QListWidget()
        self.step_list.setObjectName("stepList")
        self.step_list.setFrameShape(QFrame.Shape.NoFrame)
        self.step_list.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.step_list.setSelectionMode(QListWidget.SelectionMode.SingleSelection)
        self.step_list.setCursor(Qt.CursorShape.ArrowCursor)
        self.step_list.setSpacing(INSTRUCTOR_STEP_LIST_SPACING)
        self.step_list.setWordWrap(True)
        self.step_list.setTextElideMode(Qt.TextElideMode.ElideNone)
        self.step_list.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        for step in self.WORKFLOW_STEPS:
            item = QListWidgetItem(self._step_list_text(step))
            self.step_list.addItem(item)
            self._step_items.append(item)
        self.step_list.currentRowChanged.connect(self._on_step_row_changed)
        left_layout.addWidget(self.step_list, 1)
        left.setMaximumWidth(INSTRUCTOR_RAIL_MAX_WIDTH)

        right = QFrame()
        right.setObjectName("activeCard")
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(
            INSTRUCTOR_CARD_MARGIN,
            INSTRUCTOR_CARD_MARGIN,
            INSTRUCTOR_CARD_MARGIN,
            INSTRUCTOR_CARD_MARGIN,
        )
        right_layout.setSpacing(INSTRUCTOR_CARD_SPACING)

        top_content = QWidget()
        top_layout = QVBoxLayout(top_content)
        top_layout.setContentsMargins(*INSTRUCTOR_TOP_LAYOUT_MARGINS)
        top_layout.setSpacing(INSTRUCTOR_CARD_SPACING)

        self.active_title = QLabel()
        self.active_title.setFont(
            QFont(UI_FONT_FAMILY, INSTRUCTOR_ACTIVE_TITLE_FONT_SIZE, QFont.Weight.Bold)
        )
        self.active_title.setWordWrap(True)
        top_layout.addWidget(self.active_title)

        self.active_desc = QLabel()
        self.active_desc.setWordWrap(True)
        top_layout.addWidget(self.active_desc)

        self.active_note = QLabel(t("instructor.note.default"))
        self.active_note.setWordWrap(True)
        self.active_note.setObjectName("hintText")
        top_layout.addWidget(self.active_note)

        self.primary_action = QPushButton()
        self.primary_action.setObjectName("primaryAction")
        self.primary_action.clicked.connect(self._run_current_step_action)
        top_layout.addWidget(self.primary_action, alignment=Qt.AlignmentFlag.AlignLeft)

        self.step2_action_row = QWidget()
        step2_action_layout = QHBoxLayout(self.step2_action_row)
        step2_action_layout.setContentsMargins(
            INSTRUCTOR_STEP2_ACTION_MARGIN,
            INSTRUCTOR_STEP2_ACTION_MARGIN,
            INSTRUCTOR_STEP2_ACTION_MARGIN,
            INSTRUCTOR_STEP2_ACTION_MARGIN,
        )
        step2_action_layout.setSpacing(INSTRUCTOR_STEP2_ACTION_SPACING)

        self.step2_upload_action = QPushButton()
        self.step2_upload_action.clicked.connect(self._on_step2_upload_clicked)
        step2_action_layout.addWidget(self.step2_upload_action)

        self.step2_prepare_action = QPushButton()
        self.step2_prepare_action.setObjectName("primaryAction")
        self.step2_prepare_action.clicked.connect(self._on_step2_prepare_clicked)
        step2_action_layout.addWidget(self.step2_prepare_action)
        step2_action_layout.addStretch(1)
        top_layout.addWidget(self.step2_action_row)

        self.step3_action_row = QWidget()
        step3_action_layout = QHBoxLayout(self.step3_action_row)
        step3_action_layout.setContentsMargins(
            INSTRUCTOR_STEP2_ACTION_MARGIN,
            INSTRUCTOR_STEP2_ACTION_MARGIN,
            INSTRUCTOR_STEP2_ACTION_MARGIN,
            INSTRUCTOR_STEP2_ACTION_MARGIN,
        )
        step3_action_layout.setSpacing(INSTRUCTOR_STEP2_ACTION_SPACING)

        self.step3_upload_action = QPushButton()
        self.step3_upload_action.clicked.connect(self._on_step3_upload_clicked)
        step3_action_layout.addWidget(self.step3_upload_action)

        self.step3_generate_action = QPushButton()
        self.step3_generate_action.setObjectName("primaryAction")
        self.step3_generate_action.clicked.connect(self._on_step3_generate_clicked)
        step3_action_layout.addWidget(self.step3_generate_action)
        step3_action_layout.addStretch(1)
        top_layout.addWidget(self.step3_action_row)
        top_layout.addStretch(1)

        right_layout.addWidget(top_content, 1)

        self.info_tabs = QTabWidget()
        self.info_tabs.setObjectName("instructorInfoTabs")
        self.info_tabs.setFixedHeight(INSTRUCTOR_INFO_TAB_FIXED_HEIGHT)
        self.info_tabs.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.info_tabs.tabBar().setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.info_tabs.currentChanged.connect(self._on_info_tab_changed)

        log_tab = QWidget()
        log_tab_layout = QVBoxLayout(log_tab)
        log_tab_layout.setContentsMargins(*INSTRUCTOR_INFO_TAB_LAYOUT_MARGINS)
        log_tab_layout.setSpacing(INSTRUCTOR_INFO_TAB_LAYOUT_SPACING)

        self.user_log_view = QPlainTextEdit()
        self.user_log_view.setReadOnly(True)
        self.user_log_view.setObjectName("userLogView")
        self.user_log_view.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth)
        self.user_log_view.setFrameShape(QFrame.Shape.NoFrame)
        self.user_log_view.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        log_tab_layout.addWidget(self.user_log_view)

        links_tab = QWidget()
        links_tab_layout = QVBoxLayout(links_tab)
        links_tab_layout.setContentsMargins(*INSTRUCTOR_INFO_TAB_LAYOUT_MARGINS)
        links_tab_layout.setSpacing(INSTRUCTOR_INFO_TAB_LAYOUT_SPACING)

        self.generated_outputs_view = QTextBrowser()
        self.generated_outputs_view.setObjectName("generatedOutputsView")
        self.generated_outputs_view.setOpenExternalLinks(False)
        self.generated_outputs_view.setOpenLinks(False)
        self.generated_outputs_view.setFrameShape(QFrame.Shape.NoFrame)
        self.generated_outputs_view.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.generated_outputs_view.anchorClicked.connect(
            lambda url: self._on_quick_link_activated(url.toString())
        )
        links_tab_layout.addWidget(self.generated_outputs_view)

        # Retained for test doubles that call _refresh_quick_links directly.
        self.quick_link_labels: dict[str, QLabel] = {}

        self.info_tabs.addTab(log_tab, t("instructor.log.title"))
        self.info_tabs.addTab(links_tab, t(self.RAIL_LINK_TITLE_KEY))
        right_layout.addWidget(self.info_tabs)

        root.addWidget(left)
        root.addWidget(right, 1)

        self.setStyleSheet(INSTRUCTOR_PANEL_STYLESHEET)

    def _quick_link_items(self) -> tuple[tuple[str, str | None], ...]:
        return tuple((label_key, getattr(self, attr)) for label_key, attr in self.RAIL_LINKS)

    def _quick_link_markup(self, label_key: str, path: str | None) -> str:
        label = t(label_key)
        if not path:
            return f"{label}: {t(self.RAIL_LINK_NOT_AVAILABLE_KEY)}"
        file_link = f'<a href="file::{path}">{t(self.RAIL_LINK_OPEN_FILE_KEY)}</a>'
        folder_link = f'<a href="folder::{path}">{t(self.RAIL_LINK_OPEN_FOLDER_KEY)}</a>'
        name = escape(Path(path).name)
        full_path = escape(str(Path(path)))
        return (
            f"<b>{escape(label)}</b>: {name}<br>"
            f"<span>{full_path}</span><br>"
            f"{file_link} | {folder_link}"
        )

    def _quick_links_html(self) -> str:
        rows = [
            f"<div style='margin-bottom:10px'>{self._quick_link_markup(link_key, path)}</div>"
            for link_key, path in self._quick_link_items()
        ]
        return "".join(rows)

    def _on_quick_link_activated(self, href: str) -> None:
        mode, _, raw_path = href.partition("::")
        path = raw_path.strip()
        if not path:
            return
        target = Path(path).parent if mode == "folder" else Path(path)
        opened = QDesktopServices.openUrl(QUrl.fromLocalFile(str(target)))
        if opened:
            return
        show_toast(
            self,
            t(self.RAIL_LINK_OPEN_FAILED_KEY),
            title=t("instructor.msg.error_title"),
            level="error",
        )

    def _refresh_quick_links(self) -> None:
        generated_outputs_view = getattr(self, "generated_outputs_view", None)
        if generated_outputs_view is not None:
            generated_outputs_view.setHtml(self._quick_links_html())
        for link_key, path in self._quick_link_items():
            link_label = self.quick_link_labels.get(link_key)
            if link_label is None:
                continue
            link_label.setText(self._quick_link_markup(link_key, path))

    def _step_path(self, step: int) -> str | None:
        return getattr(self, self.PATH_ATTRS[step])

    def _step_done(self, step: int) -> bool:
        return bool(getattr(self, self.DONE_ATTRS[step]))

    def _step_outdated(self, step: int) -> bool:
        outdated_attr = self.OUTDATED_ATTRS.get(step)
        return bool(getattr(self, outdated_attr)) if outdated_attr else False

    def _step_state_text(self, step: int) -> str:
        done = self._step_done(step)
        outdated = self._step_outdated(step)
        if done and outdated:
            return t("instructor.badge.needs_update")
        return t("instructor.badge.done") if done else t("instructor.badge.pending")

    def _step_list_text(self, step: int) -> str:
        title = t(self.STEP_TITLE_KEYS[step])
        state = self._step_state_text(step)
        return f"{step}. {title}  {state}"

    def _action_text_for_step(self, step: int) -> str:
        return t(
            self.ACTION_REDO_KEYS[step]
            if self._step_done(step)
            else self.ACTION_DEFAULT_KEYS[step]
        )

    def _can_run_step(self, step: int) -> tuple[bool, str]:
        if step in (1, 2, 3):
            return True, ""
        return True, ""

    def _on_step_selected(self, step: int) -> None:
        self.current_step = step
        self.state.current_step = step
        self._refresh_ui()

    def _on_step_row_changed(self, row: int) -> None:
        if row < 0:
            return
        self._on_step_selected(row + 1)

    def _clear_info_text_selection(self) -> None:
        for view in (self.user_log_view, self.generated_outputs_view):
            cursor = view.textCursor()
            if cursor.hasSelection():
                cursor.clearSelection()
                view.setTextCursor(cursor)

    def _on_info_tab_changed(self, _index: int) -> None:
        self._clear_info_text_selection()

    def _refresh_ui(self) -> None:
        completed = sum(
            1
            for done, outdated in (
                (self.step1_done, False),
                (self.step2_done, False),
                (self.step4_done, self.step4_outdated),
            )
            if done and not outdated
        )
        self.progress_label.setText(
            t("instructor.progress", completed=completed, total=len(self.WORKFLOW_STEPS))
        )

        self.step_list.blockSignals(True)
        for index, item in enumerate(self._step_items, start=1):
            item.setText(self._step_list_text(index))
        self.step_list.setCurrentRow(self.current_step - 1)
        self.step_list.blockSignals(False)

        self.active_title.setText(
            t(
                "instructor.active_title",
                number=self.current_step,
                title=t(self.STEP_TITLE_KEYS[self.current_step]),
            )
        )
        self.active_desc.setText(t(self.STEP_DESC_KEYS[self.current_step]))
        self.info_tabs.setTabText(0, t("instructor.log.title"))
        self.info_tabs.setTabText(1, t(self.RAIL_LINK_TITLE_KEY))
        self._refresh_quick_links()
        can_run, reason = self._can_run_step(self.current_step)
        if self.current_step == 2:
            self.primary_action.setVisible(False)
            self.step2_action_row.setVisible(True)
            self.step3_action_row.setVisible(False)
            self.step2_upload_action.setText(t("instructor.action.step2.upload"))
            self.step2_prepare_action.setText(t("instructor.action.step2.prepare"))
            self.step2_prepare_action.setEnabled(self.step2_upload_ready)
            self.step2_upload_action.setDefault(not self.step2_upload_ready)
            self.step2_prepare_action.setDefault(self.step2_upload_ready)
        elif self.current_step == 3:
            self.primary_action.setVisible(False)
            self.step2_action_row.setVisible(False)
            self.step3_action_row.setVisible(True)
            self.step3_upload_action.setText(
                t("instructor.action.step4.redo")
                if self.step3_done and not self.step3_outdated
                else t("instructor.action.step4.default")
            )
            self.step3_generate_action.setText(
                t("instructor.action.step5.redo")
                if self.step4_done
                else t("instructor.action.step5.default")
            )
            self.step3_upload_action.setEnabled(True)
            self.step3_generate_action.setEnabled(
                self.step3_done and not self.step3_outdated
            )
            self.step3_upload_action.setDefault(not (self.step3_done and not self.step3_outdated))
            self.step3_generate_action.setDefault(self.step3_done and not self.step3_outdated)
        else:
            self.primary_action.setVisible(True)
            self.primary_action.setText(self._action_text_for_step(self.current_step))
            self.step2_action_row.setVisible(False)
            self.step3_action_row.setVisible(False)

        if self.current_step == 2:
            self.step2_upload_action.setEnabled(True)
        elif self.current_step == 3:
            pass
        else:
            self.primary_action.setEnabled(can_run)
        if not can_run:
            self.active_note.setText(reason)
        elif self.step3_outdated or self.step4_outdated:
            if self._step_outdated(self.current_step):
                self.active_note.setText(t("instructor.note.outdated_current"))
            else:
                self.active_note.setText(t("instructor.note.outdated_downstream"))
        else:
            self.active_note.setText(t("instructor.note.default"))

        if self.state.busy:
            self.primary_action.setEnabled(False)
            self.step2_upload_action.setEnabled(False)
            self.step2_prepare_action.setEnabled(False)
            self.step3_upload_action.setEnabled(False)
            self.step3_generate_action.setEnabled(False)

    def retranslate_ui(self) -> None:
        self._refresh_ui()
        self._clear_info_text_selection()

    def _run_current_step_action(self) -> None:
        if self.current_step == 1:
            self._download_course_template_async()
            self._refresh_ui()
            return
        if self.current_step == 2:
            self._upload_course_details_async()
            self._refresh_ui()
            return
        if self.current_step == 3:
            self._generate_final_report_async()
            self._refresh_ui()
            return
        self._refresh_ui()

    def _on_step2_upload_clicked(self) -> None:
        self._upload_course_details_async()
        self._refresh_ui()

    def _on_step2_prepare_clicked(self) -> None:
        self._prepare_marks_template_async()
        self._refresh_ui()

    def _on_step3_upload_clicked(self) -> None:
        self._upload_filled_marks_async()
        self._refresh_ui()

    def _on_step3_generate_clicked(self) -> None:
        self._generate_final_report_async()
        self._refresh_ui()

    def _remember_dialog_dir_safe(self, selected_path: str) -> None:
        try:
            remember_dialog_dir(selected_path, app_name=APP_NAME)
        except OSError:
            remember_dialog_dir_safe(
                selected_path,
                app_name=APP_NAME,
                logger=_logger,
            )

    def _setup_ui_logging(self) -> None:
        if self._ui_log_handler is not None:
            return
        self._ui_log_handler = _UILogHandler(self._append_user_log)
        _logger.addHandler(self._ui_log_handler)
        self._append_user_log(t("instructor.log.ready"))

    def _append_user_log(self, message: str) -> None:
        if not message or not message.strip():
            return
        text = message.strip()
        if not (
            len(text) >= 10
            and text[0] == "["
            and text[3] == ":"
            and text[6] == ":"
            and text[9] == "]"
        ):
            text = f"[{datetime.now().strftime('%H:%M:%S')}] {text}"
        self.user_log_view.appendPlainText(text)

    def _publish_status(self, message: str) -> None:
        self._append_user_log(message)
        emit_user_status(getattr(self, "status_changed", None), message, logger=_logger)

    def _set_busy(self, busy: bool, *, job_id: str | None = None) -> None:
        self.state.set_busy(busy, job_id=job_id)
        host_window = self.window()
        set_switch = getattr(host_window, "set_language_switch_enabled", None)
        if callable(set_switch):
            set_switch(not busy)
        self._refresh_ui()

    def _start_async_operation(
        self,
        *,
        token: CancellationToken,
        job_id: str | None,
        work,
        on_success,
        on_failure,
    ) -> None:
        self._cancel_token = token
        self._set_busy(True, job_id=job_id)
        job_ref: dict[str, object] = {}

        def _finalize() -> None:
            tracked_job = job_ref.get("job")
            if tracked_job in self._active_jobs:
                self._active_jobs.remove(tracked_job)
            self._cancel_token = None
            self._set_busy(False)
            if not self._is_closing:
                self._refresh_ui()

        def _on_finished(result: object) -> None:
            try:
                on_success(result)
            finally:
                _finalize()

        def _on_failed(exc: Exception) -> None:
            try:
                on_failure(exc)
            finally:
                _finalize()

        job = run_in_background(work, on_finished=_on_finished, on_failed=_on_failed)
        job_ref["job"] = job
        self._active_jobs.append(job)

    def closeEvent(self, event) -> None:
        self._is_closing = True
        if self._cancel_token is not None:
            self._cancel_token.cancel()
            self._cancel_token = None
        self._active_jobs.clear()
        if self._ui_log_handler is not None:
            _logger.removeHandler(self._ui_log_handler)
            self._ui_log_handler = None
        super().closeEvent(event)

    def _prepare_marks_template_async(self) -> None:
        if self.state.busy:
            return

        process_name = t("instructor.log.process.generate_marks_template")
        user_success_message, user_error_message = _localized_log_messages(
            "instructor.log.process.generate_marks_template"
        )
        if not self.step2_upload_ready or not self.step2_course_details_path:
            show_toast(
                self,
                t("instructor.require.step2"),
                title=t("instructor.msg.step_required_title"),
                level="info",
            )
            return

        source_path = self.step2_course_details_path
        default_name = getattr(
            self,
            "_step2_marks_default_name",
            t("instructor.dialog.step3.default_name"),
        ) or t("instructor.dialog.step3.default_name")
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            t("instructor.dialog.step3.title"),
            resolve_dialog_start_path(APP_NAME, default_name),
            t("instructor.dialog.filter.excel"),
        )
        if not save_path:
            return

        workflow_service = getattr(self, "_workflow_service", None)
        token = CancellationToken()
        job_context = (
            workflow_service.create_job_context(
                step_id="step2_generate_marks_template",
                payload={"source": source_path, "output": save_path},
            )
            if workflow_service is not None
            else None
        )

        def _on_finished(_result: object) -> None:
            self.step2_path = save_path
            self.step2_done = True
            self.step3_outdated = self.step3_done
            self.step4_outdated = self.step4_done
            self._remember_dialog_dir_safe(save_path)
            _publish_status_compat(self, t("instructor.status.step2_uploaded"))
            log_process_message(
                process_name,
                logger=_logger,
                success_message=f"{process_name} completed successfully.",
                user_success_message=user_success_message,
                job_id=job_context.job_id if job_context else None,
                step_id=job_context.step_id if job_context else None,
            )
            self._show_step_success_toast(2)

        def _on_failed(exc: Exception) -> None:
            if isinstance(exc, JobCancelledError):
                user_message = t("instructor.status.operation_cancelled")
                _publish_status_compat(self, user_message)
                _logger.info(
                    "%s cancelled by user/system request.",
                    process_name,
                    extra={
                        "user_message": user_message,
                        "job_id": job_context.job_id if job_context else None,
                        "step_id": job_context.step_id if job_context else None,
                    },
                )
                return
            if isinstance(exc, ValidationError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    notify=lambda message, _level: self._show_validation_error_toast(message),
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
            elif isinstance(exc, AppSystemError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(2)
            else:
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(2)

        def _work() -> Path:
            if workflow_service is not None and job_context is not None:
                return workflow_service.generate_marks_template(
                    source_path,
                    save_path,
                    context=job_context,
                    cancel_token=token,
                )
            generate_marks_template_from_course_details(source_path, save_path)
            return Path(save_path)

        _start_async_operation_compat(
            self,
            token=token,
            job_id=job_context.job_id if job_context else None,
            work=_work,
            on_success=_on_finished,
            on_failure=_on_failed,
        )

    def _download_course_template_async(self) -> None:
        if self.state.busy:
            return

        template_id = ID_COURSE_SETUP
        process_name = t("instructor.log.process.generate_course_details_template")
        user_success_message, user_error_message = _localized_log_messages(
            "instructor.log.process.generate_course_details_template"
        )

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            t("instructor.dialog.step1.title"),
            resolve_dialog_start_path(APP_NAME, t("instructor.dialog.step1.default_name")),
            t("instructor.dialog.filter.excel"),
        )
        if not save_path:
            return

        workflow_service = getattr(self, "_workflow_service", None)
        token = CancellationToken()
        job_context = (
            workflow_service.create_job_context(
                step_id="step1_generate_course_template",
                payload={"template_id": template_id, "output": save_path},
            )
            if workflow_service is not None
            else None
        )

        def _on_finished(_result: object) -> None:
            self.step1_path = save_path
            self.step1_done = True
            self._remember_dialog_dir_safe(save_path)
            _publish_status_compat(self, t("instructor.status.step1_selected"))
            log_process_message(
                process_name,
                logger=_logger,
                success_message=f"{process_name} completed successfully.",
                user_success_message=user_success_message,
                job_id=job_context.job_id if job_context else None,
                step_id=job_context.step_id if job_context else None,
            )
            self._show_step_success_toast(1)

        def _on_failed(exc: Exception) -> None:
            if isinstance(exc, JobCancelledError):
                user_message = t("instructor.status.operation_cancelled")
                _publish_status_compat(self, user_message)
                _logger.info(
                    "%s cancelled by user/system request.",
                    process_name,
                    extra={
                        "user_message": user_message,
                        "job_id": job_context.job_id if job_context else None,
                        "step_id": job_context.step_id if job_context else None,
                    },
                )
                return
            if isinstance(exc, ValidationError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    notify=lambda message, _level: self._show_validation_error_toast(message),
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
            elif isinstance(exc, AppSystemError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(1)
            else:
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(1)

        def _work() -> Path:
            if workflow_service is not None and job_context is not None:
                return workflow_service.generate_course_details_template(
                    save_path,
                    context=job_context,
                    cancel_token=token,
                )
            generate_course_details_template(save_path, template_id=template_id)
            return Path(save_path)

        _start_async_operation_compat(
            self,
            token=token,
            job_id=job_context.job_id if job_context else None,
            work=_work,
            on_success=_on_finished,
            on_failure=_on_failed,
        )

    def _upload_course_details_async(self) -> None:
        if self.state.busy:
            return

        process_name = t("instructor.log.process.validate_course_details_workbook")
        user_success_message, user_error_message = _localized_log_messages(
            "instructor.log.process.validate_course_details_workbook"
        )
        open_path, _ = QFileDialog.getOpenFileName(
            self,
            t("instructor.dialog.step2.title"),
            resolve_dialog_start_path(APP_NAME),
            t("instructor.dialog.filter.excel_open"),
        )
        if not open_path:
            return

        workflow_service = getattr(self, "_workflow_service", None)
        token = CancellationToken()
        job_context = (
            workflow_service.create_job_context(
                step_id="step2_validate_course_details",
                payload={"path": open_path},
            )
            if workflow_service is not None
            else None
        )

        def _on_finished(result: object) -> None:
            replacing = self.step2_done or self.step2_upload_ready
            self.step2_course_details_path = open_path
            self.step2_upload_ready = True
            self.step2_done = False
            self.step2_path = None
            self._step2_marks_default_name = (
                result.get("default_marks_name")
                if isinstance(result, dict)
                else t("instructor.dialog.step3.default_name")
            ) or t("instructor.dialog.step3.default_name")
            self._remember_dialog_dir_safe(open_path)

            if replacing:
                self.step4_outdated = self.step4_done
                self.step3_outdated = self.step3_done
                if self.step3_outdated or self.step4_outdated:
                    _publish_status_compat(self, t("instructor.status.step2_changed"))
            else:
                _publish_status_compat(self, t("instructor.status.step2_validated"))
            log_process_message(
                process_name,
                logger=_logger,
                success_message=f"{process_name} completed successfully.",
                user_success_message=user_success_message,
                job_id=job_context.job_id if job_context else None,
                step_id=job_context.step_id if job_context else None,
            )
            self._show_step_success_toast(2)

        def _on_failed(exc: Exception) -> None:
            if isinstance(exc, JobCancelledError):
                user_message = t("instructor.status.operation_cancelled")
                _publish_status_compat(self, user_message)
                _logger.info(
                    "%s cancelled by user/system request.",
                    process_name,
                    extra={
                        "user_message": user_message,
                        "job_id": job_context.job_id if job_context else None,
                        "step_id": job_context.step_id if job_context else None,
                    },
                )
                return
            if isinstance(exc, ValidationError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    notify=lambda message, _level: self._show_validation_error_toast(message),
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
            elif isinstance(exc, AppSystemError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(2)
            else:
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(2)

        def _work() -> dict[str, str]:
            if workflow_service is not None and job_context is not None:
                workflow_service.validate_course_details_workbook(
                    open_path,
                    context=job_context,
                    cancel_token=token,
                )
            else:
                validate_course_details_workbook(open_path)
            token.raise_if_cancelled()
            return {
                "default_marks_name": _build_marks_template_default_name(open_path),
            }

        _start_async_operation_compat(
            self,
            token=token,
            job_id=job_context.job_id if job_context else None,
            work=_work,
            on_success=_on_finished,
            on_failure=_on_failed,
        )

    def _upload_filled_marks_async(self) -> None:
        if self.state.busy:
            return

        process_name = t("instructor.log.process.upload_filled_marks_workbook")
        user_success_message, user_error_message = _localized_log_messages(
            "instructor.log.process.upload_filled_marks_workbook"
        )
        open_path, _ = QFileDialog.getOpenFileName(
            self,
            t("instructor.dialog.step3.title"),
            resolve_dialog_start_path(APP_NAME),
            t("instructor.dialog.filter.excel_open"),
        )
        if not open_path:
            return

        token = CancellationToken()
        workflow_service = getattr(self, "_workflow_service", None)
        job_context = (
            workflow_service.create_job_context(
                step_id="step3_upload_filled_marks",
                payload={"path": open_path},
            )
            if workflow_service is not None
            else None
        )

        def _on_finished(_result: object) -> None:
            replacing = self.step3_done
            self.step3_path = open_path
            self.step3_done = True
            self.step3_outdated = False
            self._remember_dialog_dir_safe(open_path)

            if replacing and self.step3_done:
                self.step4_outdated = True
                _publish_status_compat(self, t("instructor.status.step3_changed"))
            else:
                _publish_status_compat(self, t("instructor.status.step3_uploaded"))
            log_process_message(
                process_name,
                logger=_logger,
                success_message=f"{process_name} completed successfully.",
                user_success_message=user_success_message,
                job_id=job_context.job_id if job_context else None,
                step_id=job_context.step_id if job_context else None,
            )
            self._show_step_success_toast(3)

        def _on_failed(exc: Exception) -> None:
            if isinstance(exc, JobCancelledError):
                user_message = t("instructor.status.operation_cancelled")
                _publish_status_compat(self, user_message)
                _logger.info(
                    "%s cancelled by user/system request.",
                    process_name,
                    extra={
                        "user_message": user_message,
                        "job_id": job_context.job_id if job_context else None,
                        "step_id": job_context.step_id if job_context else None,
                    },
                )
                return
            if isinstance(exc, ValidationError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    notify=lambda message, _level: self._show_validation_error_toast(message),
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
            elif isinstance(exc, AppSystemError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(3)
            else:
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(3)

        def _work() -> bool:
            token.raise_if_cancelled()
            _validate_uploaded_filled_marks_workbook(open_path)
            token.raise_if_cancelled()
            return True

        _start_async_operation_compat(
            self,
            token=token,
            job_id=job_context.job_id if job_context else None,
            work=_work,
            on_success=_on_finished,
            on_failure=_on_failed,
        )

    def _generate_final_report_async(self) -> None:
        if self.state.busy:
            return

        process_name = t("instructor.log.process.generate_final_co_report")
        user_success_message, user_error_message = _localized_log_messages(
            "instructor.log.process.generate_final_co_report"
        )
        can_run, reason = self._can_run_step(3)
        if not can_run:
            show_toast(
                self,
                reason,
                title=t("instructor.msg.step_required_title"),
                level="info",
            )
            return
        if not self.step3_done or self.step3_outdated:
            show_toast(
                self,
                t("instructor.require.step3"),
                title=t("instructor.msg.step_required_title"),
                level="info",
            )
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            t("instructor.dialog.step4.title"),
            resolve_dialog_start_path(APP_NAME, t("instructor.dialog.step4.default_name")),
            t("instructor.dialog.filter.excel"),
        )
        if not save_path:
            return

        if not self.step3_path or not Path(self.step3_path).exists():
            _logger.warning("Step 4 failed: Step 3 file is missing. step3_path=%s", self.step3_path)
            show_toast(
                self,
                t("instructor.require.step3"),
                title=t("instructor.msg.step_required_title"),
                level="error",
            )
            return
        source_path = self.step3_path

        workflow_service = getattr(self, "_workflow_service", None)
        token = CancellationToken()
        job_context = (
            workflow_service.create_job_context(
                step_id="step3_generate_final_report",
                payload={"source": source_path, "output": save_path},
            )
            if workflow_service is not None
            else None
        )

        def _on_finished(_result: object) -> None:
            self.step4_path = save_path
            self.step4_done = True
            self.step4_outdated = False
            self._remember_dialog_dir_safe(save_path)
            _publish_status_compat(self, t("instructor.status.step4_selected"))
            log_process_message(
                process_name,
                logger=_logger,
                success_message=f"{process_name} completed successfully.",
                user_success_message=user_success_message,
                job_id=job_context.job_id if job_context else None,
                step_id=job_context.step_id if job_context else None,
            )
            self._show_step_success_toast(3)

        def _on_failed(exc: Exception) -> None:
            if isinstance(exc, JobCancelledError):
                user_message = t("instructor.status.operation_cancelled")
                _publish_status_compat(self, user_message)
                _logger.info(
                    "%s cancelled by user/system request.",
                    process_name,
                    extra={
                        "user_message": user_message,
                        "job_id": job_context.job_id if job_context else None,
                        "step_id": job_context.step_id if job_context else None,
                    },
                )
                return
            if isinstance(exc, ValidationError):
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    notify=lambda message, _level: self._show_validation_error_toast(message),
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
            else:
                log_process_message(
                    process_name,
                    logger=_logger,
                    error=exc,
                    user_error_message=user_error_message,
                    job_id=job_context.job_id if job_context else None,
                    step_id=job_context.step_id if job_context else None,
                )
                self._show_system_error_toast(3)

        def _work() -> Path:
            if workflow_service is not None and job_context is not None:
                return workflow_service.generate_final_report(
                    source_path,
                    save_path,
                    context=job_context,
                    cancel_token=token,
                )
            return _atomic_copy_file(source_path, save_path)

        _start_async_operation_compat(
            self,
            token=token,
            job_id=job_context.job_id if job_context else None,
            work=_work,
            on_success=_on_finished,
            on_failure=_on_failed,
        )

    def _show_step_success_toast(self, step: int) -> None:
        show_toast(
            self,
            t("instructor.msg.step_completed", step=step, title=t(self.STEP_TITLE_KEYS[step])),
            title=t("instructor.msg.success_title"),
            level="success",
        )

    def _show_validation_error_toast(self, message: str) -> None:
        show_toast(
            self,
            message,
            title=t("instructor.msg.validation_title"),
            level="error",
        )

    def _show_system_error_toast(self, step: int) -> None:
        show_toast(
            self,
            t("instructor.msg.failed_to_do", action=t(self.STEP_TITLE_KEYS[step])),
            title=t("instructor.msg.error_title"),
            level="error",
        )

    def _download_course_template(self) -> None:
        InstructorModule._download_course_template_async(self)

    def _upload_course_details(self) -> None:
        InstructorModule._upload_course_details_async(self)

    def _prepare_marks_template(self) -> None:
        InstructorModule._prepare_marks_template_async(self)

    def _upload_filled_marks(self) -> None:
        InstructorModule._upload_filled_marks_async(self)

    def _generate_final_report(self) -> None:
        InstructorModule._generate_final_report_async(self)
