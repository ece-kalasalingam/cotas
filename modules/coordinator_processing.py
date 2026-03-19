"""Processing helpers for the coordinator module."""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import sqlite3
import tempfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from common.constants import (
    APP_NAME,
    CO_REPORT_ABSENT_TOKEN,
    CO_REPORT_DIRECT_SHEET_SUFFIX,
    CO_REPORT_HEADER_REG_NO,
    CO_REPORT_HEADER_SERIAL,
    CO_REPORT_HEADER_STUDENT_NAME,
    CO_REPORT_HEADER_TOTAL_RATIO_TEMPLATE,
    CO_REPORT_INDIRECT_SHEET_SUFFIX,
    CO_REPORT_MAX_DECIMAL_PLACES,
    CO_REPORT_NOT_APPLICABLE_TOKEN,
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_HEADERS,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_SHEET,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    DIRECT_RATIO,
    ID_COURSE_SETUP,
    INDIRECT_RATIO,
    LEVEL_1_THRESHOLD,
    LEVEL_2_THRESHOLD,
    LEVEL_3_THRESHOLD,
    SYSTEM_HASH_SHEET,
    SYSTEM_HASH_TEMPLATE_HASH_HEADER,
    SYSTEM_HASH_TEMPLATE_ID_HEADER,
    SYSTEM_REPORT_INTEGRITY_HASH_HEADER,
    SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER,
    SYSTEM_REPORT_INTEGRITY_SHEET,
)
from common.excel_sheet_layout import color_without_hash as _color_without_hash
from common.excel_sheet_layout import (
    compute_sampled_column_widths as _compute_sampled_column_widths,
)
from common.excel_sheet_layout import (
    style_registry_for_setup as _style_registry_for_setup,
)
from common.jobs import CancellationToken
from common.utils import (
    app_runtime_storage_dir,
    coerce_excel_number,
    create_app_runtime_sqlite_file,
    normalize,
)
from common.workbook_signing import sign_payload, verify_payload_signature

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}

_logger = logging.getLogger(__name__)
_COURSE_METADATA_COURSE_NAME_KEY = "course_name"
_STYLE_CACHE_ATTR = "_focus_coordinator_style_cache"
_STYLE_KEY_BORDER = "border"
_STYLE_KEY_BG_COLOR = "bg_color"
_STYLE_KEY_ALIGN = "align"
_STYLE_KEY_VALIGN = "valign"
_STYLE_KEY_BOLD = "bold"
_ALIGN_CENTER = "center"
_ALIGN_VCENTER = "vcenter"
_CO_REPORT_NAME_TOKEN_RE = re.compile(r"(?:[_\-\s]*co[_\-\s]*report)+$", re.IGNORECASE)
_HEADER_SCAN_MAX_ROWS = 200
_COORDINATOR_STUDENTS_PER_SHEET = 150
_NORM_SYSTEM_HASH_TEMPLATE_ID_HEADER = normalize(SYSTEM_HASH_TEMPLATE_ID_HEADER)
_NORM_SYSTEM_HASH_TEMPLATE_HASH_HEADER = normalize(SYSTEM_HASH_TEMPLATE_HASH_HEADER)
_NORM_SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER = normalize(SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER)
_NORM_SYSTEM_REPORT_INTEGRITY_HASH_HEADER = normalize(SYSTEM_REPORT_INTEGRITY_HASH_HEADER)
_NORM_COURSE_SETUP_ID = normalize(ID_COURSE_SETUP)
_NORM_COURSE_CODE_KEY = normalize(COURSE_METADATA_COURSE_CODE_KEY)
_NORM_TOTAL_OUTCOMES_KEY = normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY)
_NORM_SECTION_KEY = normalize(COURSE_METADATA_SECTION_KEY)
_DEDUP_SQLITE_THRESHOLD_ENTRIES = 1
_DEDUP_SQLITE_PREFIX = "focus_co_dedup_"
_DEDUP_SQLITE_SUFFIX = ".sqlite3"
_INTEGRITY_SCHEMA_VERSION = 1
_SIGNATURE_VALIDATION_MAX_WORKERS = 8


@dataclass(slots=True, frozen=True)
class _FinalReportSignature:
    template_id: str
    course_code: str
    total_outcomes: int
    section: str
    direct_sheet_count: int
    indirect_sheet_count: int


@dataclass(slots=True, frozen=True)
class _CoAttainmentRow:
    reg_hash: int
    reg_no: str
    student_name: str
    direct_score: float | str
    indirect_score: float | str
    worksheet_name: str
    workbook_name: str


@dataclass(slots=True, frozen=True)
class _ParsedScoreRow:
    reg_hash: int
    reg_key: str
    reg_no: str
    student_name: str
    score: float | str


@dataclass(slots=True)
class _CoOutputSheetState:
    sheet: Any
    header_row_index: int
    formats: dict[str, Any]
    next_row_index: int
    next_serial: int
    on_roll: int
    attended: int
    level_counts: dict[int, int]


@dataclass(slots=True, frozen=True)
class _CoAttainmentWorkbookResult:
    output_path: Path
    duplicate_reg_count: int
    duplicate_entries: tuple[tuple[str, str, str], ...]


class _RegisterDedupStore:
    def __init__(self, *, total_outcomes: int, use_sqlite: bool) -> None:
        self._total_outcomes = total_outcomes
        self._use_sqlite = use_sqlite
        self._memory_sets: dict[int, set[int]] = {}
        self._conn: sqlite3.Connection | None = None
        self._db_path: str | None = None
        if use_sqlite:
            _cleanup_stale_dedup_sqlite_files()
            fd, db_path = create_app_runtime_sqlite_file(
                APP_NAME,
                prefix=_DEDUP_SQLITE_PREFIX,
                suffix=_DEDUP_SQLITE_SUFFIX,
            )
            self._db_path = db_path
            self._conn = sqlite3.connect(db_path)
            self._conn.execute("PRAGMA journal_mode=OFF")
            self._conn.execute("PRAGMA synchronous=OFF")
            self._conn.execute("PRAGMA temp_store=MEMORY")
            self._conn.execute("PRAGMA secure_delete=ON")
            self._conn.execute(
                "CREATE TABLE IF NOT EXISTS dedup (co_index INTEGER NOT NULL, reg_hash INTEGER NOT NULL, PRIMARY KEY(co_index, reg_hash))"
            )
            self._conn.commit()
            try:
                import os

                os.close(fd)
            except OSError:
                pass
        else:
            self._memory_sets = {co: set() for co in range(1, total_outcomes + 1)}

    def add_if_absent(self, *, co_index: int, reg_hash: int) -> bool:
        if self._use_sqlite and self._conn is not None:
            cursor = self._conn.execute(
                "INSERT OR IGNORE INTO dedup (co_index, reg_hash) VALUES (?, ?)",
                (co_index, int(reg_hash)),
            )
            return cursor.rowcount == 1
        bucket = self._memory_sets.setdefault(co_index, set())
        if reg_hash in bucket:
            return False
        bucket.add(reg_hash)
        return True

    def close(self) -> None:
        if self._conn is not None:
            try:
                # Explicitly wipe the sqlite table before close in normal paths.
                self._conn.execute("DELETE FROM dedup")
                self._conn.commit()
            except Exception:
                _logger.debug("Failed to clear sqlite dedup table before close.", exc_info=True)
            self._conn.close()
            self._conn = None
        if self._db_path:
            try:
                Path(self._db_path).unlink(missing_ok=True)
            except OSError:
                _logger.debug("Unable to remove dedup sqlite db: %s", self._db_path, exc_info=True)
            self._db_path = None


def _cleanup_stale_dedup_sqlite_files() -> None:
    temp_root = Path(tempfile.gettempdir())
    storage_sqlite_root = app_runtime_storage_dir(APP_NAME) / "sqlite"
    roots: tuple[Path, ...] = (temp_root, storage_sqlite_root)
    patterns = (
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}",
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}-wal",
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}-shm",
    )
    for root in roots:
        if not root.exists():
            continue
        for pattern in patterns:
            for path in root.glob(pattern):
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    _logger.debug("Unable to remove stale dedup sqlite file: %s", path, exc_info=True)


def _path_key(path: Path) -> str:
    return str(path.resolve()).casefold()


def _build_co_attainment_default_name(source_path: Path, *, section: str = "") -> str:
    stem = source_path.stem.strip()
    cleaned = _CO_REPORT_NAME_TOKEN_RE.sub("", stem).rstrip("_- ").strip()
    section_token = section.strip()
    if section_token:
        parts = [part for part in cleaned.split("_") if normalize(part) != normalize(section_token)]
        cleaned = "_".join(parts).strip("_- ")
    base = cleaned if cleaned else stem
    return f"{base}_CO_Attainment.xlsx"


def _is_supported_excel_file(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES


def _filter_excel_paths(paths: Iterable[str]) -> list[Path]:
    collected: list[Path] = []
    seen: set[str] = set()
    for value in paths:
        path = Path(value)
        if not _is_supported_excel_file(path):
            continue
        key = _path_key(path)
        if key in seen:
            continue
        seen.add(key)
        collected.append(path.resolve())
    return collected


def _has_valid_final_co_report(path: Path) -> bool:
    return _extract_final_report_signature(path) is not None


def _read_template_id_from_hash_sheet(workbook: Any) -> str | None:
    if SYSTEM_HASH_SHEET not in workbook.sheetnames:
        return None
    hash_sheet = workbook[SYSTEM_HASH_SHEET]
    if normalize(hash_sheet["A1"].value) != _NORM_SYSTEM_HASH_TEMPLATE_ID_HEADER:
        return None
    if normalize(hash_sheet["B1"].value) != _NORM_SYSTEM_HASH_TEMPLATE_HASH_HEADER:
        return None
    template_id_raw = hash_sheet["A2"].value
    template_hash_raw = hash_sheet["B2"].value
    template_id = str(template_id_raw).strip() if template_id_raw is not None else ""
    template_hash = str(template_hash_raw).strip() if template_hash_raw is not None else ""
    if not template_id or not template_hash:
        return None
    if not verify_payload_signature(template_id, template_hash):
        return None
    if normalize(template_id) != _NORM_COURSE_SETUP_ID:
        return None
    return template_id


def _read_report_sheet_counts(workbook: Any) -> tuple[int, int] | None:
    if SYSTEM_REPORT_INTEGRITY_SHEET not in workbook.sheetnames:
        return None
    integrity_sheet = workbook[SYSTEM_REPORT_INTEGRITY_SHEET]
    if normalize(integrity_sheet["A1"].value) != _NORM_SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER:
        return None
    if normalize(integrity_sheet["B1"].value) != _NORM_SYSTEM_REPORT_INTEGRITY_HASH_HEADER:
        return None
    manifest_text_raw = integrity_sheet["A2"].value
    manifest_hash_raw = integrity_sheet["B2"].value
    manifest_text = str(manifest_text_raw).strip() if manifest_text_raw is not None else ""
    manifest_hash = str(manifest_hash_raw).strip() if manifest_hash_raw is not None else ""
    if not manifest_text or not manifest_hash:
        return None
    if not verify_payload_signature(manifest_text, manifest_hash):
        return None

    manifest = json.loads(manifest_text)
    if not isinstance(manifest, dict):
        return None
    sheet_order = manifest.get("sheet_order")
    if not isinstance(sheet_order, list):
        return None
    sheet_names = [str(name).strip() for name in sheet_order if isinstance(name, str)]
    if not sheet_names:
        return None
    direct_sheet_count = sum(1 for name in sheet_names if name.endswith(CO_REPORT_DIRECT_SHEET_SUFFIX))
    indirect_sheet_count = sum(1 for name in sheet_names if name.endswith(CO_REPORT_INDIRECT_SHEET_SUFFIX))
    if direct_sheet_count <= 0 or indirect_sheet_count <= 0:
        return None
    if direct_sheet_count != indirect_sheet_count:
        return None
    return direct_sheet_count, indirect_sheet_count


def _read_signature_metadata(workbook: Any) -> tuple[str, int, str] | None:
    if COURSE_METADATA_SHEET not in workbook.sheetnames:
        return None
    metadata_sheet = workbook[COURSE_METADATA_SHEET]
    course_code = ""
    total_outcomes_text = ""
    section = ""
    row = 2
    while True:
        key = metadata_sheet.cell(row=row, column=1).value
        value = metadata_sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        key_text = str(key).strip() if key is not None else ""
        value_text = str(value).strip() if value is not None else ""
        if key_text:
            normalized_key = normalize(key_text)
            if normalized_key == _NORM_COURSE_CODE_KEY:
                course_code = value_text
            elif normalized_key == _NORM_TOTAL_OUTCOMES_KEY:
                total_outcomes_text = value_text
            elif normalized_key == _NORM_SECTION_KEY:
                section = value_text
            if course_code and total_outcomes_text and section:
                break
        row += 1

    course_code = course_code.strip()
    total_outcomes_text = total_outcomes_text.strip()
    section = section.strip()
    if not course_code or not total_outcomes_text or not section:
        return None
    try:
        total_outcomes = int(float(total_outcomes_text))
    except (TypeError, ValueError):
        return None
    if total_outcomes <= 0:
        return None
    return course_code, total_outcomes, section


def _extract_final_report_signature(path: Path) -> _FinalReportSignature | None:
    if path.suffix.lower() == ".xls":
        return None
    try:
        from openpyxl import load_workbook
    except Exception:
        _logger.exception("openpyxl is unavailable while validating '%s'.", path)
        return None
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            template_id = _read_template_id_from_hash_sheet(workbook)
            if template_id is None:
                return None
            sheet_counts = _read_report_sheet_counts(workbook)
            if sheet_counts is None:
                return None
            metadata = _read_signature_metadata(workbook)
            if metadata is None:
                return None
            direct_sheet_count, indirect_sheet_count = sheet_counts
            course_code, total_outcomes, section = metadata

            return _FinalReportSignature(
                template_id=template_id,
                course_code=course_code,
                total_outcomes=total_outcomes,
                section=section,
                direct_sheet_count=direct_sheet_count,
                indirect_sheet_count=indirect_sheet_count,
            )
        finally:
            workbook.close()
    except Exception:
        _logger.exception("Failed to validate final CO report workbook '%s'.", path)
        return None


def _analyze_dropped_files(
    dropped_files: list[str],
    *,
    existing_keys: set[str],
    existing_paths: list[str],
    token: CancellationToken,
) -> dict[str, object]:
    accepted = _filter_excel_paths(dropped_files)
    seen = set(existing_keys)
    added: list[str] = []
    duplicates = 0
    invalid_final_report: list[str] = []
    existing_resolved = [Path(path).resolve() for path in existing_paths if path]
    baseline_signature: _FinalReportSignature | None = None
    seen_sections: set[str] = set()
    signature_cache: dict[str, _FinalReportSignature | None] = {}

    unique_paths: list[Path] = []
    seen_signature_keys: set[str] = set()
    for path in [*existing_resolved, *accepted]:
        key = _path_key(path)
        if key in seen_signature_keys:
            continue
        seen_signature_keys.add(key)
        unique_paths.append(path)

    max_workers = min(
        _SIGNATURE_VALIDATION_MAX_WORKERS,
        max(1, os.cpu_count() or 1),
        len(unique_paths) or 1,
    )
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_pairs = [(path, executor.submit(_extract_final_report_signature, path)) for path in unique_paths]
        for path, future in future_pairs:
            token.raise_if_cancelled()
            signature_cache[_path_key(path)] = future.result()

    def _cached_signature(path: Path) -> _FinalReportSignature | None:
        return signature_cache.get(_path_key(path))

    for path in existing_resolved:
        token.raise_if_cancelled()
        signature = _cached_signature(path)
        if signature is None:
            continue
        if baseline_signature is None:
            baseline_signature = signature
        section_key = normalize(signature.section)
        if section_key:
            seen_sections.add(section_key)

    for path in accepted:
        token.raise_if_cancelled()
        key = _path_key(path)
        if key in seen:
            duplicates += 1
            continue
        signature = _cached_signature(path)
        if signature is None:
            invalid_final_report.append(str(path))
            continue
        if baseline_signature is None:
            baseline_signature = signature
        else:
            is_mismatch = (
                signature.template_id != baseline_signature.template_id
                or signature.course_code != baseline_signature.course_code
                or signature.total_outcomes != baseline_signature.total_outcomes
                or signature.direct_sheet_count != baseline_signature.direct_sheet_count
                or signature.indirect_sheet_count != baseline_signature.indirect_sheet_count
            )
            if is_mismatch or normalize(signature.section) in seen_sections:
                invalid_final_report.append(str(path))
                continue
        seen_sections.add(normalize(signature.section))
        seen.add(key)
        added.append(str(path))

    ignored = (len(dropped_files) - len(accepted)) + duplicates + len(invalid_final_report)
    return {
        "added": added,
        "duplicates": duplicates,
        "invalid_final_report": invalid_final_report,
        "ignored": ignored,
    }


def _ratio_percent_token(ratio: float) -> str:
    percent = ratio * 100.0
    if abs(percent - round(percent)) <= 1e-9:
        return f"{int(round(percent))}"
    return f"{percent:g}"


def _ratio_total_header(ratio: float) -> str:
    return CO_REPORT_HEADER_TOTAL_RATIO_TEMPLATE.format(ratio=_ratio_percent_token(ratio))


def _coerce_numeric_score(value: Any) -> float | str | None:
    if normalize(value) == normalize(CO_REPORT_NOT_APPLICABLE_TOKEN):
        return CO_REPORT_ABSENT_TOKEN
    parsed = coerce_excel_number(value)
    if isinstance(parsed, bool):
        return None
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


def _metadata_rows_for_output(metadata: dict[str, str], co_index: int) -> list[tuple[str, str]]:
    return [
        ("Course Code", metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")),
        ("Course Name", metadata.get(normalize(_COURSE_METADATA_COURSE_NAME_KEY), "")),
        ("Semester", metadata.get(normalize(COURSE_METADATA_SEMESTER_KEY), "")),
        ("Academic Year", metadata.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")),
        ("CO Number", f"CO{co_index}"),
    ]


def _extract_course_metadata_fields(sheet: Any) -> dict[str, str]:
    metadata: dict[str, str] = {}
    row = 2
    while True:
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        key_text = str(key).strip() if key is not None else ""
        if key_text:
            coerced = coerce_excel_number(value)
            metadata[normalize(key_text)] = str(coerced).strip() if coerced is not None else ""
        row += 1
    return metadata


def _stable_reg_hash(reg_key: str) -> int:
    # Use a compact, stable 48-bit ID to reduce sqlite footprint while keeping collision risk very low.
    return int.from_bytes(hashlib.blake2b(reg_key.encode("utf-8"), digest_size=6).digest(), byteorder="big")


def _iter_score_rows(sheet: Any, *, ratio: float) -> Iterable[_ParsedScoreRow]:
    required_headers = {
        normalize(CO_REPORT_HEADER_SERIAL),
        normalize(CO_REPORT_HEADER_REG_NO),
        normalize(CO_REPORT_HEADER_STUDENT_NAME),
        normalize(_ratio_total_header(ratio)),
    }
    header_row = 0
    column_map: dict[str, int] = {}
    max_row = int(sheet.max_row)
    max_col = int(sheet.max_column)
    header_scan_limit = min(max_row, _HEADER_SCAN_MAX_ROWS)

    def _find_header_row(start_row: int, end_row: int) -> tuple[int, dict[str, int]]:
        for row_offset, values in enumerate(
            sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ),
            start=0,
        ):
            row_idx = start_row + row_offset
            row_map: dict[str, int] = {}
            missing = set(required_headers)
            for col_idx, value in enumerate(values, start=1):
                key = normalize(value)
                if key and key not in row_map:
                    row_map[key] = col_idx
                    if key in missing:
                        missing.remove(key)
                        if not missing:
                            return row_idx, row_map
        return 0, {}

    header_row, column_map = _find_header_row(1, header_scan_limit)
    if header_row <= 0 and header_scan_limit < max_row:
        header_row, column_map = _find_header_row(header_scan_limit + 1, max_row)
    if header_row <= 0:
        raise ValueError(f"Required headers are missing in sheet '{sheet.title}'.")

    reg_col = column_map[normalize(CO_REPORT_HEADER_REG_NO)]
    name_col = column_map[normalize(CO_REPORT_HEADER_STUDENT_NAME)]
    score_col = column_map[normalize(_ratio_total_header(ratio))]

    seen_in_sheet: set[int] = set()
    for values in sheet.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if len(values) < max(reg_col, name_col, score_col):
            continue
        reg_raw = values[reg_col - 1]
        reg_value = coerce_excel_number(reg_raw)
        reg_no = str(reg_value).strip() if reg_value is not None else ""
        if not reg_no:
            continue
        reg_key = normalize(reg_no)
        reg_hash = _stable_reg_hash(reg_key)
        if reg_hash in seen_in_sheet:
            continue

        score = _coerce_numeric_score(values[score_col - 1])
        if score is None:
            continue
        student_raw = values[name_col - 1]
        student_name = str(student_raw).strip() if student_raw is not None else ""
        normalized_score = round(score, CO_REPORT_MAX_DECIMAL_PLACES) if isinstance(score, (int, float)) else score
        seen_in_sheet.add(reg_hash)
        yield _ParsedScoreRow(
            reg_hash=reg_hash,
            reg_key=reg_key,
            reg_no=reg_no,
            student_name=student_name,
            score=normalized_score,
        )


def _iter_co_rows_from_workbook(workbook: Any, *, co_index: int, workbook_name: str) -> Iterable[_CoAttainmentRow]:
    direct_name = f"CO{co_index}{CO_REPORT_DIRECT_SHEET_SUFFIX}"
    indirect_name = f"CO{co_index}{CO_REPORT_INDIRECT_SHEET_SUFFIX}"
    if direct_name not in workbook.sheetnames or indirect_name not in workbook.sheetnames:
        raise ValueError(f"Missing CO sheets for CO{co_index}.")

    indirect_lookup: dict[tuple[int, str], _ParsedScoreRow] = {}
    for item in _iter_score_rows(workbook[indirect_name], ratio=INDIRECT_RATIO):
        indirect_lookup.setdefault((item.reg_hash, item.reg_key), item)

    for direct_row in _iter_score_rows(workbook[direct_name], ratio=DIRECT_RATIO):
        key = (direct_row.reg_hash, direct_row.reg_key)
        match = indirect_lookup.get(key)
        if match is None:
            continue

        student_name = direct_row.student_name or match.student_name
        yield _CoAttainmentRow(
            reg_hash=direct_row.reg_hash,
            reg_no=direct_row.reg_no,
            student_name=student_name,
            direct_score=direct_row.score,
            indirect_score=match.score,
            worksheet_name=direct_name,
            workbook_name=workbook_name,
        )


def _xlsxwriter_formats(workbook: Any) -> dict[str, Any]:
    cached = getattr(workbook, _STYLE_CACHE_ATTR, None)
    if isinstance(cached, dict):
        return cached
    header_style, body_style = _style_registry_for_setup()
    header_bg = _color_without_hash(str(header_style.get(_STYLE_KEY_BG_COLOR, ""))) or "D9EAD3"
    border_enabled = int(body_style.get(_STYLE_KEY_BORDER, 1)) > 0
    header_border_enabled = int(header_style.get(_STYLE_KEY_BORDER, 1)) > 0
    border_value = 1 if border_enabled else 0
    header_border_value = 1 if header_border_enabled else 0
    formats = {
        "header": workbook.add_format(
            {
                "bold": bool(header_style.get(_STYLE_KEY_BOLD, True)),
                "border": header_border_value,
                "align": str(header_style.get(_STYLE_KEY_ALIGN, _ALIGN_CENTER)),
                "valign": str(header_style.get(_STYLE_KEY_VALIGN, _ALIGN_VCENTER)),
                "text_wrap": True,
                "fg_color": header_bg,
                "pattern": 1,
            }
        ),
        "body": workbook.add_format(
            {
                "border": border_value,
                "valign": _ALIGN_VCENTER,
            }
        ),
        "body_center": workbook.add_format(
            {
                "border": border_value,
                "align": _ALIGN_CENTER,
                "valign": _ALIGN_VCENTER,
            }
        ),
        "body_wrap": workbook.add_format(
            {
                "border": border_value,
                "align": "left",
                "valign": _ALIGN_VCENTER,
                "text_wrap": True,
            }
        ),
        "column_wrap": workbook.add_format(
            {
                "align": "left",
                "valign": _ALIGN_VCENTER,
                "text_wrap": True,
            }
        ),
    }
    setattr(workbook, _STYLE_CACHE_ATTR, formats)
    return formats


def _create_co_attainment_sheet(
    workbook: Any,
    *,
    co_index: int,
    metadata: dict[str, str],
    ) -> _CoOutputSheetState:
    sheet = workbook.add_worksheet(f"CO{co_index}")
    formats = _xlsxwriter_formats(workbook)
    metadata_rows = _metadata_rows_for_output(metadata, co_index)
    for row_idx, (label, value) in enumerate(metadata_rows, start=0):
        sheet.write(row_idx, 1, label, formats["body"])
        sheet.write(row_idx, 2, value, formats["body_wrap"])

    header_row_index = len(metadata_rows) + 1
    headers = [
        "#",
        "Regno",
        "Student name",
        f"Direct ({_ratio_percent_token(DIRECT_RATIO)}%)",
        f"Indirect ({_ratio_percent_token(INDIRECT_RATIO)}%)",
        "Total (100%)",
        "Level",
    ]
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, col_idx, header, formats["header"])

    sampled_rows: list[list[Any]] = [["", COURSE_METADATA_HEADERS[0], COURSE_METADATA_HEADERS[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    sampled_rows.append(["", headers[1], headers[2]])
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    sheet.set_column(1, 1, widths.get(1, 8))
    sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    sheet.set_landscape()
    sheet.set_paper(9)  # A4
    sheet.fit_to_pages(1, 0)
    sheet.repeat_rows(0, header_row_index)
    # Freeze rows through the header and columns through Student name (A:C).
    sheet.freeze_panes(header_row_index + 1, 3)
    sheet.protect()
    sheet.set_selection(header_row_index, 0, header_row_index, 0)
    return _CoOutputSheetState(
        sheet=sheet,
        header_row_index=header_row_index,
        formats=formats,
        next_row_index=header_row_index + 1,
        next_serial=1,
        on_roll=0,
        attended=0,
        level_counts={0: 0, 1: 0, 2: 0, 3: 0},
    )


def _append_co_attainment_row(
    state: _CoOutputSheetState,
    row: _CoAttainmentRow,
    *,
    thresholds: tuple[float, float, float],
) -> None:
    if isinstance(row.direct_score, (int, float)) and isinstance(row.indirect_score, (int, float)):
        total: float | str = round(row.direct_score + row.indirect_score, CO_REPORT_MAX_DECIMAL_PLACES)
    else:
        total = CO_REPORT_ABSENT_TOKEN
    level = _score_to_attainment_level(total, thresholds=thresholds)
    left = [state.next_serial, row.reg_no]
    right = [row.direct_score, row.indirect_score, total, level]
    state.sheet.write_row(state.next_row_index, 0, left, state.formats["body"])
    state.sheet.write(state.next_row_index, 2, row.student_name, state.formats["body_wrap"])
    state.sheet.write_row(state.next_row_index, 3, right, state.formats["body_center"])
    state.on_roll += 1
    if total != CO_REPORT_ABSENT_TOKEN:
        state.attended += 1
    if isinstance(level, int) and level in state.level_counts:
        state.level_counts[level] += 1
    state.next_row_index += 1
    state.next_serial += 1


def _append_co_attainment_summary(state: _CoOutputSheetState) -> None:
    # Keep one visual spacer row after data and then write label/value summary rows.
    state.next_row_index += 1
    summary_rows = [
        ("On Roll:", state.on_roll),
        ("Attended:", state.attended),
        ("Level 0:", state.level_counts.get(0, 0)),
        ("Level 1:", state.level_counts.get(1, 0)),
        ("Level 2:", state.level_counts.get(2, 0)),
        ("Level 3:", state.level_counts.get(3, 0)),
    ]
    for label, value in summary_rows:
        state.sheet.write(state.next_row_index, 1, label, state.formats["body"])
        state.sheet.write(state.next_row_index, 2, value, state.formats["body_center"])
        state.next_row_index += 1


def _co_percentage(*, level_2: int, level_3: int, attended: int) -> float | str:
    if attended <= 0:
        return CO_REPORT_NOT_APPLICABLE_TOKEN
    return round(((level_2 + level_3) / float(attended)) * 100.0, CO_REPORT_MAX_DECIMAL_PLACES)


def _create_summary_sheet(
    workbook: Any,
    *,
    metadata: dict[str, str],
    output_states: dict[int, _CoOutputSheetState],
    total_outcomes: int,
) -> tuple[int, int]:
    sheet = workbook.add_worksheet("Summary")
    formats = _xlsxwriter_formats(workbook)
    metadata_rows = _metadata_rows_for_output(metadata, co_index=0)
    for row_idx, (label, value) in enumerate(metadata_rows, start=0):
        display_value = "All COs" if label == "CO Number" else value
        sheet.write(row_idx, 1, label, formats["body"])
        sheet.write(row_idx, 2, display_value, formats["body_wrap"])

    header_row_index = len(metadata_rows) + 1
    headers = ["CO", "Level 0", "Level 1", "Level 2", "Level 3", "Attended", "CO%"]
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, col_idx, header, formats["header"])

    first_data_row = header_row_index + 1
    for co_index in range(1, total_outcomes + 1):
        state = output_states.get(co_index)
        level_counts = state.level_counts if state is not None else {0: 0, 1: 0, 2: 0, 3: 0}
        attended = state.attended if state is not None else 0
        row_values: list[Any] = [
            f"CO{co_index}",
            level_counts.get(0, 0),
            level_counts.get(1, 0),
            level_counts.get(2, 0),
            level_counts.get(3, 0),
            attended,
            _co_percentage(level_2=level_counts.get(2, 0), level_3=level_counts.get(3, 0), attended=attended),
        ]
        row_index = header_row_index + co_index
        sheet.write_row(row_index, 0, row_values, formats["body_center"])

    sampled_rows: list[list[Any]] = [["", COURSE_METADATA_HEADERS[0], COURSE_METADATA_HEADERS[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    sampled_rows.append(headers)
    widths = _compute_sampled_column_widths(sampled_rows, len(headers))
    for col_idx in range(len(headers)):
        sheet.set_column(col_idx, col_idx, widths.get(col_idx, 10))
    sheet.set_column(1, 1, widths.get(1, 8))
    sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    sheet.set_landscape()
    sheet.set_paper(9)  # A4
    sheet.fit_to_pages(1, 0)
    sheet.repeat_rows(0, header_row_index)
    sheet.protect()
    sheet.set_selection(header_row_index, 0, header_row_index, 0)
    return first_data_row, first_data_row + max(0, total_outcomes - 1)


def _create_graph_sheet(
    workbook: Any,
    *,
    metadata: dict[str, str],
    summary_first_data_row: int,
    summary_last_data_row: int,
) -> None:
    graph_sheet = workbook.add_worksheet("Graph")
    formats = _xlsxwriter_formats(workbook)
    metadata_rows = _metadata_rows_for_output(metadata, co_index=0)
    for row_idx, (label, value) in enumerate(metadata_rows, start=0):
        display_value = "All COs" if label == "CO Number" else value
        graph_sheet.write(row_idx, 1, label, formats["body"])
        graph_sheet.write(row_idx, 2, display_value, formats["body_wrap"])
    sampled_rows: list[list[Any]] = [["", COURSE_METADATA_HEADERS[0], COURSE_METADATA_HEADERS[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    graph_sheet.set_column(1, 1, widths.get(1, 8))
    graph_sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    chart = workbook.add_chart({"type": "column"})
    chart.add_series(
        {
            "name": "CO%",
            "categories": ["Summary", summary_first_data_row, 0, summary_last_data_row, 0],
            "values": ["Summary", summary_first_data_row, 6, summary_last_data_row, 6],
            "data_labels": {"value": True},
        }
    )
    x_axis_name = "CO"
    y_axis_name = "% Attainment"
    chart.set_title({"name": f"{x_axis_name} {y_axis_name}"})
    chart.set_x_axis({"name": x_axis_name})
    chart.set_y_axis({"name": y_axis_name, "min": 0, "max": 100, "major_unit": 10})
    chart.set_legend({"none": True})
    chart_anchor_row = len(metadata_rows) + 2
    graph_sheet.insert_chart(f"B{chart_anchor_row + 1}", chart, {"x_scale": 1.4, "y_scale": 1.4})
    graph_sheet.set_landscape()
    graph_sheet.set_paper(9)  # A4
    graph_sheet.fit_to_pages(1, 0)
    graph_sheet.repeat_rows(0, max(0, len(metadata_rows) - 1))
    graph_sheet.protect()


def _write_system_integrity_sheets(
    workbook: Any,
    *,
    template_id: str,
    sheet_order: list[str],
) -> None:
    template_hash = sign_payload(template_id)
    hash_ws = workbook.add_worksheet(SYSTEM_HASH_SHEET)
    hash_ws.write(0, 0, SYSTEM_HASH_TEMPLATE_ID_HEADER)
    hash_ws.write(0, 1, SYSTEM_HASH_TEMPLATE_HASH_HEADER)
    hash_ws.write(1, 0, template_id)
    hash_ws.write(1, 1, template_hash)
    hash_ws.hide()

    signed_sheet_order = [*sheet_order, SYSTEM_HASH_SHEET]
    manifest = {
        "schema_version": _INTEGRITY_SCHEMA_VERSION,
        "template_id": template_id,
        "template_hash": template_hash,
        "sheet_order": signed_sheet_order,
        "sheets": [{"name": name, "hash": sign_payload(name)} for name in signed_sheet_order],
    }
    manifest_text = json.dumps(manifest, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    manifest_hash = sign_payload(manifest_text)
    integrity_ws = workbook.add_worksheet(SYSTEM_REPORT_INTEGRITY_SHEET)
    integrity_ws.write(0, 0, SYSTEM_REPORT_INTEGRITY_MANIFEST_HEADER)
    integrity_ws.write(0, 1, SYSTEM_REPORT_INTEGRITY_HASH_HEADER)
    integrity_ws.write(1, 0, manifest_text)
    integrity_ws.write(1, 1, manifest_hash)
    integrity_ws.hide()


def _attainment_thresholds(
    thresholds: tuple[float, float, float] | None = None,
) -> tuple[float, float, float]:
    if thresholds is None:
        return (float(LEVEL_1_THRESHOLD), float(LEVEL_2_THRESHOLD), float(LEVEL_3_THRESHOLD))
    l1, l2, l3 = thresholds
    return (float(l1), float(l2), float(l3))


def _score_to_attainment_level(
    score: float | str,
    *,
    thresholds: tuple[float, float, float],
) -> int | str:
    if not isinstance(score, (int, float)) or isinstance(score, bool):
        return CO_REPORT_NOT_APPLICABLE_TOKEN

    total = float(score)
    l1, l2, l3 = thresholds
    if 0.0 <= total < l1:
        return 0
    if l1 <= total < l2:
        return 1
    if l2 <= total < l3:
        return 2
    if l3 <= total <= 100.0:
        return 3
    return CO_REPORT_NOT_APPLICABLE_TOKEN


def _reg_no_sort_key(reg_no: str) -> tuple[tuple[int, int | str], ...]:
    tokens = re.split(r"(\d+)", reg_no.strip())
    key_parts: list[tuple[int, int | str]] = []
    for token in tokens:
        if not token:
            continue
        if token.isdigit():
            key_parts.append((0, int(token)))
        else:
            key_parts.append((1, token.casefold()))
    return tuple(key_parts)


def _generate_co_attainment_workbook(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    thresholds: tuple[float, float, float] | None = None,
) -> _CoAttainmentWorkbookResult:
    if not source_paths:
        raise ValueError("No source files provided for CO attainment calculation.")

    first_signature = _extract_final_report_signature(source_paths[0])
    if first_signature is None:
        raise ValueError(f"Invalid final CO report file: {source_paths[0]}")
    if normalize(first_signature.template_id) != normalize(ID_COURSE_SETUP):
        raise ValueError(f"Unsupported template id '{first_signature.template_id}'. Expected '{ID_COURSE_SETUP}'.")
    return _generate_co_attainment_workbook_course_setup_v1(
        source_paths,
        output_path,
        token=token,
        total_outcomes=first_signature.total_outcomes,
        template_id=first_signature.template_id,
        thresholds=thresholds,
    )


def _generate_co_attainment_workbook_course_setup_v1(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    total_outcomes: int,
    template_id: str,
    thresholds: tuple[float, float, float] | None = None,
) -> _CoAttainmentWorkbookResult:
    try:
        import xlsxwriter
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - guarded by runtime dependency availability
        raise RuntimeError("openpyxl and xlsxwriter are required for CO attainment calculation.") from exc

    if total_outcomes <= 0:
        raise ValueError("No CO outcomes available in the uploaded reports.")

    metadata: dict[str, str] = {}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook = xlsxwriter.Workbook(str(output_path), {"constant_memory": True})
    workbook_closed = False
    output_states: dict[int, _CoOutputSheetState] = {}
    pending_rows: dict[int, list[_CoAttainmentRow]] = {}
    duplicate_reg_count = 0
    duplicate_entries: list[tuple[str, str, str]] = []
    level_thresholds = _attainment_thresholds(thresholds)
    dedup_store = _RegisterDedupStore(
        total_outcomes=total_outcomes,
        use_sqlite=(len(source_paths) * total_outcomes * _COORDINATOR_STUDENTS_PER_SHEET) >= _DEDUP_SQLITE_THRESHOLD_ENTRIES,
    )

    try:
        for source in source_paths:
            token.raise_if_cancelled()
            workbook = load_workbook(filename=source, data_only=True, read_only=True)
            try:
                if not metadata and COURSE_METADATA_SHEET in workbook.sheetnames:
                    metadata = _extract_course_metadata_fields(workbook[COURSE_METADATA_SHEET])
                    for co_index in range(1, total_outcomes + 1):
                        output_states[co_index] = _create_co_attainment_sheet(
                            output_workbook,
                            co_index=co_index,
                            metadata=metadata,
                        )
                        pending_rows[co_index] = []
                if not output_states:
                    for co_index in range(1, total_outcomes + 1):
                        output_states[co_index] = _create_co_attainment_sheet(
                            output_workbook,
                            co_index=co_index,
                            metadata=metadata,
                        )
                        pending_rows[co_index] = []
                for co_index in range(1, total_outcomes + 1):
                    row_bucket = pending_rows.setdefault(co_index, [])
                    for row in _iter_co_rows_from_workbook(workbook, co_index=co_index, workbook_name=source.name):
                        if not dedup_store.add_if_absent(co_index=co_index, reg_hash=row.reg_hash):
                            duplicate_reg_count += 1
                            duplicate_entries.append((row.reg_no, row.worksheet_name, row.workbook_name))
                            continue
                        row_bucket.append(row)
            finally:
                workbook.close()
        for co_index, state in output_states.items():
            for row in sorted(
                pending_rows.get(co_index, []),
                key=lambda item: (_reg_no_sort_key(item.reg_no), item.reg_hash),
            ):
                _append_co_attainment_row(state, row, thresholds=level_thresholds)
            _append_co_attainment_summary(state)
        summary_first_data_row, summary_last_data_row = _create_summary_sheet(
            output_workbook,
            metadata=metadata,
            output_states=output_states,
            total_outcomes=total_outcomes,
        )
        _create_graph_sheet(
            output_workbook,
            metadata=metadata,
            summary_first_data_row=summary_first_data_row,
            summary_last_data_row=summary_last_data_row,
        )
        sheet_order = [f"CO{co_index}" for co_index in range(1, total_outcomes + 1)]
        sheet_order.extend(["Summary", "Graph"])
        _write_system_integrity_sheets(
            output_workbook,
            template_id=template_id,
            sheet_order=sheet_order,
        )
        output_workbook.close()
        workbook_closed = True
    finally:
        dedup_store.close()
        if not workbook_closed:
            try:
                output_workbook.close()
            except Exception:
                _logger.debug("Suppressing workbook close error during cleanup.", exc_info=True)
    return _CoAttainmentWorkbookResult(
        output_path=output_path,
        duplicate_reg_count=duplicate_reg_count,
        duplicate_entries=tuple(duplicate_entries),
    )

