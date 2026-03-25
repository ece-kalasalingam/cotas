"""Domain engine for CO Analysis workbook validation and generation."""

from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Callable

from common.constants import (
    CO_REPORT_HEADER_REG_NO,
    CO_REPORT_HEADER_SERIAL,
    ID_COURSE_SETUP,
)
from common.error_catalog import validation_error_from_key
from common.excel_sheet_layout import color_without_hash as _color_without_hash
from common.excel_sheet_layout import style_registry_for_template as _style_registry_for_template
from common.excel_sheet_layout import (
    copy_openpyxl_cell_style as _copy_openpyxl_cell_style,
    copy_openpyxl_sheet as _copy_openpyxl_sheet,
    find_header_row_by_value as _find_header_row_by_value,
    protect_openpyxl_sheet as _protect_openpyxl_sheet,
)
from common.exceptions import AppSystemError, ValidationError
from common.jobs import CancellationToken
from common.utils import canonical_path_key, coerce_excel_number, normalize
from common.registry import (
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    COURSE_SETUP_SHEET_KEY_STUDENTS,
    SYSTEM_HASH_SHEET_NAME as SYSTEM_HASH_SHEET,
    get_sheet_headers_by_key,
    get_sheet_name_by_key,
)
from common.workbook_integrity.constants import SYSTEM_LAYOUT_SHEET
from domain.co_report_sheet_generator import co_direct_sheet_name, co_indirect_sheet_name
from domain.template_strategy_router import (
    assert_template_id_matches,
    get_template_strategy,
    read_valid_template_id_from_system_hash_sheet,
)

_SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
_COURSE_METADATA_DUPLICATE_FIELDS = (
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
)
_COURSE_METADATA_COURSE_NAME_KEY = "course_name"
_ORDERED_METADATA_FIELDS: tuple[tuple[str, str], ...] = (
    (COURSE_METADATA_COURSE_CODE_KEY, "Course Code"),
    (_COURSE_METADATA_COURSE_NAME_KEY, "Course Name"),
    (COURSE_METADATA_TOTAL_OUTCOMES_KEY, "Total Outcomes"),
    (COURSE_METADATA_ACADEMIC_YEAR_KEY, "Academic Year"),
    (COURSE_METADATA_SEMESTER_KEY, "Semester"),
)
_VALIDATION_REASON_SYSTEM_HASH = "system_hash"
_VALIDATION_REASON_MARKS_UNFILLED = "marks_unfilled"
_VALIDATION_REASON_LAYOUT_OR_MANIFEST = "layout_or_manifest"
_VALIDATION_REASON_TEMPLATE_MISMATCH = "template_mismatch"
_VALIDATION_REASON_MARK_VALUE = "mark_value"
_VALIDATION_REASON_OTHER = "other_validation"
COURSE_METADATA_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
STUDENTS_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_STUDENTS)


def _course_metadata_headers(template_id: str) -> tuple[str, ...]:
    return get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)


def _students_headers(template_id: str) -> tuple[str, ...]:
    return get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)


def analyze_uploaded_workbooks(
    candidate_paths: list[str],
    *,
    existing_paths: list[str],
    validate_uploaded_source_workbook: Callable[[str | Path], None],
    consume_last_source_anomaly_warnings: Callable[[], list[str]],
    token: CancellationToken | None = None,
) -> dict[str, object]:
    existing_keys = {canonical_path_key(Path(path)) for path in existing_paths}
    seen_metadata_signatures: set[tuple[str, ...]] = set()
    seen_register_numbers: set[str] = set()
    for existing_path in existing_paths:
        students, metadata_map = extract_course_metadata_and_students(Path(existing_path))
        seen_register_numbers.update(students)
        signature = _course_metadata_signature_from_map(metadata_map)
        if signature is not None:
            seen_metadata_signatures.add(signature)

    added_paths: list[str] = []
    duplicates = 0
    invalid = 0
    unsupported_or_missing_files = 0
    invalid_source_workbook_files = 0
    duplicate_reg_number_files = 0
    co_count_mismatch_files = 0
    invalid_system_hash_files = 0
    invalid_marks_unfilled_files = 0
    invalid_layout_manifest_files = 0
    invalid_template_mismatch_files = 0
    invalid_mark_value_files = 0
    invalid_other_validation_files = 0
    anomaly_warnings: list[str] = []
    validation_failures: list[dict[str, str]] = []
    baseline_total_outcomes: int | None = None
    validated_candidates: list[tuple[Path, str, tuple[str, ...] | None, set[str], int | None]] = []
    for raw_path in candidate_paths:
        if token is not None:
            token.raise_if_cancelled()
        path = Path(raw_path)
        suffix = path.suffix.lower()
        if suffix not in _SUPPORTED_EXTENSIONS or not path.exists():
            unsupported_or_missing_files += 1
            invalid += 1
            continue
        key = canonical_path_key(path)
        try:
            validate_uploaded_source_workbook(path)
            warnings = consume_last_source_anomaly_warnings()
            if warnings:
                anomaly_warnings.extend([f"{path.name} -> {msg}" for msg in warnings])
        except ValidationError as exc:
            invalid_source_workbook_files += 1
            invalid += 1
            reason = _classify_validation_reason(exc)
            if reason == _VALIDATION_REASON_SYSTEM_HASH:
                invalid_system_hash_files += 1
            elif reason == _VALIDATION_REASON_MARKS_UNFILLED:
                invalid_marks_unfilled_files += 1
            elif reason == _VALIDATION_REASON_LAYOUT_OR_MANIFEST:
                invalid_layout_manifest_files += 1
            elif reason == _VALIDATION_REASON_TEMPLATE_MISMATCH:
                invalid_template_mismatch_files += 1
            elif reason == _VALIDATION_REASON_MARK_VALUE:
                invalid_mark_value_files += 1
            else:
                invalid_other_validation_files += 1
            validation_failures.append(
                {
                    "file": path.name,
                    "reason": str(exc),
                    "code": str(exc.code),
                    "category": reason,
                    "context": dict(exc.context or {}),
                }
            )
            continue
        except Exception as exc:
            invalid_source_workbook_files += 1
            invalid += 1
            invalid_other_validation_files += 1
            validation_failures.append(
                {
                    "file": path.name,
                    "reason": str(exc),
                    "code": type(exc).__name__,
                    "category": _VALIDATION_REASON_OTHER,
                    "context": {},
                }
            )
            continue
        students, metadata_map = extract_course_metadata_and_students(path)
        metadata_signature = _course_metadata_signature_from_map(metadata_map)
        total_outcomes = _total_outcomes_from_metadata_map(metadata_map)
        validated_candidates.append((path, key, metadata_signature, students, total_outcomes))

    path_counts: dict[str, int] = {}
    metadata_counts: dict[tuple[str, ...], int] = {}
    register_owners: dict[str, list[int]] = {}
    for idx, (_path, key, metadata_signature, students, total_outcomes) in enumerate(validated_candidates):
        path_counts[key] = path_counts.get(key, 0) + 1
        if metadata_signature is not None:
            metadata_counts[metadata_signature] = metadata_counts.get(metadata_signature, 0) + 1
        for reg_no in students:
            register_owners.setdefault(reg_no, []).append(idx)
        if baseline_total_outcomes is None and total_outcomes is not None:
            baseline_total_outcomes = total_outcomes

    invalid_due_to_reg_duplicates: set[int] = set()
    for idx, (_path, _key, _metadata_signature, students, _total_outcomes) in enumerate(validated_candidates):
        if any(reg_no in seen_register_numbers for reg_no in students):
            invalid_due_to_reg_duplicates.add(idx)
    for reg_no, owner_indices in register_owners.items():
        if len(owner_indices) > 1:
            invalid_due_to_reg_duplicates.update(owner_indices)

    for idx, (path, key, metadata_signature, students, total_outcomes) in enumerate(validated_candidates):
        if idx in invalid_due_to_reg_duplicates:
            duplicate_reg_number_files += 1
            invalid += 1
            continue
        if baseline_total_outcomes is not None and total_outcomes is not None and total_outcomes != baseline_total_outcomes:
            co_count_mismatch_files += 1
            invalid += 1
            continue
        is_duplicate = False
        if key in existing_keys:
            is_duplicate = True
        elif metadata_signature is not None and metadata_signature in seen_metadata_signatures:
            is_duplicate = True
        elif path_counts.get(key, 0) > 1:
            is_duplicate = True
        elif metadata_signature is not None and metadata_counts.get(metadata_signature, 0) > 1:
            is_duplicate = True
        if is_duplicate:
            duplicates += 1
            continue
        added_paths.append(str(path))
        seen_register_numbers.update(students)
    return {
        "added": added_paths,
        "duplicates": duplicates,
        "invalid": invalid,
        "ignored": duplicates + invalid,
        "unsupported_or_missing_files": unsupported_or_missing_files,
        "invalid_source_workbook_files": invalid_source_workbook_files,
        "anomaly_warnings": anomaly_warnings,
        "duplicate_reg_number_files": duplicate_reg_number_files,
        "co_count_mismatch_files": co_count_mismatch_files,
        "invalid_system_hash_files": invalid_system_hash_files,
        "invalid_marks_unfilled_files": invalid_marks_unfilled_files,
        "invalid_layout_manifest_files": invalid_layout_manifest_files,
        "invalid_template_mismatch_files": invalid_template_mismatch_files,
        "invalid_mark_value_files": invalid_mark_value_files,
        "invalid_other_validation_files": invalid_other_validation_files,
        "validation_failures": validation_failures,
    }


def _classify_validation_reason(exc: ValidationError) -> str:
    code = normalize(getattr(exc, "code", ""))
    if code in {
        "coa_system_hash_mismatch",
        "coa_system_hash_header_template_id_missing",
        "coa_system_hash_header_template_hash_missing",
        "coa_system_hash_template_id_missing",
        "coa_system_sheet_missing",
    }:
        return _VALIDATION_REASON_SYSTEM_HASH
    if code in {"coa_mark_entry_empty"}:
        return _VALIDATION_REASON_MARKS_UNFILLED
    if code in {
        "coa_layout_sheet_missing",
        "coa_layout_header_mismatch",
        "coa_layout_manifest_missing",
        "coa_layout_hash_mismatch",
        "coa_layout_manifest_json_invalid",
    }:
        return _VALIDATION_REASON_LAYOUT_OR_MANIFEST
    if code in {"unknown_template", "coa_template_validator_missing"}:
        return _VALIDATION_REASON_TEMPLATE_MISMATCH
    if code in {
        "coa_mark_value_invalid",
        "coa_mark_precision_invalid",
        "coa_indirect_mark_integer_required",
        "coa_absence_policy_violation",
    }:
        return _VALIDATION_REASON_MARK_VALUE
    return _VALIDATION_REASON_OTHER


def _course_metadata_signature_from_map(metadata_map: dict[str, str]) -> tuple[str, ...] | None:
    signature = tuple(
        metadata_map.get(normalize(field_name), "").strip()
        for field_name in _COURSE_METADATA_DUPLICATE_FIELDS
    )
    return signature if any(signature) else None


def _total_outcomes_from_metadata_map(metadata_map: dict[str, str]) -> int | None:
    raw = metadata_map.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY), "").strip()
    if not raw:
        return None
    try:
        value = int(float(raw))
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def extract_course_metadata_signature(path: Path) -> tuple[str, ...] | None:
    try:
        import openpyxl
    except Exception:
        return None
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        template_id = ID_COURSE_SETUP
        try:
            template_id = read_valid_template_id_from_system_hash_sheet(workbook)
        except Exception:
            template_id = ID_COURSE_SETUP
        metadata_headers = _course_metadata_headers(template_id)
        if COURSE_METADATA_SHEET not in workbook.sheetnames:
            return None
        sheet = workbook[COURSE_METADATA_SHEET]
        field_header = normalize(sheet.cell(row=1, column=1).value)
        value_header = normalize(sheet.cell(row=1, column=2).value)
        if field_header != normalize(metadata_headers[0]):
            return None
        if value_header != normalize(metadata_headers[1]):
            return None
        metadata: dict[str, str] = {}
        row = 2
        while True:
            field_raw = sheet.cell(row=row, column=1).value
            value_raw = sheet.cell(row=row, column=2).value
            if normalize(field_raw) == "" and normalize(value_raw) == "":
                break
            field_key = normalize(field_raw)
            if field_key:
                metadata[field_key] = str(value_raw).strip() if value_raw is not None else ""
            row += 1
        signature = tuple(metadata.get(normalize(field_name), "").strip() for field_name in _COURSE_METADATA_DUPLICATE_FIELDS)
        return signature if any(signature) else None
    finally:
        workbook.close()


def extract_course_metadata_and_students(path: Path) -> tuple[set[str], dict[str, str]]:
    try:
        import openpyxl
    except Exception:
        return set(), {}
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return set(), {}
    unique_students: set[str] = set()
    metadata_map: dict[str, str] = {}
    try:
        template_id = ID_COURSE_SETUP
        try:
            template_id = read_valid_template_id_from_system_hash_sheet(workbook)
        except Exception:
            template_id = ID_COURSE_SETUP
        students_headers = _students_headers(template_id)
        if COURSE_METADATA_SHEET in workbook.sheetnames:
            metadata_sheet = workbook[COURSE_METADATA_SHEET]
            row = 2
            while True:
                key_raw = metadata_sheet.cell(row=row, column=1).value
                value_raw = metadata_sheet.cell(row=row, column=2).value
                if normalize(key_raw) == "" and normalize(value_raw) == "":
                    break
                key_text = str(key_raw).strip() if key_raw is not None else ""
                if key_text:
                    coerced = coerce_excel_number(value_raw) if value_raw is not None else None
                    metadata_map[normalize(key_text)] = str(coerced).strip() if coerced is not None else ""
                row += 1
        if STUDENTS_SHEET in workbook.sheetnames:
            students_sheet = workbook[STUDENTS_SHEET]
            header_map: dict[str, int] = {}
            max_col = int(students_sheet.max_column)
            for col in range(1, max_col + 1):
                key = normalize(students_sheet.cell(row=1, column=col).value)
                if key and key not in header_map:
                    header_map[key] = col
            reg_col = header_map.get(normalize(students_headers[0])) or header_map.get(normalize(CO_REPORT_HEADER_REG_NO))
            if reg_col is not None:
                max_row = int(students_sheet.max_row)
                for row in range(2, max_row + 1):
                    reg_raw = students_sheet.cell(row=row, column=reg_col).value
                    coerced = coerce_excel_number(reg_raw)
                    reg_text = str(coerced).strip() if coerced is not None else ""
                    if reg_text:
                        unique_students.add(normalize(reg_text))
        if not unique_students:
            unique_students = _extract_students_from_report_sheets(workbook, template_id=template_id)
    finally:
        workbook.close()
    return unique_students, metadata_map


def generate_co_analysis_workbook(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken | None = None,
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float | None = None,
    co_attainment_level: int | None = None,
) -> Path:
    extracted: list[tuple[set[str], dict[str, str]]] = []
    for path in source_paths:
        if token is not None:
            token.raise_if_cancelled()
        extracted.append(extract_course_metadata_and_students(path))
    return build_co_analysis_workbook(
        output_path,
        extracted,
        source_paths=source_paths,
        token=token,
        thresholds=thresholds,
        co_attainment_percent=co_attainment_percent,
        co_attainment_level=co_attainment_level,
    )


def build_co_analysis_workbook(
    output_path: Path,
    extracted: list[tuple[set[str], dict[str, str]]],
    *,
    source_paths: list[Path],
    token: CancellationToken | None = None,
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float | None = None,
    co_attainment_level: int | None = None,
) -> Path:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if not source_paths:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COA_SOURCE_WORKBOOK_REQUIRED",
        )
    probe_wb = load_workbook(source_paths[0], data_only=False, read_only=True)
    try:
        resolved_template_id = read_valid_template_id_from_system_hash_sheet(probe_wb)
    finally:
        probe_wb.close()

    workbook = Workbook()
    metadata_headers = _course_metadata_headers(resolved_template_id)
    course_metadata_sheet = workbook.active
    if course_metadata_sheet is None:
        raise AppSystemError("Failed to create Course Metadata worksheet.")
    course_metadata_sheet.title = COURSE_METADATA_SHEET
    course_metadata_sheet.cell(row=1, column=1, value=metadata_headers[0])
    course_metadata_sheet.cell(row=1, column=2, value=metadata_headers[1])
    header_style, _ = _style_registry_for_template(resolved_template_id)
    header_bg = _color_without_hash(str(header_style.get("bg_color", "")))
    header_fill = PatternFill(fill_type="solid", fgColor=header_bg)
    header_font = Font(bold=True)
    thin = Side(border_style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in (1, 2):
        header_cell = course_metadata_sheet.cell(row=1, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.border = header_border
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    row_cursor = 2
    max_col1_len = max(len(metadata_headers[0]), len("Students On-Roll"))
    max_col2_len = len(metadata_headers[1])
    seen_dedup_values: dict[str, set[str]] = {}
    for unique_students, metadata_map in extracted:
        for field_key_raw, label in _ORDERED_METADATA_FIELDS:
            field_key = normalize(field_key_raw)
            field_value = metadata_map.get(field_key, "")
            if not field_value:
                continue
            value_key = normalize(field_value)
            seen_values = seen_dedup_values.setdefault(field_key, set())
            if value_key in seen_values:
                continue
            seen_values.add(value_key)
            course_metadata_sheet.cell(row=row_cursor, column=1, value=label)
            course_metadata_sheet.cell(row=row_cursor, column=2, value=field_value)
            course_metadata_sheet.cell(row=row_cursor, column=1).border = body_border
            course_metadata_sheet.cell(row=row_cursor, column=2).border = body_border
            max_col1_len = max(max_col1_len, len(str(label)))
            max_col2_len = max(max_col2_len, len(str(field_value)))
            row_cursor += 1

        section_value = metadata_map.get(normalize(COURSE_METADATA_SECTION_KEY), "")
        course_metadata_sheet.cell(row=row_cursor, column=1, value="Section")
        course_metadata_sheet.cell(row=row_cursor, column=2, value=section_value)
        course_metadata_sheet.cell(row=row_cursor, column=1).border = body_border
        course_metadata_sheet.cell(row=row_cursor, column=2).border = body_border
        max_col1_len = max(max_col1_len, len("Section"))
        max_col2_len = max(max_col2_len, len(str(section_value)))
        row_cursor += 1

        students_on_roll = len(unique_students)
        course_metadata_sheet.cell(row=row_cursor, column=1, value="Students On-Roll")
        course_metadata_sheet.cell(row=row_cursor, column=2, value=students_on_roll)
        course_metadata_sheet.cell(row=row_cursor, column=1).border = body_border
        course_metadata_sheet.cell(row=row_cursor, column=2).border = body_border
        max_col1_len = max(max_col1_len, len("Students On-Roll"))
        max_col2_len = max(max_col2_len, len(str(students_on_roll)))
        row_cursor += 2

    course_metadata_sheet.column_dimensions["A"].width = min(max(18, max_col1_len + 2), 48)
    course_metadata_sheet.column_dimensions["B"].width = min(max(20, max_col2_len + 2), 120)
    for row in range(2, max(2, row_cursor)):
        for col in (1, 2):
            cell = course_metadata_sheet.cell(row=row, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                continue
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    workbook.security.lockStructure = True
    _protect_openpyxl_sheet(course_metadata_sheet)

    with TemporaryDirectory(prefix="focus_co_analysis_") as temp_root_raw:
        temp_root = Path(temp_root_raw)
        generated_final_reports: list[Path] = []
        for index, source_path in enumerate(source_paths, start=1):
            if token is not None:
                token.raise_if_cancelled()
            final_report_path = temp_root / f"co_analysis_source_{index}.xlsx"
            try:
                import openpyxl
            except Exception as exc:
                raise validation_error_from_key(
                    "validation.dependency.openpyxl_missing",
                    code="OPENPYXL_MISSING",
                ) from exc
            source_wb = openpyxl.load_workbook(source_path, data_only=False, read_only=True)
            try:
                template_id = read_valid_template_id_from_system_hash_sheet(source_wb)
            finally:
                source_wb.close()
            strategy = get_template_strategy(template_id)
            strategy.generate_final_report(source_path, final_report_path, cancel_token=token)
            try:
                assert_template_id_matches(
                    actual_template_id=template_id,
                    expected_template_id=resolved_template_id,
                )
            except ValidationError as exc:
                raise validation_error_from_key(
                    "common.validation_failed_invalid_data",
                    code="COA_TEMPLATE_MIXED",
                    expected=resolved_template_id,
                    found=template_id,
                ) from exc
            generated_final_reports.append(final_report_path)

        if not generated_final_reports:
            workbook.save(output_path)
            return output_path

        first_report = load_workbook(generated_final_reports[0], data_only=False)
        try:
            total_outcomes = _read_total_outcomes_from_course_metadata(first_report)
        finally:
            first_report.close()

        coordinator_output = temp_root / "co_analysis_co_sheets.xlsx"
        strategy = get_template_strategy(resolved_template_id)
        strategy.generate_co_attainment(
            generated_final_reports,
            coordinator_output,
            token=token or CancellationToken(),
            thresholds=thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
        )

        coordinator_wb = load_workbook(coordinator_output, data_only=False)
        try:
            for co_index in range(1, total_outcomes + 1):
                if token is not None:
                    token.raise_if_cancelled()
                direct_name = co_direct_sheet_name(co_index)
                indirect_source_name = co_indirect_sheet_name(co_index)
                indirect_target_name = co_indirect_sheet_name(co_index)
                co_name = f"CO{co_index}"
                merged_direct = workbook.create_sheet(title=direct_name)
                _merge_report_sheets(
                    report_paths=generated_final_reports,
                    source_sheet_name=direct_name,
                    target_sheet=merged_direct,
                )
                _protect_sheet(merged_direct)
                merged_indirect = workbook.create_sheet(title=indirect_target_name)
                _merge_report_sheets(
                    report_paths=generated_final_reports,
                    source_sheet_name=indirect_source_name,
                    target_sheet=merged_indirect,
                )
                _protect_sheet(merged_indirect)
                if co_name in coordinator_wb.sheetnames:
                    co_sheet = workbook.create_sheet(title=co_name)
                    _copy_sheet(coordinator_wb[co_name], co_sheet)
                    _protect_sheet(co_sheet)
        finally:
            coordinator_wb.close()

        workbook.save(output_path)
        return output_path


def _extract_students_from_report_sheets(workbook: object, *, template_id: str = ID_COURSE_SETUP) -> set[str]:
    unique_students: set[str] = set()
    students_headers = _students_headers(template_id)
    try:
        sheets = getattr(workbook, "worksheets", [])
    except Exception:
        return unique_students
    for sheet in sheets:
        if sheet is None:
            continue
        title = str(getattr(sheet, "title", "") or "")
        if title in {SYSTEM_HASH_SHEET, SYSTEM_LAYOUT_SHEET, COURSE_METADATA_SHEET}:
            continue
        max_row = int(getattr(sheet, "max_row", 0) or 0)
        max_col = int(getattr(sheet, "max_column", 0) or 0)
        if max_row <= 0 or max_col <= 0:
            continue
        scan_rows = min(max_row, 30)
        scan_cols = min(max_col, 80)
        reg_col: int | None = None
        header_row = 0
        for row in range(1, scan_rows + 1):
            for col in range(1, scan_cols + 1):
                key = normalize(sheet.cell(row=row, column=col).value)
                if key in {normalize(CO_REPORT_HEADER_REG_NO), normalize(students_headers[0])}:
                    reg_col = col
                    header_row = row
                    break
            if reg_col is not None:
                break
        if reg_col is None:
            continue
        for row in range(header_row + 1, max_row + 1):
            reg_raw = sheet.cell(row=row, column=reg_col).value
            coerced = coerce_excel_number(reg_raw)
            reg_text = str(coerced).strip() if coerced is not None else ""
            if reg_text:
                unique_students.add(normalize(reg_text))
    return unique_students


def _read_total_outcomes_from_course_metadata(workbook: object) -> int:
    try:
        metadata_sheet = workbook[COURSE_METADATA_SHEET]  # type: ignore[index]
    except Exception:
        return 0
    row = 2
    while True:
        key = metadata_sheet.cell(row=row, column=1).value
        value = metadata_sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        if normalize(key) == normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY):
            parsed = coerce_excel_number(value)
            if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
                return max(0, int(parsed))
        row += 1
    return 0


def _protect_sheet(sheet: object) -> None:
    _protect_openpyxl_sheet(sheet)


def _copy_sheet(source_sheet: object, target_sheet: object) -> None:
    _copy_openpyxl_sheet(source_sheet, target_sheet)


def _find_header_row_by_serial(sheet: object) -> int:
    return _find_header_row_by_value(sheet, header_value=CO_REPORT_HEADER_SERIAL, header_col=1, max_scan_rows=300)


def _merge_report_sheets(
    *,
    report_paths: list[Path],
    source_sheet_name: str,
    target_sheet: object,
) -> None:
    from openpyxl import load_workbook

    if not report_paths:
        return

    seed_wb = load_workbook(report_paths[0], data_only=False)
    try:
        if source_sheet_name not in seed_wb.sheetnames:
            return
        source_sheet = seed_wb[source_sheet_name]
        header_row = _find_header_row_by_serial(source_sheet)
        if header_row <= 0:
            return
        for row in range(1, header_row + 1):
            for col in range(1, int(source_sheet.max_column) + 1):
                source_cell = source_sheet.cell(row=row, column=col)
                target_cell = target_sheet.cell(row=row, column=col, value=source_cell.value)
                _copy_openpyxl_cell_style(source_cell, target_cell)
        target_sheet.freeze_panes = source_sheet.freeze_panes
        target_sheet.print_title_rows = source_sheet.print_title_rows
        target_sheet.print_title_cols = source_sheet.print_title_cols
        target_sheet.page_setup.orientation = source_sheet.page_setup.orientation
        target_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
        target_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
        target_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
        for key, dimension in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].width = dimension.width
    finally:
        seed_wb.close()

    next_row = header_row + 1
    next_serial = 1
    for report_path in report_paths:
        wb = load_workbook(report_path, data_only=False)
        try:
            if source_sheet_name not in wb.sheetnames:
                continue
            sheet = wb[source_sheet_name]
            local_header_row = _find_header_row_by_serial(sheet)
            if local_header_row <= 0:
                continue
            row = local_header_row + 1
            while row <= int(sheet.max_row):
                reg_cell = sheet.cell(row=row, column=2).value
                name_cell = sheet.cell(row=row, column=3).value
                if normalize(reg_cell) == "" and normalize(name_cell) == "":
                    break
                target_sheet.cell(row=next_row, column=1, value=next_serial)
                for col in range(2, int(sheet.max_column) + 1):
                    source_cell = sheet.cell(row=row, column=col)
                    target_cell = target_sheet.cell(row=next_row, column=col, value=source_cell.value)
                    _copy_openpyxl_cell_style(source_cell, target_cell)
                next_serial += 1
                next_row += 1
                row += 1
        finally:
            wb.close()
