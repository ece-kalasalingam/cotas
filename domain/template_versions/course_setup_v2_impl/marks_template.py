"""COURSE_SETUP_V2 marks-template generation."""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any, Mapping, Sequence
from uuid import uuid4

from common.constants import (
    LAYOUT_MANIFEST_KEY_SHEET_ORDER,
    LAYOUT_MANIFEST_KEY_SHEETS,
    LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
    LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
    LAYOUT_SHEET_KIND_INDIRECT,
    LAYOUT_SHEET_SPEC_KEY_KIND,
    LAYOUT_SHEET_SPEC_KEY_NAME,
    WORKBOOK_INTEGRITY_SCHEMA_VERSION,
    WORKBOOK_TEMP_SUFFIX,
    MARKS_ENTRY_ROW_HEADERS,
)
from common.error_catalog import validation_error_from_key
from common.exceptions import AppSystemError, JobCancelledError, ValidationError
from common.excel_sheet_layout import build_marks_template_xlsxwriter_formats
from common.i18n import t
from common.jobs import CancellationToken
from common.registry import (
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    COURSE_SETUP_SHEET_KEY_QUESTION_MAP,
    COURSE_SETUP_SHEET_KEY_STUDENTS,
    SYSTEM_HASH_SHEET_NAME,
    get_blueprint,
    get_sheet_headers_by_key,
    get_sheet_name_by_key,
    get_sheet_schema_by_key,
)
from common.utils import canonical_path_key, coerce_excel_number, normalize
from common.workbook_integrity import (
    add_system_layout_sheet,
    copy_system_hash_sheet,
    read_valid_template_id_from_system_hash_sheet,
)
from common.workbook_integrity.constants import SYSTEM_LAYOUT_SHEET
from domain.template_versions.course_setup_v2_impl.co_token_parser import parse_co_tokens
from domain.template_versions.course_setup_v2_impl.assessment_semantics import (
    parse_assessment_components,
)
from domain.template_versions.course_setup_v2_impl.course_semantics import (
    build_marks_template_filename_base_from_identity,
)
from domain.template_versions.course_setup_v2_impl.course_template_validator import (
    validate_course_details_rules as _validate_course_details_rules_v2,
)
from domain.template_versions.course_setup_v2_impl import instructor_engine_sheetops as _sheetops

_logger = logging.getLogger(__name__)


def _ve(translation_key: str, *, code: str, **context: Any) -> ValidationError:
    return validation_error_from_key(translation_key, code=code, **context)


def _prepare_marks_generation_from_workbook(
    source_workbook: Any,
    *,
    cancel_token: CancellationToken | None = None,
) -> tuple[str, dict[str, Any]]:
    if cancel_token is not None:
        cancel_token.raise_if_cancelled()
    # Fail fast: validate full course-details workbook before template generation starts.
    _validate_course_details_rules_v2(source_workbook)
    if cancel_token is not None:
        cancel_token.raise_if_cancelled()
    template_id = read_valid_template_id_from_system_hash_sheet(source_workbook)
    context = _extract_marks_template_context_by_template(source_workbook, template_id)
    return template_id, context


def _render_marks_template_to_output(
    *,
    source_workbook: Any,
    output_path: Path,
    template_id: str,
    context: dict[str, Any],
    cancel_token: CancellationToken | None = None,
) -> Path:
    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.xlsxwriter_missing",
            code="XLSXWRITER_MISSING",
        ) from exc

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output.with_name(f"{output.name}.{uuid4().hex}{WORKBOOK_TEMP_SUFFIX}")
    target_workbook = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    target_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal target_closed
        if not target_closed:
            try:
                target_workbook.close()
                target_closed = True
            except Exception:
                _logger.debug("Suppressing target close error during cleanup.", exc_info=True)
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp marks template file: %s", temp_output)

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()

        layout_manifest = _write_marks_template_workbook_by_template(
            target_workbook,
            context,
            template_id=template_id,
            cancel_token=cancel_token,
        )

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        copy_system_hash_sheet(source_workbook, target_workbook)
        add_system_layout_sheet(target_workbook, layout_manifest)

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        target_workbook.close()
        target_closed = True

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except JobCancelledError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate marks template. output=%s",
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc

    return output


def generate_marks_template_from_course_details(
    course_details_path: str | Path,
    output_path: str | Path,
    *,
    allow_overwrite: bool = False,
    cancel_token: CancellationToken | None = None,
) -> Path:
    """Generate marks-entry workbook from a validated course-details workbook."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.openpyxl_missing",
            code="OPENPYXL_MISSING",
        ) from exc

    source_file = Path(course_details_path)
    if not source_file.exists():
        raise _ve(
            "instructor.validation.workbook_not_found",
            code="WORKBOOK_NOT_FOUND",
            workbook=str(source_file),
        )
    output_target = Path(output_path)
    if output_target.exists() and not allow_overwrite:
        raise _ve(
            "common.validation_failed_invalid_data",
            code="OUTPUT_PATH_ALREADY_EXISTS",
            output_path=str(output_target),
        )

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        # Open once in formula-visible mode so validator checks remain authoritative.
        source_workbook = openpyxl.load_workbook(source_file, data_only=False)
    except JobCancelledError:
        raise
    except Exception as exc:
        raise _ve(
            "instructor.validation.workbook_open_failed",
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(source_file),
        ) from exc

    try:
        template_id, context = _prepare_marks_generation_from_workbook(
            source_workbook,
            cancel_token=cancel_token,
        )
        return _render_marks_template_to_output(
            source_workbook=source_workbook,
            output_path=output_target,
            template_id=template_id,
            context=context,
            cancel_token=cancel_token,
        )
    finally:
        source_workbook.close()


def _marks_output_base_from_context(context: dict[str, Any]) -> str:
    metadata_rows = context.get("metadata_rows")
    if not isinstance(metadata_rows, list):
        raise _ve(
            "common.validation_failed_invalid_data",
            code="COURSE_METADATA_MISSING",
        )
    return build_marks_template_filename_base_from_identity(
        academic_year=_metadata_value_for_key(metadata_rows, COURSE_METADATA_ACADEMIC_YEAR_KEY),
        course_code=_metadata_value_for_key(metadata_rows, COURSE_METADATA_COURSE_CODE_KEY),
        semester=_metadata_value_for_key(metadata_rows, COURSE_METADATA_SEMESTER_KEY),
        section=_metadata_value_for_key(metadata_rows, COURSE_METADATA_SECTION_KEY),
    )


def _metadata_value_for_key(metadata_rows: Sequence[Sequence[Any]], key: str) -> str:
    wanted = normalize(key)
    for row in metadata_rows:
        if len(row) < 2:
            continue
        if normalize(row[0]) == wanted:
            return str(row[1] or "").strip()
    return ""


def generate_marks_templates_from_course_details_batch(
    *,
    workbook_paths: Sequence[str | Path],
    output_dir: str | Path,
    allow_overwrite: bool = False,
    output_path_overrides: Mapping[str, str | Path] | None = None,
    cancel_token: CancellationToken | None = None,
) -> dict[str, object]:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.openpyxl_missing",
            code="OPENPYXL_MISSING",
        ) from exc

    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    normalized_output_overrides: dict[str, Path] = {}
    for raw_key, raw_output in dict(output_path_overrides or {}).items():
        source_key = canonical_path_key(raw_key)
        output_value = str(raw_output).strip()
        if not source_key or not output_value:
            continue
        normalized_output_overrides[source_key] = Path(output_value)

    seen_source_keys: set[str] = set()
    seen_output_path_keys: set[str] = set()
    results: dict[str, object] = {}
    generated = 0
    failed = 0
    skipped = 0

    for raw_path in workbook_paths:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()

        source = Path(raw_path)
        source_key = canonical_path_key(source)
        source_value = str(source)

        if source_key in seen_source_keys:
            results[source_key] = {
                "status": "skipped",
                "source_path": source_value,
                "workbook_path": None,
                "output": None,
                "output_path": None,
                "output_url": None,
                "reason": "duplicate_source",
            }
            skipped += 1
            continue
        seen_source_keys.add(source_key)

        try:
            if cancel_token is not None:
                cancel_token.raise_if_cancelled()
            source_workbook = openpyxl.load_workbook(source, data_only=False)
        except JobCancelledError:
            raise
        except Exception as exc:
            results[source_key] = {
                "status": "failed",
                "source_path": source_value,
                "workbook_path": None,
                "output": None,
                "output_path": None,
                "output_url": None,
                "reason": str(exc),
            }
            failed += 1
            continue

        try:
            template_id, context = _prepare_marks_generation_from_workbook(
                source_workbook,
                cancel_token=cancel_token,
            )
            output_base = _marks_output_base_from_context(context)
            output_name = f"{output_base}.xlsx"
            output_path = normalized_output_overrides.get(source_key, output_root / output_name)
            output_path_key = canonical_path_key(output_path)
            if output_path_key in seen_output_path_keys:
                results[source_key] = {
                    "status": "failed",
                    "source_path": source_value,
                    "workbook_path": None,
                    "output": str(output_path),
                    "output_path": str(output_path),
                    "output_url": str(output_path),
                    "reason": "output_name_collision",
                }
                failed += 1
                continue
            seen_output_path_keys.add(output_path_key)

            if output_path.exists() and not allow_overwrite:
                results[source_key] = {
                    "status": "failed",
                    "source_path": source_value,
                    "workbook_path": None,
                    "output": str(output_path),
                    "output_path": str(output_path),
                    "output_url": str(output_path),
                    "reason": "output_already_exists",
                    "existing_output_path": str(output_path),
                }
                failed += 1
                continue
            generated_path = _render_marks_template_to_output(
                source_workbook=source_workbook,
                output_path=output_path,
                template_id=template_id,
                context=context,
                cancel_token=cancel_token,
            )
            output_value = str(generated_path)
            results[source_key] = {
                "status": "generated",
                "source_path": source_value,
                "workbook_path": output_value,
                "output": output_value,
                "output_path": output_value,
                "output_url": output_value,
                "reason": None,
            }
            generated += 1
        except JobCancelledError:
            raise
        except Exception as exc:
            results[source_key] = {
                "status": "failed",
                "source_path": source_value,
                "workbook_path": None,
                "output": None,
                "output_path": None,
                "output_url": None,
                "reason": str(exc),
            }
            failed += 1
        finally:
            source_workbook.close()

    generated_paths = [
        str(entry.get("workbook_path"))
        for entry in results.values()
        if isinstance(entry, dict) and str(entry.get("status")) == "generated" and entry.get("workbook_path")
    ]

    return {
        "total": len(seen_source_keys) + skipped,
        "generated": generated,
        "failed": failed,
        "skipped": skipped,
        "generated_workbook_paths": generated_paths,
        "output_urls": list(generated_paths),
        "results": results,
    }


def _extract_marks_template_context_by_template(workbook: Any, template_id: str) -> dict[str, Any]:
    blueprint = get_blueprint(template_id)

    if blueprint is None:
        raise _ve(
            "validation.template.unknown",
            code="UNKNOWN_TEMPLATE",
            template_id=template_id,
        )

    course_metadata_sheet = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
    assessment_sheet = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
    question_map_sheet = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_QUESTION_MAP)
    students_sheet = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)

    course_metadata_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
    assessment_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
    question_map_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_QUESTION_MAP)
    students_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)

    metadata_rows = _iter_data_rows(workbook[course_metadata_sheet], len(course_metadata_headers))
    metadata_rows = _sheetops._filter_marks_template_metadata_rows(metadata_rows)
    assessment_rows = _iter_data_rows(workbook[assessment_sheet], len(assessment_headers))
    question_rows = _iter_data_rows(workbook[question_map_sheet], len(question_map_headers))
    student_rows = _iter_data_rows(workbook[students_sheet], len(students_headers))

    total_outcomes = _extract_total_outcomes(metadata_rows)
    students = _extract_students(student_rows)
    students_output_headers = (MARKS_ENTRY_ROW_HEADERS[0],) + tuple(students_headers)
    students_output_rows = [[index, reg_no, student_name] for index, (reg_no, student_name) in enumerate(students, start=1)]
    components = _extract_components(
        assessment_rows,
        assessment_sheet=assessment_sheet,
    )
    questions_by_component = _extract_questions(question_rows)

    return {
        "course_metadata_sheet": course_metadata_sheet,
        "assessment_sheet": assessment_sheet,
        "question_map_sheet": question_map_sheet,
        "students_sheet": students_sheet,
        "course_metadata_headers": course_metadata_headers,
        "assessment_headers": assessment_headers,
        "question_map_headers": question_map_headers,
        "students_headers": students_headers,
        "students_output_headers": students_output_headers,
        "metadata_rows": metadata_rows,
        "assessment_rows": assessment_rows,
        "question_rows": question_rows,
        "student_rows": student_rows,
        "students_output_rows": students_output_rows,
        "total_outcomes": total_outcomes,
        "students": students,
        "components": components,
        "questions_by_component": questions_by_component,
    }


def _write_marks_template_workbook_by_template(
    workbook: Any,
    context: dict[str, Any],
    *,
    template_id: str,
    cancel_token: CancellationToken | None = None,
) -> dict[str, Any]:
    return _write_marks_template_workbook(
        workbook,
        context,
        template_id=template_id,
        cancel_token=cancel_token,
    )


def _extract_total_outcomes(metadata_rows: Sequence[Sequence[Any]]) -> int:
    required_key = normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY)
    for row in metadata_rows:
        if len(row) < 2:
            continue
        if normalize(row[0]) == required_key:
            value = coerce_excel_number(row[1])
            if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
                break
            return value
    raise validation_error_from_key("instructor.validation.course_metadata_total_outcomes_invalid")


def _extract_students(student_rows: Sequence[Sequence[Any]]) -> list[tuple[str, str]]:
    students: list[tuple[str, str]] = []
    for row in student_rows:
        reg_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if reg_no and name:
            students.append((reg_no, name))
    if not students:
        raise validation_error_from_key("instructor.validation.students_row_required_one")
    return students


def _extract_components(
    assessment_rows: Sequence[Sequence[Any]],
    *,
    assessment_sheet: str,
) -> list[dict[str, Any]]:
    parsed = parse_assessment_components(
        assessment_rows,
        sheet_name=assessment_sheet,
        row_start=2,
        on_blank_component="skip",
        duplicate_policy="keep_first",
        require_non_empty=True,
        validate_allowed_options=False,
    )
    components: list[dict[str, Any]] = []
    for component in parsed:
        components.append(
            {
                "key": component.component_key,
                "display_name": component.component_name,
                "co_wise_breakup": component.co_wise_breakup,
                "is_direct": component.is_direct,
            }
        )
    if not components:
        raise validation_error_from_key("instructor.validation.assessment_component_required_one")
    return components


def _extract_questions(question_rows: Sequence[Sequence[Any]]) -> dict[str, list[dict[str, Any]]]:
    questions_by_component: dict[str, list[dict[str, Any]]] = {}
    for row in question_rows:
        if len(row) < 4:
            continue
        component_key = normalize(row[0])
        if not component_key:
            continue
        max_marks = coerce_excel_number(row[2])
        if isinstance(max_marks, bool) or not isinstance(max_marks, (int, float)):
            continue
        co_values = parse_co_tokens(row[3])
        if not co_values:
            continue
        questions_by_component.setdefault(component_key, []).append(
            {
                "max_marks": float(max_marks),
                "co_values": co_values,
            }
        )
    return questions_by_component


def _write_marks_template_workbook(
    workbook: Any,
    context: dict[str, Any],
    *,
    template_id: str,
    cancel_token: CancellationToken | None = None,
) -> dict[str, Any]:
    blueprint = get_blueprint(template_id)

    if blueprint is None:
        raise _ve(
            "validation.template.unknown",
            code="UNKNOWN_TEMPLATE",
            template_id=template_id,
        )

    format_bundle = build_marks_template_xlsxwriter_formats(
        workbook,
        template_id=template_id,
        include_column_wrap=True,
        normalize_header_valign_to_center=True,
    )

    header_fmt = format_bundle["header"]
    body_fmt = format_bundle["body"]
    wrapped_body_fmt = format_bundle["body_wrap"]
    wrapped_column_fmt = format_bundle["column_wrap"]
    num_fmt = format_bundle["num"]
    header_num_fmt = format_bundle["header_num"]
    unlocked_body_fmt = format_bundle["unlocked_body"]

    layout_sheets, component_plans = _precompute_marks_layout_manifest(
        context=context,
        cancel_token=cancel_token,
    )

    students = context["students"]
    assessment_wrap_columns = _assessment_wrapped_columns_from_schema(template_id)
    _sheetops._write_two_column_copy_sheet(
        workbook=workbook,
        title=context["course_metadata_sheet"],
        header=context["course_metadata_headers"],
        rows=context["metadata_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
    )
    _sheetops._write_multi_column_copy_sheet(
        workbook=workbook,
        title=context["assessment_sheet"],
        header=context["assessment_headers"],
        rows=context["assessment_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
        num_fmt=num_fmt,
        metadata_rows=context["metadata_rows"],
        wrapped_body_fmt=wrapped_body_fmt,
        wrap_columns=assessment_wrap_columns,
        fit_all_columns_single_page=True,
    )
    _sheetops._write_multi_column_copy_sheet(
        workbook=workbook,
        title=context["question_map_sheet"],
        header=context["question_map_headers"],
        rows=context["question_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
        num_fmt=num_fmt,
        metadata_rows=context["metadata_rows"],
        wrapped_body_fmt=wrapped_body_fmt,
        fit_all_columns_single_page=True,
    )
    _sheetops._write_multi_column_copy_sheet(
        workbook=workbook,
        title=context["students_sheet"],
        header=context["students_output_headers"],
        rows=context["students_output_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
        num_fmt=num_fmt,
        metadata_rows=context["metadata_rows"],
        wrapped_body_fmt=wrapped_body_fmt,
        wrapped_column_fmt=wrapped_column_fmt,
        use_common_student_columns=True,
    )

    for plan in component_plans:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        if plan[LAYOUT_SHEET_SPEC_KEY_KIND] == LAYOUT_SHEET_KIND_DIRECT_CO_WISE:
            _sheetops._write_direct_co_wise_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                plan["questions"],
                header_fmt,
                body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
                num_fmt,
                header_num_fmt,
                unlocked_body_fmt,
            )
        elif plan[LAYOUT_SHEET_SPEC_KEY_KIND] == LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE:
            _sheetops._write_direct_non_co_wise_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                plan["questions"],
                header_fmt,
                body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
                num_fmt,
                header_num_fmt,
                unlocked_body_fmt,
            )
        else:
            _sheetops._write_indirect_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                context["total_outcomes"],
                header_fmt,
                body_fmt,
                unlocked_body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
            )

    return {
        "schema_version": WORKBOOK_INTEGRITY_SCHEMA_VERSION,
        LAYOUT_MANIFEST_KEY_SHEET_ORDER: [entry[LAYOUT_SHEET_SPEC_KEY_NAME] for entry in layout_sheets]
        + [SYSTEM_HASH_SHEET_NAME, SYSTEM_LAYOUT_SHEET],
        LAYOUT_MANIFEST_KEY_SHEETS: layout_sheets,
    }


def _assessment_wrapped_columns_from_schema(template_id: str) -> tuple[int, ...]:
    schema = get_sheet_schema_by_key(template_id, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
    if schema is None:
        return ()
    column_keys_raw = schema.sheet_rules.get("column_keys")
    if not isinstance(column_keys_raw, (tuple, list)):
        return ()
    column_keys = [normalize(value) for value in column_keys_raw]
    wrapped_keys = {normalize("mode"), normalize("participation")}
    return tuple(index for index, key in enumerate(column_keys) if key in wrapped_keys)


def _precompute_marks_layout_manifest(
    *,
    context: dict[str, Any],
    cancel_token: CancellationToken | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    students = context["students"]
    metadata_rows = context["metadata_rows"]
    assessment_rows = context["assessment_rows"]
    questions_by_component = context["questions_by_component"]
    total_outcomes = context["total_outcomes"]
    student_identity_hash = _sheetops._student_identity_hash(students)

    layout_sheets: list[dict[str, Any]] = [
        _sheetops._build_two_column_copy_sheet_spec(
            title=context["course_metadata_sheet"],
            header=context["course_metadata_headers"],
            rows=metadata_rows,
        ),
        _sheetops._build_multi_column_copy_sheet_spec(
            title=context["assessment_sheet"],
            header=context["assessment_headers"],
            rows=assessment_rows,
            metadata_rows=metadata_rows,
        ),
        _sheetops._build_multi_column_copy_sheet_spec(
            title=context["question_map_sheet"],
            header=context["question_map_headers"],
            rows=context["question_rows"],
            metadata_rows=metadata_rows,
        ),
        _sheetops._build_multi_column_copy_sheet_spec(
            title=context["students_sheet"],
            header=context["students_output_headers"],
            rows=context["students_output_rows"],
            metadata_rows=metadata_rows,
        ),
    ]
    component_plans: list[dict[str, Any]] = []

    used_sheet_names = {
        normalize(context["course_metadata_sheet"]),
        normalize(context["assessment_sheet"]),
        normalize(context["question_map_sheet"]),
        normalize(context["students_sheet"]),
    }

    for component in context["components"]:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        component_name = component["display_name"]
        sheet_name = _sheetops._safe_sheet_name(component_name, used_sheet_names)
        questions = questions_by_component.get(component["key"], [])
        if component["is_direct"]:
            if not questions:
                raise validation_error_from_key("instructor.validation.question_map_row_required_one")
            if component["co_wise_breakup"]:
                layout_sheets.append(
                    _sheetops._build_direct_co_wise_sheet_spec(
                        sheet_name=sheet_name,
                        metadata_rows=metadata_rows,
                        component_name=component_name,
                        students=students,
                        questions=questions,
                        student_identity_hash=student_identity_hash,
                    )
                )
                component_plans.append(
                    {
                        LAYOUT_SHEET_SPEC_KEY_KIND: LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
                        "sheet_name": sheet_name,
                        "component_name": component_name,
                        "questions": questions,
                    }
                )
            else:
                layout_sheets.append(
                    _sheetops._build_direct_non_co_wise_sheet_spec(
                        sheet_name=sheet_name,
                        metadata_rows=metadata_rows,
                        component_name=component_name,
                        students=students,
                        questions=questions,
                        student_identity_hash=student_identity_hash,
                    )
                )
                component_plans.append(
                    {
                        LAYOUT_SHEET_SPEC_KEY_KIND: LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
                        "sheet_name": sheet_name,
                        "component_name": component_name,
                        "questions": questions,
                    }
                )
        else:
            layout_sheets.append(
                _sheetops._build_indirect_sheet_spec(
                    sheet_name=sheet_name,
                    metadata_rows=metadata_rows,
                    component_name=component_name,
                    students=students,
                    total_outcomes=total_outcomes,
                    student_identity_hash=student_identity_hash,
                )
            )
            component_plans.append(
                {
                    LAYOUT_SHEET_SPEC_KEY_KIND: LAYOUT_SHEET_KIND_INDIRECT,
                    "sheet_name": sheet_name,
                    "component_name": component_name,
                    "questions": [],
                }
            )

    return layout_sheets, component_plans


def _iter_data_rows(worksheet: Any, expected_col_count: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=2, max_col=expected_col_count, values_only=True):
        values = list(row)
        if any(normalize(value) != "" for value in values):
            rows.append(values)
    return rows


__all__ = [
    "generate_marks_template_from_course_details",
    "generate_marks_templates_from_course_details_batch",
]
