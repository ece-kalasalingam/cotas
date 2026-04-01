"""COURSE_SETUP_V2 course-template workbook validation."""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any

from common.error_catalog import resolve_validation_issue, validation_error_from_key
from common.exceptions import JobCancelledError, ValidationError
from common.jobs import CancellationToken
from common.registry import (
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    COURSE_SETUP_ASSESSMENT_FORMAT_OPTIONS,
    COURSE_SETUP_ASSESSMENT_MODE_OPTIONS,
    COURSE_SETUP_ASSESSMENT_PARTICIPATION_OPTIONS,
    COURSE_SETUP_ASSESSMENT_TYPE_OPTIONS,
    COURSE_SETUP_QUESTION_DOMAIN_LEVEL_OPTIONS,
    COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    COURSE_SETUP_SHEET_KEY_QUESTION_MAP,
    COURSE_SETUP_SHEET_KEY_STUDENTS,
    WEIGHT_TOTAL_EXPECTED,
    WEIGHT_TOTAL_ROUND_DIGITS,
    get_blueprint,
    get_sheet_name_by_key,
    get_sheet_schema_by_key,
)
from common.sample_setup_data import SAMPLE_SETUP_DATA
from common.sheet_schema import SheetSchema, ValidationRule
from common.utils import (
    assert_not_symlink_path,
    coerce_excel_number,
    normalize,
)
from common.workbook_integrity import read_valid_template_id_from_system_hash_sheet
from domain.template_strategy_router import (
    assert_template_id_matches,
)
from domain.template_versions.course_setup_v2_impl.assessment_semantics import (
    parse_assessment_components,
)
from domain.template_versions.course_setup_v2_impl.co_token_parser import (
    parse_co_tokens,
)
from domain.template_versions.course_setup_v2_impl.schema_columns import (
    column_index_by_key,
    required_column_index,
)
from domain.template_versions.course_setup_v2_impl.validation_batch_runner import (
    BatchValidationAccumulator,
    BatchValidationRunner,
)

_TEMPLATE_ID = "COURSE_SETUP_V2"
_COL_COMPONENT = "component"
_COL_WEIGHT_PERCENT = "weight_percent"
_COL_QUESTION_LABEL = "question_label"
_COL_MAX_MARKS = "max_marks"
_COL_CO = "co"
_COL_BLOOM_LEVEL = "bloom_level"
_COL_REG_NO = "reg_no"
_COL_STUDENT_NAME = "student_name"


class _CourseIdentity:
    def __init__(
        self,
        *,
        template_id: str,
        course_code: str,
        semester: str,
        academic_year: str,
        total_outcomes: int,
        section: str,
    ) -> None:
        """Initialize normalized course identity fields.

        Args:
            template_id: Template id for the workbook.
            course_code: Course code value.
            semester: Semester value.
            academic_year: Academic year value.
            total_outcomes: Total configured outcomes.
            section: Section identifier.
        """
        self.template_id = template_id
        self.course_code = course_code
        self.semester = semester
        self.academic_year = academic_year
        self.total_outcomes = total_outcomes
        self.section = section

    def cohort_key(self) -> tuple[str, str, str, int]:
        """Build normalized cohort key tuple for cross-workbook comparison.

        Returns:
            Tuple of normalized cohort identity fields.
        """
        return (
            normalize(self.course_code),
            normalize(self.semester),
            normalize(self.academic_year),
            int(self.total_outcomes),
        )


def _issue_dict(*, code: str, context: dict[str, Any], fallback_message: str) -> dict[str, object]:
    """Build a normalized issue payload from validation issue metadata.

    Args:
        code: Validation issue code.
        context: Structured issue context payload.
        fallback_message: Default message when catalog resolution is unavailable.

    Returns:
        Normalized issue dictionary used by batch validation outputs.
    """
    resolved = resolve_validation_issue(code, context, fallback_message=fallback_message)
    return {
        "code": resolved.code,
        "category": resolved.category,
        "severity": resolved.severity,
        "translation_key": resolved.translation_key,
        "message": resolved.message,
        "context": dict(resolved.context),
    }


class _ValidationCollector:
    def __init__(self) -> None:
        """Initialize collector for accumulating structured validation issues."""
        self._issues: list[dict[str, object]] = []

    def add(self, exc: ValidationError) -> None:
        """Capture one validation exception as a normalized issue payload.

        Args:
            exc: Validation exception to record.
        """
        self._issues.append(
            _issue_dict(
                code=str(getattr(exc, "code", "VALIDATION_ERROR")),
                context=dict(getattr(exc, "context", {}) or {}),
                fallback_message=str(exc).strip() or "Validation failed.",
            )
        )

    def capture(self, fn, *args, **kwargs) -> Any | None:
        """Run callable and capture validation exceptions as issues.

        Args:
            fn: Callable to execute.
            *args: Positional arguments for callable.
            **kwargs: Keyword arguments for callable.

        Returns:
            Callable result, or `None` when a validation exception is captured.
        """
        try:
            return fn(*args, **kwargs)
        except ValidationError as exc:
            self.add(exc)
            return None

    def raise_if_any(self) -> None:
        """Raise aggregated validation error when any issues were collected.

        Raises:
            ValidationError: If one or more issues were recorded.
        """
        if not self._issues:
            return
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COURSE_DETAILS_VALIDATION_FAILED",
            issue_count=len(self._issues),
            issues=list(self._issues),
        )


def validate_course_details_rules(workbook: Any) -> None:
    """Validate COURSE_SETUP_V2 course-details workbook structure and content.

    This validates template identity, sheet order, headers, formula restrictions,
    schema-driven rules, percentage columns, and domain rules for course metadata,
    assessment config, question map, and students.

    Args:
        workbook: Open workbook object (typically openpyxl workbook) to validate.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If any structural or business-rule validation fails.
    """
    blueprint = get_blueprint(_TEMPLATE_ID)
    if blueprint is None:
        raise validation_error_from_key(
            "validation.template.unknown",
            code="UNKNOWN_TEMPLATE",
            template_id=_TEMPLATE_ID,
        )

    collector = _ValidationCollector()
    template_id = collector.capture(read_valid_template_id_from_system_hash_sheet, workbook)
    if isinstance(template_id, str):
        try:
            assert_template_id_matches(
                actual_template_id=template_id,
                expected_template_id=_TEMPLATE_ID,
            )
        except ValidationError as exc:
            collector.add(exc)

    collector.capture(_validate_sheet_order, workbook, blueprint.sheets)
    collector.capture(_reject_any_formula_cells, workbook, blueprint.sheets)
    collector.capture(_validate_sheet_headers, workbook, blueprint.sheets)

    row_data_by_sheet: dict[str, list[tuple[int, list[Any]]]] = {}
    for sheet_schema in blueprint.sheets:
        worksheet = workbook[sheet_schema.name] if sheet_schema.name in workbook.sheetnames else None
        if worksheet is None:
            continue
        row_data = collector.capture(_validated_non_empty_data_rows, worksheet, sheet_schema)
        if not isinstance(row_data, list):
            continue
        row_data_by_sheet[sheet_schema.name] = row_data
        collector.capture(_validate_sheet_rules_from_schema, sheet_schema, row_data)
        collector.capture(_validate_percentage_columns, sheet_schema, row_data)

    identity = collector.capture(_validate_course_metadata_rules, row_data_by_sheet)
    component_config = collector.capture(_validate_assessment_config_rules, row_data_by_sheet)
    if isinstance(identity, _CourseIdentity) and isinstance(component_config, dict):
        collector.capture(_validate_question_map_rules, row_data_by_sheet, component_config, identity.total_outcomes)
    collector.capture(_validate_students_rules, row_data_by_sheet)
    collector.raise_if_any()


def validate_course_details_workbooks(
    workbook_paths: Sequence[str | Path],
    *,
    cancel_token: CancellationToken | None = None,
) -> dict[str, object]:
    """Validate a batch of COURSE_SETUP_V2 course-detail workbooks.

    Args:
        workbook_paths: Source workbook paths to validate.
        cancel_token: Optional cancellation token for cooperative cancellation.

    Returns:
        Batch validation summary with accepted/rejected paths and issue payloads.

    Raises:
        JobCancelledError: If cancellation is requested during batch processing.
    """
    def _validate_path(path: str) -> _CourseIdentity:
        """Validate one workbook path and return extracted identity.

        Args:
            path: Workbook path text.

        Returns:
            Parsed course identity.
        """
        return _validate_course_details_workbook_impl(
            workbook_path=path,
            cancel_token=cancel_token,
        )

    def _on_validated(acc: BatchValidationAccumulator, path: str, identity: _CourseIdentity) -> None:
        """Record successful validation result into batch accumulator.

        Args:
            acc: Batch validation accumulator.
            path: Workbook path text.
            identity: Parsed course identity from validation.
        """
        acc.add_valid(path=path, template_id=identity.template_id)

    runner = BatchValidationRunner[_CourseIdentity](
        issue_builder=_issue_dict,
        duplicate_path_issue_code="COURSE_DETAILS_DUPLICATE_PATH",
        unexpected_issue_code="COURSE_DETAILS_UNEXPECTED_REJECTION",
    )
    return runner.run(
        workbook_paths=workbook_paths,
        validate_path=_validate_path,
        on_validated=_on_validated,
        cancel_token=cancel_token,
    )


def _validate_course_details_workbook_impl(
    *,
    workbook_path: str | Path,
    cancel_token: CancellationToken | None = None,
) -> _CourseIdentity:
    """Validate one course-details workbook and return normalized cohort identity.

    Args:
        workbook_path: Path to the workbook file.
        cancel_token: Optional cancellation token for cooperative cancellation.

    Returns:
        Parsed course identity used for downstream cohort consistency checks.

    Raises:
        ValidationError: If dependency, file access, template, schema, or
            business-rule validation fails.
        JobCancelledError: If cancellation is requested while validating.
    """
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise validation_error_from_key(
            "validation.dependency.openpyxl_missing",
            code="OPENPYXL_MISSING",
        ) from exc

    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise validation_error_from_key(
            "validation.workbook.not_found",
            code="WORKBOOK_NOT_FOUND",
            workbook=str(workbook_file),
        )
    assert_not_symlink_path(workbook_file, context_key="workbook")

    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=False, read_only=False)
    except Exception as exc:
        raise validation_error_from_key(
            "validation.workbook.open_failed",
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(workbook_file),
        ) from exc

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        validate_course_details_rules(workbook)
        metadata_schema = get_sheet_schema_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
        metadata_sheet = get_sheet_name_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
        if metadata_schema is None:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="SCHEMA_MISSING",
                sheet_name=metadata_sheet,
            )
        if metadata_sheet not in workbook.sheetnames:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="SHEET_DATA_REQUIRED",
                sheet_name=metadata_sheet,
            )
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        metadata_rows = _validated_non_empty_data_rows(workbook[metadata_sheet], metadata_schema)
        return _validate_course_metadata_rules({metadata_sheet: metadata_rows})
    except JobCancelledError:
        raise
    finally:
        workbook.close()


def _validate_sheet_order(workbook: Any, sheet_schemas: Sequence[SheetSchema]) -> None:
    """Ensure workbook sheets match expected schema order plus system hash sheet.

    Args:
        workbook: Open workbook object to inspect.
        sheet_schemas: Ordered sheet schemas for the template.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If actual sheet order differs from expected order.
    """
    expected = [sheet.name for sheet in sheet_schemas] + ["__SYSTEM_HASH__"]
    actual = list(workbook.sheetnames)
    if actual != expected:
        raise validation_error_from_key(
            "instructor.validation.workbook_sheet_mismatch",
            template_id=_TEMPLATE_ID,
            expected=expected,
            found=actual,
        )


def _validate_sheet_headers(
    workbook: Any,
    sheet_schemas: Sequence[SheetSchema],
    *,
    cancel_token: CancellationToken | None = None,
) -> None:
    """Validate fixed header rows for all configured template sheets.

    Args:
        workbook: Open workbook object to inspect.
        sheet_schemas: Sheet schemas declaring expected headers.
        cancel_token: Optional cancellation token for cooperative cancellation.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If header row shape or values are invalid.
        JobCancelledError: If cancellation is requested during iteration.
    """
    for sheet_schema in sheet_schemas:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        if len(sheet_schema.header_matrix) != 1:
            raise validation_error_from_key(
                "instructor.validation.sheet_single_header_row",
                sheet_name=sheet_schema.name,
            )
        expected_headers = list(sheet_schema.header_matrix[0])
        worksheet = workbook[sheet_schema.name]
        for col_index, expected in enumerate(expected_headers, start=1):
            actual = worksheet.cell(row=1, column=col_index).value
            if normalize(actual) != normalize(expected):
                raise validation_error_from_key(
                    "instructor.validation.header_mismatch",
                    sheet_name=sheet_schema.name,
                    expected=expected_headers,
                )


def _reject_any_formula_cells(
    workbook: Any,
    sheet_schemas: Sequence[SheetSchema],
    *,
    cancel_token: CancellationToken | None = None,
) -> None:
    """Reject formula cells in template input sheets.

    Args:
        workbook: Open workbook object to inspect.
        sheet_schemas: Sheet schemas to scan.
        cancel_token: Optional cancellation token for cooperative cancellation.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If any cell contains a formula expression.
        JobCancelledError: If cancellation is requested during iteration.
    """
    for sheet_schema in sheet_schemas:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        worksheet = workbook[sheet_schema.name]
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=max(1, int(worksheet.max_row)),
            min_col=1,
            max_col=max(1, int(worksheet.max_column)),
        ):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip().startswith("="):
                    raise validation_error_from_key(
                        "common.validation_failed_invalid_data",
                        code="FORMULA_NOT_ALLOWED",
                        sheet_name=sheet_schema.name,
                        cell=cell.coordinate,
                    )


def _validated_non_empty_data_rows(
    worksheet: Any,
    sheet_schema: SheetSchema,
) -> list[tuple[int, list[Any]]]:
    """Read non-empty data rows and enforce no blank cells within active rows.

    Args:
        worksheet: Worksheet to read.
        sheet_schema: Schema for expected header width and sheet identity.

    Returns:
        List of tuples containing row number and row values.

    Raises:
        ValidationError: If any active row has blank cells or sheet has no data.
    """
    header_count = len(sheet_schema.header_matrix[0])
    rows: list[tuple[int, list[Any]]] = []
    for row_number in range(2, int(worksheet.max_row) + 1):
        values = [worksheet.cell(row=row_number, column=col).value for col in range(1, header_count + 1)]
        has_any = any(normalize(value) != "" for value in values)
        if not has_any:
            continue
        for value in values:
            if normalize(value) == "":
                raise validation_error_from_key(
                    "common.validation_failed_invalid_data",
                    code="CELL_EMPTY_NOT_ALLOWED",
                    sheet_name=sheet_schema.name,
                    row=row_number,
                )
        rows.append((row_number, values))
    if not rows:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="SHEET_DATA_REQUIRED",
            sheet_name=sheet_schema.name,
        )
    return rows


def _validate_sheet_rules_from_schema(
    sheet_schema: SheetSchema,
    rows: list[tuple[int, list[Any]]],
) -> None:
    """Apply schema-declared validation rules to data rows.

    Args:
        sheet_schema: Sheet schema containing validation rule declarations.
        rows: Parsed row-number/value tuples for the sheet.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If any rule check fails.
    """
    headers = list(sheet_schema.header_matrix[0])
    for rule in sheet_schema.validations:
        _apply_schema_rule(sheet_schema.name, headers, rows, rule)


def _apply_schema_rule(
    sheet_name: str,
    headers: list[str],
    rows: list[tuple[int, list[Any]]],
    rule: ValidationRule,
) -> None:
    """Apply one schema validation rule across relevant row/column cells.

    Args:
        sheet_name: Display name of the sheet under validation.
        headers: Header labels for field-name mapping.
        rows: Parsed row-number/value tuples.
        rule: Validation rule to enforce.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If any cell violates the configured rule.
    """
    options = dict(rule.options)
    validation_type = normalize(options.get("validate"))
    if not validation_type:
        return

    first_col = int(rule.first_col) + 1
    last_col = int(rule.last_col) + 1
    allowed_raw = options.get("source")
    allowed_tokens = {
        normalize(value)
        for value in (allowed_raw if isinstance(allowed_raw, list) else [])
    }
    allowed_display = ", ".join(str(value) for value in (allowed_raw if isinstance(allowed_raw, list) else []))
    for row_number, values in rows:
        for col_index in range(first_col, last_col + 1):
            if col_index <= 0 or col_index > len(values):
                continue
            value = values[col_index - 1]
            field_name = headers[col_index - 1] if col_index - 1 < len(headers) else f"col_{col_index}"
            if validation_type == "list":
                if normalize(value) not in allowed_tokens:
                    raise validation_error_from_key(
                        "instructor.validation.allowed_values_required",
                        sheet_name=sheet_name,
                        row=row_number,
                        field=field_name,
                        allowed=allowed_display,
                    )
                continue
            if validation_type == "integer":
                numeric = coerce_excel_number(value)
                if isinstance(numeric, bool) or not isinstance(numeric, int):
                    raise validation_error_from_key(
                        "common.validation_failed_invalid_data",
                        code="INTEGER_VALUE_REQUIRED",
                        sheet_name=sheet_name,
                        row=row_number,
                        field=field_name,
                    )
                criteria = str(options.get("criteria") or "").strip()
                threshold = coerce_excel_number(options.get("value"))
                if criteria == ">" and isinstance(threshold, (int, float)):
                    if numeric <= float(threshold):
                        raise validation_error_from_key(
                            "common.validation_failed_invalid_data",
                            code="INTEGER_VALUE_OUT_OF_RANGE",
                            sheet_name=sheet_name,
                            row=row_number,
                            field=field_name,
                        )
                continue
            if validation_type == "length":
                criteria = str(options.get("criteria") or "").strip().lower()
                text_len = len(str(value).strip())
                minimum = coerce_excel_number(options.get("minimum"))
                maximum = coerce_excel_number(options.get("maximum"))
                if criteria == "between" and isinstance(minimum, (int, float)) and isinstance(maximum, (int, float)):
                    if not (int(minimum) <= text_len <= int(maximum)):
                        raise validation_error_from_key(
                            "common.validation_failed_invalid_data",
                            code="TEXT_LENGTH_OUT_OF_RANGE",
                            sheet_name=sheet_name,
                            row=row_number,
                            field=field_name,
                        )


def _validate_percentage_columns(
    sheet_schema: SheetSchema,
    rows: list[tuple[int, list[Any]]],
) -> None:
    """Validate percentage-like columns for numeric type and 0-100 range.

    Args:
        sheet_schema: Sheet schema with optional percentage key hints.
        rows: Parsed row-number/value tuples.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If percentage cells are non-numeric or out of range.
    """
    headers = list(sheet_schema.header_matrix[0])
    configured_percent_keys = sheet_schema.sheet_rules.get("percentage_column_keys")
    percent_columns: list[int] = []
    if isinstance(configured_percent_keys, (list, tuple)):
        for key in configured_percent_keys:
            if not isinstance(key, str):
                continue
            index = column_index_by_key(sheet_schema, key)
            if index is not None:
                percent_columns.append(index)
    if not percent_columns:
        percent_columns = [
            index
            for index, header in enumerate(headers)
            if "%" in str(header)
        ]
    for row_number, values in rows:
        for col_index in percent_columns:
            value = coerce_excel_number(values[col_index])
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                raise validation_error_from_key(
                    "common.validation_failed_invalid_data",
                    code="PERCENTAGE_NUMERIC_REQUIRED",
                    sheet_name=sheet_schema.name,
                    row=row_number,
                    field=headers[col_index],
                )
            numeric = float(value)
            if numeric < 0.0 or numeric > 100.0:
                raise validation_error_from_key(
                    "common.validation_failed_invalid_data",
                    code="PERCENTAGE_RANGE_INVALID",
                    sheet_name=sheet_schema.name,
                    row=row_number,
                    field=headers[col_index],
                )


def _required_sheet_rows(
    row_data_by_sheet: dict[str, list[tuple[int, list[Any]]]],
    *,
    sheet_name: str,
) -> list[tuple[int, list[Any]]]:
    """Return required sheet rows or raise when sheet data is unavailable.

    Args:
        row_data_by_sheet: Parsed rows keyed by sheet name.
        sheet_name: Required sheet name.

    Returns:
        Row tuples for the required sheet.

    Raises:
        ValidationError: If the required sheet has no parsed data rows.
    """
    rows = row_data_by_sheet.get(sheet_name)
    if isinstance(rows, list):
        return rows
    raise validation_error_from_key(
        "common.validation_failed_invalid_data",
        code="SHEET_DATA_REQUIRED",
        sheet_name=sheet_name,
    )


def _validate_course_metadata_rules(row_data_by_sheet: dict[str, list[tuple[int, list[Any]]]]) -> _CourseIdentity:
    """Validate course metadata sheet rows and derive canonical cohort identity.

    Args:
        row_data_by_sheet: Parsed sheet rows keyed by sheet name.

    Returns:
        Canonical course identity extracted from validated metadata fields.

    Raises:
        ValidationError: If metadata fields are missing, duplicated, or invalid.
    """
    metadata_sheet = get_sheet_name_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
    metadata_schema = get_sheet_schema_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
    if metadata_schema is None:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="SCHEMA_MISSING",
            sheet_name=metadata_sheet,
        )
    rows = _required_sheet_rows(row_data_by_sheet, sheet_name=metadata_sheet)
    fields: dict[str, Any] = {}
    row_by_key: dict[str, int] = {}
    required_keys_raw = metadata_schema.sheet_rules.get("required_field_keys")
    required_keys = {
        normalize(key)
        for key in (required_keys_raw if isinstance(required_keys_raw, (list, tuple)) else [])
        if isinstance(key, str) and key.strip()
    }
    if not required_keys:
        required_keys = {
            normalize(COURSE_METADATA_COURSE_CODE_KEY),
            normalize(COURSE_METADATA_SEMESTER_KEY),
            normalize(COURSE_METADATA_SECTION_KEY),
            normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY),
            normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY),
        }

    field_index = required_column_index(metadata_schema, "field")
    value_index = required_column_index(metadata_schema, "value")
    expected_field_types: dict[str, type] = {}
    for field_name, sample_value in SAMPLE_SETUP_DATA.get(metadata_sheet, []):
        field_key = normalize(field_name)
        expected_field_types[field_key] = int if isinstance(sample_value, int) else str

    for row_number, values in rows:
        key_raw = values[field_index] if field_index < len(values) else ""
        value_raw = values[value_index] if value_index < len(values) else ""
        key = normalize(key_raw)
        if not key:
            raise validation_error_from_key("instructor.validation.course_metadata_field_empty", row=row_number)
        if key in row_by_key:
            raise validation_error_from_key(
                "instructor.validation.course_metadata_duplicate_field",
                row=row_number,
                field=key_raw,
            )
        if key not in expected_field_types:
            raise validation_error_from_key(
                "instructor.validation.course_metadata_unknown_field",
                row=row_number,
                field=key_raw,
            )
        row_by_key[key] = row_number
        if normalize(value_raw) == "":
            raise validation_error_from_key(
                "instructor.validation.course_metadata_value_required",
                row=row_number,
                field=key_raw,
            )
        fields[key] = coerce_excel_number(value_raw)

    missing = sorted(key for key in required_keys if key not in fields or normalize(fields[key]) == "")
    if missing:
        raise validation_error_from_key(
            "instructor.validation.course_metadata_missing_fields",
            fields=", ".join(missing),
        )

    for field_key, expected_type in expected_field_types.items():
        if field_key not in fields:
            continue
        value = fields[field_key]
        if expected_type is int:
            if isinstance(value, bool) or not isinstance(value, int):
                raise validation_error_from_key(
                    "instructor.validation.course_metadata_field_must_be_int",
                    field=field_key,
                )
            continue
        if not isinstance(value, str) or normalize(value) == "":
            raise validation_error_from_key(
                "instructor.validation.course_metadata_field_must_be_non_empty_str",
                field=field_key,
            )

    total_outcomes_key = normalize(str(metadata_schema.sheet_rules.get("total_outcomes_key", "")))
    if not total_outcomes_key:
        total_outcomes_key = normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY)
    total_outcomes_value = fields.get(total_outcomes_key)
    if isinstance(total_outcomes_value, bool) or not isinstance(total_outcomes_value, int) or total_outcomes_value <= 0:
        raise validation_error_from_key("instructor.validation.course_metadata_total_outcomes_invalid")
    return _CourseIdentity(
        template_id=_TEMPLATE_ID,
        course_code=str(fields.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")).strip(),
        semester=str(fields.get(normalize(COURSE_METADATA_SEMESTER_KEY), "")).strip(),
        academic_year=str(fields.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")).strip(),
        total_outcomes=total_outcomes_value,
        section=str(fields.get(normalize(COURSE_METADATA_SECTION_KEY), "")).strip(),
    )


def _validate_assessment_config_rules(
    row_data_by_sheet: dict[str, list[tuple[int, list[Any]]]],
) -> dict[str, dict[str, Any]]:
    """Validate assessment configuration and build component configuration map.

    Args:
        row_data_by_sheet: Parsed sheet rows keyed by sheet name.

    Returns:
        Mapping of normalized component keys to parsed component attributes.

    Raises:
        ValidationError: If component rows, constraints, or totals are invalid.
    """
    assessment_sheet = get_sheet_name_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
    assessment_schema = get_sheet_schema_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
    if assessment_schema is None:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="SCHEMA_MISSING",
            sheet_name=assessment_sheet,
        )
    rows = _required_sheet_rows(row_data_by_sheet, sheet_name=assessment_sheet)
    row_numbers = [row_number for row_number, _values in rows]
    row_values = [values for _row_number, values in rows]

    direct_weight_total = 0.0
    indirect_weight_total = 0.0
    direct_count = 0
    indirect_count = 0
    component_config: dict[str, dict[str, Any]] = {}
    allowed_assessment_type = {normalize(value) for value in COURSE_SETUP_ASSESSMENT_TYPE_OPTIONS}
    allowed_assessment_format = {normalize(value) for value in COURSE_SETUP_ASSESSMENT_FORMAT_OPTIONS}
    allowed_mode = {normalize(value) for value in COURSE_SETUP_ASSESSMENT_MODE_OPTIONS}
    allowed_participation = {normalize(value) for value in COURSE_SETUP_ASSESSMENT_PARTICIPATION_OPTIONS}

    parsed_components = parse_assessment_components(
        row_values,
        sheet_name=assessment_sheet,
        row_numbers=row_numbers,
        row_start=row_numbers[0] if row_numbers else 2,
        on_blank_component="error",
        duplicate_policy="error",
        require_non_empty=True,
        validate_allowed_options=True,
        assessment_type_allowed_tokens=allowed_assessment_type,
        assessment_type_allowed_display=COURSE_SETUP_ASSESSMENT_TYPE_OPTIONS,
        assessment_format_allowed_tokens=allowed_assessment_format,
        assessment_format_allowed_display=COURSE_SETUP_ASSESSMENT_FORMAT_OPTIONS,
        mode_allowed_tokens=allowed_mode,
        mode_allowed_display=COURSE_SETUP_ASSESSMENT_MODE_OPTIONS,
        participation_allowed_tokens=allowed_participation,
        participation_allowed_display=COURSE_SETUP_ASSESSMENT_PARTICIPATION_OPTIONS,
    )

    for component in parsed_components:
        component_key = component.component_key
        if component.is_direct:
            direct_weight_total += component.weight
            direct_count += 1
        else:
            indirect_weight_total += component.weight
            indirect_count += 1
        component_config[component_key] = {
            "display_name": component.component_name,
            "co_wise_breakup": component.co_wise_breakup,
            "is_direct": component.is_direct,
        }

    if direct_count <= 0:
        raise validation_error_from_key("instructor.validation.assessment_direct_missing")
    if indirect_count <= 0:
        raise validation_error_from_key("instructor.validation.assessment_indirect_missing")
    expected_total = coerce_excel_number(assessment_schema.sheet_rules.get("weight_total_expected"))
    if isinstance(expected_total, bool) or not isinstance(expected_total, (int, float)):
        expected_total = WEIGHT_TOTAL_EXPECTED
    round_digits = coerce_excel_number(assessment_schema.sheet_rules.get("weight_total_round_digits"))
    if isinstance(round_digits, bool) or not isinstance(round_digits, int):
        round_digits = WEIGHT_TOTAL_ROUND_DIGITS
    indirect_tools_min = coerce_excel_number(assessment_schema.sheet_rules.get("indirect_tools_min"))
    if isinstance(indirect_tools_min, bool) or not isinstance(indirect_tools_min, int):
        indirect_tools_min = 1
    indirect_tools_max = coerce_excel_number(assessment_schema.sheet_rules.get("indirect_tools_max"))
    if isinstance(indirect_tools_max, bool) or not isinstance(indirect_tools_max, int):
        indirect_tools_max = 3

    if indirect_count < indirect_tools_min or indirect_count > indirect_tools_max:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="INDIRECT_TOOL_COUNT_INVALID",
            minimum=indirect_tools_min,
            maximum=indirect_tools_max,
            found=indirect_count,
        )
    if round(direct_weight_total, round_digits) != float(expected_total):
        raise validation_error_from_key(
            "instructor.validation.assessment_direct_total_invalid",
            found=direct_weight_total,
        )
    if round(indirect_weight_total, round_digits) != float(expected_total):
        raise validation_error_from_key(
            "instructor.validation.assessment_indirect_total_invalid",
            found=indirect_weight_total,
        )
    return component_config


def _validate_question_map_rules(
    row_data_by_sheet: dict[str, list[tuple[int, list[Any]]]],
    component_config: dict[str, dict[str, Any]],
    total_outcomes: int,
) -> None:
    """Validate question-map rows against component config and CO constraints.

    Args:
        row_data_by_sheet: Parsed sheet rows keyed by sheet name.
        component_config: Parsed assessment component configuration.
        total_outcomes: Maximum allowed CO index from metadata.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If question map rows violate business constraints.
    """
    question_sheet = get_sheet_name_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_QUESTION_MAP)
    question_schema = get_sheet_schema_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_QUESTION_MAP)
    if question_schema is None:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="SCHEMA_MISSING",
            sheet_name=question_sheet,
        )
    question_headers = list(question_schema.header_matrix[0])
    component_idx = required_column_index(question_schema, _COL_COMPONENT)
    question_idx = required_column_index(question_schema, _COL_QUESTION_LABEL)
    max_marks_idx = required_column_index(question_schema, _COL_MAX_MARKS)
    co_idx = required_column_index(question_schema, _COL_CO)
    bloom_idx = required_column_index(question_schema, _COL_BLOOM_LEVEL)
    rows = _required_sheet_rows(row_data_by_sheet, sheet_name=question_sheet)
    allowed_bloom_levels = {normalize(value) for value in COURSE_SETUP_QUESTION_DOMAIN_LEVEL_OPTIONS}
    question_count_by_component: dict[str, int] = {}
    seen_co_wise_questions: set[tuple[str, str]] = set()

    for row_number, values in rows:
        component_raw = values[component_idx]
        question_raw = values[question_idx]
        max_marks_raw = values[max_marks_idx]
        co_raw = values[co_idx]
        bloom_raw = values[bloom_idx]

        component_key = normalize(component_raw)
        if component_key not in component_config:
            raise validation_error_from_key(
                "instructor.validation.question_component_unknown",
                row=row_number,
                component=component_raw,
            )
        question_key = normalize(question_raw)
        if not question_key:
            raise validation_error_from_key(
                "instructor.validation.question_label_required",
                row=row_number,
            )
        max_marks = coerce_excel_number(max_marks_raw)
        if isinstance(max_marks, bool) or not isinstance(max_marks, (int, float)):
            raise validation_error_from_key("instructor.validation.question_max_marks_numeric", row=row_number)
        if float(max_marks) <= 0.0:
            raise validation_error_from_key("instructor.validation.question_max_marks_positive", row=row_number)

        co_values = parse_co_tokens(co_raw, dedupe=False)
        if not co_values:
            raise validation_error_from_key("instructor.validation.question_co_required", row=row_number)
        if len(set(co_values)) != len(co_values):
            raise validation_error_from_key("instructor.validation.question_co_no_repeat", row=row_number)
        if any(value <= 0 or value > total_outcomes for value in co_values):
            raise validation_error_from_key(
                "instructor.validation.question_co_out_of_range",
                row=row_number,
                total_outcomes=total_outcomes,
            )
        if normalize(bloom_raw) not in allowed_bloom_levels:
            raise validation_error_from_key(
                "instructor.validation.allowed_values_required",
                sheet_name=question_sheet,
                row=row_number,
                field=question_headers[bloom_idx],
                allowed=", ".join(COURSE_SETUP_QUESTION_DOMAIN_LEVEL_OPTIONS),
            )

        is_co_wise = bool(component_config[component_key]["co_wise_breakup"])
        question_count_by_component[component_key] = question_count_by_component.get(component_key, 0) + 1
        if is_co_wise:
            if len(co_values) != 1:
                raise validation_error_from_key(
                    "instructor.validation.question_co_wise_requires_one",
                    row=row_number,
                    component=component_raw,
                )
            question_id = (component_key, question_key)
            if question_id in seen_co_wise_questions:
                raise validation_error_from_key(
                    "instructor.validation.question_duplicate_for_component",
                    row=row_number,
                    question=question_raw,
                    component=component_raw,
                )
            seen_co_wise_questions.add(question_id)

    for component_key, config in component_config.items():
        count = question_count_by_component.get(component_key, 0)
        if bool(config.get("is_direct")) and count <= 0:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="QUESTION_MAP_COMPONENT_MISSING",
                component=config.get("display_name", component_key),
            )


def _validate_students_rules(row_data_by_sheet: dict[str, list[tuple[int, list[Any]]]]) -> None:
    """Validate student rows for required identity fields and uniqueness.

    Args:
        row_data_by_sheet: Parsed sheet rows keyed by sheet name.

    Returns:
        None. Validation succeeds silently.

    Raises:
        ValidationError: If student identity fields are missing or duplicated.
    """
    students_sheet = get_sheet_name_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_STUDENTS)
    students_schema = get_sheet_schema_by_key(_TEMPLATE_ID, COURSE_SETUP_SHEET_KEY_STUDENTS)
    if students_schema is None:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="SCHEMA_MISSING",
            sheet_name=students_sheet,
        )
    reg_no_index = required_column_index(students_schema, _COL_REG_NO)
    student_name_index = required_column_index(students_schema, _COL_STUDENT_NAME)
    rows = _required_sheet_rows(row_data_by_sheet, sheet_name=students_sheet)
    seen_reg_numbers: set[str] = set()
    for row_number, values in rows:
        reg_no = str(values[reg_no_index]).strip() if reg_no_index < len(values) else ""
        student_name = str(values[student_name_index]).strip() if student_name_index < len(values) else ""
        if not reg_no or not student_name:
            raise validation_error_from_key(
                "instructor.validation.students_reg_and_name_required",
                row=row_number,
            )
        reg_key = normalize(reg_no)
        if reg_key in seen_reg_numbers:
            raise validation_error_from_key(
                "instructor.validation.students_duplicate_reg_no",
                row=row_number,
                reg_no=reg_no,
            )
        seen_reg_numbers.add(reg_key)


__all__ = [
    "validate_course_details_rules",
    "validate_course_details_workbooks",
]


