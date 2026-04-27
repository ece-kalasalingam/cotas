"""Shared CO attainment generation/signature helpers for template strategy routing."""

from __future__ import annotations

import hashlib
import json
import logging
import os
import random
import re
import shutil
import sqlite3
import tempfile
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from common.constants import (
    APP_NAME,
    CO_ANALYSIS_INSTITUTION_ROWS,
    CO_ANALYSIS_SHEET_FOOTER_TEXT,
    CO_ATTAINMENT_LEVEL_DEFAULT,
    CO_ATTAINMENT_PERCENT_DEFAULT,
    CO_REPORT_ABSENT_TOKEN,
    CO_REPORT_DIRECT_SHEET_SUFFIX,
    CO_REPORT_HEADER_REG_NO,
    CO_REPORT_HEADER_SERIAL,
    CO_REPORT_HEADER_STUDENT_NAME,
    CO_REPORT_INDIRECT_SHEET_SUFFIX,
    CO_REPORT_MAX_DECIMAL_PLACES,
    CO_REPORT_NOT_APPLICABLE_TOKEN,
    DIRECT_RATIO,
    ID_COURSE_SETUP,
    INDIRECT_RATIO,
    LAYOUT_MANIFEST_KEY_SHEETS,
    LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
    LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
    LAYOUT_SHEET_KIND_INDIRECT,
    LEVEL_1_THRESHOLD,
    LEVEL_2_THRESHOLD,
    LEVEL_3_THRESHOLD,
    LIKERT_MAX,
    LIKERT_MIN,
    WORKBOOK_INTEGRITY_SCHEMA_VERSION,
)
from common.error_catalog import validation_error_from_key
from common.excel_sheet_layout import (
    apply_xlsxwriter_column_widths as _apply_xlsxwriter_column_widths,
)
from common.excel_sheet_layout import (
    apply_xlsxwriter_layout as _apply_xlsxwriter_layout_shared,
)
from common.excel_sheet_layout import (
    build_template_xlsxwriter_formats as _build_template_xlsxwriter_formats,
)
from common.excel_sheet_layout import (
    compute_sampled_column_widths as _compute_sampled_column_widths,
)
from common.excel_sheet_layout import (
    write_sheet_footer_xlsxwriter,
    write_two_column_metadata_rows,
)
from common.exceptions import AppSystemError, ValidationError
from common.jobs import CancellationToken
from common.registry import (
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG,
    COURSE_SETUP_SHEET_KEY_CO_DESCRIPTION,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    COURSE_SETUP_SHEET_KEY_STUDENTS,
)
from common.registry import (
    SYSTEM_HASH_HEADER_TEMPLATE_HASH as SYSTEM_HASH_TEMPLATE_HASH_HEADER,
)
from common.registry import (
    SYSTEM_HASH_HEADER_TEMPLATE_ID as SYSTEM_HASH_TEMPLATE_ID_HEADER,
)
from common.registry import SYSTEM_HASH_SHEET_NAME as SYSTEM_HASH_SHEET
from common.registry import (
    get_sheet_headers_by_key,
    get_sheet_name_by_key,
)
from common.runtime_dependency_guard import import_runtime_dependency
from common.utils import (
    app_runtime_storage_dir,
    canonical_path_key,
    coerce_excel_number,
    create_app_runtime_sqlite_file,
    normalize,
    ratio_percent_token,
)
from common.workbook_integrity.constants import (
    SYSTEM_LAYOUT_MANIFEST_HASH_HEADER,
    SYSTEM_LAYOUT_MANIFEST_HEADER,
    SYSTEM_LAYOUT_SHEET,
)
from common.workbook_integrity.workbook_signing import (
    sign_payload,
    verify_payload_signature,
)
from domain.template_strategy_router import (
    read_template_id_from_system_hash_sheet_if_valid,
    read_valid_system_workbook_payload,
    read_valid_template_id_from_system_hash_sheet,
)
from domain.template_versions.course_setup_v2_impl.co_report_sheet_generator import (
    co_direct_sheet_name,
    co_indirect_sheet_name,
)
from domain.template_versions.course_setup_v2_impl.co_report_sheet_generator import (
    course_metadata_headers as _course_metadata_headers,
)
from domain.template_versions.course_setup_v2_impl.co_report_sheet_generator import (
    ratio_total_header as _ratio_total_header,
)
from domain.template_versions.course_setup_v2_impl.co_report_sheet_generator import (
    write_co_outcome_sheets,
    write_co_outcome_sheets_openpyxl,
)
from domain.template_versions.course_setup_v2_impl.co_cip_json_builder import (
    build_cip_payload,
)
from domain.template_versions.course_setup_v2_impl.co_description_template_validator import (
    CoDescriptionRecord as _CoDescriptionRecord,
    read_co_description_records,
    validate_co_description_workbooks,
)

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}

_logger = logging.getLogger(__name__)
_COURSE_METADATA_COURSE_NAME_KEY = "course_name"
_STYLE_CACHE_ATTR = "_focus_coordinator_style_cache"
_CO_REPORT_NAME_TOKEN_RE = re.compile(r"(?:[_\-\s]*co[_\-\s]*report)+$", re.IGNORECASE)
_SEMESTER_PREFIX_TOKEN_RE = re.compile(r"^sem(?:ester)?[\s\-_]*([0-9]{1,2}|[ivxlcdm]+)$", re.IGNORECASE)
_SEMESTER_VALUE_TOKEN_RE = re.compile(r"^(?:[0-9]{1,2}|[ivxlcdm]+)$", re.IGNORECASE)
_HEADER_SCAN_MAX_ROWS = 200
_AGGREGATION_STUDENTS_PER_SHEET = 150
_DEDUP_SQLITE_THRESHOLD_ENTRIES = 10_000
_DEDUP_SQLITE_PREFIX = "focus_co_dedup_"
_DEDUP_SQLITE_SUFFIX = ".sqlite3"
_SIGNATURE_VALIDATION_MAX_WORKERS = 8
_PASS_PERCENTAGE_SHEET_NAME = "Pass_Percentage"


@dataclass(slots=True, frozen=True)
class FinalReportWorkbookSignature:
    template_id: str
    course_code: str
    total_outcomes: int
    section: str
    direct_sheet_count: int
    indirect_sheet_count: int


_FinalReportSignature = FinalReportWorkbookSignature
COURSE_METADATA_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)


@dataclass(slots=True, frozen=True)
class _CoAttainmentRow:
    reg_hash: int
    reg_no: str
    student_name: str
    direct_score: float | str
    indirect_score: float | str
    worksheet_name: str
    workbook_name: str
    direct_component_scores: tuple[float | str, ...] = ()
    indirect_component_scores: tuple[float | str, ...] = ()


@dataclass(slots=True, frozen=True)
class _ParsedScoreRow:
    reg_hash: int
    reg_key: str
    reg_no: str
    student_name: str
    score: float | str


@dataclass(slots=True, frozen=True)
class _DirectComponentColumn:
    name: str
    max_marks: float
    weight: float
    score_column: int


@dataclass(slots=True, frozen=True)
class _IndirectComponentColumn:
    name: str
    weight: float
    score_column: int


@dataclass(slots=True)
class _CoReportComponent:
    name: str
    weight: float
    max_by_co: dict[int, float]
    marks_by_co: dict[int, list[Any]]


@dataclass(slots=True)
class _CoOutputSheetState:
    sheet: Any
    header_row_index: int
    formats: dict[str, Any]
    next_row_index: int
    next_serial: int
    on_roll: int
    attended: int
    level_counts: dict[int, int]


@dataclass(slots=True, frozen=True)
class _CoAttainmentWorkbookResult:
    output_path: Path
    duplicate_reg_count: int
    duplicate_entries: tuple[tuple[str, str, str], ...]
    inner_join_drop_count: int = 0
    inner_join_drop_details: tuple[str, ...] = ()
    word_report_path: Path | None = None
    word_report_error_key: str | None = None
    cip_payload: dict[str, object] | None = None


class _RegisterDedupStore:
    def __init__(self, *, total_outcomes: int, use_sqlite: bool) -> None:
        """Init.
        
        Args:
            total_outcomes: Parameter value (int).
            use_sqlite: Parameter value (bool).
        
        Returns:
            None.
        
        Raises:
            None.
        """
        self._total_outcomes = total_outcomes
        self._use_sqlite = use_sqlite
        self._memory_sets: dict[int, set[int]] = {}
        self._conn: sqlite3.Connection | None = None
        self._db_path: str | None = None
        if use_sqlite:
            _cleanup_stale_dedup_sqlite_files()
            fd, db_path = create_app_runtime_sqlite_file(
                APP_NAME,
                prefix=_DEDUP_SQLITE_PREFIX,
                suffix=_DEDUP_SQLITE_SUFFIX,
            )
            self._db_path = db_path
            self._conn = sqlite3.connect(db_path)
            self._conn.execute("PRAGMA journal_mode=OFF")
            self._conn.execute("PRAGMA synchronous=OFF")
            self._conn.execute("PRAGMA temp_store=MEMORY")
            self._conn.execute("PRAGMA secure_delete=ON")
            self._conn.execute(
                "CREATE TABLE IF NOT EXISTS dedup (co_index INTEGER NOT NULL, reg_hash INTEGER NOT NULL, PRIMARY KEY(co_index, reg_hash))"
            )
            self._conn.commit()
            try:
                import os

                os.close(fd)
            except OSError:
                pass
        else:
            self._memory_sets = {co: set() for co in range(1, total_outcomes + 1)}

    def add_if_absent(self, *, co_index: int, reg_hash: int) -> bool:
        """Add if absent.
        
        Args:
            co_index: Parameter value (int).
            reg_hash: Parameter value (int).
        
        Returns:
            bool: Return value.
        
        Raises:
            None.
        """
        if self._use_sqlite and self._conn is not None:
            cursor = self._conn.execute(
                "INSERT OR IGNORE INTO dedup (co_index, reg_hash) VALUES (?, ?)",
                (co_index, int(reg_hash)),
            )
            return cursor.rowcount == 1
        bucket = self._memory_sets.setdefault(co_index, set())
        if reg_hash in bucket:
            return False
        bucket.add(reg_hash)
        return True

    def close(self) -> None:
        """Close.
        
        Args:
            None.
        
        Returns:
            None.
        
        Raises:
            None.
        """
        if self._conn is not None:
            try:
                # Explicitly wipe the sqlite table before close in normal paths.
                self._conn.execute("DELETE FROM dedup")
                self._conn.commit()
            except Exception:
                _logger.debug("Failed to clear sqlite dedup table before close.", exc_info=True)
            self._conn.close()
            self._conn = None
        if self._db_path:
            try:
                Path(self._db_path).unlink(missing_ok=True)
            except OSError:
                _logger.debug("Unable to remove dedup sqlite db: %s", self._db_path, exc_info=True)
            self._db_path = None


def _should_use_sqlite_dedup(*, source_count: int, total_outcomes: int) -> bool:
    """Should use sqlite dedup.
    
    Args:
        source_count: Parameter value (int).
        total_outcomes: Parameter value (int).
    
    Returns:
        bool: Return value.
    
    Raises:
        None.
    """
    estimated_entries = source_count * total_outcomes * _AGGREGATION_STUDENTS_PER_SHEET
    return estimated_entries >= _DEDUP_SQLITE_THRESHOLD_ENTRIES


def _cleanup_stale_dedup_sqlite_files() -> None:
    """Cleanup stale dedup sqlite files.
    
    Args:
        None.
    
    Returns:
        None.
    
    Raises:
        None.
    """
    temp_root = Path(tempfile.gettempdir())
    storage_sqlite_root = app_runtime_storage_dir(APP_NAME) / "sqlite"
    roots: tuple[Path, ...] = (temp_root, storage_sqlite_root)
    patterns = (
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}",
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}-wal",
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}-shm",
    )
    for root in roots:
        if not root.exists():
            continue
        for pattern in patterns:
            for path in root.glob(pattern):
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    _logger.debug("Unable to remove stale dedup sqlite file: %s", path, exc_info=True)


def _build_co_attainment_default_name(source_path: Path, *, section: str = "") -> str:
    """Build co attainment default name.
    
    Args:
        source_path: Parameter value (Path).
        section: Parameter value (str).
    
    Returns:
        str: Return value.
    
    Raises:
        None.
    """
    stem = source_path.stem.strip()
    cleaned = _CO_REPORT_NAME_TOKEN_RE.sub("", stem).rstrip("_- ").strip()
    section_token = section.strip()
    if section_token:
        parts = [part for part in cleaned.split("_") if normalize(part) != normalize(section_token)]
        cleaned = "_".join(parts).strip("_- ")
    tokens = [token.strip() for token in cleaned.split("_") if token.strip()]
    filtered_tokens: list[str] = []
    semester_removed = False
    for index, token in enumerate(tokens):
        normalized_token = normalize(token)
        if _SEMESTER_PREFIX_TOKEN_RE.match(normalized_token):
            semester_removed = True
            continue
        if (
            not semester_removed
            and index > 0
            and _SEMESTER_VALUE_TOKEN_RE.match(normalized_token)
            and len(normalized_token) <= 4
        ):
            semester_removed = True
            continue
        filtered_tokens.append(token)
    cleaned = "_".join(filtered_tokens).strip("_- ")
    base = cleaned if cleaned else stem
    return f"{base}_CO_Attainment.xlsx"


def _is_supported_excel_file(path: Path) -> bool:
    """Is supported excel file.
    
    Args:
        path: Parameter value (Path).
    
    Returns:
        bool: Return value.
    
    Raises:
        None.
    """
    return path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES


def _filter_excel_paths(paths: Iterable[str]) -> list[Path]:
    """Filter excel paths.
    
    Args:
        paths: Parameter value (Iterable[str]).
    
    Returns:
        list[Path]: Return value.
    
    Raises:
        None.
    """
    collected: list[Path] = []
    seen: set[str] = set()
    for value in paths:
        path = Path(value)
        if not _is_supported_excel_file(path):
            continue
        key = canonical_path_key(path)
        if key in seen:
            continue
        seen.add(key)
        collected.append(path.resolve())
    return collected


_VERIFY_SIGNATURE = Callable[[str, str], bool]


def _count_co_sheets_from_manifest(manifest: dict[str, Any]) -> tuple[int, int] | None:
    """Count co sheets from manifest.
    
    Args:
        manifest: Parameter value (dict[str, Any]).
    
    Returns:
        tuple[int, int] | None: Return value.
    
    Raises:
        None.
    """
    sheets = manifest.get(LAYOUT_MANIFEST_KEY_SHEETS, [])
    if not isinstance(sheets, list):
        return None
    direct = 0
    indirect = 0
    for entry in sheets:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip()
        if not name:
            continue
        if name.endswith(CO_REPORT_DIRECT_SHEET_SUFFIX):
            direct += 1
        elif name.endswith(CO_REPORT_INDIRECT_SHEET_SUFFIX):
            indirect += 1
    if direct <= 0 and indirect <= 0:
        return None
    return direct, indirect


def read_layout_manifest_co_sheet_counts(
    workbook: Any,
    *,
    verify_signature: _VERIFY_SIGNATURE = verify_payload_signature,
) -> tuple[int, int] | None:
    """Read layout manifest co sheet counts.
    
    Args:
        workbook: Parameter value (Any).
        verify_signature: Parameter value (_VERIFY_SIGNATURE).
    
    Returns:
        tuple[int, int] | None: Return value.
    
    Raises:
        None.
    """
    try:
        payload = read_valid_system_workbook_payload(workbook, verify_signature=verify_signature)
    except Exception:
        return None
    return _count_co_sheets_from_manifest(payload.manifest)


def read_course_metadata_signature(
    workbook: Any,
    *,
    course_code_key: str,
    total_outcomes_key: str,
    section_key: str,
) -> tuple[str, int, str] | None:
    """Read course metadata signature.
    
    Args:
        workbook: Parameter value (Any).
        course_code_key: Parameter value (str).
        total_outcomes_key: Parameter value (str).
        section_key: Parameter value (str).
    
    Returns:
        tuple[str, int, str] | None: Return value.
    
    Raises:
        None.
    """
    template_id = read_template_id_from_system_hash_sheet_if_valid(workbook)
    if not template_id:
        return None
    metadata_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
    if metadata_sheet_name not in getattr(workbook, "sheetnames", []):
        return None
    metadata = _extract_course_metadata_fields(workbook[metadata_sheet_name])

    course_code = metadata.get(normalize(course_code_key), "").strip()
    section = metadata.get(normalize(section_key), "").strip()
    total_token = metadata.get(normalize(total_outcomes_key), "").strip()
    if not course_code or not section or not total_token:
        return None
    try:
        total_outcomes = int(float(total_token))
    except (TypeError, ValueError):
        return None
    if total_outcomes <= 0:
        return None
    return course_code, total_outcomes, section


def extract_final_report_signature_from_path(
    path: Path,
    *,
    verify_signature: _VERIFY_SIGNATURE = verify_payload_signature,
) -> FinalReportWorkbookSignature | None:
    """Extract final report signature from path.
    
    Args:
        path: Parameter value (Path).
        verify_signature: Parameter value (_VERIFY_SIGNATURE).
    
    Returns:
        FinalReportWorkbookSignature | None: Return value.
    
    Raises:
        None.
    """
    try:
        import openpyxl
    except Exception:
        return None
    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return None
    try:
        template_id = read_template_id_from_system_hash_sheet_if_valid(
            workbook,
            verify_signature=verify_signature,
        )
        if not template_id:
            return None
        sheet_counts = read_layout_manifest_co_sheet_counts(workbook, verify_signature=verify_signature)
        if sheet_counts is None:
            return None
        metadata = read_course_metadata_signature(
            workbook,
            course_code_key=COURSE_METADATA_COURSE_CODE_KEY,
            total_outcomes_key=COURSE_METADATA_TOTAL_OUTCOMES_KEY,
            section_key=COURSE_METADATA_SECTION_KEY,
        )
        if metadata is None:
            return None
        course_code, total_outcomes, section = metadata
        direct_sheet_count, indirect_sheet_count = sheet_counts
        return FinalReportWorkbookSignature(
            template_id=template_id,
            course_code=course_code,
            total_outcomes=total_outcomes,
            section=section,
            direct_sheet_count=direct_sheet_count,
            indirect_sheet_count=indirect_sheet_count,
        )
    finally:
        workbook.close()


def _layout_sheet_specs_by_name(workbook: Any) -> dict[str, dict[str, Any]]:
    """Layout sheet specs by name.
    
    Args:
        workbook: Parameter value (Any).
    
    Returns:
        dict[str, dict[str, Any]]: Return value.
    
    Raises:
        None.
    """
    try:
        payload = read_valid_system_workbook_payload(workbook, verify_signature=verify_payload_signature)
    except Exception:
        return {}
    manifest = payload.manifest if isinstance(payload.manifest, dict) else {}
    raw_sheets = manifest.get(LAYOUT_MANIFEST_KEY_SHEETS, [])
    if not isinstance(raw_sheets, list):
        return {}
    by_name: dict[str, dict[str, Any]] = {}
    for entry in raw_sheets:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip()
        if not name:
            continue
        by_name[name] = entry
    return by_name


def _has_valid_final_co_report(path: Path) -> bool:
    """Has valid final co report.
    
    Args:
        path: Parameter value (Path).
    
    Returns:
        bool: Return value.
    
    Raises:
        None.
    """
    return _extract_final_report_signature(path) is not None


def _read_template_id_from_hash_sheet(workbook: Any) -> str | None:
    """Read template id from hash sheet.
    
    Args:
        workbook: Parameter value (Any).
    
    Returns:
        str | None: Return value.
    
    Raises:
        None.
    """
    return read_template_id_from_system_hash_sheet_if_valid(
        workbook,
        verify_signature=verify_payload_signature,
    )


def _read_report_sheet_counts(workbook: Any) -> tuple[int, int] | None:
    """Read report sheet counts.
    
    Args:
        workbook: Parameter value (Any).
    
    Returns:
        tuple[int, int] | None: Return value.
    
    Raises:
        None.
    """
    return read_layout_manifest_co_sheet_counts(
        workbook,
        verify_signature=verify_payload_signature,
    )


def _read_signature_metadata(workbook: Any) -> tuple[str, int, str] | None:
    """Read signature metadata.
    
    Args:
        workbook: Parameter value (Any).
    
    Returns:
        tuple[str, int, str] | None: Return value.
    
    Raises:
        None.
    """
    return read_course_metadata_signature(
        workbook,
        course_code_key=COURSE_METADATA_COURSE_CODE_KEY,
        total_outcomes_key=COURSE_METADATA_TOTAL_OUTCOMES_KEY,
        section_key=COURSE_METADATA_SECTION_KEY,
    )


def _extract_final_report_signature(path: Path) -> _FinalReportSignature | None:
    """Extract final report signature.
    
    Args:
        path: Parameter value (Path).
    
    Returns:
        _FinalReportSignature | None: Return value.
    
    Raises:
        None.
    """
    return extract_final_report_signature_from_path(
        path,
        verify_signature=verify_payload_signature,
    )


def _analyze_dropped_files(
    dropped_files: list[str],
    *,
    existing_keys: set[str],
    existing_paths: list[str],
    token: CancellationToken,
) -> dict[str, object]:
    """Analyze dropped files.
    
    Args:
        dropped_files: Parameter value (list[str]).
        existing_keys: Parameter value (set[str]).
        existing_paths: Parameter value (list[str]).
        token: Parameter value (CancellationToken).
    
    Returns:
        dict[str, object]: Return value.
    
    Raises:
        None.
    """
    accepted = _filter_excel_paths(dropped_files)
    seen = set(existing_keys)
    added: list[str] = []
    duplicates = 0
    invalid_final_report: list[str] = []
    invalid_final_report_details: list[dict[str, str]] = []
    existing_resolved = [Path(path).resolve() for path in existing_paths if path]
    baseline_signature: _FinalReportSignature | None = None
    seen_sections: set[str] = set()
    signature_cache: dict[str, _FinalReportSignature | None] = {}

    unique_paths: list[Path] = []
    path_key_by_path: dict[Path, str] = {}
    seen_signature_keys: set[str] = set()
    for path in [*existing_resolved, *accepted]:
        key = canonical_path_key(path)
        if key in seen_signature_keys:
            continue
        seen_signature_keys.add(key)
        path_key_by_path[path] = key
        unique_paths.append(path)

    max_workers = min(
        _SIGNATURE_VALIDATION_MAX_WORKERS,
        max(1, os.cpu_count() or 1),
        len(unique_paths) or 1,
    )
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_path = {
            executor.submit(_extract_final_report_signature, path): path
            for path in unique_paths
        }
        for future in as_completed(future_to_path):
            token.raise_if_cancelled()
            path = future_to_path[future]
            signature_cache[path_key_by_path[path]] = future.result()

    def _cached_signature(path: Path) -> _FinalReportSignature | None:
        """Cached signature.
        
        Args:
            path: Parameter value (Path).
        
        Returns:
            _FinalReportSignature | None: Return value.
        
        Raises:
            None.
        """
        return signature_cache.get(path_key_by_path.get(path, canonical_path_key(path)))

    for path in existing_resolved:
        token.raise_if_cancelled()
        signature = _cached_signature(path)
        if signature is None:
            continue
        if baseline_signature is None:
            baseline_signature = signature
        section_key = normalize(signature.section)
        if section_key:
            seen_sections.add(section_key)

    for path in accepted:
        token.raise_if_cancelled()
        key = canonical_path_key(path)
        if key in seen:
            duplicates += 1
            continue
        signature = _cached_signature(path)
        if signature is None:
            invalid_final_report.append(str(path))
            invalid_final_report_details.append(
                {
                    "path": str(path),
                    "reason": "Invalid final CO report workbook (signature/manifest/metadata validation failed).",
                }
            )
            continue
        if baseline_signature is None:
            baseline_signature = signature
        else:
            is_mismatch = (
                signature.template_id != baseline_signature.template_id
                or signature.course_code != baseline_signature.course_code
                or signature.total_outcomes != baseline_signature.total_outcomes
                or signature.direct_sheet_count != baseline_signature.direct_sheet_count
                or signature.indirect_sheet_count != baseline_signature.indirect_sheet_count
            )
            section_already_seen = normalize(signature.section) in seen_sections
            if is_mismatch or section_already_seen:
                invalid_final_report.append(str(path))
                if is_mismatch:
                    reason = (
                        "File does not match selected batch (template/course/CO-sheet structure mismatch)."
                    )
                else:
                    reason = f"Duplicate section '{signature.section}' is not allowed in one batch."
                invalid_final_report_details.append(
                    {
                        "path": str(path),
                        "reason": reason,
                    }
                )
                continue
        seen_sections.add(normalize(signature.section))
        seen.add(key)
        added.append(str(path))

    ignored = (len(dropped_files) - len(accepted)) + duplicates + len(invalid_final_report)
    return {
        "added": added,
        "duplicates": duplicates,
        "invalid_final_report": invalid_final_report,
        "invalid_final_report_details": invalid_final_report_details,
        "ignored": ignored,
    }


def _coerce_numeric_score(value: Any) -> float | str | None:
    """Coerce numeric score.
    
    Args:
        value: Parameter value (Any).
    
    Returns:
        float | str | None: Return value.
    
    Raises:
        None.
    """
    if normalize(value) == normalize(CO_REPORT_NOT_APPLICABLE_TOKEN):
        return CO_REPORT_ABSENT_TOKEN
    parsed = coerce_excel_number(value)
    if isinstance(parsed, bool):
        return None
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


def _metadata_rows_for_output(metadata: dict[str, str], co_index: int) -> list[tuple[str, str]]:
    """Metadata rows for output.
    
    Args:
        metadata: Parameter value (dict[str, str]).
        co_index: Parameter value (int).
    
    Returns:
        list[tuple[str, str]]: Return value.
    
    Raises:
        None.
    """
    institution_rows = [(line, "") for line in CO_ANALYSIS_INSTITUTION_ROWS]
    return [
        *institution_rows,
        ("Course Code", metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")),
        ("Course Name", metadata.get(normalize(_COURSE_METADATA_COURSE_NAME_KEY), "")),
        ("Semester", metadata.get(normalize(COURSE_METADATA_SEMESTER_KEY), "")),
        ("Academic Year", metadata.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")),
        ("CO Number", f"CO{co_index}"),
    ]


def _metadata_rows_for_summary_graph(
    metadata: dict[str, str],
    *,
    total_outcomes: int,
) -> list[tuple[str, str]]:
    """Metadata rows for summary graph.
    
    Args:
        metadata: Parameter value (dict[str, str]).
        total_outcomes: Parameter value (int).
    
    Returns:
        list[tuple[str, str]]: Return value.
    
    Raises:
        None.
    """
    total_outcomes_value = metadata.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY), "").strip()
    if not total_outcomes_value:
        total_outcomes_value = str(total_outcomes)
    institution_rows = [(line, "") for line in CO_ANALYSIS_INSTITUTION_ROWS]
    return [
        *institution_rows,
        ("Course Code", metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")),
        ("Course Name", metadata.get(normalize(_COURSE_METADATA_COURSE_NAME_KEY), "")),
        ("Academic Year", metadata.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")),
        ("Total Outcomes", total_outcomes_value),
    ]


def _threshold_rows_for_output(
    thresholds: tuple[float, float, float] | None,
    *,
    co_attainment_percent: float,
    co_attainment_level: int,
    include: bool,
) -> list[tuple[str, str]]:
    """Threshold rows for output.
    
    Args:
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float).
        co_attainment_level: Parameter value (int).
        include: Parameter value (bool).
    
    Returns:
        list[tuple[str, str]]: Return value.
    
    Raises:
        None.
    """
    if not include or thresholds is None:
        return []
    l1, l2, l3 = thresholds
    return [
        ("Threshold L1", f"{l1:g}"),
        ("Threshold L2", f"{l2:g}"),
        ("Threshold L3", f"{l3:g}"),
        ("CO AT%", f"{co_attainment_percent:g}"),
        ("CO AT Level", f"L{co_attainment_level}"),
    ]


def _summary_graph_metadata_rows(
    *,
    metadata: dict[str, str],
    total_outcomes: int,
    thresholds: tuple[float, float, float] | None,
    co_attainment_percent: float,
    co_attainment_level: int,
    include_thresholds: bool,
) -> list[tuple[str, str]]:
    """Unified metadata rows for Summary and Graph sheets.

    Args:
        metadata: Parameter value (dict[str, str]).
        total_outcomes: Parameter value (int).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float).
        co_attainment_level: Parameter value (int).
        include_thresholds: Parameter value (bool).

    Returns:
        list[tuple[str, str]]: Return value.

    Raises:
        None.
    """
    rows = _metadata_rows_for_summary_graph(metadata, total_outcomes=total_outcomes)
    rows.extend(
        _threshold_rows_for_output(
            thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
            include=include_thresholds,
        )
    )
    return rows


def _metadata_rows_for_co_analysis_outcome_sheets(
    *,
    metadata: dict[str, str],
    co_index: int,
    thresholds: tuple[float, float, float] | None,
    co_attainment_percent: float,
    co_attainment_level: int,
) -> list[tuple[str, str]]:
    """Metadata rows for co analysis outcome sheets.
    
    Args:
        metadata: Parameter value (dict[str, str]).
        co_index: Parameter value (int).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float).
        co_attainment_level: Parameter value (int).
    
    Returns:
        list[tuple[str, str]]: Return value.
    
    Raises:
        None.
    """
    metadata_rows = [
        (field, value)
        for field, value in _metadata_rows_for_output(metadata, co_index)
        if normalize(field) != normalize("CO Number")
    ]
    metadata_rows.extend(
        _threshold_rows_for_output(
            thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
            include=bool(metadata),
        )
    )
    return metadata_rows


def _extract_course_metadata_fields(sheet: Any) -> dict[str, str]:
    """Extract course metadata fields.
    
    Args:
        sheet: Parameter value (Any).
    
    Returns:
        dict[str, str]: Return value.
    
    Raises:
        None.
    """
    metadata: dict[str, str] = {}
    row = 2
    while True:
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        key_text = str(key).strip() if key is not None else ""
        if key_text:
            coerced = coerce_excel_number(value)
            metadata[normalize(key_text)] = str(coerced).strip() if coerced is not None else ""
        row += 1
    return metadata


def _students_sheet_header_row(
    *,
    sheet_specs_by_name: dict[str, dict[str, Any]],
    sheet_name: str,
) -> int:
    """Students sheet header row.
    
    Args:
        sheet_specs_by_name: Parameter value (dict[str, dict[str, Any]]).
        sheet_name: Parameter value (str).
    
    Returns:
        int: Return value.
    
    Raises:
        None.
    """
    students_spec = sheet_specs_by_name.get(sheet_name, {})
    header_row = 1
    header_row_value = students_spec.get("header_row")
    if isinstance(header_row_value, int) and header_row_value > 0:
        header_row = header_row_value
    return header_row


def _header_map_for_row(sheet: Any, *, header_row: int) -> dict[str, int]:
    """Header map for row.
    
    Args:
        sheet: Parameter value (Any).
        header_row: Parameter value (int).
    
    Returns:
        dict[str, int]: Return value.
    
    Raises:
        None.
    """
    header_map: dict[str, int] = {}
    max_col = int(sheet.max_column or 0)
    for col_index in range(1, max_col + 1):
        header_token = normalize(sheet.cell(row=header_row, column=col_index).value)
        if header_token and header_token not in header_map:
            header_map[header_token] = col_index
    return header_map


def _resolve_students_reg_col(
    *,
    header_map: dict[str, int],
    students_headers: tuple[str, ...],
) -> int | None:
    """Resolve students reg col.
    
    Args:
        header_map: Parameter value (dict[str, int]).
        students_headers: Parameter value (tuple[str, ...]).
    
    Returns:
        int | None: Return value.
    
    Raises:
        None.
    """
    return header_map.get(normalize(students_headers[0])) or header_map.get(normalize(CO_REPORT_HEADER_REG_NO))


def _resolve_students_name_col(
    *,
    header_map: dict[str, int],
    students_headers: tuple[str, ...],
) -> int | None:
    """Resolve students name col.
    
    Args:
        header_map: Parameter value (dict[str, int]).
        students_headers: Parameter value (tuple[str, ...]).
    
    Returns:
        int | None: Return value.
    
    Raises:
        None.
    """
    return header_map.get(normalize(students_headers[1])) or header_map.get(normalize(CO_REPORT_HEADER_STUDENT_NAME))


def _iter_students_sheet_rows(
    *,
    students_sheet: Any,
    header_row: int,
    reg_col: int,
    name_col: int,
) -> Iterable[tuple[str, str]]:
    """Iter students sheet rows.
    
    Args:
        students_sheet: Parameter value (Any).
        header_row: Parameter value (int).
        reg_col: Parameter value (int).
        name_col: Parameter value (int).
    
    Returns:
        Iterable[tuple[str, str]]: Return value.
    
    Raises:
        None.
    """
    max_row = int(students_sheet.max_row or 0)
    for row_index in range(header_row + 1, max_row + 1):
        reg_value = students_sheet.cell(row=row_index, column=reg_col).value
        name_value = students_sheet.cell(row=row_index, column=name_col).value
        reg_no = str(coerce_excel_number(reg_value) or "").strip()
        student_name = str(name_value or "").strip()
        if not reg_no and not student_name:
            break
        if reg_no:
            yield reg_no, student_name


def _iter_data_rows(worksheet: Any, expected_col_count: int, *, header_row: int = 1) -> list[list[Any]]:
    """Iter data rows.

    Args:
        worksheet: Parameter value (Any).
        expected_col_count: Parameter value (int).

    Returns:
        list[list[Any]]: Return value.

    Raises:
        None.
    """
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=max(1, int(header_row) + 1), max_col=expected_col_count, values_only=True):
        values = list(row)
        if any(normalize(value) != "" for value in values):
            rows.append(values)
    return rows


def _extract_component_name_from_marks_sheet(sheet: Any, *, header_row: int) -> str:
    """Extract component name from marks sheet.

    Args:
        sheet: Parameter value (Any).
        header_row: Parameter value (int).

    Returns:
        str: Return value.

    Raises:
        None.
    """
    scan_to = max(1, min(header_row, int(getattr(sheet, "max_row", 0) or 0)))
    for row_index in range(1, scan_to + 1):
        label = normalize(sheet.cell(row=row_index, column=2).value)
        if label == normalize("Component Name"):
            return str(sheet.cell(row=row_index, column=3).value or "").strip()
    return ""


def _co_index_from_label(value: Any) -> int | None:
    """Co index from label.

    Args:
        value: Parameter value (Any).

    Returns:
        int | None: Return value.

    Raises:
        None.
    """
    token = str(value or "").strip()
    match = re.fullmatch(r"CO\s*([0-9]+)", token, re.IGNORECASE)
    if match is None:
        return None
    try:
        parsed = int(match.group(1))
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _extract_direct_sheet_max_by_co(
    sheet: Any,
    *,
    header_row: int,
    kind: str,
) -> dict[int, float]:
    """Extract direct sheet max by co.

    Args:
        sheet: Parameter value (Any).
        header_row: Parameter value (int).
        kind: Parameter value (str).

    Returns:
        dict[int, float]: Return value.

    Raises:
        None.
    """
    result: dict[int, float] = {}
    co_row = header_row + 1
    max_row = header_row + 2
    max_col = int(getattr(sheet, "max_column", 0) or 0)
    if normalize(kind) == normalize(LAYOUT_SHEET_KIND_DIRECT_CO_WISE):
        for col_index in range(4, max_col + 1):
            co_index = _co_index_from_label(sheet.cell(row=co_row, column=col_index).value)
            if co_index is None:
                continue
            max_value = coerce_excel_number(sheet.cell(row=max_row, column=col_index).value)
            if isinstance(max_value, bool) or not isinstance(max_value, (int, float)):
                continue
            result[co_index] = round(result.get(co_index, 0.0) + float(max_value), CO_REPORT_MAX_DECIMAL_PLACES)
        return result
    if normalize(kind) == normalize(LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE):
        for col_index in range(5, max_col + 1):
            co_index = _co_index_from_label(sheet.cell(row=co_row, column=col_index).value)
            if co_index is None:
                continue
            max_value = coerce_excel_number(sheet.cell(row=max_row, column=col_index).value)
            if isinstance(max_value, bool) or not isinstance(max_value, (int, float)):
                continue
            result[co_index] = round(float(max_value), CO_REPORT_MAX_DECIMAL_PLACES)
    return result


def _stable_reg_hash(reg_key: str) -> int:
    # Use a compact, stable 48-bit ID to reduce sqlite footprint while keeping collision risk very low.
    """Stable reg hash.
    
    Args:
        reg_key: Parameter value (str).
    
    Returns:
        int: Return value.
    
    Raises:
        None.
    """
    return int.from_bytes(hashlib.blake2b(reg_key.encode("utf-8"), digest_size=6).digest(), byteorder="big")


def _iter_score_rows(sheet: Any, *, ratio: float) -> Iterable[_ParsedScoreRow]:
    """Iter score rows.
    
    Args:
        sheet: Parameter value (Any).
        ratio: Parameter value (float).
    
    Returns:
        Iterable[_ParsedScoreRow]: Return value.
    
    Raises:
        None.
    """
    required_headers = {
        normalize(CO_REPORT_HEADER_SERIAL),
        normalize(CO_REPORT_HEADER_REG_NO),
        normalize(CO_REPORT_HEADER_STUDENT_NAME),
        normalize(_ratio_total_header(ratio)),
    }
    header_row = 0
    column_map: dict[str, int] = {}
    max_row = int(sheet.max_row)
    max_col = int(sheet.max_column)
    header_scan_limit = min(max_row, _HEADER_SCAN_MAX_ROWS)

    def _find_header_row(start_row: int, end_row: int) -> tuple[int, dict[str, int]]:
        """Find header row.
        
        Args:
            start_row: Parameter value (int).
            end_row: Parameter value (int).
        
        Returns:
            tuple[int, dict[str, int]]: Return value.
        
        Raises:
            None.
        """
        for row_offset, values in enumerate(
            sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ),
            start=0,
        ):
            row_idx = start_row + row_offset
            row_map: dict[str, int] = {}
            missing = set(required_headers)
            for col_idx, value in enumerate(values, start=1):
                key = normalize(value)
                if key and key not in row_map:
                    row_map[key] = col_idx
                    if key in missing:
                        missing.remove(key)
                        if not missing:
                            return row_idx, row_map
        return 0, {}

    header_row, column_map = _find_header_row(1, header_scan_limit)
    if header_row <= 0 and header_scan_limit < max_row:
        header_row, column_map = _find_header_row(header_scan_limit + 1, max_row)
    if header_row <= 0:
        raise validation_error_from_key(
            "validation.layout.header_mismatch",
            code="COA_LAYOUT_HEADER_MISMATCH",
            sheet_name=str(getattr(sheet, "title", "") or ""),
        )

    reg_col = column_map[normalize(CO_REPORT_HEADER_REG_NO)]
    name_col = column_map[normalize(CO_REPORT_HEADER_STUDENT_NAME)]
    score_col = column_map[normalize(_ratio_total_header(ratio))]

    seen_in_sheet: set[int] = set()
    for values in sheet.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if len(values) < max(reg_col, name_col, score_col):
            continue
        reg_raw = values[reg_col - 1]
        reg_value = coerce_excel_number(reg_raw)
        reg_no = str(reg_value).strip() if reg_value is not None else ""
        if not reg_no:
            continue
        reg_key = normalize(reg_no)
        reg_hash = _stable_reg_hash(reg_key)
        if reg_hash in seen_in_sheet:
            continue

        score = _coerce_numeric_score(values[score_col - 1])
        if score is None:
            continue
        student_raw = values[name_col - 1]
        student_name = str(student_raw).strip() if student_raw is not None else ""
        normalized_score = round(score, CO_REPORT_MAX_DECIMAL_PLACES) if isinstance(score, (int, float)) else score
        seen_in_sheet.add(reg_hash)
        yield _ParsedScoreRow(
            reg_hash=reg_hash,
            reg_key=reg_key,
            reg_no=reg_no,
            student_name=student_name,
            score=normalized_score,
        )


def _parse_direct_component_header(
    value: Any,
) -> tuple[str, float, bool] | None:
    """Parse direct component header.

    Args:
        value: Parameter value (Any).

    Returns:
        tuple[str, float, bool] | None: Return value.

    Raises:
        None.
    """
    token = str(value or "").strip()
    match = re.fullmatch(r"(.+)\(([^()]*)\)", token)
    if match is None:
        return None
    name = str(match.group(1) or "").strip()
    suffix = str(match.group(2) or "").strip()
    if not name or not suffix:
        return None
    is_percent = suffix.endswith("%")
    numeric_token = suffix[:-1].strip() if is_percent else suffix
    parsed = coerce_excel_number(numeric_token)
    if isinstance(parsed, bool) or not isinstance(parsed, (int, float)):
        return None
    return name, float(parsed), is_percent


def _direct_component_columns(
    *,
    sheet: Any,
    header_row: int,
    max_col: int,
) -> tuple[list[_DirectComponentColumn], dict[tuple[int, str], tuple[float | str, ...]]]:
    """Direct component columns.

    Args:
        sheet: Parameter value (Any).
        header_row: Parameter value (int).
        max_col: Parameter value (int).

    Returns:
        tuple[list[_DirectComponentColumn], dict[tuple[int, str], tuple[float | str, ...]]]:
        Return value.

    Raises:
        None.
    """
    component_columns: list[_DirectComponentColumn] = []
    column = 4
    while column + 1 <= max_col:
        first = _parse_direct_component_header(sheet.cell(row=header_row, column=column).value)
        second = _parse_direct_component_header(sheet.cell(row=header_row, column=column + 1).value)
        if first is None or second is None:
            break
        first_name, max_marks, first_is_percent = first
        second_name, weight, second_is_percent = second
        if first_is_percent or not second_is_percent or normalize(first_name) != normalize(second_name):
            break
        component_columns.append(
            _DirectComponentColumn(
                name=first_name,
                max_marks=round(float(max_marks), CO_REPORT_MAX_DECIMAL_PLACES),
                weight=round(float(weight), CO_REPORT_MAX_DECIMAL_PLACES),
                score_column=column,
            )
        )
        column += 2

    marks_lookup: dict[tuple[int, str], tuple[float | str, ...]] = {}
    if not component_columns:
        return component_columns, marks_lookup
    reg_col: int | None = None
    for col_index in range(1, max_col + 1):
        token = normalize(sheet.cell(row=header_row, column=col_index).value)
        if token == normalize(CO_REPORT_HEADER_REG_NO):
            reg_col = col_index
            break
    if reg_col is None:
        return component_columns, marks_lookup

    for values in sheet.iter_rows(
        min_row=header_row + 1,
        max_row=int(sheet.max_row or 0),
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if len(values) < reg_col:
            continue
        reg_raw = values[reg_col - 1]
        reg_value = coerce_excel_number(reg_raw)
        reg_no = str(reg_value).strip() if reg_value is not None else ""
        if not reg_no:
            continue
        reg_key = normalize(reg_no)
        reg_hash = _stable_reg_hash(reg_key)
        component_scores: list[float | str] = []
        for item in component_columns:
            raw = values[item.score_column - 1] if len(values) >= item.score_column else None
            if normalize(raw) == normalize(CO_REPORT_ABSENT_TOKEN):
                component_scores.append(CO_REPORT_ABSENT_TOKEN)
                continue
            numeric = coerce_excel_number(raw)
            if isinstance(numeric, bool) or not isinstance(numeric, (int, float)):
                component_scores.append(0.0)
                continue
            component_scores.append(round(float(numeric), CO_REPORT_MAX_DECIMAL_PLACES))
        marks_lookup[(reg_hash, reg_key)] = tuple(component_scores)
    return component_columns, marks_lookup


def _indirect_component_columns(
    *,
    sheet: Any,
    header_row: int,
    max_col: int,
) -> tuple[list[_IndirectComponentColumn], dict[tuple[int, str], tuple[float | str, ...]]]:
    """Indirect component columns.

    Args:
        sheet: Parameter value (Any).
        header_row: Parameter value (int).
        max_col: Parameter value (int).

    Returns:
        tuple[list[_IndirectComponentColumn], dict[tuple[int, str], tuple[float | str, ...]]]:
        Return value.

    Raises:
        None.
    """
    component_columns: list[_IndirectComponentColumn] = []
    column = 4
    expected_raw_suffix = normalize(f"{LIKERT_MIN}-{LIKERT_MAX}")
    expected_scaled_prefix = normalize("scaled 0-")
    while column + 2 <= max_col:
        raw_header = str(sheet.cell(row=header_row, column=column).value or "").strip()
        scaled_header = str(sheet.cell(row=header_row, column=column + 1).value or "").strip()
        weighted = _parse_direct_component_header(sheet.cell(row=header_row, column=column + 2).value)
        raw_match = re.fullmatch(r"(.+)\(([^()]*)\)", raw_header)
        scaled_match = re.fullmatch(r"(.+)\(([^()]*)\)", scaled_header)
        if raw_match is None or scaled_match is None or weighted is None:
            break
        raw_name = str(raw_match.group(1) or "").strip()
        raw_suffix = normalize(str(raw_match.group(2) or "").strip())
        scaled_name = str(scaled_match.group(1) or "").strip()
        scaled_suffix = normalize(str(scaled_match.group(2) or "").strip())
        weighted_name, weight, weighted_is_percent = weighted
        if (
            not raw_name
            or raw_suffix != expected_raw_suffix
            or not scaled_name
            or not scaled_suffix.startswith(expected_scaled_prefix)
            or not weighted_is_percent
            or normalize(raw_name) != normalize(scaled_name)
            or normalize(raw_name) != normalize(weighted_name)
        ):
            break
        component_columns.append(
            _IndirectComponentColumn(
                name=raw_name,
                weight=round(float(weight), CO_REPORT_MAX_DECIMAL_PLACES),
                score_column=column,
            )
        )
        column += 3

    marks_lookup: dict[tuple[int, str], tuple[float | str, ...]] = {}
    if not component_columns:
        return component_columns, marks_lookup
    reg_col: int | None = None
    for col_index in range(1, max_col + 1):
        token = normalize(sheet.cell(row=header_row, column=col_index).value)
        if token == normalize(CO_REPORT_HEADER_REG_NO):
            reg_col = col_index
            break
    if reg_col is None:
        return component_columns, marks_lookup

    for values in sheet.iter_rows(
        min_row=header_row + 1,
        max_row=int(sheet.max_row or 0),
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if len(values) < reg_col:
            continue
        reg_raw = values[reg_col - 1]
        reg_value = coerce_excel_number(reg_raw)
        reg_no = str(reg_value).strip() if reg_value is not None else ""
        if not reg_no:
            continue
        reg_key = normalize(reg_no)
        reg_hash = _stable_reg_hash(reg_key)
        component_scores: list[float | str] = []
        for item in component_columns:
            raw = values[item.score_column - 1] if len(values) >= item.score_column else None
            if normalize(raw) == normalize(CO_REPORT_ABSENT_TOKEN):
                component_scores.append(CO_REPORT_ABSENT_TOKEN)
                continue
            numeric = coerce_excel_number(raw)
            if isinstance(numeric, bool) or not isinstance(numeric, (int, float)):
                component_scores.append(0.0)
                continue
            component_scores.append(round(float(numeric), CO_REPORT_MAX_DECIMAL_PLACES))
        marks_lookup[(reg_hash, reg_key)] = tuple(component_scores)
    return component_columns, marks_lookup


def _iter_co_rows_from_workbook(
    workbook: Any,
    *,
    co_index: int,
    workbook_name: str,
) -> tuple[list[_CoAttainmentRow], int, int, int, list[_DirectComponentColumn], list[_IndirectComponentColumn]]:
    """Collect joined CO rows and join/drop counts from one workbook.
    
    Args:
        workbook: Parameter value (Any).
        co_index: Parameter value (int).
        workbook_name: Parameter value (str).
    
    Returns:
        tuple[list[_CoAttainmentRow], int, int, int, list[_DirectComponentColumn], list[_IndirectComponentColumn]]:
        Matched rows, direct row count, indirect row count, unmatched drop count,
        and parsed direct/indirect component columns.
    
    Raises:
        None.
    """
    direct_name = co_direct_sheet_name(co_index)
    indirect_name = co_indirect_sheet_name(co_index)
    if direct_name not in workbook.sheetnames or indirect_name not in workbook.sheetnames:
        raise validation_error_from_key(
            "validation.layout.sheet_missing",
            code="COA_LAYOUT_SHEET_MISSING",
            sheet_name=f"CO{co_index}",
        )

    direct_sheet = workbook[direct_name]
    direct_lookup: dict[tuple[int, str], _ParsedScoreRow] = {}
    for item in _iter_score_rows(direct_sheet, ratio=DIRECT_RATIO):
        direct_lookup.setdefault((item.reg_hash, item.reg_key), item)
    direct_component_columns: list[_DirectComponentColumn] = []
    direct_component_lookup: dict[tuple[int, str], tuple[float | str, ...]] = {}
    if hasattr(direct_sheet, "cell") and hasattr(direct_sheet, "iter_rows"):
        direct_header_row = 0
        max_row_value = int(getattr(direct_sheet, "max_row", 0) or 0)
        for row_index in range(1, min(max_row_value, _HEADER_SCAN_MAX_ROWS) + 1):
            if normalize(direct_sheet.cell(row=row_index, column=1).value) == normalize(CO_REPORT_HEADER_SERIAL):
                direct_header_row = row_index
                break
        if direct_header_row > 0:
            direct_component_columns, direct_component_lookup = _direct_component_columns(
                sheet=direct_sheet,
                header_row=direct_header_row,
                max_col=int(getattr(direct_sheet, "max_column", 0) or 0),
            )

    indirect_sheet = workbook[indirect_name]
    indirect_lookup: dict[tuple[int, str], _ParsedScoreRow] = {}
    for item in _iter_score_rows(indirect_sheet, ratio=INDIRECT_RATIO):
        indirect_lookup.setdefault((item.reg_hash, item.reg_key), item)
    indirect_component_columns: list[_IndirectComponentColumn] = []
    indirect_component_lookup: dict[tuple[int, str], tuple[float | str, ...]] = {}
    if hasattr(indirect_sheet, "cell") and hasattr(indirect_sheet, "iter_rows"):
        indirect_header_row = 0
        max_row_value = int(getattr(indirect_sheet, "max_row", 0) or 0)
        for row_index in range(1, min(max_row_value, _HEADER_SCAN_MAX_ROWS) + 1):
            if normalize(indirect_sheet.cell(row=row_index, column=1).value) == normalize(CO_REPORT_HEADER_SERIAL):
                indirect_header_row = row_index
                break
        if indirect_header_row > 0:
            indirect_component_columns, indirect_component_lookup = _indirect_component_columns(
                sheet=indirect_sheet,
                header_row=indirect_header_row,
                max_col=int(getattr(indirect_sheet, "max_column", 0) or 0),
            )

    direct_only_count = 0
    direct_preview_rows: list[str] = []
    for key, row in direct_lookup.items():
        if key in indirect_lookup:
            continue
        direct_only_count += 1
        if len(direct_preview_rows) < 5:
            direct_preview_rows.append(row.reg_no)

    indirect_only_count = 0
    indirect_preview_rows: list[str] = []
    for key, row in indirect_lookup.items():
        if key in direct_lookup:
            continue
        indirect_only_count += 1
        if len(indirect_preview_rows) < 5:
            indirect_preview_rows.append(row.reg_no)

    dropped_count = direct_only_count + indirect_only_count
    if dropped_count:
        direct_preview = ", ".join(direct_preview_rows) if direct_preview_rows else "-"
        indirect_preview = ", ".join(indirect_preview_rows) if indirect_preview_rows else "-"
        _logger.warning(
            "CO join dropped unmatched students. workbook=%s, co_index=%s, dropped=%s, direct_only=%s, indirect_only=%s",
            workbook_name,
            co_index,
            dropped_count,
            direct_preview,
            indirect_preview,
        )

    matched_rows: list[_CoAttainmentRow] = []
    for key, direct_row in direct_lookup.items():
        match = indirect_lookup.get(key)
        if match is None:
            continue
        student_name = direct_row.student_name or match.student_name
        matched_rows.append(
            _CoAttainmentRow(
                reg_hash=direct_row.reg_hash,
                reg_no=direct_row.reg_no,
                student_name=student_name,
                direct_score=direct_row.score,
                indirect_score=match.score,
                worksheet_name=direct_name,
                workbook_name=workbook_name,
                direct_component_scores=direct_component_lookup.get(key, ()),
                indirect_component_scores=indirect_component_lookup.get(key, ()),
            )
        )
    return (
        matched_rows,
        len(direct_lookup),
        len(indirect_lookup),
        dropped_count,
        direct_component_columns,
        indirect_component_columns,
    )


def _xlsxwriter_formats(workbook: Any, *, template_id: str) -> dict[str, Any]:
    """Xlsxwriter formats.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
    
    Returns:
        dict[str, Any]: Return value.
    
    Raises:
        None.
    """
    return _build_template_xlsxwriter_formats(
        workbook,
        template_id=template_id,
        cache_attr=_STYLE_CACHE_ATTR,
        include_column_wrap=True,
        normalize_header_valign_to_center=False,
    )


def _create_co_attainment_sheet(
    workbook: Any,
    *,
    template_id: str,
    co_index: int,
    metadata: dict[str, str],
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float,
    co_attainment_level: int,
) -> _CoOutputSheetState:
    """Create co attainment sheet.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
        co_index: Parameter value (int).
        metadata: Parameter value (dict[str, str]).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float).
        co_attainment_level: Parameter value (int).
    
    Returns:
        _CoOutputSheetState: Return value.
    
    Raises:
        None.
    """
    sheet = workbook.add_worksheet(f"CO{co_index}")
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_headers = _course_metadata_headers(template_id)
    metadata_rows = _metadata_rows_for_output(metadata, co_index)
    metadata_rows.extend(
        _threshold_rows_for_output(
            thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
            include=bool(metadata),
        )
    )
    write_two_column_metadata_rows(
        sheet,
        rows=metadata_rows,
        label_col_index=1,
        value_col_index=2,
        label_format=formats["body"],
        value_format=formats["body_wrap"],
        centered_row_labels=CO_ANALYSIS_INSTITUTION_ROWS,
        centered_label_format=formats["body_center"],
        centered_value_format=formats["body_center"],
    )

    header_row_index = len(metadata_rows) + 1
    headers = [
        "#",
        "Student Name",
        "Reg. No.",
        f"Direct ({ratio_percent_token(DIRECT_RATIO)}%)",
        f"Indirect ({ratio_percent_token(INDIRECT_RATIO)}%)",
        "Total (100%)",
        "Level",
    ]
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, col_idx, header, formats["header"])

    sampled_rows: list[list[Any]] = [["", metadata_headers[0], metadata_headers[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    sampled_rows.append(["", headers[1], headers[2]])
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    sheet.set_column(1, 1, widths.get(1, 8))
    sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    _apply_xlsxwriter_layout_shared(
        sheet,
        header_row_index=header_row_index,
        paper_size=9,
        landscape=True,
    )
    sheet.repeat_rows(0, header_row_index)
    # Freeze rows through the header and columns through Reg. No. (A:C).
    sheet.freeze_panes(header_row_index + 1, 3)
    return _CoOutputSheetState(
        sheet=sheet,
        header_row_index=header_row_index,
        formats=formats,
        next_row_index=header_row_index + 1,
        next_serial=1,
        on_roll=0,
        attended=0,
        level_counts={level: 0 for level in range(0, (len(thresholds) if thresholds is not None else 3) + 1)},
    )


def _append_co_attainment_row(
    state: _CoOutputSheetState,
    row: _CoAttainmentRow,
    *,
    thresholds: tuple[float, float, float],
) -> None:
    """Append co attainment row.
    
    Args:
        state: Parameter value (_CoOutputSheetState).
        row: Parameter value (_CoAttainmentRow).
        thresholds: Parameter value (tuple[float, float, float]).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    if isinstance(row.direct_score, (int, float)) and isinstance(row.indirect_score, (int, float)):
        total: float | str = round(row.direct_score + row.indirect_score, CO_REPORT_MAX_DECIMAL_PLACES)
    else:
        total = CO_REPORT_ABSENT_TOKEN
    level = _score_to_attainment_level(total, thresholds=thresholds)
    state.sheet.write(state.next_row_index, 0, state.next_serial, state.formats["body"])
    state.sheet.write(state.next_row_index, 1, row.student_name, state.formats["body_wrap"])
    right = [row.direct_score, row.indirect_score, total, level]
    state.sheet.write(state.next_row_index, 2, row.reg_no, state.formats["body"])
    state.sheet.write_row(state.next_row_index, 3, right, state.formats["body_center"])
    state.on_roll += 1
    if total != CO_REPORT_ABSENT_TOKEN:
        state.attended += 1
    if isinstance(level, int) and level in state.level_counts:
        state.level_counts[level] += 1
    state.next_row_index += 1
    state.next_serial += 1


def _append_co_attainment_summary(state: _CoOutputSheetState) -> None:
    # Keep one visual spacer row after data and then write label/value summary rows.
    """Append co attainment summary.
    
    Args:
        state: Parameter value (_CoOutputSheetState).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    state.next_row_index += 1
    summary_rows: list[tuple[str, int]] = [
        ("On Roll:", state.on_roll),
        ("Attended:", state.attended),
    ]
    summary_rows.extend(
        (f"Level {level}:", state.level_counts.get(level, 0))
        for level in sorted(state.level_counts)
    )
    for label, value in summary_rows:
        state.sheet.write(state.next_row_index, 1, label, state.formats["body"])
        state.sheet.write(state.next_row_index, 2, value, state.formats["body_center"])
        state.next_row_index += 1
    write_sheet_footer_xlsxwriter(
        state.sheet,
        footer_text=CO_ANALYSIS_SHEET_FOOTER_TEXT,
        row_index=state.next_row_index + 1,
        col_index=0,
        cell_format=state.formats["body"],
    )


def _direct_total_100_from_direct_score(score: float | str) -> float | None:
    """Direct total 100 from direct score.

    Args:
        score: Parameter value (float | str).

    Returns:
        float | None: Return value.

    Raises:
        None.
    """
    if isinstance(score, str):
        token = normalize(score)
        if token in {
            normalize(CO_REPORT_ABSENT_TOKEN),
            normalize(CO_REPORT_NOT_APPLICABLE_TOKEN),
        }:
            return 0.0
        return None
    if isinstance(score, bool) or not isinstance(score, (int, float)):
        return None
    if DIRECT_RATIO <= 0:
        return 0.0
    total_100 = round(float(score) / float(DIRECT_RATIO), CO_REPORT_MAX_DECIMAL_PLACES)
    return max(0.0, min(100.0, total_100))


def _co_direct_total_100_for_row(
    *,
    co_index: int,
    row: _CoAttainmentRow,
    direct_columns_by_co: dict[int, list[_DirectComponentColumn]],
    direct_scores_by_co: dict[int, dict[int, dict[str, float | str]]],
) -> float | None:
    """Compute direct total (100%) for one row in one CO.

    Args:
        co_index: Parameter value (int).
        row: Parameter value (_CoAttainmentRow).
        direct_columns_by_co: Parameter value (dict[int, list[_DirectComponentColumn]]).
        direct_scores_by_co: Parameter value (dict[int, dict[int, dict[str, float | str]]]).

    Returns:
        float | None: Return value.

    Raises:
        None.
    """
    columns = list(direct_columns_by_co.get(co_index, []))
    if not columns:
        return _direct_total_100_from_direct_score(row.direct_score)

    score_map = direct_scores_by_co.get(co_index, {}).get(row.reg_hash, {})
    weighted_total = 0.0
    total_weight = 0.0
    for column in columns:
        max_marks = float(column.max_marks)
        weight = float(column.weight)
        if max_marks <= 0 or weight <= 0:
            continue
        raw = score_map.get(normalize(column.name), CO_REPORT_NOT_APPLICABLE_TOKEN)
        if normalize(raw) == normalize(CO_REPORT_NOT_APPLICABLE_TOKEN):
            continue
        total_weight += weight
        if isinstance(raw, str) and normalize(raw) == normalize(CO_REPORT_ABSENT_TOKEN):
            raw_numeric = 0.0
        elif isinstance(raw, (int, float)) and not isinstance(raw, bool):
            raw_numeric = float(raw)
        else:
            raw_numeric = 0.0
        bounded_raw = max(0.0, min(raw_numeric, max_marks))
        weighted_total += (bounded_raw * weight / max_marks)
    if total_weight <= 0:
        return None
    total_100 = round((weighted_total * 100.0 / total_weight), CO_REPORT_MAX_DECIMAL_PLACES)
    return max(0.0, min(100.0, total_100))


def _create_pass_percentage_sheet(
    workbook: Any,
    *,
    template_id: str,
    metadata: dict[str, str],
    thresholds: tuple[float, float, float],
    pending_rows: dict[int, list[_CoAttainmentRow]],
    pending_direct_columns: dict[int, list[_DirectComponentColumn]],
    pending_direct_scores: dict[int, dict[int, dict[str, float | str]]],
    total_outcomes: int,
) -> None:
    """Create pass percentage sheet.

    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
        metadata: Parameter value (dict[str, str]).
        thresholds: Parameter value (tuple[float, float, float]).
        pending_rows: Parameter value (dict[int, list[_CoAttainmentRow]]).
        pending_direct_columns: Parameter value (dict[int, list[_DirectComponentColumn]]).
        pending_direct_scores: Parameter value (dict[int, dict[int, dict[str, float | str]]]).
        total_outcomes: Parameter value (int).

    Returns:
        None.

    Raises:
        None.
    """
    sheet = workbook.add_worksheet(_PASS_PERCENTAGE_SHEET_NAME)
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_headers = _course_metadata_headers(template_id)
    metadata_rows = _metadata_rows_for_summary_graph(metadata, total_outcomes=total_outcomes)
    l1_threshold = float(thresholds[0]) if thresholds else 0.0
    metadata_rows.append(("Pass Threshold (L1)", f"{l1_threshold:g}"))
    write_two_column_metadata_rows(
        sheet,
        rows=metadata_rows,
        label_col_index=1,
        value_col_index=2,
        label_format=formats["body"],
        value_format=formats["body_wrap"],
        centered_row_labels=CO_ANALYSIS_INSTITUTION_ROWS,
        centered_label_format=formats["body_center"],
        centered_value_format=formats["body_center"],
    )

    header_row_index = len(metadata_rows) + 1
    headers = ["#", "Student Name", "Reg. No.", "Total (100%)", "Result"]
    table_start_col = 0
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, table_start_col + col_idx, header, formats["header"])

    student_rows: dict[int, dict[str, Any]] = {}
    for co_index in range(1, total_outcomes + 1):
        for row in pending_rows.get(co_index, []):
            item = student_rows.setdefault(
                row.reg_hash,
                {"reg_no": row.reg_no, "student_name": row.student_name, "direct_totals_100": []},
            )
            direct_total_100 = _co_direct_total_100_for_row(
                co_index=co_index,
                row=row,
                direct_columns_by_co=pending_direct_columns,
                direct_scores_by_co=pending_direct_scores,
            )
            if direct_total_100 is not None:
                item["direct_totals_100"].append(direct_total_100)

    ordered_students = sorted(
        student_rows.values(),
        key=lambda item: (_reg_no_sort_key(str(item.get("reg_no", ""))), str(item.get("reg_no", "")).casefold()),
    )

    pass_count = 0
    fail_count = 0
    evaluated_count = 0
    for idx, item in enumerate(ordered_students, start=1):
        row_index = header_row_index + idx
        totals = list(item.get("direct_totals_100", []))
        if totals:
            total_100: float | str = round(sum(totals) / float(len(totals)), CO_REPORT_MAX_DECIMAL_PLACES)
            result = "Pass" if float(total_100) >= l1_threshold else "Fail"
            evaluated_count += 1
            if result == "Pass":
                pass_count += 1
            else:
                fail_count += 1
        else:
            total_100 = CO_REPORT_NOT_APPLICABLE_TOKEN
            result = CO_REPORT_NOT_APPLICABLE_TOKEN
        row_values: list[Any] = [
            idx,
            str(item.get("student_name", "")),
            str(item.get("reg_no", "")),
            total_100,
            result,
        ]
        for col_offset, value in enumerate(row_values, start=0):
            col = table_start_col + col_offset
            if col_offset == 1:
                fmt = formats["body_wrap"]
            else:
                fmt = formats["body_center"]
            sheet.write(row_index, col, value, fmt)

    summary_row_index = header_row_index + len(ordered_students) + 2
    pass_percentage = (
        round((pass_count / float(evaluated_count)) * 100.0, CO_REPORT_MAX_DECIMAL_PLACES)
        if evaluated_count > 0
        else CO_REPORT_NOT_APPLICABLE_TOKEN
    )
    summary_rows: list[tuple[str, Any]] = [
        ("On Roll", len(ordered_students)),
        ("Evaluated", evaluated_count),
        ("Pass", pass_count),
        ("Fail", fail_count),
        ("Pass Percentage (%)", pass_percentage),
    ]
    for label, value in summary_rows:
        sheet.write(summary_row_index, 1, label, formats["body"])
        sheet.write(summary_row_index, 2, value, formats["body_center"])
        summary_row_index += 1
    write_sheet_footer_xlsxwriter(
        sheet,
        footer_text=CO_ANALYSIS_SHEET_FOOTER_TEXT,
        row_index=summary_row_index + 1,
        col_index=0,
        cell_format=formats["body"],
    )

    sampled_rows: list[list[Any]] = [["", metadata_headers[0], metadata_headers[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    sampled_rows.append(["", headers[0], headers[1], headers[2], headers[3], headers[4]])
    sampled_rows.extend(
        ["", item["reg_no"], item["student_name"], "100", "Pass"]
        for item in ordered_students[: min(len(ordered_students), 20)]
    )
    sampled_rows.extend([["", label, value] for label, value in summary_rows])
    widths = _compute_sampled_column_widths(sampled_rows, 4)
    for col in range(0, 5):
        if col in {2}:
            sheet.set_column(col, col, widths.get(col, 8), formats["column_wrap"])
        else:
            sheet.set_column(col, col, widths.get(col, 8))

    _apply_xlsxwriter_layout_shared(
        sheet,
        header_row_index=header_row_index,
        paper_size=9,
        landscape=True,
        selection_col=table_start_col,
    )
    sheet.repeat_rows(0, header_row_index)
    # Freeze through Reg. No. with table starting at column A.
    sheet.freeze_panes(header_row_index + 1, 3)


def _co_percentage(
    *,
    level_counts: dict[int, int],
    attended: int,
    co_attainment_level: int,
) -> float | str:
    """Co percentage.
    
    Args:
        level_counts: Parameter value (dict[int, int]).
        attended: Parameter value (int).
        co_attainment_level: Parameter value (int).
    
    Returns:
        float | str: Return value.
    
    Raises:
        None.
    """
    if attended <= 0:
        return CO_REPORT_NOT_APPLICABLE_TOKEN
    attained_count = sum(
        int(count)
        for level, count in level_counts.items()
        if isinstance(level, int) and level >= co_attainment_level
    )
    return round((attained_count / float(attended)) * 100.0, CO_REPORT_MAX_DECIMAL_PLACES)


def _indirect_total_100_from_indirect_score(score: float | str) -> float | None:
    """Indirect total 100 from indirect score.

    Args:
        score: Parameter value (float | str).

    Returns:
        float | None: Return value.

    Raises:
        None.
    """
    if isinstance(score, str):
        token = normalize(score)
        if token in {
            normalize(CO_REPORT_ABSENT_TOKEN),
            normalize(CO_REPORT_NOT_APPLICABLE_TOKEN),
        }:
            return 0.0
        return None
    if isinstance(score, bool) or not isinstance(score, (int, float)):
        return None
    if INDIRECT_RATIO <= 0:
        return 0.0
    total_100 = round(float(score) / float(INDIRECT_RATIO), CO_REPORT_MAX_DECIMAL_PLACES)
    return max(0.0, min(100.0, total_100))


def _format_report_percent(value: float | str) -> str:
    """Format report percent.

    Args:
        value: Parameter value (float | str).

    Returns:
        str: Return value.

    Raises:
        None.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{float(value):g}%"
    return str(value)


def _shortfall_severity(
    *,
    overall_percent: float | None,
    shortfall_percent: float,
) -> str:
    """Shortfall severity classification based on CO AT student-percentage gap."""
    if overall_percent is None:
        return "Data Gap (Insufficient Evidence)"
    if shortfall_percent <= 0:
        return "On Target (Meets CO AT Target)"
    if shortfall_percent <= 5.0:
        return "Low (Shortfall up to 5%)"
    if shortfall_percent <= 15.0:
        return "Moderate (Shortfall 5%-15%)"
    return "High (Shortfall above 15%)"


def _recommended_corrective_action(
    *,
    direct_value: float | None,
    indirect_value: float | None,
    severity: str,
) -> str:
    """Recommended corrective action line for one CO."""
    if severity.startswith("On Target"):
        return (
            "Sustain current attainment with continuous formative assessment and periodic CO-mapping review as per OBE cycle."
        )
    if severity.startswith("Data Gap"):
        return (
            "Complete missing direct/indirect evidence and validate data integrity before CO attainment review."
        )
    if severity.startswith("High"):
        return (
            "Initiate immediate remedial instruction, bridge assignments, and compulsory reassessment for units mapped to this CO."
        )
    if severity.startswith("Moderate"):
        return (
            "Plan focused tutorial/problem-solving sessions and strengthen question-to-CO alignment in direct assessments."
        )
    if severity.startswith("Low"):
        return (
            "Run targeted improvement activities and additional formative checks to close the remaining CO AT shortfall."
        )
    if direct_value is None and indirect_value is None:
        return (
            "Review evidence completeness and ensure both direct and indirect assessments are captured."
        )
    if direct_value is None:
        return "Strengthen direct assessment design and improve question-to-CO alignment for this CO."
    if indirect_value is None:
        return "Improve indirect feedback coverage and instrument quality for this CO."
    if direct_value + 5.0 < indirect_value:
        return "Direct attainment is lagging. Recalibrate teaching-learning activities and map assessments more tightly to this CO."
    if indirect_value + 5.0 < direct_value:
        return "Indirect attainment is lagging. Improve survey administration, response quality, and student awareness for this CO."
    return "Apply micro-interventions and targeted reassessment to close the CO AT shortfall."


def _build_co_word_summary_rows(
    *,
    pending_rows: dict[int, list[_CoAttainmentRow]],
    output_states: dict[int, _CoOutputSheetState],
    total_outcomes: int,
    course_code: str,
    co_attainment_level: int,
    co_attainment_percent: float,
) -> list[dict[str, str]]:
    """Build CO summary rows for Word report."""
    rows: list[dict[str, str]] = []
    for co_index in range(1, total_outcomes + 1):
        co_rows = list(pending_rows.get(co_index, []))
        direct_values: list[float] = []
        indirect_values: list[float] = []
        for row in co_rows:
            direct_value = _direct_total_100_from_direct_score(row.direct_score)
            if direct_value is not None:
                direct_values.append(float(direct_value))
            indirect_value = _indirect_total_100_from_indirect_score(row.indirect_score)
            if indirect_value is not None:
                indirect_values.append(float(indirect_value))
        direct_avg: float | str = (
            round(sum(direct_values) / len(direct_values), CO_REPORT_MAX_DECIMAL_PLACES)
            if direct_values
            else CO_REPORT_NOT_APPLICABLE_TOKEN
        )
        indirect_avg: float | str = (
            round(sum(indirect_values) / len(indirect_values), CO_REPORT_MAX_DECIMAL_PLACES)
            if indirect_values
            else CO_REPORT_NOT_APPLICABLE_TOKEN
        )
        state = output_states.get(co_index)
        level_counts = state.level_counts if state is not None else {}
        attended = state.attended if state is not None else 0
        overall_attainment = _co_percentage(
            level_counts=level_counts,
            attended=attended,
            co_attainment_level=co_attainment_level,
        )
        result_text = (
            "Attained"
            if isinstance(overall_attainment, (int, float))
            and float(overall_attainment) >= float(co_attainment_percent)
            else "Yet to Attain"
        )
        direct_value = float(direct_avg) if isinstance(direct_avg, (int, float)) and not isinstance(direct_avg, bool) else None
        indirect_value = (
            float(indirect_avg)
            if isinstance(indirect_avg, (int, float)) and not isinstance(indirect_avg, bool)
            else None
        )
        overall_value = (
            float(overall_attainment)
            if isinstance(overall_attainment, (int, float)) and not isinstance(overall_attainment, bool)
            else None
        )
        shortfall_percent = (
            max(0.0, float(co_attainment_percent) - overall_value)
            if overall_value is not None
            else float(co_attainment_percent)
        )
        severity = _shortfall_severity(
            overall_percent=overall_value,
            shortfall_percent=shortfall_percent,
        )
        action = _recommended_corrective_action(
            direct_value=direct_value,
            indirect_value=indirect_value,
            severity=severity,
        )
        co_label = f"{course_code}.{co_index}" if course_code else f"CO{co_index}"
        rows.append(
            {
                "co": co_label,
                "direct": _format_report_percent(direct_avg),
                "indirect": _format_report_percent(indirect_avg),
                "overall": _format_report_percent(overall_attainment),
                "result": result_text,
                "shortfall": _format_report_percent(round(shortfall_percent, CO_REPORT_MAX_DECIMAL_PLACES)),
                "severity": severity,
                "recommended_action": action,
            }
        )
    return rows


def _metadata_report_value(metadata: dict[str, str], key: str) -> str:
    """Read normalized metadata value for reporting."""
    return str(metadata.get(normalize(key), "")).strip()


def _raise_first_validation_issue_from_result(*, result: dict[str, object], workbook_path: str) -> None:
    """Raise first validation issue for workbook from batch validator payload."""
    workbook_key = canonical_path_key(workbook_path)
    valid_paths_raw = result.get("valid_paths", [])
    valid_paths = [str(path) for path in valid_paths_raw] if isinstance(valid_paths_raw, list) else []
    valid_keys = {canonical_path_key(path) for path in valid_paths}
    if workbook_key in valid_keys:
        return
    rejections_raw = result.get("rejections", [])
    rejection_items = [item for item in rejections_raw if isinstance(item, dict)] if isinstance(rejections_raw, list) else []
    for item in rejection_items:
        if canonical_path_key(str(item.get("path", "")).strip()) != workbook_key:
            continue
        issue = item.get("issue", {})
        if not isinstance(issue, dict):
            continue
        code = str(issue.get("code", "VALIDATION_ERROR")).strip() or "VALIDATION_ERROR"
        message = str(issue.get("message", code)).strip() or code
        context = issue.get("context", {})
        context_dict = dict(context) if isinstance(context, dict) else {}
        raise ValidationError(message, code=code, context=context_dict)
    raise validation_error_from_key(
        "validation.workbook.open_failed",
        code="WORKBOOK_OPEN_FAILED",
        workbook=workbook_path,
    )


def _validated_co_description_sentences(
    *,
    co_description_path: Path,
    template_id: str,
    total_outcomes: int,
    token: CancellationToken | None = None,
) -> list[str]:
    """Validate CO-description workbook and return CO description sentences ordered by CO#."""
    if token is not None:
        token.raise_if_cancelled()
    result = validate_co_description_workbooks(
        workbook_paths=[str(co_description_path)],
        template_id=template_id,
        cancel_token=token,
    )
    _raise_first_validation_issue_from_result(result=result, workbook_path=str(co_description_path))
    metadata_map = extract_course_metadata_and_students_from_workbook_path(co_description_path)[1]
    total_token = metadata_map.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY), "")
    parsed_total = coerce_excel_number(total_token)
    if isinstance(parsed_total, bool) or not isinstance(parsed_total, (int, float)):
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COA_TOTAL_OUTCOMES_MISSING",
        )
    if int(parsed_total) != int(total_outcomes):
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="CO_DESCRIPTION_MARKS_COHORT_MISMATCH",
            fields=COURSE_METADATA_TOTAL_OUTCOMES_KEY,
        )

    openpyxl = import_runtime_dependency("openpyxl")
    workbook = openpyxl.load_workbook(co_description_path, data_only=True, read_only=True)
    try:
        validated_template_id = read_valid_template_id_from_system_hash_sheet(workbook)
        sheet_name = get_sheet_name_by_key(validated_template_id, COURSE_SETUP_SHEET_KEY_CO_DESCRIPTION)
        if sheet_name not in workbook.sheetnames:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="SHEET_DATA_REQUIRED",
                sheet_name=sheet_name,
            )
        sheet = workbook[sheet_name]
        headers = {
            normalize(sheet.cell(row=1, column=col).value): col
            for col in range(1, int(sheet.max_column or 0) + 1)
        }
        co_col = int(headers.get("co#", 0) or 0)
        description_col = int(headers.get("description", 0) or 0)
        if co_col <= 0 or description_col <= 0:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="SCHEMA_COLUMN_KEY_MISSING",
                sheet_name=sheet_name,
            )
        by_co: dict[int, str] = {}
        for row in range(2, int(sheet.max_row or 0) + 1):
            raw_co = sheet.cell(row=row, column=co_col).value
            raw_description = sheet.cell(row=row, column=description_col).value
            if normalize(raw_co) == "" and normalize(raw_description) == "":
                continue
            parsed_co = coerce_excel_number(raw_co)
            if isinstance(parsed_co, bool) or not isinstance(parsed_co, (int, float)):
                continue
            co_value = int(parsed_co)
            if co_value <= 0:
                continue
            description = str(raw_description or "").strip()
            if not description:
                continue
            by_co[co_value] = description
    finally:
        workbook.close()
    expected = set(range(1, int(total_outcomes) + 1))
    found = set(by_co)
    if found != expected:
        missing = sorted(expected - found)
        extras = sorted(found - expected)
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="CO_DESCRIPTION_CO_NUMBER_SET_MISMATCH",
            expected=f"1..{int(total_outcomes)}",
            missing=", ".join(str(value) for value in missing) if missing else "",
            extras=", ".join(str(value) for value in extras) if extras else "",
        )
    return [by_co[index] for index in range(1, int(total_outcomes) + 1)]


def _generate_co_attainment_word_report(
    *,
    output_path: Path,
    metadata: dict[str, str],
    thresholds: tuple[float, float, float],
    co_attainment_percent: float,
    co_attainment_level: int,
    total_outcomes: int,
    co_rows: list[dict[str, str]],
    co_sentences: list[str] | None = None,
) -> Path:
    """Generate CO attainment Word report document."""
    docx_mod = import_runtime_dependency("docx")
    Document = docx_mod.Document
    Pt = docx_mod.shared.Pt
    Inches = docx_mod.shared.Inches
    WD_PARAGRAPH_ALIGNMENT = docx_mod.enum.text.WD_PARAGRAPH_ALIGNMENT

    document = Document()
    qn = docx_mod.oxml.ns.qn
    times_new_roman = "Times New Roman"

    def _set_run_font(run: Any) -> None:
        run.font.name = times_new_roman
        run.font.size = Pt(12)
        r_pr = run._element.get_or_add_rPr()
        r_fonts = r_pr.get_or_add_rFonts()
        r_fonts.set(qn("w:ascii"), times_new_roman)
        r_fonts.set(qn("w:hAnsi"), times_new_roman)
        r_fonts.set(qn("w:eastAsia"), times_new_roman)
        r_fonts.set(qn("w:cs"), times_new_roman)

    def _style_paragraph(paragraph: Any, *, bold: bool = False, justify: bool = False) -> None:
        if justify:
            paragraph.alignment = WD_PARAGRAPH_ALIGNMENT.JUSTIFY
        if not paragraph.runs:
            _set_run_font(paragraph.add_run(""))
        for run in paragraph.runs:
            _set_run_font(run)
            if bold:
                run.bold = True

    def _add_section_heading(text: str, *, centered: bool = False) -> Any:
        paragraph = document.add_paragraph(text)
        if centered:
            paragraph.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
        _style_paragraph(paragraph, bold=True)
        return paragraph

    for institution_line in CO_ANALYSIS_INSTITUTION_ROWS:
        para = document.add_paragraph(str(institution_line).upper())
        para.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
        para.paragraph_format.space_before = Pt(0)
        para.paragraph_format.space_after = Pt(0)
        _style_paragraph(para, bold=True)

    _add_section_heading("Course Coordinator Report", centered=True)
    generated_on = document.add_paragraph(f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    generated_on.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
    _style_paragraph(generated_on)

    enriched_rows: list[dict[str, str]] = []
    for item in co_rows:
        result_text = str(item.get("result", "")).strip() or "Yet to Attain"
        shortfall_text = str(item.get("shortfall", "")).strip()
        if not shortfall_text:
            shortfall_text = "0%" if result_text == "Attained" else f"{co_attainment_percent:g}%"
        severity_text = str(item.get("severity", "")).strip()
        if not severity_text:
            severity_text = "None" if result_text == "Attained" else "Moderate"
        action_text = str(item.get("recommended_action", "")).strip()
        if not action_text:
            action_text = (
                "Sustain current delivery and monitor consistency in subsequent assessments."
                if result_text == "Attained"
                else "Use targeted remedial sessions, focused practice, and reassessment on key weak units mapped to this CO."
            )
        enriched_rows.append(
            {
                "co": str(item.get("co", "")).strip(),
                "direct": str(item.get("direct", "")).strip(),
                "indirect": str(item.get("indirect", "")).strip(),
                "overall": str(item.get("overall", "")).strip(),
                "result": result_text,
                "shortfall": shortfall_text,
                "severity": severity_text,
                "recommended_action": action_text,
            }
        )

    _add_section_heading("Course Details")
    course_code = _metadata_report_value(metadata, COURSE_METADATA_COURSE_CODE_KEY) or "N/A"
    semester = _metadata_report_value(metadata, COURSE_METADATA_SEMESTER_KEY) or "N/A"
    academic_year = _metadata_report_value(metadata, COURSE_METADATA_ACADEMIC_YEAR_KEY) or "N/A"
    course_details_templates = (
        (
            "This Course Coordinator report presents the CO attainment analysis for course {course_code}, "
            "offered in Semester {semester}, during the academic year {academic_year}. "
            "The course is mapped to {total_outcomes} course outcomes (COs), and the sections below summarize "
            "attainment status, identified shortfalls, severity classification, and recommended corrective actions."
        ),
        (
            "For the academic year {academic_year}, this report reviews course {course_code} in Semester {semester} "
            "through the lens of CO attainment. A total of {total_outcomes} COs are covered, and this document "
            "captures attainment outcomes, shortfall severity, and suggested corrective interventions."
        ),
        (
            "Course {course_code} (Semester {semester}, {academic_year}) is evaluated here using the CO attainment framework. "
            "With {total_outcomes} mapped COs, the report provides a concise view of attainment performance, "
            "areas below target, severity levels, and follow-up action points."
        ),
        (
            "This document summarizes the Course Coordinator analysis for {course_code} in Semester {semester} "
            "for the academic cycle {academic_year}. Across {total_outcomes} course outcomes, the report highlights "
            "attainment trends, shortfall conditions, severity categorization, and practical corrective actions."
        ),
        (
            "As part of continuous OBE monitoring, this report analyzes CO attainment for course {course_code}, "
            "offered in Semester {semester} during {academic_year}. The evaluation spans {total_outcomes} COs and "
            "presents attainment status, the extent of shortfall, severity classification, and recommended improvements."
        ),
        (
            "Presented here is the CO attainment review for course {course_code} in Semester {semester} for {academic_year}. "
            "The course includes {total_outcomes} COs, and the following sections outline attainment position, "
            "shortfall severity, and corrective actions to strengthen outcome achievement."
        ),
    )
    course_details_paragraph = document.add_paragraph(
        random.choice(course_details_templates).format(
            course_code=course_code,
            semester=semester,
            academic_year=academic_year,
            total_outcomes=total_outcomes,
        )
    )
    _style_paragraph(course_details_paragraph, justify=True)

    if co_sentences:
        _add_section_heading("Course Outcomes")
        intro_paragraph = document.add_paragraph("The students will be able to:")
        _style_paragraph(intro_paragraph, justify=True)
        intro_paragraph.paragraph_format.space_before = Pt(0)
        intro_paragraph.paragraph_format.space_after = Pt(0)
        intro_paragraph.paragraph_format.line_spacing = 1.0
        last_sentence_paragraph: Any | None = None
        for index, sentence in enumerate(co_sentences, start=1):
            normalized_sentence = str(sentence or "").strip()
            if not normalized_sentence:
                continue
            if normalized_sentence.endswith("."):
                line_text = f"CO{index}: {normalized_sentence}"
            else:
                line_text = f"CO{index}: {normalized_sentence}."
            sentence_line_paragraph = document.add_paragraph(f"\t{line_text}")
            _style_paragraph(sentence_line_paragraph, justify=True)
            sentence_line_paragraph.paragraph_format.space_before = Pt(0)
            sentence_line_paragraph.paragraph_format.space_after = Pt(0)
            sentence_line_paragraph.paragraph_format.line_spacing = 1.0
            last_sentence_paragraph = sentence_line_paragraph
        if last_sentence_paragraph is not None:
            # Restore normal visual separation after the final CO item.
            last_sentence_paragraph.paragraph_format.space_after = Pt(12)

    _add_section_heading("Threshold Details")
    l1, l2, l3 = thresholds
    l2_policy_text = (
        "Level L2 is fixed at 60% as configured for this run."
        if abs(float(l2) - 60.0) < 1e-9
        else (
            f"Level L2 is set to {l2:g}% for this run, and this value is treated as "
            "the average of the last three offerings in the previous regulations."
        )
    )
    threshold_templates = (
        (
            "The threshold policy applied in this analysis is as follows: Level L1 is set to "
            "{l1:g}% for this run. {l2_policy} Level L3 is set to 75%. "
            "For attainment judgement, the configured CO attainment target is {co_at:g}% at or above Level L{co_level}."
        ),
        (
            "For this course run, Level L1 corresponds to the configured value of {l1:g}%. {l2_policy} "
            "Level L3 is considered as 75%. The CO attainment decision is evaluated against "
            "{co_at:g}% of students at or above Level L{co_level}."
        ),
        (
            "Threshold interpretation used in this report: Level L1 equals {l1:g}% for this run. "
            "{l2_policy} Level L3 represents 75%. CO attainment is finalized using the target "
            "{co_at:g}% at or above Level L{co_level}."
        ),
        (
            "As per the configured OBE threshold settings, Level L1 is {l1:g}%. "
            "{l2_policy} Level L3 is maintained at 75%. In addition, the run uses a CO attainment target of "
            "{co_at:g}% at or above Level L{co_level}."
        ),
        (
            "The present analysis applies these thresholds: Level L1 is provided as {l1:g}% for this run. "
            "{l2_policy} Level L3 is 75%. CO-wise attainment status is then computed using "
            "{co_at:g}% at or above Level L{co_level}."
        ),
        (
            "For threshold-based classification in this report, Level L1 is taken as "
            "{l1:g}%. {l2_policy} Level L3 is fixed at 75%. CO attainment is assessed against "
            "the configured target of {co_at:g}% at or above Level L{co_level}."
        ),
    )
    threshold_paragraph = document.add_paragraph(
        random.choice(threshold_templates).format(
            l1=l1,
            l2_policy=l2_policy_text,
            co_at=co_attainment_percent,
            co_level=co_attainment_level,
        )
    )
    _style_paragraph(threshold_paragraph, justify=True)

    formula_image_path = Path("assets") / "formula.jpg"
    if formula_image_path.is_file():
        image_paragraph = document.add_paragraph()
        image_paragraph.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
        image_paragraph.add_run().add_picture(str(formula_image_path), width=Inches(5.8))
    else:
        equation_fallback = document.add_paragraph("Formula image not available.")
        equation_fallback.alignment = WD_PARAGRAPH_ALIGNMENT.CENTER
        _style_paragraph(equation_fallback)

    attained_count = sum(1 for item in enriched_rows if item["result"] == "Attained")
    _add_section_heading("CO-wise Attainment Summary")
    summary_para = document.add_paragraph(
        "This report summarizes CO-wise attainment using configured thresholds and target attainment policy. "
        f"Out of {total_outcomes} COs, {attained_count} attained and {max(0, total_outcomes - attained_count)} are below target."
    )
    _style_paragraph(summary_para, justify=True)

    co_table = document.add_table(rows=1, cols=4)
    co_table.style = "Table Grid"
    co_table.rows[0].cells[0].text = "CO"
    co_table.rows[0].cells[1].text = "Overall"
    co_table.rows[0].cells[2].text = "Shortfall"
    co_table.rows[0].cells[3].text = "Result"
    for item in enriched_rows:
        cells = co_table.add_row().cells
        cells[0].text = item["co"]
        cells[1].text = item["overall"]
        cells[2].text = item["shortfall"]
        cells[3].text = item["result"]

    _add_section_heading("Continuous Improvement Action Suggestions")

    for table in document.tables:
        for row in table.rows:
            for cell in row.cells:
                for paragraph in cell.paragraphs:
                    _style_paragraph(paragraph)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    document.save(output_path)
    return output_path


def _summary_table_headers(*, max_level: int) -> list[str]:
    """Build Summary sheet table headers.

    Args:
        max_level: Parameter value (int).

    Returns:
        list[str]: Return value.

    Raises:
        None.
    """
    level_headers = [f"Level {level}" for level in range(0, max_level + 1)]
    return ["CO", *level_headers, "CO%", "Result"]


def _summary_table_start_col() -> int:
    """Summary table start column index (0-based)."""
    return 2


def _format_level_count_with_percent(*, level_count: int, attended: int) -> str:
    """Format level count with in-row level percentage.

    Args:
        level_count: Parameter value (int).
        attended: Parameter value (int).

    Returns:
        str: Return value.

    Raises:
        None.
    """
    if attended <= 0:
        level_percent = 0.0
    else:
        level_percent = round((float(level_count) / float(attended)) * 100.0, CO_REPORT_MAX_DECIMAL_PLACES)
    return f"{int(level_count)} ({level_percent:g}%)"


def _create_summary_sheet(
    workbook: Any,
    *,
    sheet: Any | None = None,
    template_id: str,
    metadata: dict[str, str],
    thresholds: tuple[float, float, float] | None,
    co_attainment_percent: float,
    co_attainment_level: int,
    output_states: dict[int, _CoOutputSheetState],
    total_outcomes: int,
) -> tuple[int, int, list[str]]:
    """Create summary sheet.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
        metadata: Parameter value (dict[str, str]).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float).
        co_attainment_level: Parameter value (int).
        output_states: Parameter value (dict[int, _CoOutputSheetState]).
        total_outcomes: Parameter value (int).
    
    Returns:
        tuple[int, int, list[str]]: Return value.
    
    Raises:
        None.
    """
    if sheet is None:
        sheet = workbook.add_worksheet("Summary")
    if sheet is None:
        raise AppSystemError("Unable to create Summary worksheet.")
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_rows = _summary_graph_metadata_rows(
        metadata=metadata,
        total_outcomes=total_outcomes,
        thresholds=thresholds,
        co_attainment_percent=co_attainment_percent,
        co_attainment_level=co_attainment_level,
        include_thresholds=bool(metadata),
    )
    write_two_column_metadata_rows(
        sheet,
        rows=metadata_rows,
        label_col_index=1,
        value_col_index=2,
        label_format=formats["body"],
        value_format=formats["body_wrap"],
        centered_row_labels=CO_ANALYSIS_INSTITUTION_ROWS,
        centered_label_format=formats["body_center"],
        centered_value_format=formats["body_center"],
    )

    header_row_index = len(metadata_rows) + 1
    max_level = max((max(state.level_counts) for state in output_states.values()), default=3)
    headers = _summary_table_headers(max_level=max_level)
    table_start_col = _summary_table_start_col()  # Start summary table at column C.
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, table_start_col + col_idx, header, formats["header"])

    first_data_row = header_row_index + 1
    course_code = metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "").strip()
    table_row_values: list[list[Any]] = []
    for co_index in range(1, total_outcomes + 1):
        state = output_states.get(co_index)
        level_counts = state.level_counts if state is not None else {level: 0 for level in range(0, max_level + 1)}
        attended = state.attended if state is not None else 0
        co_label = f"{course_code}.{co_index}" if course_code else f"CO{co_index}"
        co_percentage = _co_percentage(
            level_counts=level_counts,
            attended=attended,
            co_attainment_level=co_attainment_level,
        )
        result_text = (
            "Attained"
            if isinstance(co_percentage, (int, float)) and float(co_percentage) >= float(co_attainment_percent)
            else "Yet to Attain"
        )
        row_values: list[Any] = [
            co_label,
            *[
                _format_level_count_with_percent(level_count=int(level_counts.get(level, 0)), attended=int(attended))
                for level in range(0, max_level + 1)
            ],
            co_percentage,
            result_text,
        ]
        table_row_values.append(row_values)
        row_index = header_row_index + co_index
        sheet.write_row(row_index, table_start_col, row_values, formats["body_center"])

    last_col = table_start_col + len(headers) - 1
    max_sample_row = 30
    sampled_rows: list[list[Any]] = [[""] * (last_col + 1) for _ in range(max_sample_row)]
    for row_index, (field, value) in enumerate(metadata_rows):
        if row_index >= max_sample_row:
            break
        sampled_rows[row_index][1] = field
        sampled_rows[row_index][2] = value
    if header_row_index < max_sample_row:
        for col_offset, header in enumerate(headers):
            sampled_rows[header_row_index][table_start_col + col_offset] = header
    for row_offset, row_values in enumerate(table_row_values, start=1):
        row_index = header_row_index + row_offset
        if row_index >= max_sample_row:
            break
        for col_offset, value in enumerate(row_values):
            sampled_rows[row_index][table_start_col + col_offset] = value
    widths = _compute_sampled_column_widths(sampled_rows, last_col)
    _apply_xlsxwriter_column_widths(
        sheet,
        widths,
        wrap_columns=(1, 2, last_col),
        wrap_format=formats["column_wrap"],
    )

    _apply_xlsxwriter_layout_shared(
        sheet,
        header_row_index=header_row_index,
        paper_size=9,
        landscape=True,
        selection_col=table_start_col,
    )
    sheet.repeat_rows(0, header_row_index)
    # Freeze columns A and B; summary table starts at column C.
    sheet.freeze_panes(header_row_index + 1, table_start_col)
    summary_last_data_row = first_data_row + max(0, total_outcomes - 1)
    write_sheet_footer_xlsxwriter(
        sheet,
        footer_text=CO_ANALYSIS_SHEET_FOOTER_TEXT,
        row_index=summary_last_data_row + 2,
        col_index=0,
        cell_format=formats["body"],
    )
    return first_data_row, first_data_row + max(0, total_outcomes - 1), headers


def _create_graph_sheet(
    workbook: Any,
    *,
    sheet: Any | None = None,
    template_id: str,
    metadata: dict[str, str],
    total_outcomes: int,
    thresholds: tuple[float, float, float] | None,
    co_attainment_percent: float,
    co_attainment_level: int,
    summary_first_data_row: int,
    summary_headers: list[str],
) -> None:
    """Create graph sheet.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
        metadata: Parameter value (dict[str, str]).
        total_outcomes: Parameter value (int).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float).
        co_attainment_level: Parameter value (int).
        summary_first_data_row: Parameter value (int).
        summary_headers: Parameter value (list[str]).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    graph_sheet = sheet if sheet is not None else workbook.add_worksheet("Graph")
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_headers = _course_metadata_headers(template_id)
    metadata_rows = _summary_graph_metadata_rows(
        metadata=metadata,
        total_outcomes=total_outcomes,
        thresholds=thresholds,
        co_attainment_percent=co_attainment_percent,
        co_attainment_level=co_attainment_level,
        include_thresholds=bool(metadata),
    )
    metadata_rows = [
        (field, value)
        for field, value in metadata_rows
        if field not in CO_ANALYSIS_INSTITUTION_ROWS
    ]
    write_two_column_metadata_rows(
        graph_sheet,
        rows=metadata_rows,
        label_col_index=1,
        value_col_index=2,
        label_format=formats["body"],
        value_format=formats["body_wrap"],
        centered_row_labels=CO_ANALYSIS_INSTITUTION_ROWS,
        centered_label_format=formats["body_center"],
        centered_value_format=formats["body_center"],
    )
    sampled_rows: list[list[Any]] = [["", metadata_headers[0], metadata_headers[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    graph_sheet.set_column(1, 1, widths.get(1, 8))
    graph_sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    chart = workbook.add_chart({"type": "column"})
    summary_table_start_col = _summary_table_start_col()
    summary_co_header = "CO"
    summary_co_percent_header = "Attainment %"
    summary_co_col = summary_table_start_col
    summary_co_percent_col_index = max(0, len(summary_headers) - 2)
    if summary_headers:
        for index, header in enumerate(summary_headers):
            if str(header).strip().casefold() == "co%":
                summary_co_percent_col_index = index
    summary_co_percent_col = summary_table_start_col + summary_co_percent_col_index
    summary_last_data_row = summary_first_data_row + max(0, total_outcomes - 1)
    chart.add_series(
        {
            "name": summary_co_percent_header,
            "categories": [
                "Summary",
                summary_first_data_row,
                summary_co_col,
                summary_last_data_row,
                summary_co_col,
            ],
            "values": [
                "Summary",
                summary_first_data_row,
                summary_co_percent_col,
                summary_last_data_row,
                summary_co_percent_col,
            ],
            "fill": {"color": "#4F81BD"},
            "border": {"color": "#4F81BD"},
            "data_labels": {
                "value": True,
                "series_name": False,
                "category": False,
                "legend_key": False,
                "position": "outside_end",
                "num_format": '0.##"%"',
            },
        }
    )
    chart.set_title({"name": "CO Attainment %"})
    chart.set_x_axis(
        {
            "name": summary_co_header,
            "label_position": "low",
            "interval_unit": 1,
        }
    )
    chart.set_y_axis(
        {
            "name": summary_co_percent_header,
            "min": 0,
            "max": 100,
            "major_unit": 10,
            "num_format": '0"%"',
            "major_gridlines": {"visible": True},
        }
    )
    chart.set_legend({"none": True})
    chart.set_style(10)
    chart_anchor_row = len(metadata_rows) + 2
    graph_sheet.insert_chart(f"B{chart_anchor_row + 1}", chart, {"x_scale": 1.4, "y_scale": 1.4})
    write_sheet_footer_xlsxwriter(
        graph_sheet,
        footer_text=CO_ANALYSIS_SHEET_FOOTER_TEXT,
        row_index=chart_anchor_row + 20,
        col_index=0,
        cell_format=formats["body"],
    )
    graph_header_row = max(0, len(metadata_rows) - 1)
    _apply_xlsxwriter_layout_shared(
        graph_sheet,
        header_row_index=graph_header_row,
        paper_size=9,
        landscape=True,
    )
    graph_sheet.repeat_rows(0, graph_header_row)


def _add_system_hash_sheet(workbook: Any, template_id: str) -> None:
    """Add system hash sheet.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    template_hash = sign_payload(template_id)
    hash_ws = workbook.add_worksheet(SYSTEM_HASH_SHEET)
    hash_ws.write(0, 0, SYSTEM_HASH_TEMPLATE_ID_HEADER)
    hash_ws.write(0, 1, SYSTEM_HASH_TEMPLATE_HASH_HEADER)
    hash_ws.write(1, 0, template_id)
    hash_ws.write(1, 1, template_hash)
    hash_ws.hide()


def _add_system_layout_sheet(workbook: Any, manifest_text: str, manifest_hash: str) -> None:
    """Add system layout sheet.
    
    Args:
        workbook: Parameter value (Any).
        manifest_text: Parameter value (str).
        manifest_hash: Parameter value (str).
    
    Returns:
        None.
    
    Raises:
        None.
    """
    layout_ws = workbook.add_worksheet(SYSTEM_LAYOUT_SHEET)
    layout_ws.write(0, 0, SYSTEM_LAYOUT_MANIFEST_HEADER)
    layout_ws.write(0, 1, SYSTEM_LAYOUT_MANIFEST_HASH_HEADER)
    layout_ws.write(1, 0, manifest_text)
    layout_ws.write(1, 1, manifest_hash)
    layout_ws.hide()


def _build_system_layout_manifest(
    *,
    template_id: str,
    sheet_order: list[str],
) -> tuple[str, str]:
    """Build system layout manifest.
    
    Args:
        template_id: Parameter value (str).
        sheet_order: Parameter value (list[str]).
    
    Returns:
        tuple[str, str]: Return value.
    
    Raises:
        None.
    """
    template_hash = sign_payload(template_id)
    signed_sheet_order = [*sheet_order, SYSTEM_HASH_SHEET]
    manifest = {
        "schema_version": WORKBOOK_INTEGRITY_SCHEMA_VERSION,
        "template_id": template_id,
        "template_hash": template_hash,
        "sheet_order": signed_sheet_order,
        "sheets": [{"name": name, "hash": sign_payload(name)} for name in signed_sheet_order],
    }
    manifest_text = json.dumps(manifest, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    manifest_hash = sign_payload(manifest_text)
    return manifest_text, manifest_hash


def _attainment_thresholds(
    thresholds: tuple[float, float, float] | None = None,
) -> tuple[float, float, float]:
    """Attainment thresholds.
    
    Args:
        thresholds: Parameter value (tuple[float, float, float] | None).
    
    Returns:
        tuple[float, float, float]: Return value.
    
    Raises:
        None.
    """
    if thresholds is None:
        return (float(LEVEL_1_THRESHOLD), float(LEVEL_2_THRESHOLD), float(LEVEL_3_THRESHOLD))
    l1, l2, l3 = thresholds
    return (float(l1), float(l2), float(l3))


def _co_attainment_target(
    *,
    co_attainment_percent: float | None,
    co_attainment_level: int | None,
    threshold_count: int,
) -> tuple[float, int]:
    """Co attainment target.
    
    Args:
        co_attainment_percent: Parameter value (float | None).
        co_attainment_level: Parameter value (int | None).
        threshold_count: Parameter value (int).
    
    Returns:
        tuple[float, int]: Return value.
    
    Raises:
        None.
    """
    percent = float(CO_ATTAINMENT_PERCENT_DEFAULT if co_attainment_percent is None else co_attainment_percent)
    level = int(CO_ATTAINMENT_LEVEL_DEFAULT if co_attainment_level is None else co_attainment_level)
    percent = max(0.0, min(100.0, percent))
    level = max(1, min(max(1, threshold_count), level))
    return percent, level


def _score_to_attainment_level(
    score: float | str,
    *,
    thresholds: tuple[float, float, float],
) -> int | str:
    """Score to attainment level.
    
    Args:
        score: Parameter value (float | str).
        thresholds: Parameter value (tuple[float, float, float]).
    
    Returns:
        int | str: Return value.
    
    Raises:
        None.
    """
    if not isinstance(score, (int, float)) or isinstance(score, bool):
        return CO_REPORT_NOT_APPLICABLE_TOKEN

    total = float(score)
    tolerance = 10 ** (-CO_REPORT_MAX_DECIMAL_PLACES)
    if 100.0 < total <= (100.0 + tolerance):
        total = 100.0
    if (-tolerance) <= total < 0.0:
        total = 0.0
    sorted_thresholds = tuple(float(value) for value in thresholds)
    if not sorted_thresholds:
        return CO_REPORT_NOT_APPLICABLE_TOKEN
    if 0.0 <= total < sorted_thresholds[0]:
        return 0
    for index in range(1, len(sorted_thresholds)):
        if sorted_thresholds[index - 1] <= total < sorted_thresholds[index]:
            return index
    if sorted_thresholds[-1] <= total <= 100.0:
        return len(sorted_thresholds)
    return CO_REPORT_NOT_APPLICABLE_TOKEN


def _reg_no_sort_key(reg_no: str) -> tuple[tuple[int, int | str], ...]:
    """Reg no sort key.
    
    Args:
        reg_no: Parameter value (str).
    
    Returns:
        tuple[tuple[int, int | str], ...]: Return value.
    
    Raises:
        None.
    """
    tokens = re.split(r"(\d+)", reg_no.strip())
    key_parts: list[tuple[int, int | str]] = []
    for token in tokens:
        if not token:
            continue
        if token.isdigit():
            key_parts.append((0, int(token)))
        else:
            key_parts.append((1, token.casefold()))
    return tuple(key_parts)


def _generate_co_attainment_workbook_course_setup_v2(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    total_outcomes: int | None = None,
    template_id: str = ID_COURSE_SETUP,
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float | None = None,
    co_attainment_level: int | None = None,
    generate_word_report: bool = False,
    word_output_path: Path | None = None,
    co_description_path: Path | None = None,
) -> _CoAttainmentWorkbookResult:
    """Generate co attainment workbook course setup v2.
    
    Args:
        source_paths: Parameter value (list[Path]).
        output_path: Parameter value (Path).
        token: Parameter value (CancellationToken).
        total_outcomes: Parameter value (int | None).
        template_id: Parameter value (str).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float | None).
        co_attainment_level: Parameter value (int | None).
        generate_word_report: Parameter value (bool).
        word_output_path: Parameter value (Path | None).
        co_description_path: Parameter value (Path | None).
    
    Returns:
        _CoAttainmentWorkbookResult: Return value.
    
    Raises:
        None.
    """
    xlsxwriter = import_runtime_dependency("xlsxwriter")
    openpyxl = import_runtime_dependency("openpyxl")
    load_workbook = openpyxl.load_workbook

    resolved_total_outcomes = int(total_outcomes or 0)
    if resolved_total_outcomes <= 0:
        first_signature = _extract_final_report_signature(source_paths[0]) if source_paths else None
        if first_signature is not None:
            resolved_total_outcomes = int(first_signature.total_outcomes)

    if resolved_total_outcomes <= 0:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COA_TOTAL_OUTCOMES_MISSING",
        )

    metadata: dict[str, str] = {}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook = xlsxwriter.Workbook(str(output_path), {"constant_memory": True})
    workbook_closed = False
    output_states: dict[int, _CoOutputSheetState] = {}
    pending_rows: dict[int, list[_CoAttainmentRow]] = {}
    pending_direct_columns: dict[int, list[_DirectComponentColumn]] = {}
    pending_direct_scores: dict[int, dict[int, dict[str, float | str]]] = {}
    pending_indirect_columns: dict[int, list[_IndirectComponentColumn]] = {}
    pending_indirect_scores: dict[int, dict[int, dict[str, float | str]]] = {}
    duplicate_reg_count = 0
    duplicate_entries: list[tuple[str, str, str]] = []
    inner_join_drop_count = 0
    inner_join_drop_details: list[str] = []
    generated_word_report_path: Path | None = None
    word_report_error_key: str | None = None
    generated_cip_payload: dict[str, object] | None = None
    level_thresholds = _attainment_thresholds(thresholds)
    target_percent, target_level = _co_attainment_target(
        co_attainment_percent=co_attainment_percent,
        co_attainment_level=co_attainment_level,
        threshold_count=len(level_thresholds),
    )
    dedup_store = _RegisterDedupStore(
        total_outcomes=resolved_total_outcomes,
        use_sqlite=_should_use_sqlite_dedup(
            source_count=len(source_paths),
            total_outcomes=resolved_total_outcomes,
        ),
    )

    try:
        for source in source_paths:
            token.raise_if_cancelled()
            workbook = load_workbook(filename=source, data_only=True, read_only=True)
            try:
                if not metadata and COURSE_METADATA_SHEET in workbook.sheetnames:
                    metadata = _extract_course_metadata_fields(workbook[COURSE_METADATA_SHEET])
                for co_index in range(1, resolved_total_outcomes + 1):
                    row_bucket = pending_rows.setdefault(co_index, [])
                    (
                        rows,
                        direct_total,
                        indirect_total,
                        dropped_for_sheet,
                        direct_columns,
                        indirect_columns,
                    ) = _iter_co_rows_from_workbook(workbook, co_index=co_index, workbook_name=source.name)
                    if direct_columns:
                        merged_columns = pending_direct_columns.setdefault(co_index, [])
                        merged_index = {normalize(item.name): idx for idx, item in enumerate(merged_columns)}
                        for item in direct_columns:
                            key = normalize(item.name)
                            existing_idx = merged_index.get(key)
                            if existing_idx is None:
                                merged_columns.append(item)
                                merged_index[key] = len(merged_columns) - 1
                                continue
                            existing = merged_columns[existing_idx]
                            merged_columns[existing_idx] = _DirectComponentColumn(
                                name=existing.name,
                                max_marks=max(existing.max_marks, item.max_marks),
                                weight=max(existing.weight, item.weight),
                                score_column=existing.score_column,
                            )
                    if indirect_columns:
                        merged_columns = pending_indirect_columns.setdefault(co_index, [])
                        merged_index = {normalize(item.name): idx for idx, item in enumerate(merged_columns)}
                        for item in indirect_columns:
                            key = normalize(item.name)
                            existing_idx = merged_index.get(key)
                            if existing_idx is None:
                                merged_columns.append(item)
                                merged_index[key] = len(merged_columns) - 1
                                continue
                            existing = merged_columns[existing_idx]
                            merged_columns[existing_idx] = _IndirectComponentColumn(
                                name=existing.name,
                                weight=max(existing.weight, item.weight),
                                score_column=existing.score_column,
                            )
                    if dropped_for_sheet:
                        inner_join_drop_count += dropped_for_sheet
                        inner_join_drop_details.append(
                            f"{source.name} CO{co_index}: dropped={dropped_for_sheet} (direct_rows={direct_total}, indirect_rows={indirect_total})"
                        )
                    for row in rows:
                        if not dedup_store.add_if_absent(co_index=co_index, reg_hash=row.reg_hash):
                            duplicate_reg_count += 1
                            duplicate_entries.append((row.reg_no, row.worksheet_name, row.workbook_name))
                            continue
                        if direct_columns and row.direct_component_scores:
                            score_map = pending_direct_scores.setdefault(co_index, {}).setdefault(row.reg_hash, {})
                            for idx, component in enumerate(direct_columns):
                                if idx >= len(row.direct_component_scores):
                                    break
                                score_map[normalize(component.name)] = row.direct_component_scores[idx]
                        if indirect_columns and row.indirect_component_scores:
                            score_map = pending_indirect_scores.setdefault(co_index, {}).setdefault(row.reg_hash, {})
                            for idx, component in enumerate(indirect_columns):
                                if idx >= len(row.indirect_component_scores):
                                    break
                                score_map[normalize(component.name)] = row.indirect_component_scores[idx]
                        row_bucket.append(row)
            finally:
                workbook.close()
        shared_formats = _xlsxwriter_formats(output_workbook, template_id=template_id)
        _create_pass_percentage_sheet(
            output_workbook,
            template_id=template_id,
            metadata=metadata,
            thresholds=level_thresholds,
            pending_rows=pending_rows,
            pending_direct_columns=pending_direct_columns,
            pending_direct_scores=pending_direct_scores,
            total_outcomes=resolved_total_outcomes,
        )
        summary_sheet = output_workbook.add_worksheet("Summary")
        graph_sheet = output_workbook.add_worksheet("Graph")
        for co_index in range(1, resolved_total_outcomes + 1):
            sorted_rows = sorted(
                pending_rows.get(co_index, []),
                key=lambda item: (_reg_no_sort_key(item.reg_no), item.reg_hash),
            )
            direct_columns = pending_direct_columns.get(co_index, [])
            indirect_columns = pending_indirect_columns.get(co_index, [])
            direct_components: list[_CoReportComponent] = []
            indirect_components: list[_CoReportComponent] = []
            if direct_columns:
                for item in direct_columns:
                    component_marks: list[Any] = []
                    for row in sorted_rows:
                        component_value = (
                            pending_direct_scores.get(co_index, {})
                            .get(row.reg_hash, {})
                            .get(normalize(item.name), CO_REPORT_NOT_APPLICABLE_TOKEN)
                        )
                        component_marks.append(component_value)
                    direct_components.append(
                        _CoReportComponent(
                            name=item.name,
                            weight=item.weight,
                            max_by_co={co_index: item.max_marks},
                            marks_by_co={co_index: component_marks},
                        )
                    )
            else:
                synthetic_marks: list[Any] = []
                for row in sorted_rows:
                    if isinstance(row.direct_score, (int, float)):
                        synthetic_marks.append(
                            round(float(row.direct_score) / DIRECT_RATIO, CO_REPORT_MAX_DECIMAL_PLACES)
                            if DIRECT_RATIO > 0
                            else 0.0
                        )
                    else:
                        synthetic_marks.append(CO_REPORT_ABSENT_TOKEN)
                direct_components.append(
                    _CoReportComponent(
                        name="Direct",
                        weight=100.0,
                        max_by_co={co_index: 100.0},
                        marks_by_co={co_index: synthetic_marks},
                    )
                )
            if indirect_columns:
                for item in indirect_columns:
                    component_marks: list[Any] = []
                    for row in sorted_rows:
                        component_value = (
                            pending_indirect_scores.get(co_index, {})
                            .get(row.reg_hash, {})
                            .get(normalize(item.name), CO_REPORT_NOT_APPLICABLE_TOKEN)
                        )
                        component_marks.append(component_value)
                    indirect_components.append(
                        _CoReportComponent(
                            name=item.name,
                            weight=item.weight,
                            max_by_co={co_index: float(LIKERT_MAX)},
                            marks_by_co={co_index: component_marks},
                        )
                    )
            write_co_outcome_sheets(
                output_workbook,
                template_id=template_id,
                co_index=co_index,
                metadata_rows=_metadata_rows_for_co_analysis_outcome_sheets(
                    metadata=metadata,
                    co_index=co_index,
                    thresholds=level_thresholds,
                    co_attainment_percent=target_percent,
                    co_attainment_level=target_level,
                ),
                rows=sorted_rows,
                direct_components=direct_components,
                indirect_components=indirect_components,
                formats=shared_formats,
            )
            state = _create_co_attainment_sheet(
                output_workbook,
                template_id=template_id,
                co_index=co_index,
                metadata=metadata,
                thresholds=level_thresholds,
                co_attainment_percent=target_percent,
                co_attainment_level=target_level,
            )
            output_states[co_index] = state
            for row in sorted_rows:
                _append_co_attainment_row(state, row, thresholds=level_thresholds)
            _append_co_attainment_summary(state)
        summary_first_data_row, _summary_last_data_row, summary_headers = _create_summary_sheet(
            output_workbook,
            sheet=summary_sheet,
            template_id=template_id,
            metadata=metadata,
            thresholds=level_thresholds,
            co_attainment_percent=target_percent,
            co_attainment_level=target_level,
            output_states=output_states,
            total_outcomes=resolved_total_outcomes,
        )
        _create_graph_sheet(
            output_workbook,
            sheet=graph_sheet,
            template_id=template_id,
            metadata=metadata,
            total_outcomes=resolved_total_outcomes,
            thresholds=level_thresholds,
            co_attainment_percent=target_percent,
            co_attainment_level=target_level,
            summary_first_data_row=summary_first_data_row,
            summary_headers=summary_headers,
        )
        sheet_order: list[str] = [_PASS_PERCENTAGE_SHEET_NAME, "Summary", "Graph"]
        for co_index in range(1, resolved_total_outcomes + 1):
            sheet_order.extend(
                [
                    co_direct_sheet_name(co_index),
                    co_indirect_sheet_name(co_index),
                    f"CO{co_index}",
                ]
            )
        _add_system_hash_sheet(output_workbook, template_id)
        manifest_text, manifest_hash = _build_system_layout_manifest(
            template_id=template_id,
            sheet_order=sheet_order,
        )
        _add_system_layout_sheet(output_workbook, manifest_text, manifest_hash)
        output_workbook.close()
        workbook_closed = True
    finally:
        dedup_store.close()
        if not workbook_closed:
            try:
                output_workbook.close()
            except Exception:
                _logger.debug("Suppressing workbook close error during cleanup.", exc_info=True)
    if generate_word_report:
        try:
            if co_description_path is None:
                raise validation_error_from_key(
                    "common.validation_failed_invalid_data",
                    code="CO_DESCRIPTION_SELECTION_INVALID",
                )
            course_code = str(metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")).strip()
            co_rows = _build_co_word_summary_rows(
                pending_rows=pending_rows,
                output_states=output_states,
                total_outcomes=resolved_total_outcomes,
                course_code=course_code,
                co_attainment_level=target_level,
                co_attainment_percent=target_percent,
            )
            co_sentences = _validated_co_description_sentences(
                co_description_path=co_description_path,
                template_id=template_id,
                total_outcomes=resolved_total_outcomes,
                token=token,
            )
            co_level_data: dict[int, tuple[int, dict[int, int]]] = {
                co_idx: (state.attended, dict(state.level_counts))
                for co_idx, state in output_states.items()
            }
            seen_direct: dict[str, float] = {}
            for _cols in pending_direct_columns.values():
                for _col in _cols:
                    if _col.name not in seen_direct:
                        seen_direct[_col.name] = _col.weight
            seen_indirect: dict[str, float] = {}
            for _cols in pending_indirect_columns.values():
                for _col in _cols:
                    if _col.name not in seen_indirect:
                        seen_indirect[_col.name] = _col.weight
            cip_assessments: list[dict[str, object]] = [
                {"name": name, "wt": round(wt, 2), "d": True}
                for name, wt in seen_direct.items()
            ] + [
                {"name": name, "wt": round(wt, 2), "d": False}
                for name, wt in seen_indirect.items()
            ]
            co_desc_records: list[_CoDescriptionRecord] = read_co_description_records(
                co_description_path,
                template_id=template_id,
                total_outcomes=resolved_total_outcomes,
                cancel_token=token,
            )
            generated_cip_payload = build_cip_payload(
                metadata=metadata,
                thresholds=level_thresholds,
                co_attainment_percent=target_percent,
                co_attainment_level=target_level,
                total_outcomes=resolved_total_outcomes,
                co_rows=co_rows,
                co_level_data=co_level_data,
                assessments=cip_assessments,
                co_description_records=co_desc_records,
            )
            resolved_word_output = word_output_path or output_path.with_name(f"{output_path.stem}_Report.docx")
            cip_json_path = resolved_word_output.with_name(
                resolved_word_output.stem + "_CIP_Payload.json"
            )
            cip_json_path.write_text(
                json.dumps(generated_cip_payload, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
            generated_word_report_path = _generate_co_attainment_word_report(
                output_path=resolved_word_output,
                metadata=metadata,
                thresholds=level_thresholds,
                co_attainment_percent=target_percent,
                co_attainment_level=target_level,
                total_outcomes=resolved_total_outcomes,
                co_rows=co_rows,
                co_sentences=co_sentences,
            )
        except ValidationError:
            raise
        except Exception:
            word_report_error_key = "co_analysis.status.word_report_generate_failed"
            _logger.exception("CO analysis Word report generation failed.")
    return _CoAttainmentWorkbookResult(
        output_path=output_path,
        duplicate_reg_count=duplicate_reg_count,
        duplicate_entries=tuple(duplicate_entries),
        inner_join_drop_count=inner_join_drop_count,
        inner_join_drop_details=tuple(inner_join_drop_details),
        word_report_path=generated_word_report_path,
        word_report_error_key=word_report_error_key,
        cip_payload=generated_cip_payload,
    )


def generate_co_attainment_workbook(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    total_outcomes: int | None = None,
    template_id: str = ID_COURSE_SETUP,
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float | None = None,
    co_attainment_level: int | None = None,
    generate_word_report: bool = False,
    word_output_path: Path | None = None,
    co_description_path: Path | None = None,
) -> _CoAttainmentWorkbookResult:
    """Generate co attainment workbook.
    
    Args:
        source_paths: Parameter value (list[Path]).
        output_path: Parameter value (Path).
        token: Parameter value (CancellationToken).
        total_outcomes: Parameter value (int | None).
        template_id: Parameter value (str).
        thresholds: Parameter value (tuple[float, float, float] | None).
        co_attainment_percent: Parameter value (float | None).
        co_attainment_level: Parameter value (int | None).
        generate_word_report: Parameter value (bool).
        word_output_path: Parameter value (Path | None).
        co_description_path: Parameter value (Path | None).
    
    Returns:
        _CoAttainmentWorkbookResult: Return value.
    
    Raises:
        None.
    """
    if not source_paths:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COA_SOURCE_WORKBOOK_REQUIRED",
        )
    prepared_source_paths: list[Path] = list(source_paths)
    temp_root_dir: Path | None = None
    source_is_final_report: list[bool] = []
    generated_source_paths: list[Path] = []
    try:
        for path in source_paths:
            token.raise_if_cancelled()
            is_final_report = extract_final_report_signature_from_path(path) is not None
            source_is_final_report.append(is_final_report)
            if is_final_report:
                continue
            if temp_root_dir is None:
                temp_root_base = app_runtime_storage_dir(APP_NAME) / ".co_attainment_tmp"
                temp_root_base.mkdir(parents=True, exist_ok=True)
                temp_root_dir = temp_root_base / f"focus_co_attainment_src_{uuid.uuid4().hex}"
                temp_root_dir.mkdir(parents=True, exist_ok=True)
            output_name = f"co_attainment_source_{len(generated_source_paths) + 1}.xlsx"
            generated_path = temp_root_dir / output_name
            generate_final_report_workbook(
                filled_marks_path=path,
                output_path=generated_path,
                cancel_token=token,
            )
            generated_source_paths.append(generated_path)
        if temp_root_dir is not None:
            resolved_paths: list[Path] = []
            generated_index = 0
            for original, is_final_report in zip(source_paths, source_is_final_report, strict=True):
                if is_final_report:
                    resolved_paths.append(original)
                else:
                    resolved_paths.append(generated_source_paths[generated_index])
                    generated_index += 1
            prepared_source_paths = resolved_paths

        normalized_template_id = normalize(template_id)
        if normalized_template_id == normalize("COURSE_SETUP_V2"):
            return _generate_co_attainment_workbook_course_setup_v2(
                source_paths=prepared_source_paths,
                output_path=output_path,
                token=token,
                total_outcomes=total_outcomes,
                template_id="COURSE_SETUP_V2",
                thresholds=thresholds,
                co_attainment_percent=co_attainment_percent,
                co_attainment_level=co_attainment_level,
                generate_word_report=generate_word_report,
                word_output_path=word_output_path,
                co_description_path=co_description_path,
            )
        raise validation_error_from_key(
            "validation.template.unknown",
            code="UNKNOWN_TEMPLATE",
            template_id=template_id,
        )
    finally:
        if temp_root_dir is not None:
            shutil.rmtree(temp_root_dir, ignore_errors=True)


def generate_final_report_workbook(
    *,
    filled_marks_path: Path,
    output_path: Path,
    cancel_token: CancellationToken | None = None,
) -> Path:
    """Generate final report workbook.
    
    Args:
        filled_marks_path: Parameter value (Path).
        output_path: Parameter value (Path).
        cancel_token: Parameter value (CancellationToken | None).
    
    Returns:
        Path: Return value.
    
    Raises:
        None.
    """
    if cancel_token is not None:
        cancel_token.raise_if_cancelled()
    if not filled_marks_path.exists():
        raise validation_error_from_key(
            "validation.workbook.not_found",
            code="WORKBOOK_NOT_FOUND",
            workbook=str(filled_marks_path),
        )
    openpyxl = import_runtime_dependency("openpyxl")

    source_wb = openpyxl.load_workbook(filled_marks_path, data_only=False)
    try:
        from domain.template_strategy_router import (
            read_template_id_from_system_hash_sheet_if_valid,
        )

        template_id = read_template_id_from_system_hash_sheet_if_valid(source_wb) or ID_COURSE_SETUP
        metadata_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
        sheet_specs_by_name = _layout_sheet_specs_by_name(source_wb)
        total_outcomes = extract_total_outcomes_from_workbook_path(filled_marks_path)
        if total_outcomes is None:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="COA_TOTAL_OUTCOMES_MISSING",
            )
        students_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)
        students_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)
        students: list[tuple[str, str]] = []
        if students_sheet_name in source_wb.sheetnames:
            students_sheet = source_wb[students_sheet_name]
            header_row = _students_sheet_header_row(
                sheet_specs_by_name=sheet_specs_by_name,
                sheet_name=students_sheet_name,
            )
            header_map = _header_map_for_row(students_sheet, header_row=header_row)
            reg_col = _resolve_students_reg_col(header_map=header_map, students_headers=students_headers)
            name_col = _resolve_students_name_col(header_map=header_map, students_headers=students_headers)
            if reg_col is not None and name_col is not None:
                students.extend(
                    _iter_students_sheet_rows(
                        students_sheet=students_sheet,
                        header_row=header_row,
                        reg_col=reg_col,
                        name_col=name_col,
                    )
                )
        if not students:
            raise validation_error_from_key(
                "common.validation_failed_invalid_data",
                code="SHEET_DATA_REQUIRED",
                sheet_name=students_sheet_name,
            )
        direct_component_order: list[str] = []
        direct_component_weights: dict[str, float] = {}
        indirect_component_order: list[str] = []
        indirect_component_weights: dict[str, float] = {}
        assessment_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
        if assessment_sheet_name in source_wb.sheetnames:
            from domain.template_versions.course_setup_v2_impl.assessment_semantics import (
                parse_assessment_components,
            )

            assessment_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
            assessment_header_row = int(sheet_specs_by_name.get(assessment_sheet_name, {}).get("header_row") or 1)
            assessment_rows = _iter_data_rows(
                source_wb[assessment_sheet_name],
                len(assessment_headers),
                header_row=assessment_header_row,
            )
            parsed_components = parse_assessment_components(
                assessment_rows,
                sheet_name=assessment_sheet_name,
                row_start=assessment_header_row + 1,
                on_blank_component="skip",
                duplicate_policy="keep_first",
                require_non_empty=False,
                validate_allowed_options=False,
            )
            for component in parsed_components:
                name = str(component.component_name or "").strip()
                if not name:
                    continue
                key = normalize(name)
                if component.is_direct:
                    if key in direct_component_weights:
                        continue
                    direct_component_order.append(name)
                    direct_component_weights[key] = round(float(component.weight), CO_REPORT_MAX_DECIMAL_PLACES)
                else:
                    if key in indirect_component_weights:
                        continue
                    indirect_component_order.append(name)
                    indirect_component_weights[key] = round(float(component.weight), CO_REPORT_MAX_DECIMAL_PLACES)

        direct_components_by_co: dict[int, list[tuple[str, float, float]]] = {}
        seen_by_co: dict[int, set[str]] = {}
        direct_scores_by_co: dict[int, dict[str, dict[str, float | str]]] = {}
        indirect_components_by_co: dict[int, list[tuple[str, float]]] = {}
        indirect_seen_by_co: dict[int, set[str]] = {}
        indirect_scores_by_co: dict[int, dict[str, dict[str, float | str]]] = {}
        valid_student_keys = {normalize(reg_no): reg_no for reg_no, _name in students}

        def _score_cell(value: Any) -> float | str:
            token = normalize(value)
            if token == normalize(CO_REPORT_ABSENT_TOKEN):
                return CO_REPORT_ABSENT_TOKEN
            parsed = coerce_excel_number(value)
            if isinstance(parsed, bool) or not isinstance(parsed, (int, float)):
                return 0.0
            return round(float(parsed), CO_REPORT_MAX_DECIMAL_PLACES)

        for sheet_name, spec in sheet_specs_by_name.items():
            kind = normalize(spec.get("kind"))
            if sheet_name not in source_wb.sheetnames:
                continue
            ws = source_wb[sheet_name]
            header_row = int(spec.get("header_row") or 1)
            component_name = _extract_component_name_from_marks_sheet(ws, header_row=header_row)
            if not component_name:
                continue
            component_key = normalize(component_name)
            if kind in {normalize(LAYOUT_SHEET_KIND_DIRECT_CO_WISE), normalize(LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE)}:
                component_weight = direct_component_weights.get(component_key)
                if component_weight is None:
                    continue
                max_by_co = _extract_direct_sheet_max_by_co(ws, header_row=header_row, kind=kind)
                for co_value, max_marks in max_by_co.items():
                    if max_marks <= 0:
                        continue
                    items = direct_components_by_co.setdefault(co_value, [])
                    seen = seen_by_co.setdefault(co_value, set())
                    if component_key in seen:
                        continue
                    seen.add(component_key)
                    items.append(
                        (
                            component_name,
                            round(float(max_marks), CO_REPORT_MAX_DECIMAL_PLACES),
                            round(float(component_weight), CO_REPORT_MAX_DECIMAL_PLACES),
                        )
                    )
                max_col = int(ws.max_column or 0)
                if kind == normalize(LAYOUT_SHEET_KIND_DIRECT_CO_WISE):
                    co_by_col: dict[int, int] = {}
                    for col_index in range(4, max_col + 1):
                        co_index = _co_index_from_label(ws.cell(row=header_row + 1, column=col_index).value)
                        if co_index is None:
                            continue
                        co_by_col[col_index] = co_index
                    for row_index in range(header_row + 3, int(ws.max_row or 0) + 1):
                        reg_no_raw = ws.cell(row=row_index, column=2).value
                        name_raw = ws.cell(row=row_index, column=3).value
                        if reg_no_raw is None and name_raw is None:
                            break
                        reg_key = normalize(coerce_excel_number(reg_no_raw))
                        if reg_key not in valid_student_keys:
                            continue
                        per_co_numeric: dict[int, float] = {}
                        per_co_absent: set[int] = set()
                        for col_index, co_index in co_by_col.items():
                            value = _score_cell(ws.cell(row=row_index, column=col_index).value)
                            if isinstance(value, str):
                                per_co_absent.add(co_index)
                                continue
                            per_co_numeric[co_index] = round(
                                per_co_numeric.get(co_index, 0.0) + float(value),
                                CO_REPORT_MAX_DECIMAL_PLACES,
                            )
                        for co_index in set(per_co_numeric) | set(per_co_absent):
                            co_map = direct_scores_by_co.setdefault(co_index, {})
                            comp_map = co_map.setdefault(component_key, {})
                            comp_map[reg_key] = (
                                CO_REPORT_ABSENT_TOKEN if co_index in per_co_absent else per_co_numeric.get(co_index, 0.0)
                            )
                else:
                    co_columns: list[tuple[int, int]] = []
                    for col_index in range(5, max_col + 1):
                        co_index = _co_index_from_label(ws.cell(row=header_row + 1, column=col_index).value)
                        if co_index is None:
                            continue
                        co_columns.append((col_index, co_index))
                    divisor = len(co_columns)
                    for row_index in range(header_row + 3, int(ws.max_row or 0) + 1):
                        reg_no_raw = ws.cell(row=row_index, column=2).value
                        name_raw = ws.cell(row=row_index, column=3).value
                        if reg_no_raw is None and name_raw is None:
                            break
                        reg_key = normalize(coerce_excel_number(reg_no_raw))
                        if reg_key not in valid_student_keys:
                            continue
                        total_value = _score_cell(ws.cell(row=row_index, column=4).value)
                        if isinstance(total_value, str):
                            for _col, co_index in co_columns:
                                co_map = direct_scores_by_co.setdefault(co_index, {})
                                comp_map = co_map.setdefault(component_key, {})
                                comp_map[reg_key] = CO_REPORT_ABSENT_TOKEN
                            continue
                        total_numeric = float(total_value)
                        if divisor <= 0:
                            continue
                        base = round(total_numeric / float(divisor), CO_REPORT_MAX_DECIMAL_PLACES)
                        running = 0.0
                        for idx, (_col, co_index) in enumerate(co_columns):
                            if idx == divisor - 1:
                                split_value = round(total_numeric - running, CO_REPORT_MAX_DECIMAL_PLACES)
                            else:
                                split_value = base
                                running = round(running + split_value, CO_REPORT_MAX_DECIMAL_PLACES)
                            co_map = direct_scores_by_co.setdefault(co_index, {})
                            comp_map = co_map.setdefault(component_key, {})
                            comp_map[reg_key] = split_value
            elif kind == normalize(LAYOUT_SHEET_KIND_INDIRECT):
                component_weight = indirect_component_weights.get(component_key)
                if component_weight is None:
                    continue
                max_col = int(ws.max_column or 0)
                co_columns: list[tuple[int, int]] = []
                for col_index in range(4, max_col + 1):
                    co_index = _co_index_from_label(ws.cell(row=header_row, column=col_index).value)
                    if co_index is None:
                        continue
                    co_columns.append((col_index, co_index))
                    items = indirect_components_by_co.setdefault(co_index, [])
                    seen = indirect_seen_by_co.setdefault(co_index, set())
                    if component_key in seen:
                        continue
                    seen.add(component_key)
                    items.append((component_name, round(float(component_weight), CO_REPORT_MAX_DECIMAL_PLACES)))
                for row_index in range(header_row + 1, int(ws.max_row or 0) + 1):
                    reg_no_raw = ws.cell(row=row_index, column=2).value
                    name_raw = ws.cell(row=row_index, column=3).value
                    if reg_no_raw is None and name_raw is None:
                        break
                    reg_key = normalize(coerce_excel_number(reg_no_raw))
                    if reg_key not in valid_student_keys:
                        continue
                    for col_index, co_index in co_columns:
                        score = _score_cell(ws.cell(row=row_index, column=col_index).value)
                        co_map = indirect_scores_by_co.setdefault(co_index, {})
                        comp_map = co_map.setdefault(component_key, {})
                        comp_map[reg_key] = score
        if direct_component_order:
            for co_value, items in direct_components_by_co.items():
                order_map = {normalize(name): idx for idx, name in enumerate(direct_component_order)}
                items.sort(key=lambda item: (order_map.get(normalize(item[0]), 10_000), normalize(item[0])))
        if indirect_component_order:
            for co_value, items in indirect_components_by_co.items():
                order_map = {normalize(name): idx for idx, name in enumerate(indirect_component_order)}
                items.sort(key=lambda item: (order_map.get(normalize(item[0]), 10_000), normalize(item[0])))

        output_wb = openpyxl.Workbook()
        default_sheet = output_wb.active
        if default_sheet is not None and default_sheet.title == "Sheet":
            output_wb.remove(default_sheet)

        sheet_order: list[str] = []
        if metadata_sheet_name in source_wb.sheetnames:
            src_meta = source_wb[metadata_sheet_name]
            dst_meta = output_wb.create_sheet(metadata_sheet_name)
            for row in src_meta.iter_rows(min_row=1, max_row=src_meta.max_row, min_col=1, max_col=src_meta.max_column):
                for cell in row:
                    dst_meta.cell(row=cell.row, column=cell.column, value=cell.value)
            sheet_order.append(metadata_sheet_name)
        sheet_order.extend(
            write_co_outcome_sheets_openpyxl(
                output_wb,
                total_outcomes=int(total_outcomes),
                students=students,
                direct_components_by_co=direct_components_by_co,
                direct_scores_by_co=direct_scores_by_co,
                indirect_components_by_co=indirect_components_by_co,
                indirect_scores_by_co=indirect_scores_by_co,
            )
        )

        hash_sheet = output_wb.create_sheet(SYSTEM_HASH_SHEET)
        template_hash = sign_payload(template_id)
        hash_sheet.cell(row=1, column=1, value=SYSTEM_HASH_TEMPLATE_ID_HEADER)
        hash_sheet.cell(row=1, column=2, value=SYSTEM_HASH_TEMPLATE_HASH_HEADER)
        hash_sheet.cell(row=2, column=1, value=template_id)
        hash_sheet.cell(row=2, column=2, value=template_hash)
        hash_sheet.sheet_state = "hidden"

        manifest_sheet_order = list(sheet_order) + [SYSTEM_HASH_SHEET]
        manifest = {
            "schema_version": WORKBOOK_INTEGRITY_SCHEMA_VERSION,
            "template_id": template_id,
            "template_hash": template_hash,
            "sheet_order": manifest_sheet_order,
            "sheets": [{"name": name, "hash": sign_payload(name)} for name in manifest_sheet_order],
        }
        manifest_text = json.dumps(manifest, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
        manifest_hash = sign_payload(manifest_text)
        layout_sheet = output_wb.create_sheet(SYSTEM_LAYOUT_SHEET)
        layout_sheet.cell(row=1, column=1, value=SYSTEM_LAYOUT_MANIFEST_HEADER)
        layout_sheet.cell(row=1, column=2, value=SYSTEM_LAYOUT_MANIFEST_HASH_HEADER)
        layout_sheet.cell(row=2, column=1, value=manifest_text)
        layout_sheet.cell(row=2, column=2, value=manifest_hash)
        layout_sheet.sheet_state = "hidden"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_wb.save(output_path)
        output_wb.close()
        return output_path
    finally:
        source_wb.close()


def extract_total_outcomes_from_workbook_path(path: Path) -> int | None:
    """Extract total outcomes from workbook path.
    
    Args:
        path: Parameter value (Path).
    
    Returns:
        int | None: Return value.
    
    Raises:
        None.
    """
    try:
        import openpyxl
    except Exception:
        return None
    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return None
    try:
        try:
            from domain.template_strategy_router import (
                read_template_id_from_system_hash_sheet_if_valid,
            )

            template_id = read_template_id_from_system_hash_sheet_if_valid(workbook) or ID_COURSE_SETUP
        except Exception:
            template_id = ID_COURSE_SETUP
        metadata_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
        if metadata_sheet_name not in workbook.sheetnames:
            return None
        metadata = _extract_course_metadata_fields(workbook[metadata_sheet_name])
        parsed = coerce_excel_number(metadata.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY), ""))
        if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
            value = int(parsed)
            return value if value > 0 else None
    finally:
        workbook.close()
    return None


def _extract_students_from_report_sheets_for_template(workbook: Any, *, template_id: str) -> set[str]:
    """Extract students from report sheets for template.
    
    Args:
        workbook: Parameter value (Any).
        template_id: Parameter value (str).
    
    Returns:
        set[str]: Return value.
    
    Raises:
        None.
    """
    unique_students: set[str] = set()
    students_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)
    accepted_reg_headers = {
        normalize(CO_REPORT_HEADER_REG_NO),
        normalize(students_headers[0]),
    }
    try:
        sheets = getattr(workbook, "worksheets", [])
    except Exception:
        return unique_students
    for sheet in sheets:
        if sheet is None:
            continue
        title = str(getattr(sheet, "title", "") or "")
        if title in {SYSTEM_HASH_SHEET, SYSTEM_LAYOUT_SHEET, COURSE_METADATA_SHEET}:
            continue
        max_row = int(getattr(sheet, "max_row", 0) or 0)
        max_col = int(getattr(sheet, "max_column", 0) or 0)
        if max_row <= 0 or max_col <= 0:
            continue
        scan_rows = min(max_row, 30)
        scan_cols = min(max_col, 80)
        reg_col: int | None = None
        header_row = 0
        for row in range(1, scan_rows + 1):
            for col in range(1, scan_cols + 1):
                key = normalize(sheet.cell(row=row, column=col).value)
                if key in accepted_reg_headers:
                    reg_col = col
                    header_row = row
                    break
            if reg_col is not None:
                break
        if reg_col is None:
            continue
        for row in range(header_row + 1, max_row + 1):
            reg_raw = sheet.cell(row=row, column=reg_col).value
            coerced = coerce_excel_number(reg_raw)
            reg_text = str(coerced).strip() if coerced is not None else ""
            if reg_text:
                unique_students.add(normalize(reg_text))
    return unique_students


def extract_course_metadata_and_students_from_workbook_path(path: Path) -> tuple[set[str], dict[str, str]]:
    """Extract course metadata and students from workbook path.
    
    Args:
        path: Parameter value (Path).
    
    Returns:
        tuple[set[str], dict[str, str]]: Return value.
    
    Raises:
        None.
    """
    try:
        import openpyxl
    except Exception:
        return set(), {}
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return set(), {}
    unique_students: set[str] = set()
    metadata_map: dict[str, str] = {}
    try:
        template_id = ID_COURSE_SETUP
        try:
            template_id = read_valid_template_id_from_system_hash_sheet(workbook)
        except Exception:
            fallback_id = read_template_id_from_system_hash_sheet_if_valid(workbook)
            if isinstance(fallback_id, str) and fallback_id.strip():
                template_id = fallback_id
        metadata_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
        students_sheet_name = get_sheet_name_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)
        students_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_STUDENTS)

        if metadata_sheet_name in workbook.sheetnames:
            metadata_map = _extract_course_metadata_fields(workbook[metadata_sheet_name])

        if students_sheet_name in workbook.sheetnames:
            students_sheet = workbook[students_sheet_name]
            sheet_specs_by_name = _layout_sheet_specs_by_name(workbook)
            header_row = _students_sheet_header_row(
                sheet_specs_by_name=sheet_specs_by_name,
                sheet_name=students_sheet_name,
            )
            header_map = _header_map_for_row(students_sheet, header_row=header_row)
            reg_col = _resolve_students_reg_col(header_map=header_map, students_headers=students_headers)
            if reg_col is not None:
                for reg_no, _student_name in _iter_students_sheet_rows(
                    students_sheet=students_sheet,
                    header_row=header_row,
                    reg_col=reg_col,
                    name_col=3,
                ):
                    unique_students.add(normalize(reg_no))
        if not unique_students:
            unique_students = _extract_students_from_report_sheets_for_template(workbook, template_id=template_id)
    finally:
        workbook.close()
    return unique_students, metadata_map


__all__ = [
    "FinalReportWorkbookSignature",
    "extract_course_metadata_and_students_from_workbook_path",
    "extract_final_report_signature_from_path",
    "extract_total_outcomes_from_workbook_path",
    "generate_final_report_workbook",
    "generate_co_attainment_workbook",
]
