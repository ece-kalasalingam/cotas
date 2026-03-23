"""Processing helpers for the coordinator module."""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import sqlite3
import tempfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from common.constants import (
    APP_NAME,
    CO_ATTAINMENT_LEVEL_DEFAULT,
    CO_ATTAINMENT_PERCENT_DEFAULT,
    CO_REPORT_ABSENT_TOKEN,
    CO_REPORT_HEADER_REG_NO,
    CO_REPORT_HEADER_SERIAL,
    CO_REPORT_HEADER_STUDENT_NAME,
    CO_REPORT_HEADER_TOTAL_RATIO_TEMPLATE,
    CO_REPORT_MAX_DECIMAL_PLACES,
    CO_REPORT_NOT_APPLICABLE_TOKEN,
    DIRECT_RATIO,
    ID_COURSE_SETUP,
    INDIRECT_RATIO,
    LEVEL_1_THRESHOLD,
    LEVEL_2_THRESHOLD,
    LEVEL_3_THRESHOLD,
    SYSTEM_LAYOUT_MANIFEST_HASH_HEADER,
    SYSTEM_LAYOUT_MANIFEST_HEADER,
    SYSTEM_LAYOUT_SHEET,
    WORKBOOK_INTEGRITY_SCHEMA_VERSION,
)
from common.registry import (
    COURSE_METADATA_ACADEMIC_YEAR_KEY,
    COURSE_METADATA_COURSE_CODE_KEY,
    COURSE_METADATA_SECTION_KEY,
    COURSE_METADATA_SEMESTER_KEY,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    SYSTEM_HASH_HEADER_TEMPLATE_HASH as SYSTEM_HASH_TEMPLATE_HASH_HEADER,
    SYSTEM_HASH_HEADER_TEMPLATE_ID as SYSTEM_HASH_TEMPLATE_ID_HEADER,
    SYSTEM_HASH_SHEET_NAME as SYSTEM_HASH_SHEET,
    get_sheet_headers_by_key,
    get_sheet_name_by_key,
)
from common.excel_sheet_layout import (
    apply_xlsxwriter_layout as _apply_xlsxwriter_layout_shared,
    build_template_xlsxwriter_formats as _build_template_xlsxwriter_formats,
    compute_sampled_column_widths as _compute_sampled_column_widths,
)
from common.error_catalog import validation_error_from_key
from common.exceptions import AppSystemError, ValidationError
from common.jobs import CancellationToken
from common.utils import (
    app_runtime_storage_dir,
    canonical_path_key,
    coerce_excel_number,
    create_app_runtime_sqlite_file,
    normalize,
)
from common.workbook_signing import sign_payload, verify_payload_signature
from domain.co_report_sheet_generator import co_direct_sheet_name, co_indirect_sheet_name
from domain.template_strategy_router import (
    FinalReportWorkbookSignature,
    extract_final_report_signature_from_path,
    generate_workbook,
    get_template_strategy,
    read_course_metadata_signature,
    read_layout_manifest_co_sheet_counts,
    read_template_id_from_system_hash_sheet_if_valid,
)

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}

_logger = logging.getLogger(__name__)
_COURSE_METADATA_COURSE_NAME_KEY = "course_name"
_STYLE_CACHE_ATTR = "_focus_coordinator_style_cache"
_CO_REPORT_NAME_TOKEN_RE = re.compile(r"(?:[_\-\s]*co[_\-\s]*report)+$", re.IGNORECASE)
_SEMESTER_PREFIX_TOKEN_RE = re.compile(r"^sem(?:ester)?[\s\-_]*([0-9]{1,2}|[ivxlcdm]+)$", re.IGNORECASE)
_SEMESTER_VALUE_TOKEN_RE = re.compile(r"^(?:[0-9]{1,2}|[ivxlcdm]+)$", re.IGNORECASE)
_HEADER_SCAN_MAX_ROWS = 200
_COORDINATOR_STUDENTS_PER_SHEET = 150
_DEDUP_SQLITE_THRESHOLD_ENTRIES = 10_000
_DEDUP_SQLITE_PREFIX = "focus_co_dedup_"
_DEDUP_SQLITE_SUFFIX = ".sqlite3"
_SIGNATURE_VALIDATION_MAX_WORKERS = 8


_FinalReportSignature = FinalReportWorkbookSignature
COURSE_METADATA_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)


def _course_metadata_headers(template_id: str) -> tuple[str, ...]:
    return get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)


@dataclass(slots=True, frozen=True)
class _CoAttainmentRow:
    reg_hash: int
    reg_no: str
    student_name: str
    direct_score: float | str
    indirect_score: float | str
    worksheet_name: str
    workbook_name: str


@dataclass(slots=True, frozen=True)
class _ParsedScoreRow:
    reg_hash: int
    reg_key: str
    reg_no: str
    student_name: str
    score: float | str


@dataclass(slots=True)
class _CoOutputSheetState:
    sheet: Any
    header_row_index: int
    formats: dict[str, Any]
    next_row_index: int
    next_serial: int
    on_roll: int
    attended: int
    level_counts: dict[int, int]


@dataclass(slots=True, frozen=True)
class _CoAttainmentWorkbookResult:
    output_path: Path
    duplicate_reg_count: int
    duplicate_entries: tuple[tuple[str, str, str], ...]
    inner_join_drop_count: int = 0
    inner_join_drop_details: tuple[str, ...] = ()


class _RegisterDedupStore:
    def __init__(self, *, total_outcomes: int, use_sqlite: bool) -> None:
        self._total_outcomes = total_outcomes
        self._use_sqlite = use_sqlite
        self._memory_sets: dict[int, set[int]] = {}
        self._conn: sqlite3.Connection | None = None
        self._db_path: str | None = None
        if use_sqlite:
            _cleanup_stale_dedup_sqlite_files()
            fd, db_path = create_app_runtime_sqlite_file(
                APP_NAME,
                prefix=_DEDUP_SQLITE_PREFIX,
                suffix=_DEDUP_SQLITE_SUFFIX,
            )
            self._db_path = db_path
            self._conn = sqlite3.connect(db_path)
            self._conn.execute("PRAGMA journal_mode=OFF")
            self._conn.execute("PRAGMA synchronous=OFF")
            self._conn.execute("PRAGMA temp_store=MEMORY")
            self._conn.execute("PRAGMA secure_delete=ON")
            self._conn.execute(
                "CREATE TABLE IF NOT EXISTS dedup (co_index INTEGER NOT NULL, reg_hash INTEGER NOT NULL, PRIMARY KEY(co_index, reg_hash))"
            )
            self._conn.commit()
            try:
                import os

                os.close(fd)
            except OSError:
                pass
        else:
            self._memory_sets = {co: set() for co in range(1, total_outcomes + 1)}

    def add_if_absent(self, *, co_index: int, reg_hash: int) -> bool:
        if self._use_sqlite and self._conn is not None:
            cursor = self._conn.execute(
                "INSERT OR IGNORE INTO dedup (co_index, reg_hash) VALUES (?, ?)",
                (co_index, int(reg_hash)),
            )
            return cursor.rowcount == 1
        bucket = self._memory_sets.setdefault(co_index, set())
        if reg_hash in bucket:
            return False
        bucket.add(reg_hash)
        return True

    def close(self) -> None:
        if self._conn is not None:
            try:
                # Explicitly wipe the sqlite table before close in normal paths.
                self._conn.execute("DELETE FROM dedup")
                self._conn.commit()
            except Exception:
                _logger.debug("Failed to clear sqlite dedup table before close.", exc_info=True)
            self._conn.close()
            self._conn = None
        if self._db_path:
            try:
                Path(self._db_path).unlink(missing_ok=True)
            except OSError:
                _logger.debug("Unable to remove dedup sqlite db: %s", self._db_path, exc_info=True)
            self._db_path = None


def _should_use_sqlite_dedup(*, source_count: int, total_outcomes: int) -> bool:
    estimated_entries = source_count * total_outcomes * _COORDINATOR_STUDENTS_PER_SHEET
    return estimated_entries >= _DEDUP_SQLITE_THRESHOLD_ENTRIES


def _cleanup_stale_dedup_sqlite_files() -> None:
    temp_root = Path(tempfile.gettempdir())
    storage_sqlite_root = app_runtime_storage_dir(APP_NAME) / "sqlite"
    roots: tuple[Path, ...] = (temp_root, storage_sqlite_root)
    patterns = (
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}",
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}-wal",
        f"{_DEDUP_SQLITE_PREFIX}*{_DEDUP_SQLITE_SUFFIX}-shm",
    )
    for root in roots:
        if not root.exists():
            continue
        for pattern in patterns:
            for path in root.glob(pattern):
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    _logger.debug("Unable to remove stale dedup sqlite file: %s", path, exc_info=True)


def _build_co_attainment_default_name(source_path: Path, *, section: str = "") -> str:
    stem = source_path.stem.strip()
    cleaned = _CO_REPORT_NAME_TOKEN_RE.sub("", stem).rstrip("_- ").strip()
    section_token = section.strip()
    if section_token:
        parts = [part for part in cleaned.split("_") if normalize(part) != normalize(section_token)]
        cleaned = "_".join(parts).strip("_- ")
    tokens = [token.strip() for token in cleaned.split("_") if token.strip()]
    filtered_tokens: list[str] = []
    semester_removed = False
    for index, token in enumerate(tokens):
        normalized_token = normalize(token)
        if _SEMESTER_PREFIX_TOKEN_RE.match(normalized_token):
            semester_removed = True
            continue
        if (
            not semester_removed
            and index > 0
            and _SEMESTER_VALUE_TOKEN_RE.match(normalized_token)
            and len(normalized_token) <= 4
        ):
            semester_removed = True
            continue
        filtered_tokens.append(token)
    cleaned = "_".join(filtered_tokens).strip("_- ")
    base = cleaned if cleaned else stem
    return f"{base}_CO_Attainment.xlsx"


def _is_supported_excel_file(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in EXCEL_SUFFIXES


def _filter_excel_paths(paths: Iterable[str]) -> list[Path]:
    collected: list[Path] = []
    seen: set[str] = set()
    for value in paths:
        path = Path(value)
        if not _is_supported_excel_file(path):
            continue
        key = canonical_path_key(path)
        if key in seen:
            continue
        seen.add(key)
        collected.append(path.resolve())
    return collected


def _has_valid_final_co_report(path: Path) -> bool:
    return _extract_final_report_signature(path) is not None


def _read_template_id_from_hash_sheet(workbook: Any) -> str | None:
    return read_template_id_from_system_hash_sheet_if_valid(
        workbook,
        verify_signature=verify_payload_signature,
    )


def _read_report_sheet_counts(workbook: Any) -> tuple[int, int] | None:
    return read_layout_manifest_co_sheet_counts(
        workbook,
        verify_signature=verify_payload_signature,
    )


def _read_signature_metadata(workbook: Any) -> tuple[str, int, str] | None:
    return read_course_metadata_signature(
        workbook,
        course_code_key=COURSE_METADATA_COURSE_CODE_KEY,
        total_outcomes_key=COURSE_METADATA_TOTAL_OUTCOMES_KEY,
        section_key=COURSE_METADATA_SECTION_KEY,
    )


def _extract_final_report_signature(path: Path) -> _FinalReportSignature | None:
    return extract_final_report_signature_from_path(
        path,
        verify_signature=verify_payload_signature,
    )


def _analyze_dropped_files(
    dropped_files: list[str],
    *,
    existing_keys: set[str],
    existing_paths: list[str],
    token: CancellationToken,
) -> dict[str, object]:
    accepted = _filter_excel_paths(dropped_files)
    seen = set(existing_keys)
    added: list[str] = []
    duplicates = 0
    invalid_final_report: list[str] = []
    invalid_final_report_details: list[dict[str, str]] = []
    existing_resolved = [Path(path).resolve() for path in existing_paths if path]
    baseline_signature: _FinalReportSignature | None = None
    seen_sections: set[str] = set()
    signature_cache: dict[str, _FinalReportSignature | None] = {}

    unique_paths: list[Path] = []
    seen_signature_keys: set[str] = set()
    for path in [*existing_resolved, *accepted]:
        key = canonical_path_key(path)
        if key in seen_signature_keys:
            continue
        seen_signature_keys.add(key)
        unique_paths.append(path)

    max_workers = min(
        _SIGNATURE_VALIDATION_MAX_WORKERS,
        max(1, os.cpu_count() or 1),
        len(unique_paths) or 1,
    )
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_pairs = [(path, executor.submit(_extract_final_report_signature, path)) for path in unique_paths]
        for path, future in future_pairs:
            token.raise_if_cancelled()
            signature_cache[canonical_path_key(path)] = future.result()

    def _cached_signature(path: Path) -> _FinalReportSignature | None:
        return signature_cache.get(canonical_path_key(path))

    for path in existing_resolved:
        token.raise_if_cancelled()
        signature = _cached_signature(path)
        if signature is None:
            continue
        if baseline_signature is None:
            baseline_signature = signature
        section_key = normalize(signature.section)
        if section_key:
            seen_sections.add(section_key)

    for path in accepted:
        token.raise_if_cancelled()
        key = canonical_path_key(path)
        if key in seen:
            duplicates += 1
            continue
        signature = _cached_signature(path)
        if signature is None:
            invalid_final_report.append(str(path))
            invalid_final_report_details.append(
                {
                    "path": str(path),
                    "reason": "Invalid final CO report workbook (signature/manifest/metadata validation failed).",
                }
            )
            continue
        if baseline_signature is None:
            baseline_signature = signature
        else:
            is_mismatch = (
                signature.template_id != baseline_signature.template_id
                or signature.course_code != baseline_signature.course_code
                or signature.total_outcomes != baseline_signature.total_outcomes
                or signature.direct_sheet_count != baseline_signature.direct_sheet_count
                or signature.indirect_sheet_count != baseline_signature.indirect_sheet_count
            )
            section_already_seen = normalize(signature.section) in seen_sections
            if is_mismatch or section_already_seen:
                invalid_final_report.append(str(path))
                if is_mismatch:
                    reason = (
                        "File does not match selected batch (template/course/CO-sheet structure mismatch)."
                    )
                else:
                    reason = f"Duplicate section '{signature.section}' is not allowed in one batch."
                invalid_final_report_details.append(
                    {
                        "path": str(path),
                        "reason": reason,
                    }
                )
                continue
        seen_sections.add(normalize(signature.section))
        seen.add(key)
        added.append(str(path))

    ignored = (len(dropped_files) - len(accepted)) + duplicates + len(invalid_final_report)
    return {
        "added": added,
        "duplicates": duplicates,
        "invalid_final_report": invalid_final_report,
        "invalid_final_report_details": invalid_final_report_details,
        "ignored": ignored,
    }


def _ratio_percent_token(ratio: float) -> str:
    percent = ratio * 100.0
    if abs(percent - round(percent)) <= 1e-9:
        return f"{int(round(percent))}"
    return f"{percent:g}"


def _ratio_total_header(ratio: float) -> str:
    return CO_REPORT_HEADER_TOTAL_RATIO_TEMPLATE.format(ratio=_ratio_percent_token(ratio))


def _coerce_numeric_score(value: Any) -> float | str | None:
    if normalize(value) == normalize(CO_REPORT_NOT_APPLICABLE_TOKEN):
        return CO_REPORT_ABSENT_TOKEN
    parsed = coerce_excel_number(value)
    if isinstance(parsed, bool):
        return None
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


def _metadata_rows_for_output(metadata: dict[str, str], co_index: int) -> list[tuple[str, str]]:
    return [
        ("Course Code", metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")),
        ("Course Name", metadata.get(normalize(_COURSE_METADATA_COURSE_NAME_KEY), "")),
        ("Semester", metadata.get(normalize(COURSE_METADATA_SEMESTER_KEY), "")),
        ("Academic Year", metadata.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")),
        ("CO Number", f"CO{co_index}"),
    ]


def _metadata_rows_for_summary_graph(
    metadata: dict[str, str],
    *,
    total_outcomes: int,
) -> list[tuple[str, str]]:
    total_outcomes_value = metadata.get(normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY), "").strip()
    if not total_outcomes_value:
        total_outcomes_value = str(total_outcomes)
    return [
        ("Course Code", metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "")),
        ("Course Name", metadata.get(normalize(_COURSE_METADATA_COURSE_NAME_KEY), "")),
        ("Academic Year", metadata.get(normalize(COURSE_METADATA_ACADEMIC_YEAR_KEY), "")),
        ("Total Outcomes", total_outcomes_value),
    ]


def _threshold_rows_for_output(
    thresholds: tuple[float, float, float] | None,
    *,
    co_attainment_percent: float,
    co_attainment_level: int,
    include: bool,
) -> list[tuple[str, str]]:
    if not include or thresholds is None:
        return []
    l1, l2, l3 = thresholds
    return [
        ("Threshold L1", f"{l1:g}"),
        ("Threshold L2", f"{l2:g}"),
        ("Threshold L3", f"{l3:g}"),
        ("CO AT%", f"{co_attainment_percent:g}"),
        ("CO AT Level", f"L{co_attainment_level}"),
    ]


def _extract_course_metadata_fields(sheet: Any) -> dict[str, str]:
    metadata: dict[str, str] = {}
    row = 2
    while True:
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        key_text = str(key).strip() if key is not None else ""
        if key_text:
            coerced = coerce_excel_number(value)
            metadata[normalize(key_text)] = str(coerced).strip() if coerced is not None else ""
        row += 1
    return metadata


def _stable_reg_hash(reg_key: str) -> int:
    # Use a compact, stable 48-bit ID to reduce sqlite footprint while keeping collision risk very low.
    return int.from_bytes(hashlib.blake2b(reg_key.encode("utf-8"), digest_size=6).digest(), byteorder="big")


def _iter_score_rows(sheet: Any, *, ratio: float) -> Iterable[_ParsedScoreRow]:
    required_headers = {
        normalize(CO_REPORT_HEADER_SERIAL),
        normalize(CO_REPORT_HEADER_REG_NO),
        normalize(CO_REPORT_HEADER_STUDENT_NAME),
        normalize(_ratio_total_header(ratio)),
    }
    header_row = 0
    column_map: dict[str, int] = {}
    max_row = int(sheet.max_row)
    max_col = int(sheet.max_column)
    header_scan_limit = min(max_row, _HEADER_SCAN_MAX_ROWS)

    def _find_header_row(start_row: int, end_row: int) -> tuple[int, dict[str, int]]:
        for row_offset, values in enumerate(
            sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=1,
                max_col=max_col,
                values_only=True,
            ),
            start=0,
        ):
            row_idx = start_row + row_offset
            row_map: dict[str, int] = {}
            missing = set(required_headers)
            for col_idx, value in enumerate(values, start=1):
                key = normalize(value)
                if key and key not in row_map:
                    row_map[key] = col_idx
                    if key in missing:
                        missing.remove(key)
                        if not missing:
                            return row_idx, row_map
        return 0, {}

    header_row, column_map = _find_header_row(1, header_scan_limit)
    if header_row <= 0 and header_scan_limit < max_row:
        header_row, column_map = _find_header_row(header_scan_limit + 1, max_row)
    if header_row <= 0:
        raise validation_error_from_key(
            "validation.layout.header_mismatch",
            code="COA_LAYOUT_HEADER_MISMATCH",
            sheet_name=str(getattr(sheet, "title", "") or ""),
        )

    reg_col = column_map[normalize(CO_REPORT_HEADER_REG_NO)]
    name_col = column_map[normalize(CO_REPORT_HEADER_STUDENT_NAME)]
    score_col = column_map[normalize(_ratio_total_header(ratio))]

    seen_in_sheet: set[int] = set()
    for values in sheet.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if len(values) < max(reg_col, name_col, score_col):
            continue
        reg_raw = values[reg_col - 1]
        reg_value = coerce_excel_number(reg_raw)
        reg_no = str(reg_value).strip() if reg_value is not None else ""
        if not reg_no:
            continue
        reg_key = normalize(reg_no)
        reg_hash = _stable_reg_hash(reg_key)
        if reg_hash in seen_in_sheet:
            continue

        score = _coerce_numeric_score(values[score_col - 1])
        if score is None:
            continue
        student_raw = values[name_col - 1]
        student_name = str(student_raw).strip() if student_raw is not None else ""
        normalized_score = round(score, CO_REPORT_MAX_DECIMAL_PLACES) if isinstance(score, (int, float)) else score
        seen_in_sheet.add(reg_hash)
        yield _ParsedScoreRow(
            reg_hash=reg_hash,
            reg_key=reg_key,
            reg_no=reg_no,
            student_name=student_name,
            score=normalized_score,
        )


def _iter_co_rows_from_workbook(workbook: Any, *, co_index: int, workbook_name: str) -> Iterable[_CoAttainmentRow]:
    direct_name = co_direct_sheet_name(co_index)
    indirect_name = co_indirect_sheet_name(co_index)
    if direct_name not in workbook.sheetnames or indirect_name not in workbook.sheetnames:
        raise validation_error_from_key(
            "validation.layout.sheet_missing",
            code="COA_LAYOUT_SHEET_MISSING",
            sheet_name=f"CO{co_index}",
        )

    direct_lookup: dict[tuple[int, str], _ParsedScoreRow] = {}
    for item in _iter_score_rows(workbook[direct_name], ratio=DIRECT_RATIO):
        direct_lookup.setdefault((item.reg_hash, item.reg_key), item)

    indirect_lookup: dict[tuple[int, str], _ParsedScoreRow] = {}
    for item in _iter_score_rows(workbook[indirect_name], ratio=INDIRECT_RATIO):
        indirect_lookup.setdefault((item.reg_hash, item.reg_key), item)

    direct_only = [row.reg_no for key, row in direct_lookup.items() if key not in indirect_lookup]
    indirect_only = [row.reg_no for key, row in indirect_lookup.items() if key not in direct_lookup]
    if direct_only or indirect_only:
        dropped_count = len(direct_only) + len(indirect_only)
        direct_preview = ", ".join(direct_only[:5]) if direct_only else "-"
        indirect_preview = ", ".join(indirect_only[:5]) if indirect_only else "-"
        _logger.warning(
            "CO join dropped unmatched students. workbook=%s, co_index=%s, dropped=%s, direct_only=%s, indirect_only=%s",
            workbook_name,
            co_index,
            dropped_count,
            direct_preview,
            indirect_preview,
        )

    for key, direct_row in direct_lookup.items():
        match = indirect_lookup.get(key)
        if match is None:
            continue
        student_name = direct_row.student_name or match.student_name
        yield _CoAttainmentRow(
            reg_hash=direct_row.reg_hash,
            reg_no=direct_row.reg_no,
            student_name=student_name,
            direct_score=direct_row.score,
            indirect_score=match.score,
            worksheet_name=direct_name,
            workbook_name=workbook_name,
        )


def _xlsxwriter_formats(workbook: Any, *, template_id: str) -> dict[str, Any]:
    return _build_template_xlsxwriter_formats(
        workbook,
        template_id=template_id,
        cache_attr=_STYLE_CACHE_ATTR,
        include_column_wrap=True,
        normalize_header_valign_to_center=False,
    )


def _create_co_attainment_sheet(
    workbook: Any,
    *,
    template_id: str,
    co_index: int,
    metadata: dict[str, str],
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float,
    co_attainment_level: int,
) -> _CoOutputSheetState:
    sheet = workbook.add_worksheet(f"CO{co_index}")
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_headers = _course_metadata_headers(template_id)
    metadata_rows = _metadata_rows_for_output(metadata, co_index)
    metadata_rows.extend(
        _threshold_rows_for_output(
            thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
            include=bool(metadata),
        )
    )
    for row_idx, (label, value) in enumerate(metadata_rows, start=0):
        sheet.write(row_idx, 1, label, formats["body"])
        sheet.write(row_idx, 2, value, formats["body_wrap"])

    header_row_index = len(metadata_rows) + 1
    headers = [
        "#",
        "Regno",
        "Student name",
        f"Direct ({_ratio_percent_token(DIRECT_RATIO)}%)",
        f"Indirect ({_ratio_percent_token(INDIRECT_RATIO)}%)",
        "Total (100%)",
        "Level",
    ]
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, col_idx, header, formats["header"])

    sampled_rows: list[list[Any]] = [["", metadata_headers[0], metadata_headers[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    sampled_rows.append(["", headers[1], headers[2]])
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    sheet.set_column(1, 1, widths.get(1, 8))
    sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    _apply_xlsxwriter_layout_shared(
        sheet,
        header_row_index=header_row_index,
        paper_size=9,
        landscape=True,
    )
    sheet.repeat_rows(0, header_row_index)
    # Freeze rows through the header and columns through Student name (A:C).
    sheet.freeze_panes(header_row_index + 1, 3)
    return _CoOutputSheetState(
        sheet=sheet,
        header_row_index=header_row_index,
        formats=formats,
        next_row_index=header_row_index + 1,
        next_serial=1,
        on_roll=0,
        attended=0,
        level_counts={level: 0 for level in range(0, (len(thresholds) if thresholds is not None else 3) + 1)},
    )


def _append_co_attainment_row(
    state: _CoOutputSheetState,
    row: _CoAttainmentRow,
    *,
    thresholds: tuple[float, float, float],
) -> None:
    if isinstance(row.direct_score, (int, float)) and isinstance(row.indirect_score, (int, float)):
        total: float | str = round(row.direct_score + row.indirect_score, CO_REPORT_MAX_DECIMAL_PLACES)
    else:
        total = CO_REPORT_ABSENT_TOKEN
    level = _score_to_attainment_level(total, thresholds=thresholds)
    left = [state.next_serial, row.reg_no]
    right = [row.direct_score, row.indirect_score, total, level]
    state.sheet.write_row(state.next_row_index, 0, left, state.formats["body"])
    state.sheet.write(state.next_row_index, 2, row.student_name, state.formats["body_wrap"])
    state.sheet.write_row(state.next_row_index, 3, right, state.formats["body_center"])
    state.on_roll += 1
    if total != CO_REPORT_ABSENT_TOKEN:
        state.attended += 1
    if isinstance(level, int) and level in state.level_counts:
        state.level_counts[level] += 1
    state.next_row_index += 1
    state.next_serial += 1


def _append_co_attainment_summary(state: _CoOutputSheetState) -> None:
    # Keep one visual spacer row after data and then write label/value summary rows.
    state.next_row_index += 1
    summary_rows: list[tuple[str, int]] = [
        ("On Roll:", state.on_roll),
        ("Attended:", state.attended),
    ]
    summary_rows.extend(
        (f"Level {level}:", state.level_counts.get(level, 0))
        for level in sorted(state.level_counts)
    )
    for label, value in summary_rows:
        state.sheet.write(state.next_row_index, 1, label, state.formats["body"])
        state.sheet.write(state.next_row_index, 2, value, state.formats["body_center"])
        state.next_row_index += 1


def _co_percentage(
    *,
    level_counts: dict[int, int],
    attended: int,
    co_attainment_level: int,
) -> float | str:
    if attended <= 0:
        return CO_REPORT_NOT_APPLICABLE_TOKEN
    attained_count = sum(
        int(count)
        for level, count in level_counts.items()
        if isinstance(level, int) and level >= co_attainment_level
    )
    return round((attained_count / float(attended)) * 100.0, CO_REPORT_MAX_DECIMAL_PLACES)


def _create_summary_sheet(
    workbook: Any,
    *,
    template_id: str,
    metadata: dict[str, str],
    thresholds: tuple[float, float, float] | None,
    co_attainment_percent: float,
    co_attainment_level: int,
    output_states: dict[int, _CoOutputSheetState],
    total_outcomes: int,
) -> tuple[int, int]:
    sheet = workbook.add_worksheet("Summary")
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_headers = _course_metadata_headers(template_id)
    metadata_rows = _metadata_rows_for_summary_graph(metadata, total_outcomes=total_outcomes)
    metadata_rows.extend(
        _threshold_rows_for_output(
            thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
            include=bool(metadata),
        )
    )
    for row_idx, (label, value) in enumerate(metadata_rows, start=0):
        sheet.write(row_idx, 1, label, formats["body"])
        sheet.write(row_idx, 2, value, formats["body_wrap"])

    header_row_index = len(metadata_rows) + 1
    max_level = max((max(state.level_counts) for state in output_states.values()), default=3)
    level_headers = [f"Level {level}" for level in range(0, max_level + 1)]
    headers = ["CO", "Attended", *level_headers, "CO%", "Result"]
    table_start_col = 1  # Start summary table at column B.
    for col_idx, header in enumerate(headers, start=0):
        sheet.write(header_row_index, table_start_col + col_idx, header, formats["header"])

    first_data_row = header_row_index + 1
    course_code = metadata.get(normalize(COURSE_METADATA_COURSE_CODE_KEY), "").strip()
    for co_index in range(1, total_outcomes + 1):
        state = output_states.get(co_index)
        level_counts = state.level_counts if state is not None else {level: 0 for level in range(0, max_level + 1)}
        attended = state.attended if state is not None else 0
        co_label = f"{course_code}.{co_index}" if course_code else f"CO{co_index}"
        co_percentage = _co_percentage(
            level_counts=level_counts,
            attended=attended,
            co_attainment_level=co_attainment_level,
        )
        result_text = (
            "Attained"
            if isinstance(co_percentage, (int, float)) and float(co_percentage) >= float(co_attainment_percent)
            else "Yet to Attain"
        )
        row_values: list[Any] = [
            co_label,
            attended,
            *[level_counts.get(level, 0) for level in range(0, max_level + 1)],
            co_percentage,
            result_text,
        ]
        row_index = header_row_index + co_index
        sheet.write_row(row_index, table_start_col, row_values, formats["body_center"])

    sampled_rows: list[list[Any]] = [["", metadata_headers[0], metadata_headers[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    sampled_rows.append(["", headers[0], headers[1]])
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    sheet.set_column(1, 1, widths.get(1, 8))
    sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])
    result_col = table_start_col + (len(headers) - 1)
    result_sample_rows: list[list[Any]] = [
        [headers[-1]],
        ["Attained"],
        ["Yet to Attain"],
    ]
    result_widths = _compute_sampled_column_widths(result_sample_rows, 0)
    sheet.set_column(result_col, result_col, result_widths.get(0, 12), formats["column_wrap"])

    _apply_xlsxwriter_layout_shared(
        sheet,
        header_row_index=header_row_index,
        paper_size=9,
        landscape=True,
        selection_col=table_start_col,
    )
    sheet.repeat_rows(0, header_row_index)
    # Match CO-sheet freeze behavior for header rows and left reference columns.
    sheet.freeze_panes(header_row_index + 1, 3)
    return first_data_row, first_data_row + max(0, total_outcomes - 1)


def _create_graph_sheet(
    workbook: Any,
    *,
    template_id: str,
    metadata: dict[str, str],
    total_outcomes: int,
    thresholds: tuple[float, float, float] | None,
    co_attainment_percent: float,
    co_attainment_level: int,
    summary_first_data_row: int,
    summary_last_data_row: int,
) -> None:
    graph_sheet = workbook.add_worksheet("Graph")
    formats = _xlsxwriter_formats(workbook, template_id=template_id)
    metadata_headers = _course_metadata_headers(template_id)
    metadata_rows = _metadata_rows_for_summary_graph(metadata, total_outcomes=total_outcomes)
    metadata_rows.extend(
        _threshold_rows_for_output(
            thresholds,
            co_attainment_percent=co_attainment_percent,
            co_attainment_level=co_attainment_level,
            include=bool(metadata),
        )
    )
    for row_idx, (label, value) in enumerate(metadata_rows, start=0):
        graph_sheet.write(row_idx, 1, label, formats["body"])
        graph_sheet.write(row_idx, 2, value, formats["body_wrap"])
    sampled_rows: list[list[Any]] = [["", metadata_headers[0], metadata_headers[1]]]
    sampled_rows.extend(["", field, value] for field, value in metadata_rows)
    widths = _compute_sampled_column_widths(sampled_rows, 2)
    graph_sheet.set_column(1, 1, widths.get(1, 8))
    graph_sheet.set_column(2, 2, widths.get(2, 8), formats["column_wrap"])

    chart = workbook.add_chart({"type": "column"})
    summary_table_start_col = 1
    level_count = (len(thresholds) if thresholds is not None else 3) + 1
    summary_co_percent_col = summary_table_start_col + 2 + level_count
    chart.add_series(
        {
            "name": "CO%",
            "categories": [
                "Summary",
                summary_first_data_row,
                summary_table_start_col,
                summary_last_data_row,
                summary_table_start_col,
            ],
            "values": [
                "Summary",
                summary_first_data_row,
                summary_co_percent_col,
                summary_last_data_row,
                summary_co_percent_col,
            ],
            "data_labels": {"value": True},
        }
    )
    x_axis_name = "CO"
    y_axis_name = "Attinment %"
    chart.set_title({"name": f"{x_axis_name} {y_axis_name}"})
    chart.set_x_axis({"name": x_axis_name})
    chart.set_y_axis({"name": y_axis_name, "min": 0, "max": 100, "major_unit": 10})
    chart.set_legend({"none": True})
    chart_anchor_row = len(metadata_rows) + 2
    graph_sheet.insert_chart(f"B{chart_anchor_row + 1}", chart, {"x_scale": 1.4, "y_scale": 1.4})
    graph_header_row = max(0, len(metadata_rows) - 1)
    _apply_xlsxwriter_layout_shared(
        graph_sheet,
        header_row_index=graph_header_row,
        paper_size=9,
        landscape=True,
    )
    graph_sheet.repeat_rows(0, graph_header_row)


def _add_system_hash_sheet(workbook: Any, template_id: str) -> None:
    template_hash = sign_payload(template_id)
    hash_ws = workbook.add_worksheet(SYSTEM_HASH_SHEET)
    hash_ws.write(0, 0, SYSTEM_HASH_TEMPLATE_ID_HEADER)
    hash_ws.write(0, 1, SYSTEM_HASH_TEMPLATE_HASH_HEADER)
    hash_ws.write(1, 0, template_id)
    hash_ws.write(1, 1, template_hash)
    hash_ws.hide()


def _add_system_layout_sheet(workbook: Any, manifest_text: str, manifest_hash: str) -> None:
    layout_ws = workbook.add_worksheet(SYSTEM_LAYOUT_SHEET)
    layout_ws.write(0, 0, SYSTEM_LAYOUT_MANIFEST_HEADER)
    layout_ws.write(0, 1, SYSTEM_LAYOUT_MANIFEST_HASH_HEADER)
    layout_ws.write(1, 0, manifest_text)
    layout_ws.write(1, 1, manifest_hash)
    layout_ws.hide()


def _build_system_layout_manifest(
    *,
    template_id: str,
    sheet_order: list[str],
) -> tuple[str, str]:
    template_hash = sign_payload(template_id)
    signed_sheet_order = [*sheet_order, SYSTEM_HASH_SHEET]
    manifest = {
        "schema_version": WORKBOOK_INTEGRITY_SCHEMA_VERSION,
        "template_id": template_id,
        "template_hash": template_hash,
        "sheet_order": signed_sheet_order,
        "sheets": [{"name": name, "hash": sign_payload(name)} for name in signed_sheet_order],
    }
    manifest_text = json.dumps(manifest, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    manifest_hash = sign_payload(manifest_text)
    return manifest_text, manifest_hash


def _attainment_thresholds(
    thresholds: tuple[float, float, float] | None = None,
) -> tuple[float, float, float]:
    if thresholds is None:
        return (float(LEVEL_1_THRESHOLD), float(LEVEL_2_THRESHOLD), float(LEVEL_3_THRESHOLD))
    l1, l2, l3 = thresholds
    return (float(l1), float(l2), float(l3))


def _co_attainment_target(
    *,
    co_attainment_percent: float | None,
    co_attainment_level: int | None,
    threshold_count: int,
) -> tuple[float, int]:
    percent = float(CO_ATTAINMENT_PERCENT_DEFAULT if co_attainment_percent is None else co_attainment_percent)
    level = int(CO_ATTAINMENT_LEVEL_DEFAULT if co_attainment_level is None else co_attainment_level)
    percent = max(0.0, min(100.0, percent))
    level = max(1, min(max(1, threshold_count), level))
    return percent, level


def _score_to_attainment_level(
    score: float | str,
    *,
    thresholds: tuple[float, float, float],
) -> int | str:
    if not isinstance(score, (int, float)) or isinstance(score, bool):
        return CO_REPORT_NOT_APPLICABLE_TOKEN

    total = float(score)
    tolerance = 10 ** (-CO_REPORT_MAX_DECIMAL_PLACES)
    if 100.0 < total <= (100.0 + tolerance):
        total = 100.0
    if (-tolerance) <= total < 0.0:
        total = 0.0
    sorted_thresholds = tuple(float(value) for value in thresholds)
    if not sorted_thresholds:
        return CO_REPORT_NOT_APPLICABLE_TOKEN
    if 0.0 <= total < sorted_thresholds[0]:
        return 0
    for index in range(1, len(sorted_thresholds)):
        if sorted_thresholds[index - 1] <= total < sorted_thresholds[index]:
            return index
    if sorted_thresholds[-1] <= total <= 100.0:
        return len(sorted_thresholds)
    return CO_REPORT_NOT_APPLICABLE_TOKEN


def _reg_no_sort_key(reg_no: str) -> tuple[tuple[int, int | str], ...]:
    tokens = re.split(r"(\d+)", reg_no.strip())
    key_parts: list[tuple[int, int | str]] = []
    for token in tokens:
        if not token:
            continue
        if token.isdigit():
            key_parts.append((0, int(token)))
        else:
            key_parts.append((1, token.casefold()))
    return tuple(key_parts)


def _generate_co_attainment_workbook(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float | None = None,
    co_attainment_level: int | None = None,
) -> _CoAttainmentWorkbookResult:
    if not source_paths:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COA_SOURCE_WORKBOOK_REQUIRED",
        )

    first_signature = _extract_final_report_signature(source_paths[0])
    if first_signature is None:
        raise validation_error_from_key(
            "validation.workbook.open_failed",
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(source_paths[0]),
        )
    try:
        get_template_strategy(first_signature.template_id)
    except ValidationError as exc:
        raise validation_error_from_key(
            "validation.template.unknown",
            code="UNKNOWN_TEMPLATE",
            template_id=first_signature.template_id,
        ) from exc
    return generate_workbook(
        template_id=first_signature.template_id,
        output_path=output_path,
        workbook_name=output_path.name,
        workbook_kind="co_attainment",
        cancel_token=token,
        context={
            "source_paths": [str(path) for path in source_paths],
            "thresholds": tuple(thresholds) if thresholds is not None else None,
            "co_attainment_percent": co_attainment_percent,
            "co_attainment_level": co_attainment_level,
        },
    )


def _generate_co_attainment_workbook_course_setup_v1(
    source_paths: list[Path],
    output_path: Path,
    *,
    token: CancellationToken,
    total_outcomes: int | None = None,
    template_id: str = ID_COURSE_SETUP,
    thresholds: tuple[float, float, float] | None = None,
    co_attainment_percent: float | None = None,
    co_attainment_level: int | None = None,
) -> _CoAttainmentWorkbookResult:
    try:
        import xlsxwriter
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - guarded by runtime dependency availability
        raise AppSystemError("openpyxl and xlsxwriter are required for CO attainment calculation.") from exc

    resolved_total_outcomes = int(total_outcomes or 0)
    if resolved_total_outcomes <= 0:
        first_signature = _extract_final_report_signature(source_paths[0]) if source_paths else None
        if first_signature is not None:
            resolved_total_outcomes = int(first_signature.total_outcomes)

    if resolved_total_outcomes <= 0:
        raise validation_error_from_key(
            "common.validation_failed_invalid_data",
            code="COA_TOTAL_OUTCOMES_MISSING",
        )

    metadata: dict[str, str] = {}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_workbook = xlsxwriter.Workbook(str(output_path), {"constant_memory": True})
    workbook_closed = False
    output_states: dict[int, _CoOutputSheetState] = {}
    pending_rows: dict[int, list[_CoAttainmentRow]] = {}
    duplicate_reg_count = 0
    duplicate_entries: list[tuple[str, str, str]] = []
    inner_join_drop_count = 0
    inner_join_drop_details: list[str] = []
    level_thresholds = _attainment_thresholds(thresholds)
    target_percent, target_level = _co_attainment_target(
        co_attainment_percent=co_attainment_percent,
        co_attainment_level=co_attainment_level,
        threshold_count=len(level_thresholds),
    )
    dedup_store = _RegisterDedupStore(
        total_outcomes=resolved_total_outcomes,
        use_sqlite=_should_use_sqlite_dedup(
            source_count=len(source_paths),
            total_outcomes=resolved_total_outcomes,
        ),
    )

    try:
        for source in source_paths:
            token.raise_if_cancelled()
            workbook = load_workbook(filename=source, data_only=True, read_only=True)
            try:
                if not metadata and COURSE_METADATA_SHEET in workbook.sheetnames:
                    metadata = _extract_course_metadata_fields(workbook[COURSE_METADATA_SHEET])
                    for co_index in range(1, resolved_total_outcomes + 1):
                        output_states[co_index] = _create_co_attainment_sheet(
                            output_workbook,
                            template_id=template_id,
                            co_index=co_index,
                            metadata=metadata,
                            thresholds=level_thresholds,
                            co_attainment_percent=target_percent,
                            co_attainment_level=target_level,
                        )
                        pending_rows[co_index] = []
                if not output_states:
                    for co_index in range(1, resolved_total_outcomes + 1):
                        output_states[co_index] = _create_co_attainment_sheet(
                            output_workbook,
                            template_id=template_id,
                            co_index=co_index,
                            metadata=metadata,
                            thresholds=level_thresholds,
                            co_attainment_percent=target_percent,
                            co_attainment_level=target_level,
                        )
                        pending_rows[co_index] = []
                for co_index in range(1, resolved_total_outcomes + 1):
                    row_bucket = pending_rows.setdefault(co_index, [])
                    rows = list(_iter_co_rows_from_workbook(workbook, co_index=co_index, workbook_name=source.name))
                    direct_name = co_direct_sheet_name(co_index)
                    indirect_name = co_indirect_sheet_name(co_index)
                    direct_total = sum(1 for _ in _iter_score_rows(workbook[direct_name], ratio=DIRECT_RATIO))
                    indirect_total = sum(1 for _ in _iter_score_rows(workbook[indirect_name], ratio=INDIRECT_RATIO))
                    dropped_for_sheet = max(0, (direct_total + indirect_total) - (2 * len(rows)))
                    if dropped_for_sheet:
                        inner_join_drop_count += dropped_for_sheet
                        inner_join_drop_details.append(
                            f"{source.name} CO{co_index}: dropped={dropped_for_sheet} (direct_rows={direct_total}, indirect_rows={indirect_total})"
                        )
                    for row in rows:
                        if not dedup_store.add_if_absent(co_index=co_index, reg_hash=row.reg_hash):
                            duplicate_reg_count += 1
                            duplicate_entries.append((row.reg_no, row.worksheet_name, row.workbook_name))
                            continue
                        row_bucket.append(row)
            finally:
                workbook.close()
        for co_index, state in output_states.items():
            for row in sorted(
                pending_rows.get(co_index, []),
                key=lambda item: (_reg_no_sort_key(item.reg_no), item.reg_hash),
            ):
                _append_co_attainment_row(state, row, thresholds=level_thresholds)
            _append_co_attainment_summary(state)
        summary_first_data_row, summary_last_data_row = _create_summary_sheet(
            output_workbook,
            template_id=template_id,
            metadata=metadata,
            thresholds=level_thresholds,
            co_attainment_percent=target_percent,
            co_attainment_level=target_level,
            output_states=output_states,
            total_outcomes=resolved_total_outcomes,
        )
        _create_graph_sheet(
            output_workbook,
            template_id=template_id,
            metadata=metadata,
            total_outcomes=resolved_total_outcomes,
            thresholds=level_thresholds,
            co_attainment_percent=target_percent,
            co_attainment_level=target_level,
            summary_first_data_row=summary_first_data_row,
            summary_last_data_row=summary_last_data_row,
        )
        sheet_order = [f"CO{co_index}" for co_index in range(1, resolved_total_outcomes + 1)]
        sheet_order.extend(["Summary", "Graph"])
        _add_system_hash_sheet(output_workbook, template_id)
        manifest_text, manifest_hash = _build_system_layout_manifest(
            template_id=template_id,
            sheet_order=sheet_order,
        )
        _add_system_layout_sheet(output_workbook, manifest_text, manifest_hash)
        output_workbook.close()
        workbook_closed = True
    finally:
        dedup_store.close()
        if not workbook_closed:
            try:
                output_workbook.close()
            except Exception:
                _logger.debug("Suppressing workbook close error during cleanup.", exc_info=True)
    return _CoAttainmentWorkbookResult(
        output_path=output_path,
        duplicate_reg_count=duplicate_reg_count,
        duplicate_entries=tuple(duplicate_entries),
        inner_join_drop_count=inner_join_drop_count,
        inner_join_drop_details=tuple(inner_join_drop_details),
    )

