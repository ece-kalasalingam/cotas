"""Final CO report generation from validated filled-marks workbooks."""

from __future__ import annotations

import json
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from common.error_catalog import validation_error_from_key
from common.constants import (
    CO_REPORT_ABSENT_TOKEN,
    CO_REPORT_MAX_DECIMAL_PLACES,
    COMPONENT_NAME_LABEL,
    COURSE_METADATA_FACULTY_NAME_KEY,
    ID_COURSE_SETUP,
    COURSE_METADATA_TOTAL_OUTCOMES_KEY,
    LAYOUT_MANIFEST_KEY_SHEETS,
    LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
    LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
    LAYOUT_SHEET_KIND_INDIRECT,
    LAYOUT_SHEET_SPEC_KEY_ANCHORS,
    LAYOUT_SHEET_SPEC_KEY_HEADER_ROW,
    LAYOUT_SHEET_SPEC_KEY_HEADERS,
    LAYOUT_SHEET_SPEC_KEY_KIND,
    LAYOUT_SHEET_SPEC_KEY_NAME,
    SYSTEM_LAYOUT_MANIFEST_HASH_HEADER,
    SYSTEM_LAYOUT_MANIFEST_HEADER,
    SYSTEM_LAYOUT_SHEET,
    WORKBOOK_INTEGRITY_SCHEMA_VERSION,
    WORKBOOK_TEMP_SUFFIX,
)
from common.registry import (
    COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    SYSTEM_HASH_HEADER_TEMPLATE_HASH as SYSTEM_HASH_TEMPLATE_HASH_HEADER,
    SYSTEM_HASH_HEADER_TEMPLATE_ID as SYSTEM_HASH_TEMPLATE_ID_HEADER,
    SYSTEM_HASH_SHEET_NAME as SYSTEM_HASH_SHEET,
    get_sheet_headers_by_key,
    get_sheet_name_by_key,
)
from common.excel_sheet_layout import (
    apply_xlsxwriter_layout as _apply_xlsxwriter_layout_shared,
    build_template_xlsxwriter_formats as _build_template_xlsxwriter_formats,
)
from common.exceptions import AppSystemError, JobCancelledError, ValidationError
from common.jobs import CancellationToken
from common.texts import t
from common.utils import coerce_excel_number, normalize
from common.workbook_signing import sign_payload
from domain.assessment_semantics import AssessmentComponent, parse_assessment_components
from domain.co_token_parser import parse_co_tokens
from domain.co_report_sheet_generator import (
    ratio_total_header as _co_sheet_ratio_total_header,
    write_co_direct_sheet as _co_sheet_write_direct,
    write_co_indirect_sheet as _co_sheet_write_indirect,
    write_co_outcome_sheets,
)
from domain.template_strategy_router import read_valid_system_workbook_payload

_logger = logging.getLogger(__name__)
COURSE_METADATA_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
ASSESSMENT_CONFIG_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
_ANCHOR_COL_B = "B"
_ANCHOR_COL_C = "C"
_EXCEL_COL_REG_NO = 2
_EXCEL_COL_STUDENT_NAME = 3
_EXCEL_COL_FIRST_MARK = 4
_EXCEL_ROW_HEADER_OFFSET_DIRECT = 3
_EXCEL_ROW_HEADER_OFFSET_INDIRECT = 1
_EXCEL_ROW_CO_OFFSET = 1
_EXCEL_ROW_MAX_OFFSET = 2
_STYLE_CACHE_ATTR = "_focus_report_style_cache"
_LOG_FINAL_REPORT_READY = "Final CO report prepared with direct and indirect sheets."
_LOG_FINAL_REPORT_FAILED = "Final CO report generation failed unexpectedly."

@dataclass(frozen=True)
class _DirectComponent:
    name: str
    weight: float

@dataclass
class _DirectComponentComputed:
    name: str
    weight: float
    max_by_co: dict[int, float]
    marks_by_co: dict[int, list[float | str]]

@dataclass(frozen=True)
class _IndirectComponent:
    name: str
    weight: float

@dataclass
class _IndirectComponentComputed:
    name: str
    weight: float
    marks_by_co: dict[int, list[float | str]]

def generate_final_co_report(
    filled_marks_path: str | Path,
    output_path: str | Path,
    *,
    cancel_token: CancellationToken | None = None,
) -> Path:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise validation_error_from_key("instructor.validation.openpyxl_missing") from exc
    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise validation_error_from_key("instructor.validation.openpyxl_missing") from exc

    source = Path(filled_marks_path)
    output = Path(output_path)
    if not source.exists():
        raise validation_error_from_key("instructor.validation.workbook_not_found", workbook=source)
    output.parent.mkdir(parents=True, exist_ok=True)

    try:
        _raise_if_cancelled(cancel_token)
        src_wb = openpyxl.load_workbook(source, data_only=False)
        system_payload = read_valid_system_workbook_payload(src_wb)
    except JobCancelledError:
        raise
    except ValidationError:
        raise
    except Exception as exc:
        raise validation_error_from_key("instructor.validation.workbook_open_failed", workbook=source) from exc

    tmp_path: Path | None = None
    try:
        _raise_if_cancelled(cancel_token)
        metadata_rows, total_outcomes = _read_course_metadata(src_wb[COURSE_METADATA_SHEET])
        template_id = system_payload.template_id
        template_hash = system_payload.template_hash
        direct_components = _read_direct_components(src_wb[ASSESSMENT_CONFIG_SHEET], template_id=template_id)
        indirect_components = _read_indirect_components(src_wb[ASSESSMENT_CONFIG_SHEET], template_id=template_id)
        if not direct_components and not indirect_components:
            raise validation_error_from_key("instructor.validation.final_report.no_direct_components")
        manifest = system_payload.manifest
        direct_sheet_specs = _component_sheet_specs_by_kind(
            manifest,
            allowed_kinds={
                LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
                LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
            },
        )
        indirect_sheet_specs = _component_sheet_specs_by_kind(
            manifest,
            allowed_kinds={LAYOUT_SHEET_KIND_INDIRECT},
        )
        students = _read_students_from_component_sheets(
            src_wb,
            direct_specs=direct_sheet_specs,
            indirect_specs=indirect_sheet_specs,
        )

        computed_components: list[_DirectComponentComputed] = []
        for component in direct_components:
            _raise_if_cancelled(cancel_token)
            computed_components.append(
                _compute_component_marks(
                    workbook=src_wb,
                    component=component,
                    students=students,
                    spec=direct_sheet_specs.get(normalize(component.name)),
                    total_outcomes=total_outcomes,
                )
            )

        computed_indirect_components: list[_IndirectComponentComputed] = []
        for component in indirect_components:
            _raise_if_cancelled(cancel_token)
            computed_indirect_components.append(
                _compute_indirect_component_marks(
                    workbook=src_wb,
                    component=component,
                    students=students,
                    spec=indirect_sheet_specs.get(normalize(component.name)),
                    total_outcomes=total_outcomes,
                )
            )

        _raise_if_cancelled(cancel_token)
        with NamedTemporaryFile(
            mode="wb",
            delete=False,
            dir=str(output.parent),
            prefix=f"{output.name}.",
            suffix=WORKBOOK_TEMP_SUFFIX,
        ) as temp_file:
            tmp_path = Path(temp_file.name)
        _write_final_report_workbook_xlsxwriter(
            xlsxwriter_module=xlsxwriter,
            output_path=tmp_path,
            metadata_rows=metadata_rows,
            total_outcomes=total_outcomes,
            students=students,
            direct_components=computed_components,
            indirect_components=computed_indirect_components,
            template_id=template_id,
            template_hash=template_hash,
        )
        _normalize_page_setup_fit(tmp_path)
        _logger.info(
            _LOG_FINAL_REPORT_READY,
            extra={
                "user_message": f"Prepared {total_outcomes} direct and {total_outcomes} indirect CO sheets.",
                "co_count": total_outcomes,
            },
        )
        _raise_if_cancelled(cancel_token)
        os.replace(tmp_path, output)
        tmp_path = None
        return output
    except ValidationError:
        raise
    except JobCancelledError:
        raise
    except Exception as exc:
        _logger.exception(_LOG_FINAL_REPORT_FAILED, exc_info=exc)
        raise AppSystemError(
            t("instructor.log.error_while_process", process=t("instructor.log.process.generate_final_co_report"))
        ) from exc
    finally:
        src_wb.close()
        if tmp_path is not None and tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp final report file: %s", tmp_path)

def _raise_if_cancelled(cancel_token: CancellationToken | None) -> None:
    if cancel_token is not None:
        cancel_token.raise_if_cancelled()

def _read_course_metadata(sheet: Any) -> tuple[list[tuple[str, Any]], int]:
    rows: list[tuple[str, Any]] = []
    total_outcomes = 0
    row = 2
    while True:
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=2).value
        if normalize(key) == "" and normalize(value) == "":
            break
        key_text = str(key).strip() if key is not None else ""
        rows.append((key_text, value))
        if normalize(key_text) == normalize(COURSE_METADATA_TOTAL_OUTCOMES_KEY):
            parsed = coerce_excel_number(value)
            if isinstance(parsed, (int, float)) and not isinstance(parsed, bool):
                total_outcomes = int(parsed)
        row += 1
    if total_outcomes <= 0:
        raise validation_error_from_key("instructor.validation.course_metadata_total_outcomes_invalid")
    return rows, total_outcomes

def _read_direct_components(sheet: Any, *, template_id: str) -> list[_DirectComponent]:
    components = _read_assessment_components(sheet, template_id=template_id)
    return [
        _DirectComponent(name=component.component_name, weight=component.weight)
        for component in components
        if component.is_direct
    ]


def _read_indirect_components(sheet: Any, *, template_id: str) -> list[_IndirectComponent]:
    components = _read_assessment_components(sheet, template_id=template_id)
    return [
        _IndirectComponent(name=component.component_name, weight=component.weight)
        for component in components
        if not component.is_direct
    ]


def _read_assessment_components(sheet: Any, *, template_id: str) -> list[AssessmentComponent]:
    expected_headers = list(get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG))
    headers = [normalize(value) for value in expected_headers]
    for idx, expected in enumerate(headers, start=1):
        actual = normalize(sheet.cell(row=1, column=idx).value)
        if actual != expected:
            raise validation_error_from_key(
                "instructor.validation.header_mismatch", sheet_name=ASSESSMENT_CONFIG_SHEET, expected=expected_headers
            )
    rows: list[list[Any]] = []
    row = 2
    while True:
        row_values = [sheet.cell(row=row, column=col_index + 1).value for col_index in range(len(expected_headers))]
        if all(normalize(value) == "" for value in row_values):
            break
        rows.append(row_values)
        row += 1
    return parse_assessment_components(
        rows,
        sheet_name=ASSESSMENT_CONFIG_SHEET,
        headers=expected_headers,
        row_start=2,
        on_blank_component="break",
        duplicate_policy="keep_all",
        require_non_empty=False,
        validate_allowed_options=False,
    )


def _parse_co_values(raw: Any) -> list[int]:
    return parse_co_tokens(raw, dedupe=True)


def _component_sheet_specs_by_kind(
    manifest: dict[str, Any],
    *,
    allowed_kinds: set[str],
) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for spec in manifest.get(LAYOUT_MANIFEST_KEY_SHEETS, []):
        if not isinstance(spec, dict):
            continue
        kind = spec.get(LAYOUT_SHEET_SPEC_KEY_KIND)
        if kind not in allowed_kinds:
            continue
        component_name = _component_name_from_spec(spec)
        if not component_name:
            continue
        out[normalize(component_name)] = spec
    return out


def _component_name_from_spec(spec: dict[str, Any]) -> str:
    anchors = spec.get(LAYOUT_SHEET_SPEC_KEY_ANCHORS, [])
    if not isinstance(anchors, list):
        return ""
    row_by_label: dict[str, int] = {}
    value_by_row: dict[int, str] = {}
    for item in anchors:
        if not isinstance(item, list) or len(item) != 2:
            continue
        cell_ref, value = item
        if not isinstance(cell_ref, str):
            continue
        row_no = _cell_row(cell_ref)
        if row_no <= 0:
            continue
        cell_col = cell_ref[:1].upper()
        text = str(value).strip() if value is not None else ""
        if cell_col == _ANCHOR_COL_B:
            row_by_label[normalize(text)] = row_no
        elif cell_col == _ANCHOR_COL_C:
            value_by_row[row_no] = text
    component_row = row_by_label.get(normalize(COMPONENT_NAME_LABEL))
    if component_row is None:
        return ""
    return value_by_row.get(component_row, "")


def _cell_row(cell_ref: str) -> int:
    digits = "".join(ch for ch in cell_ref if ch.isdigit())
    return int(digits) if digits else 0


def _read_students_from_component_sheets(
    workbook: Any,
    *,
    direct_specs: dict[str, dict[str, Any]],
    indirect_specs: dict[str, dict[str, Any]],
) -> list[tuple[str, str]]:
    specs_by_component = direct_specs or indirect_specs
    if not specs_by_component:
        raise validation_error_from_key("instructor.validation.final_report.no_direct_components")
    first_spec = next(iter(specs_by_component.values()))
    sheet_name = first_spec.get(LAYOUT_SHEET_SPEC_KEY_NAME)
    header_row = first_spec.get(LAYOUT_SHEET_SPEC_KEY_HEADER_ROW)
    if not isinstance(sheet_name, str) or not isinstance(header_row, int):
        raise validation_error_from_key("instructor.validation.final_report.layout_manifest_invalid")
    ws = workbook[sheet_name]
    first_data_row = header_row + _EXCEL_ROW_HEADER_OFFSET_DIRECT
    students: list[tuple[str, str]] = []
    row = first_data_row
    while True:
        reg = ws.cell(row=row, column=_EXCEL_COL_REG_NO).value
        name = ws.cell(row=row, column=_EXCEL_COL_STUDENT_NAME).value
        if normalize(reg) == "" and normalize(name) == "":
            break
        students.append((str(reg).strip(), str(name).strip()))
        row += 1
    return students


def _compute_component_marks(
    *,
    workbook: Any,
    component: _DirectComponent,
    students: list[tuple[str, str]],
    spec: dict[str, Any] | None,
    total_outcomes: int,
) -> _DirectComponentComputed:
    if spec is None:
        raise validation_error_from_key(
            "instructor.validation.final_report.direct_component_sheet_missing", component=component.name
        )
    sheet_name = spec.get(LAYOUT_SHEET_SPEC_KEY_NAME)
    header_row = spec.get(LAYOUT_SHEET_SPEC_KEY_HEADER_ROW)
    headers = spec.get(LAYOUT_SHEET_SPEC_KEY_HEADERS)
    kind = spec.get(LAYOUT_SHEET_SPEC_KEY_KIND)
    if (
        not isinstance(sheet_name, str)
        or not isinstance(header_row, int)
        or not isinstance(headers, list)
        or kind not in {LAYOUT_SHEET_KIND_DIRECT_CO_WISE, LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE}
    ):
        raise validation_error_from_key("instructor.validation.final_report.layout_manifest_invalid")
    ws = workbook[sheet_name]
    first_data_row = header_row + _EXCEL_ROW_HEADER_OFFSET_DIRECT
    expected_students = len(students)

    max_by_co = {co: 0.0 for co in range(1, total_outcomes + 1)}
    marks_by_co: dict[int, list[float | str]] = {
        co: [0.0] * expected_students for co in range(1, total_outcomes + 1)
    }
    if kind == LAYOUT_SHEET_KIND_DIRECT_CO_WISE:
        question_cols = list(range(_EXCEL_COL_FIRST_MARK, max(_EXCEL_COL_FIRST_MARK, len(headers))))
        co_by_col: dict[int, int] = {}
        for col in question_cols:
            co_token = ws.cell(row=header_row + _EXCEL_ROW_CO_OFFSET, column=col).value
            co_values = _parse_co_values(co_token)
            if len(co_values) != 1:
                continue
            co_by_col[col] = co_values[0]
            max_marks = _to_float(ws.cell(row=header_row + _EXCEL_ROW_MAX_OFFSET, column=col).value)
            if max_marks is not None and 1 <= co_values[0] <= total_outcomes:
                max_by_co[co_values[0]] += max_marks
        if not co_by_col:
            raise validation_error_from_key(
                "instructor.validation.final_report.direct_component_marks_shape_invalid", component=component.name
            )
        active_cos = [co for co in range(1, total_outcomes + 1) if max_by_co[co] > 0]
        for idx in range(expected_students):
            row = first_data_row + idx
            absent = False
            co_totals = {co: 0.0 for co in active_cos}
            for col, co_value in co_by_col.items():
                cell_value = ws.cell(row=row, column=col).value
                if _is_absent(cell_value):
                    absent = True
                    break
                numeric = _to_float(cell_value)
                if numeric is None:
                    continue
                if 1 <= co_value <= total_outcomes:
                    co_totals[co_value] += numeric
            if absent:
                for co in active_cos:
                    marks_by_co[co][idx] = CO_REPORT_ABSENT_TOKEN
            else:
                for co in active_cos:
                    marks_by_co[co][idx] = _round2(co_totals[co])
    else:
        covered_cos: list[int] = []
        for col in range(_EXCEL_COL_FIRST_MARK + 1, len(headers) + 1):
            co_token = ws.cell(row=header_row + _EXCEL_ROW_CO_OFFSET, column=col).value
            co_values = _parse_co_values(co_token)
            if len(co_values) != 1:
                continue
            co_value = co_values[0]
            if co_value < 1 or co_value > total_outcomes:
                continue
            covered_cos.append(co_value)
            max_marks = _to_float(ws.cell(row=header_row + _EXCEL_ROW_MAX_OFFSET, column=col).value)
            if max_marks is not None:
                max_by_co[co_value] = max_marks
        for idx in range(expected_students):
            row = first_data_row + idx
            total_cell_value = ws.cell(row=row, column=_EXCEL_COL_FIRST_MARK).value
            if _is_absent(total_cell_value):
                for co in covered_cos:
                    marks_by_co[co][idx] = CO_REPORT_ABSENT_TOKEN
                continue
            total_numeric = _to_float(total_cell_value)
            if total_numeric is None:
                total_numeric = 0.0
            split_values = _split_equal_with_residual(total_numeric, len(covered_cos))
            for co_idx, co in enumerate(covered_cos):
                marks_by_co[co][idx] = split_values[co_idx]

    return _DirectComponentComputed(
        name=component.name,
        weight=component.weight,
        max_by_co=max_by_co,
        marks_by_co=marks_by_co,
    )


def _compute_indirect_component_marks(
    *,
    workbook: Any,
    component: _IndirectComponent,
    students: list[tuple[str, str]],
    spec: dict[str, Any] | None,
    total_outcomes: int,
) -> _IndirectComponentComputed:
    if spec is None:
        raise validation_error_from_key(
            "instructor.validation.final_report.direct_component_sheet_missing", component=component.name
        )
    sheet_name = spec.get(LAYOUT_SHEET_SPEC_KEY_NAME)
    header_row = spec.get(LAYOUT_SHEET_SPEC_KEY_HEADER_ROW)
    headers = spec.get(LAYOUT_SHEET_SPEC_KEY_HEADERS)
    kind = spec.get(LAYOUT_SHEET_SPEC_KEY_KIND)
    if (
        not isinstance(sheet_name, str)
        or not isinstance(header_row, int)
        or not isinstance(headers, list)
        or kind != LAYOUT_SHEET_KIND_INDIRECT
    ):
        raise validation_error_from_key("instructor.validation.final_report.layout_manifest_invalid")
    ws = workbook[sheet_name]
    first_data_row = header_row + _EXCEL_ROW_HEADER_OFFSET_INDIRECT
    expected_students = len(students)

    marks_by_co: dict[int, list[float | str]] = {
        co: [0.0] * expected_students for co in range(1, total_outcomes + 1)
    }
    for idx in range(expected_students):
        row = first_data_row + idx
        for co in range(1, total_outcomes + 1):
            col = _EXCEL_COL_STUDENT_NAME + co
            cell_value = ws.cell(row=row, column=col).value
            if _is_absent(cell_value):
                marks_by_co[co][idx] = CO_REPORT_ABSENT_TOKEN
                continue
            numeric = _to_float(cell_value)
            marks_by_co[co][idx] = _round2(numeric if numeric is not None else 0.0)
    return _IndirectComponentComputed(
        name=component.name,
        weight=component.weight,
        marks_by_co=marks_by_co,
    )


def _split_equal_with_residual(total: float, count: int) -> list[float]:
    if count <= 0:
        return []
    if count == 1:
        return [_round2(total)]
    base = _round2(total / count)
    values = [base] * count
    values[-1] = _round2(total - sum(values[:-1]))
    return values


def _is_absent(value: Any) -> bool:
    token = normalize(value)
    return token == normalize(CO_REPORT_ABSENT_TOKEN)


def _to_float(value: Any) -> float | None:
    parsed = coerce_excel_number(value)
    if isinstance(parsed, bool):
        return None
    if isinstance(parsed, (int, float)):
        return float(parsed)
    return None


def _round2(value: float) -> float:
    return round(float(value), CO_REPORT_MAX_DECIMAL_PLACES)


def _xlsxwriter_formats(workbook: Any, *, template_id: str) -> dict[str, Any]:
    return _build_template_xlsxwriter_formats(
        workbook,
        template_id=template_id,
        cache_attr=_STYLE_CACHE_ATTR,
        include_column_wrap=False,
        normalize_header_valign_to_center=True,
    )


def _xlsxwriter_apply_layout(ws: Any, *, header_row_index: int, paper_size: int, landscape: bool) -> None:
    _apply_xlsxwriter_layout_shared(
        ws,
        header_row_index=header_row_index,
        paper_size=paper_size,
        landscape=landscape,
    )


def _xlsxwriter_write_course_metadata_sheet(
    workbook: Any,
    *,
    template_id: str,
    metadata_rows: list[tuple[str, Any]],
    formats: dict[str, Any],
) -> None:
    course_metadata_headers = get_sheet_headers_by_key(template_id, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
    ws = workbook.add_worksheet(COURSE_METADATA_SHEET)
    ws.write(0, 0, course_metadata_headers[0], formats["header"])
    ws.write(0, 1, course_metadata_headers[1], formats["header"])
    filtered_rows = [
        (field, value)
        for field, value in metadata_rows
        if normalize(field) != normalize(COURSE_METADATA_FACULTY_NAME_KEY)
    ]
    for idx, (field, value) in enumerate(filtered_rows, start=1):
        ws.write(idx, 0, field, formats["body"])
        ws.write(idx, 1, value, formats["body"])
    _xlsxwriter_apply_layout(ws, header_row_index=0, paper_size=9, landscape=False)


def _xlsxwriter_write_direct_sheet(
    workbook: Any,
    *,
    template_id: str,
    co_index: int,
    metadata_rows: list[tuple[str, Any]],
    students: list[tuple[str, str]],
    components: list[_DirectComponentComputed],
    formats: dict[str, Any],
) -> None:
    _co_sheet_write_direct(
        workbook,
        template_id=template_id,
        co_index=co_index,
        metadata_rows=metadata_rows,
        students=students,
        components=components,
        formats=formats,
    )


def _xlsxwriter_write_indirect_sheet(
    workbook: Any,
    *,
    template_id: str,
    co_index: int,
    metadata_rows: list[tuple[str, Any]],
    students: list[tuple[str, str]],
    components: list[_IndirectComponentComputed],
    formats: dict[str, Any],
) -> None:
    _co_sheet_write_indirect(
        workbook,
        template_id=template_id,
        co_index=co_index,
        metadata_rows=metadata_rows,
        students=students,
        components=components,
        formats=formats,
    )


def _add_system_layout_sheet(workbook: Any, manifest_text: str, manifest_hash: str) -> None:
    layout_ws = workbook.add_worksheet(SYSTEM_LAYOUT_SHEET)
    layout_ws.write_row(0, 0, [SYSTEM_LAYOUT_MANIFEST_HEADER, SYSTEM_LAYOUT_MANIFEST_HASH_HEADER])
    layout_ws.write_row(1, 0, [manifest_text, manifest_hash])
    layout_ws.hide()
def _write_final_report_workbook_xlsxwriter(
    *,
    xlsxwriter_module: Any,
    output_path: Path,
    metadata_rows: list[tuple[str, Any]],
    total_outcomes: int,
    students: list[tuple[str, str]],
    direct_components: list[_DirectComponentComputed],
    indirect_components: list[_IndirectComponentComputed],
    template_id: str,
    template_hash: str,
) -> None:
    workbook = xlsxwriter_module.Workbook(str(output_path), {"constant_memory": True})
    sheet_order: list[str] = []
    try:
        formats = _xlsxwriter_formats(workbook, template_id=template_id)
        _xlsxwriter_write_course_metadata_sheet(
            workbook,
            template_id=template_id,
            metadata_rows=metadata_rows,
            formats=formats,
        )
        sheet_order.append(COURSE_METADATA_SHEET)
        for co_index in range(1, total_outcomes + 1):
            direct_sheet, indirect_sheet = write_co_outcome_sheets(
                workbook,
                template_id=template_id,
                co_index=co_index,
                metadata_rows=metadata_rows,
                students=students,
                direct_components=direct_components,
                indirect_components=indirect_components,
                formats=formats,
            )
            sheet_order.append(direct_sheet)
            sheet_order.append(indirect_sheet)

        hash_ws = workbook.add_worksheet(SYSTEM_HASH_SHEET)
        hash_ws.write(0, 0, SYSTEM_HASH_TEMPLATE_ID_HEADER)
        hash_ws.write(0, 1, SYSTEM_HASH_TEMPLATE_HASH_HEADER)
        hash_ws.write(1, 0, template_id)
        hash_ws.write(1, 1, template_hash)
        hash_ws.hide()
        sheet_order.append(SYSTEM_HASH_SHEET)

        manifest = {
            "schema_version": WORKBOOK_INTEGRITY_SCHEMA_VERSION,
            "template_id": template_id,
            "template_hash": template_hash,
            "sheet_order": sheet_order,
            "sheets": [{"name": name, "hash": sign_payload(name)} for name in sheet_order],
        }
        manifest_text = json.dumps(manifest, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
        manifest_hash = sign_payload(manifest_text)
        _add_system_layout_sheet(workbook, manifest_text, manifest_hash)
    finally:
        workbook.close()

def _normalize_page_setup_fit(path: Path) -> None:
    try:
        import openpyxl
    except Exception:
        return
    with path.open("rb") as handle:
        wb = openpyxl.load_workbook(handle)
    try:
        for ws in wb.worksheets:
            if ws.title in {SYSTEM_HASH_SHEET, SYSTEM_LAYOUT_SHEET}:
                continue
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        wb.save(path)
    finally:
        wb.close()


def _ratio_total_header(ratio: float) -> str:
    return _co_sheet_ratio_total_header(ratio)
