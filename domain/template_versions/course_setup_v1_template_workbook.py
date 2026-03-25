"""Generator for the Course Details workbook template."""

from __future__ import annotations

import logging
import os
import re
from pathlib import Path
from typing import Any, Sequence
from uuid import uuid4

from common.error_catalog import validation_error_from_key
from common.constants import (
    ID_COURSE_SETUP,
    LAYOUT_MANIFEST_KEY_SHEET_ORDER,
    LAYOUT_MANIFEST_KEY_SHEETS,
    LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
    LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
    LAYOUT_SHEET_KIND_INDIRECT,
    LAYOUT_SHEET_SPEC_KEY_KIND,
    LAYOUT_SHEET_SPEC_KEY_NAME,
    SYSTEM_LAYOUT_SHEET,
    WORKBOOK_INTEGRITY_SCHEMA_VERSION,
    WORKBOOK_TEMP_SUFFIX,
)
from common.exceptions import AppSystemError, JobCancelledError, ValidationError
from common.jobs import CancellationToken
from common.registry import (
    SYSTEM_HASH_SHEET_NAME as SYSTEM_HASH_SHEET,
    COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG,
    COURSE_SETUP_SHEET_KEY_COURSE_METADATA,
    COURSE_SETUP_SHEET_KEY_QUESTION_MAP,
    COURSE_SETUP_SHEET_KEY_STUDENTS,
    get_blueprint as _registry_get_blueprint,
    get_sheet_headers_by_key,
    get_sheet_name_by_key,
)
from common.sample_setup_data import SAMPLE_SETUP_DATA
from common.sheet_schema import WorkbookBlueprint
from common.i18n import t
from common.utils import coerce_excel_number, normalize
from domain.assessment_semantics import parse_assessment_components
from domain.template_strategy_router import (
    available_template_ids,
    get_template_strategy,
    read_valid_template_id_from_system_hash_sheet,
)

_logger = logging.getLogger(__name__)
COURSE_METADATA_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
ASSESSMENT_CONFIG_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG)
QUESTION_MAP_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_QUESTION_MAP)
STUDENTS_SHEET = get_sheet_name_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_STUDENTS)
COURSE_METADATA_HEADERS = get_sheet_headers_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_COURSE_METADATA)
ASSESSMENT_CONFIG_HEADERS = get_sheet_headers_by_key(
    ID_COURSE_SETUP,
    COURSE_SETUP_SHEET_KEY_ASSESSMENT_CONFIG,
)
QUESTION_MAP_HEADERS = get_sheet_headers_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_QUESTION_MAP)
STUDENTS_HEADERS = get_sheet_headers_by_key(ID_COURSE_SETUP, COURSE_SETUP_SHEET_KEY_STUDENTS)


def _ve(translation_key: str, *, code: str, **context: object) -> ValidationError:
    return validation_error_from_key(translation_key, code=code, **context)


def _get_blueprint(template_id: str) -> WorkbookBlueprint:
    strategy = get_template_strategy(template_id)
    blueprint = _registry_get_blueprint(strategy.template_id)
    if blueprint is None:
        available = ", ".join(available_template_ids())
        raise _ve(
            "instructor.validation.unknown_template",
            code="UNKNOWN_TEMPLATE",
            template_id=strategy.template_id,
            available=available,
        )
    return blueprint


def generate_course_details_template(
    output_path: str | Path,
    template_id: str = ID_COURSE_SETUP,
    *,
    cancel_token: CancellationToken | None = None,
) -> Path:
    """Generate and save the course details template workbook with atomic replace."""
    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.xlsxwriter_missing",
            code="XLSXWRITER_MISSING",
        ) from exc

    blueprint = _get_blueprint(template_id)
    sample_data = SAMPLE_SETUP_DATA
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    temp_output = output.with_name(f"{output.name}.{uuid4().hex}{WORKBOOK_TEMP_SUFFIX}")
    workbook = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    workbook_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal workbook_closed
        if not workbook_closed:
            try:
                workbook.close()
                workbook_closed = True
            except Exception:
                _logger.debug("Suppressing workbook close error during cleanup.", exc_info=True)
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp template file: %s", temp_output)

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        header_format = _build_header_format(workbook, blueprint.style_registry.get("header", {}))
        body_format = _build_body_format(workbook, blueprint.style_registry.get("body", {}))

        for sheet_schema in blueprint.sheets:
            if cancel_token is not None:
                cancel_token.raise_if_cancelled()
            if len(sheet_schema.header_matrix) != 1:
                raise validation_error_from_key(
                    
                        "instructor.validation.sheet_single_header_row",
                        sheet_name=sheet_schema.name,
                    
                )

            worksheet = generate_worksheet(
                workbook=workbook,
                sheet_name=sheet_schema.name,
                headers=sheet_schema.header_matrix[0],
                data=sample_data.get(sheet_schema.name, []),
                header_format=header_format,
                body_format=body_format,
            )

            for validation in sheet_schema.validations:
                if cancel_token is not None:
                    cancel_token.raise_if_cancelled()
                _apply_validation(worksheet, validation)

            if sheet_schema.is_protected:
                _protect_sheet(worksheet)

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        _add_system_hash_sheet(workbook, template_id)

        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        workbook.close()
        workbook_closed = True
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except JobCancelledError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate course details template. template_id=%s output=%s",
            template_id,
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc

    return output


def generate_marks_template_from_course_details(
    course_details_path: str | Path,
    output_path: str | Path,
    *,
    cancel_token: CancellationToken | None = None,
) -> Path:
    """Generate marks-entry workbook from a validated course-details workbook."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.openpyxl_missing",
            code="OPENPYXL_MISSING",
        ) from exc

    try:
        import xlsxwriter
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.xlsxwriter_missing",
            code="XLSXWRITER_MISSING",
        ) from exc

    source_file = Path(course_details_path)
    if not source_file.exists():
        raise _ve(
            "instructor.validation.workbook_not_found",
            code="WORKBOOK_NOT_FOUND",
            workbook=str(source_file),
        )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    temp_output = output.with_name(f"{output.name}.{uuid4().hex}{WORKBOOK_TEMP_SUFFIX}")

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        workbook = openpyxl.load_workbook(source_file, data_only=True)
    except JobCancelledError:
        raise
    except Exception as exc:
        raise _ve(
            "instructor.validation.workbook_open_failed",
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(source_file),
        ) from exc

    target = xlsxwriter.Workbook(str(temp_output), {"constant_memory": True})
    target_closed = False

    def _cleanup_incomplete_output() -> None:
        nonlocal target_closed
        if not target_closed:
            try:
                target.close()
                target_closed = True
            except Exception:
                _logger.debug("Suppressing target close error during cleanup.", exc_info=True)
        if temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                _logger.warning("Failed to cleanup temp marks template file: %s", temp_output)

    try:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        template_id = _extract_and_validate_template_id(workbook)
        context = _extract_marks_template_context_by_template(workbook, template_id)
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        layout_manifest = _write_marks_template_workbook_by_template(
            target,
            context,
            template_id=template_id,
            cancel_token=cancel_token,
        )
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        _copy_system_hash_sheet(workbook, target)
        _add_system_layout_sheet(target, layout_manifest)
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        target.close()
        target_closed = True
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        os.replace(temp_output, output)
    except ValidationError:
        _cleanup_incomplete_output()
        raise
    except JobCancelledError:
        _cleanup_incomplete_output()
        raise
    except Exception as exc:
        _cleanup_incomplete_output()
        _logger.exception(
            "Failed to generate marks template. source=%s output=%s",
            source_file,
            output,
        )
        raise AppSystemError(
            t("instructor.system.template_generate_failed", output=output)
        ) from exc
    finally:
        workbook.close()

    return output


def _extract_marks_template_context(workbook: Any) -> dict[str, Any]:
    metadata_rows = _iter_data_rows(workbook[COURSE_METADATA_SHEET], len(COURSE_METADATA_HEADERS))
    metadata_rows = _filter_marks_template_metadata_rows(metadata_rows)
    assessment_rows = _iter_data_rows(workbook[ASSESSMENT_CONFIG_SHEET], len(ASSESSMENT_CONFIG_HEADERS))
    question_rows = _iter_data_rows(workbook[QUESTION_MAP_SHEET], len(QUESTION_MAP_HEADERS))
    student_rows = _iter_data_rows(workbook[STUDENTS_SHEET], len(STUDENTS_HEADERS))

    total_outcomes = _extract_total_outcomes(metadata_rows)
    students = _extract_students(student_rows)
    components = _extract_components(assessment_rows)
    questions_by_component = _extract_questions(question_rows)

    return {
        "metadata_rows": metadata_rows,
        "assessment_rows": assessment_rows,
        "total_outcomes": total_outcomes,
        "students": students,
        "components": components,
        "questions_by_component": questions_by_component,
    }


def _extract_marks_template_context_by_template(workbook: Any, template_id: str) -> dict[str, Any]:
    del template_id
    return _extract_marks_template_context(workbook)


def _write_marks_template_workbook_by_template(
    workbook: Any,
    context: dict[str, Any],
    *,
    template_id: str,
    cancel_token: CancellationToken | None = None,
) -> dict[str, Any]:
    return _write_marks_template_workbook(
        workbook,
        context,
        template_id=template_id,
        cancel_token=cancel_token,
    )


def _extract_total_outcomes(metadata_rows: Sequence[Sequence[Any]]) -> int:
    for row in metadata_rows:
        if len(row) < 2:
            continue
        if normalize(row[0]) == "total_outcomes":
            value = coerce_excel_number(row[1])
            if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
                break
            return value
    raise validation_error_from_key("instructor.validation.course_metadata_total_outcomes_invalid")


def _extract_students(student_rows: Sequence[Sequence[Any]]) -> list[tuple[str, str]]:
    students: list[tuple[str, str]] = []
    for row in student_rows:
        reg_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if reg_no and name:
            students.append((reg_no, name))
    if not students:
        raise validation_error_from_key("instructor.validation.students_row_required_one")
    return students


def _extract_components(assessment_rows: Sequence[Sequence[Any]]) -> list[dict[str, Any]]:
    parsed = parse_assessment_components(
        assessment_rows,
        template_id="COURSE_SETUP_V1",
        sheet_name=ASSESSMENT_CONFIG_SHEET,
        row_start=2,
        on_blank_component="skip",
        duplicate_policy="keep_first",
        require_non_empty=True,
        validate_allowed_options=False,
    )
    components: list[dict[str, Any]] = []
    for component in parsed:
        components.append(
            {
                "key": component.component_key,
                "display_name": component.component_name,
                "co_wise_breakup": component.co_wise_breakup,
                "is_direct": component.is_direct,
            }
        )
    if not components:
        raise validation_error_from_key("instructor.validation.assessment_component_required_one")
    return components


def _extract_questions(question_rows: Sequence[Sequence[Any]]) -> dict[str, list[dict[str, Any]]]:
    questions_by_component: dict[str, list[dict[str, Any]]] = {}
    for row in question_rows:
        if len(row) < 4:
            continue
        component_key = normalize(row[0])
        if not component_key:
            continue
        max_marks = coerce_excel_number(row[2])
        if isinstance(max_marks, bool) or not isinstance(max_marks, (int, float)):
            continue
        co_values = _co_tokens(row[3])
        if not co_values:
            continue
        questions_by_component.setdefault(component_key, []).append(
            {
                "max_marks": float(max_marks),
                "co_values": co_values,
            }
        )
    return questions_by_component


def _write_marks_template_workbook(
    workbook: Any,
    context: dict[str, Any],
    *,
    template_id: str = ID_COURSE_SETUP,
    cancel_token: CancellationToken | None = None,
) -> dict[str, Any]:
    setup_blueprint = _get_blueprint(template_id)
    header_fmt = _build_header_format(workbook, setup_blueprint.style_registry.get("header", {}))
    body_fmt = workbook.add_format({"border": 1})
    wrapped_body_fmt = workbook.add_format({"border": 1, "text_wrap": True})
    wrapped_column_fmt = workbook.add_format({"text_wrap": True})
    num_fmt = workbook.add_format({"border": 1, "num_format": "0.00"})
    header_num_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": str(setup_blueprint.style_registry.get("header", {}).get("bg_color", "")),
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "num_format": "0.00",
        }
    )
    unlocked_body_fmt = workbook.add_format({"border": 1, "locked": False})

    layout_sheets, component_plans = _precompute_marks_layout_manifest(
        context=context,
        cancel_token=cancel_token,
    )

    students = context["students"]
    _write_two_column_copy_sheet(
        workbook=workbook,
        title=COURSE_METADATA_SHEET,
        header=COURSE_METADATA_HEADERS,
        rows=context["metadata_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
    )
    _write_multi_column_copy_sheet(
        workbook=workbook,
        title=ASSESSMENT_CONFIG_SHEET,
        header=ASSESSMENT_CONFIG_HEADERS,
        rows=context["assessment_rows"],
        header_fmt=header_fmt,
        body_fmt=body_fmt,
        num_fmt=num_fmt,
    )

    for plan in component_plans:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        if plan[LAYOUT_SHEET_SPEC_KEY_KIND] == LAYOUT_SHEET_KIND_DIRECT_CO_WISE:
            _write_direct_co_wise_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                plan["questions"],
                header_fmt,
                body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
                num_fmt,
                header_num_fmt,
                unlocked_body_fmt,
            )
        elif plan[LAYOUT_SHEET_SPEC_KEY_KIND] == LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE:
            _write_direct_non_co_wise_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                plan["questions"],
                header_fmt,
                body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
                num_fmt,
                header_num_fmt,
                unlocked_body_fmt,
            )
        else:
            _write_indirect_sheet(
                workbook,
                plan["sheet_name"],
                context["metadata_rows"],
                plan["component_name"],
                students,
                context["total_outcomes"],
                header_fmt,
                body_fmt,
                unlocked_body_fmt,
                wrapped_body_fmt,
                wrapped_column_fmt,
            )

    return {
        "schema_version": WORKBOOK_INTEGRITY_SCHEMA_VERSION,
        LAYOUT_MANIFEST_KEY_SHEET_ORDER: [entry[LAYOUT_SHEET_SPEC_KEY_NAME] for entry in layout_sheets]
        + [SYSTEM_HASH_SHEET, SYSTEM_LAYOUT_SHEET],
        LAYOUT_MANIFEST_KEY_SHEETS: layout_sheets,
    }


def _precompute_marks_layout_manifest(
    *,
    context: dict[str, Any],
    cancel_token: CancellationToken | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    students = context["students"]
    metadata_rows = context["metadata_rows"]
    assessment_rows = context["assessment_rows"]
    questions_by_component = context["questions_by_component"]
    total_outcomes = context["total_outcomes"]
    student_identity_hash = _student_identity_hash(students)

    layout_sheets: list[dict[str, Any]] = [
        _build_two_column_copy_sheet_spec(
            title=COURSE_METADATA_SHEET,
            header=COURSE_METADATA_HEADERS,
            rows=metadata_rows,
        ),
        _build_multi_column_copy_sheet_spec(
            title=ASSESSMENT_CONFIG_SHEET,
            header=ASSESSMENT_CONFIG_HEADERS,
            rows=assessment_rows,
        ),
    ]
    component_plans: list[dict[str, Any]] = []

    used_sheet_names = {normalize(COURSE_METADATA_SHEET), normalize(ASSESSMENT_CONFIG_SHEET)}
    for component in context["components"]:
        if cancel_token is not None:
            cancel_token.raise_if_cancelled()
        component_name = component["display_name"]
        sheet_name = _safe_sheet_name(component_name, used_sheet_names)
        questions = questions_by_component.get(component["key"], [])
        if component["is_direct"]:
            if not questions:
                raise validation_error_from_key("instructor.validation.question_map_row_required_one")
            if component["co_wise_breakup"]:
                layout_sheets.append(
                    _build_direct_co_wise_sheet_spec(
                        sheet_name=sheet_name,
                        metadata_rows=metadata_rows,
                        component_name=component_name,
                        students=students,
                        questions=questions,
                        student_identity_hash=student_identity_hash,
                    )
                )
                component_plans.append(
                    {
                        LAYOUT_SHEET_SPEC_KEY_KIND: LAYOUT_SHEET_KIND_DIRECT_CO_WISE,
                        "sheet_name": sheet_name,
                        "component_name": component_name,
                        "questions": questions,
                    }
                )
            else:
                layout_sheets.append(
                    _build_direct_non_co_wise_sheet_spec(
                        sheet_name=sheet_name,
                        metadata_rows=metadata_rows,
                        component_name=component_name,
                        students=students,
                        questions=questions,
                        student_identity_hash=student_identity_hash,
                    )
                )
                component_plans.append(
                    {
                        LAYOUT_SHEET_SPEC_KEY_KIND: LAYOUT_SHEET_KIND_DIRECT_NON_CO_WISE,
                        "sheet_name": sheet_name,
                        "component_name": component_name,
                        "questions": questions,
                    }
                )
        else:
            layout_sheets.append(
                _build_indirect_sheet_spec(
                    sheet_name=sheet_name,
                    metadata_rows=metadata_rows,
                    component_name=component_name,
                    students=students,
                    total_outcomes=total_outcomes,
                    student_identity_hash=student_identity_hash,
                )
            )
            component_plans.append(
                {
                    LAYOUT_SHEET_SPEC_KEY_KIND: LAYOUT_SHEET_KIND_INDIRECT,
                    "sheet_name": sheet_name,
                    "component_name": component_name,
                    "questions": [],
                }
            )

    return layout_sheets, component_plans



from domain.template_versions.course_setup_v2_impl import instructor_engine_sheetops as _sheetops

_add_system_hash_sheet = _sheetops._add_system_hash_sheet
_add_system_layout_sheet = _sheetops._add_system_layout_sheet
_apply_validation = _sheetops._apply_validation
_build_body_format = _sheetops._build_body_format
_build_direct_co_wise_sheet_spec = _sheetops._build_direct_co_wise_sheet_spec
_build_direct_non_co_wise_sheet_spec = _sheetops._build_direct_non_co_wise_sheet_spec
_build_header_format = _sheetops._build_header_format
_build_indirect_sheet_spec = _sheetops._build_indirect_sheet_spec
_build_multi_column_copy_sheet_spec = _sheetops._build_multi_column_copy_sheet_spec
_build_two_column_copy_sheet_spec = _sheetops._build_two_column_copy_sheet_spec
_compute_template_hash = _sheetops._compute_template_hash
_copy_system_hash_sheet = _sheetops._copy_system_hash_sheet
_filter_marks_template_metadata_rows = _sheetops._filter_marks_template_metadata_rows
_protect_sheet = _sheetops._protect_sheet
_safe_sheet_name = _sheetops._safe_sheet_name
_student_identity_hash = _sheetops._student_identity_hash
_write_direct_co_wise_sheet = _sheetops._write_direct_co_wise_sheet
_write_direct_non_co_wise_sheet = _sheetops._write_direct_non_co_wise_sheet
_write_indirect_sheet = _sheetops._write_indirect_sheet
_write_multi_column_copy_sheet = _sheetops._write_multi_column_copy_sheet
_write_two_column_copy_sheet = _sheetops._write_two_column_copy_sheet
generate_worksheet = _sheetops.generate_worksheet


def validate_course_details_workbook(workbook_path: str | Path) -> str:
    """Validate uploaded course details workbook and return template id."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise _ve(
            "instructor.validation.openpyxl_missing",
            code="OPENPYXL_MISSING",
        ) from exc

    workbook_file = Path(workbook_path)
    if not workbook_file.exists():
        raise _ve(
            "instructor.validation.workbook_not_found",
            code="WORKBOOK_NOT_FOUND",
            workbook=str(workbook_file),
        )

    # Fast pass: stream workbook in read-only mode for identity/schema checks.
    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=True, read_only=True)
    except Exception as exc:
        raise _ve(
            "instructor.validation.workbook_open_failed",
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(workbook_file),
        ) from exc

    try:
        template_id = _extract_and_validate_template_id(workbook)
        blueprint = _get_blueprint(template_id)
        _validate_workbook_schema_read_only(workbook, blueprint)
    finally:
        workbook.close()

    # Deep pass: run template-specific business rules only after fast pass succeeds.
    try:
        workbook = openpyxl.load_workbook(workbook_file, data_only=True)
    except Exception as exc:
        raise _ve(
            "instructor.validation.workbook_open_failed",
            code="WORKBOOK_OPEN_FAILED",
            workbook=str(workbook_file),
        ) from exc
    try:
        _validate_template_specific_rules(workbook, template_id)
    finally:
        workbook.close()

    return template_id


def _extract_and_validate_template_id(workbook: Any) -> str:
    return read_valid_template_id_from_system_hash_sheet(workbook)


def _validate_workbook_schema(workbook: Any, blueprint: WorkbookBlueprint) -> None:
    expected_sheet_names = [schema.name for schema in blueprint.sheets] + [SYSTEM_HASH_SHEET]
    actual_sheet_names = list(workbook.sheetnames)
    if actual_sheet_names != expected_sheet_names:
        raise validation_error_from_key(
            
                "instructor.validation.workbook_sheet_mismatch",
                template_id=blueprint.type_id,
                expected=expected_sheet_names,
                found=actual_sheet_names,
            
        )

    for sheet_schema in blueprint.sheets:
        if len(sheet_schema.header_matrix) != 1:
            raise validation_error_from_key(
                
                    "instructor.validation.sheet_single_header_row",
                    sheet_name=sheet_schema.name,
                
            )

        expected_headers = [normalize(h) for h in sheet_schema.header_matrix[0]]
        worksheet = workbook[sheet_schema.name]
        actual_headers = [
            normalize(worksheet.cell(row=1, column=col_index + 1).value)
            for col_index in range(len(expected_headers))
        ]
        if actual_headers != expected_headers:
            raise validation_error_from_key(
                
                    "instructor.validation.header_mismatch",
                    sheet_name=sheet_schema.name,
                    expected=sheet_schema.header_matrix[0],
                
            )


def _read_header_row_values(worksheet: Any, *, count: int) -> list[str]:
    first_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=count, values_only=True),
        tuple(),
    )
    return [normalize(value) for value in first_row]


def _validate_workbook_schema_read_only(workbook: Any, blueprint: WorkbookBlueprint) -> None:
    expected_sheet_names = [schema.name for schema in blueprint.sheets] + [SYSTEM_HASH_SHEET]
    actual_sheet_names = list(workbook.sheetnames)
    if actual_sheet_names != expected_sheet_names:
        raise validation_error_from_key(
            
                "instructor.validation.workbook_sheet_mismatch",
                template_id=blueprint.type_id,
                expected=expected_sheet_names,
                found=actual_sheet_names,
            
        )

    for sheet_schema in blueprint.sheets:
        if len(sheet_schema.header_matrix) != 1:
            raise validation_error_from_key(
                
                    "instructor.validation.sheet_single_header_row",
                    sheet_name=sheet_schema.name,
                
            )
        expected_headers = [normalize(h) for h in sheet_schema.header_matrix[0]]
        worksheet = workbook[sheet_schema.name]
        actual_headers = _read_header_row_values(worksheet, count=len(expected_headers))
        if actual_headers != expected_headers:
            raise validation_error_from_key(
                
                    "instructor.validation.header_mismatch",
                    sheet_name=sheet_schema.name,
                    expected=sheet_schema.header_matrix[0],
                
            )


def _iter_data_rows(worksheet: Any, expected_col_count: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=2, max_col=expected_col_count, values_only=True):
        values = list(row)
        if any(normalize(value) != "" for value in values):
            rows.append(values)
    return rows


def _validate_template_specific_rules(workbook: Any, template_id: str) -> None:
    strategy = get_template_strategy(template_id)
    strategy.validate_course_details_rules(workbook, context=None)


def _co_tokens(value: Any) -> list[int]:
    if value is None:
        return []
    if isinstance(value, bool):
        return []
    if isinstance(value, int):
        return [value]
    if isinstance(value, float):
        return [int(value)] if value.is_integer() else []

    token = str(value).strip()
    if not token:
        return []

    numbers: list[int] = []
    for item in token.split(","):
        part = item.strip()
        if not part:
            return []
        match = re.fullmatch(r"(?:co)?\s*(\d+)", part, flags=re.IGNORECASE)
        if not match:
            return []
        numbers.append(int(match.group(1)))
    return numbers




