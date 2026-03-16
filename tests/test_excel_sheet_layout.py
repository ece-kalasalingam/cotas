from __future__ import annotations

from collections import defaultdict

import pytest

from common import excel_sheet_layout as layout
from common.sheet_schema import SheetSchema, WorkbookBlueprint


class _ColDim:
    def __init__(self) -> None:
        self.width = None


class _FakeWS:
    def __init__(self, max_row: int, values_by_col: dict[int, list[object]] | None = None) -> None:
        self.max_row = max_row
        self.column_dimensions = defaultdict(_ColDim)
        self._vals = values_by_col or {}

    def iter_cols(self, *, min_col, max_col, min_row, max_row, values_only):
        del min_row, max_row, values_only
        for col in range(min_col, max_col + 1):
            yield tuple(self._vals.get(col, []))


def test_style_registry_for_setup_missing_blueprint_returns_empty(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(layout, "BLUEPRINT_REGISTRY", {})
    assert layout.style_registry_for_setup() == ({}, {})


def test_style_registry_for_setup_returns_copies(monkeypatch: pytest.MonkeyPatch) -> None:
    bp = WorkbookBlueprint(
        type_id=layout.ID_COURSE_SETUP,
        style_registry={"header": {"bold": True}, "body": {"locked": False}},
        sheets=[SheetSchema(name="S", header_matrix=[["A"]])],
    )
    monkeypatch.setattr(layout, "BLUEPRINT_REGISTRY", {layout.ID_COURSE_SETUP: bp})

    header, body = layout.style_registry_for_setup()
    header["bold"] = False

    assert bp.style_registry["header"]["bold"] is True
    assert body == {"locked": False}


def test_thin_border_and_color_without_hash() -> None:
    border = layout.thin_border()
    assert border.left.style == "thin"
    assert border.right.style == "thin"
    assert layout.color_without_hash("#AABBCC") == "AABBCC"
    assert layout.color_without_hash("AABBCC") == "AABBCC"


def test_excel_col_name() -> None:
    assert layout.excel_col_name(1) == "A"
    assert layout.excel_col_name(26) == "Z"
    assert layout.excel_col_name(27) == "AA"
    assert layout.excel_col_name(52) == "AZ"
    assert layout.excel_col_name(53) == "BA"


def test_autosize_columns_sets_min_width_when_no_rows() -> None:
    ws = _FakeWS(max_row=0)
    layout.autosize_columns(ws, 3)

    assert ws.column_dimensions["A"].width == 8
    assert ws.column_dimensions["B"].width == 8
    assert ws.column_dimensions["C"].width == 8


def test_autosize_columns_computes_clamped_widths() -> None:
    ws = _FakeWS(
        max_row=10,
        values_by_col={
            1: ["a", "abcd"],
            2: [None, "x" * 200],
        },
    )
    layout.autosize_columns(ws, 2)

    assert ws.column_dimensions["A"].width == 8
    assert ws.column_dimensions["B"].width == 60


def test_compute_sampled_column_widths_handles_padding_and_short_rows() -> None:
    sample_rows = [
        ["abc", "  x  "],
        [None],
        ["z", "long text"],
    ]
    widths = layout.compute_sampled_column_widths(sample_rows, 2, min_width=5, max_width=20, padding=1)

    assert widths[0] == 5
    assert widths[1] == 10
    assert widths[2] == 5


def test_set_header_selected_cell_updates_active_and_sqref() -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    layout.set_header_selected_cell(ws, 3)

    assert ws.sheet_view.selection[0].activeCell == "A3"
    assert ws.sheet_view.selection[0].sqref == "A3"


def test_apply_sheet_layout_and_protection_sets_expected_flags(monkeypatch: pytest.MonkeyPatch) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "h"

    calls = {"policy": 0, "autosize": 0, "selected": 0}
    monkeypatch.setattr(layout, "ensure_workbook_secret_policy", lambda: calls.__setitem__("policy", calls["policy"] + 1))
    monkeypatch.setattr(layout, "get_workbook_password", lambda: "secret")
    monkeypatch.setattr(layout, "autosize_columns", lambda _ws, _count: calls.__setitem__("autosize", calls["autosize"] + 1))
    monkeypatch.setattr(layout, "set_header_selected_cell", lambda _ws, _row: calls.__setitem__("selected", calls["selected"] + 1))

    layout.apply_sheet_layout_and_protection(
        ws=ws,
        header_row=1,
        header_count=4,
        paper_size=9,
        orientation="landscape",
    )

    assert calls == {"policy": 1, "autosize": 1, "selected": 1}
    assert ws.freeze_panes == "D2"
    assert ws.page_setup.paperSize == 9
    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws.protection.sheet is True
    assert isinstance(ws.protection.password, str)
    assert ws.protection.password
    assert ws.protection.sort == layout.ALLOW_SORT
    assert ws.protection.autoFilter == layout.ALLOW_FILTER
    assert ws.protection.selectLockedCells == layout.ALLOW_SELECT_LOCKED
    assert ws.protection.selectUnlockedCells == layout.ALLOW_SELECT_UNLOCKED



