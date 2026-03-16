from __future__ import annotations

from pathlib import Path

import pytest

from common.exceptions import ValidationError
from modules.instructor.validators import step3_filled_marks_validator as validator


class _Cell:
    def __init__(self, value):
        self.value = value


class _Sheet:
    def __init__(self, values: dict[str, object]) -> None:
        self._values = values

    def __getitem__(self, key: str):
        return _Cell(self._values.get(key))


class _Workbook:
    def __init__(self, sheetnames: list[str], sheets: dict[str, _Sheet]) -> None:
        self.sheetnames = sheetnames
        self._sheets = sheets
        self.closed = False

    def __getitem__(self, key: str):
        return self._sheets[key]

    def close(self) -> None:
        self.closed = True


def test_validate_manifest_schema_by_template_missing_validator() -> None:
    with pytest.raises(ValidationError):
        validator.validate_filled_marks_manifest_schema_by_template(object(), {}, template_id="UNKNOWN")


def test_validate_uploaded_filled_marks_workbook_openpyxl_missing(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    path = tmp_path / "m.xlsx"
    path.write_text("x", encoding="utf-8")

    import builtins

    real_import = builtins.__import__

    def _fake_import(name, *args, **kwargs):
        if name == "openpyxl":
            raise ModuleNotFoundError("no openpyxl")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", _fake_import)

    with pytest.raises(ValidationError) as exc:
        validator.validate_uploaded_filled_marks_workbook(path)
    assert getattr(exc.value, "code", "") == "OPENPYXL_MISSING"


def test_validate_uploaded_filled_marks_workbook_not_found(tmp_path: Path) -> None:
    with pytest.raises(ValidationError) as exc:
        validator.validate_uploaded_filled_marks_workbook(tmp_path / "missing.xlsx")
    assert getattr(exc.value, "code", "") == "WORKBOOK_NOT_FOUND"


def test_validate_uploaded_filled_marks_workbook_open_failure(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    import openpyxl

    path = tmp_path / "m.xlsx"
    path.write_text("x", encoding="utf-8")

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("bad")))

    with pytest.raises(ValidationError) as exc:
        validator.validate_uploaded_filled_marks_workbook(path)
    assert getattr(exc.value, "code", "") == "WORKBOOK_OPEN_FAILED"


def test_validate_uploaded_filled_marks_workbook_header_and_manifest_errors(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    import openpyxl

    path = tmp_path / "m.xlsx"
    path.write_text("x", encoding="utf-8")

    wb = _Workbook(
        sheetnames=[validator.SYSTEM_HASH_SHEET, validator.SYSTEM_LAYOUT_SHEET],
        sheets={
            validator.SYSTEM_HASH_SHEET: _Sheet({"A1": "bad", "B1": "bad", "A2": "", "B2": ""}),
            validator.SYSTEM_LAYOUT_SHEET: _Sheet({"A1": "bad", "B1": "bad", "A2": "", "B2": ""}),
        },
    )

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: wb)

    with pytest.raises(ValidationError):
        validator.validate_uploaded_filled_marks_workbook(path)

    assert wb.closed is True


def test_validate_uploaded_filled_marks_workbook_json_invalid_and_validator_call(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    import openpyxl

    path = tmp_path / "m.xlsx"
    path.write_text("x", encoding="utf-8")

    wb = _Workbook(
        sheetnames=[validator.SYSTEM_HASH_SHEET, validator.SYSTEM_LAYOUT_SHEET],
        sheets={
            validator.SYSTEM_HASH_SHEET: _Sheet(
                {
                    "A1": validator.SYSTEM_HASH_TEMPLATE_ID_KEY,
                    "B1": validator.SYSTEM_HASH_TEMPLATE_HASH_KEY,
                    "A2": validator.ID_COURSE_SETUP,
                    "B2": "sig1",
                }
            ),
            validator.SYSTEM_LAYOUT_SHEET: _Sheet(
                {
                    "A1": validator.SYSTEM_LAYOUT_MANIFEST_KEY,
                    "B1": validator.SYSTEM_LAYOUT_MANIFEST_HASH_KEY,
                    "A2": "{bad json",
                    "B2": "sig2",
                }
            ),
        },
    )

    seen: list[tuple[object, object, str]] = []

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: wb)
    monkeypatch.setattr(validator, "verify_payload_signature", lambda payload, sig: bool(payload) and bool(sig))
    monkeypatch.setattr(
        validator,
        "validate_filled_marks_manifest_schema_by_template",
        lambda workbook, manifest, *, template_id: seen.append((workbook, manifest, template_id)),
    )

    with pytest.raises(ValidationError):
        validator.validate_uploaded_filled_marks_workbook(path)

    wb2 = _Workbook(
        sheetnames=[validator.SYSTEM_HASH_SHEET, validator.SYSTEM_LAYOUT_SHEET],
        sheets={
            validator.SYSTEM_HASH_SHEET: _Sheet(
                {
                    "A1": validator.SYSTEM_HASH_TEMPLATE_ID_KEY,
                    "B1": validator.SYSTEM_HASH_TEMPLATE_HASH_KEY,
                    "A2": validator.ID_COURSE_SETUP,
                    "B2": "sig1",
                }
            ),
            validator.SYSTEM_LAYOUT_SHEET: _Sheet(
                {
                    "A1": validator.SYSTEM_LAYOUT_MANIFEST_KEY,
                    "B1": validator.SYSTEM_LAYOUT_MANIFEST_HASH_KEY,
                    "A2": "{}",
                    "B2": "sig2",
                }
            ),
        },
    )
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: wb2)

    validator.validate_uploaded_filled_marks_workbook(path)

    assert seen and seen[-1][2] == validator.ID_COURSE_SETUP
    assert wb2.closed is True

