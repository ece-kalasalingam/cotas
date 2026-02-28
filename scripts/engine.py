from __future__ import annotations

from collections import defaultdict
import re
from typing import Any, Dict, List

import openpyxl

from scripts.constants import ID_COURSE_SETUP, SYSTEM_HASH_SHEET_NAME


class UniversalEngine:
    def __init__(self, bp_registry: Dict[str, Any]):
        self.bp_registry = bp_registry
        self.bp = None
        self.data_store: Dict[str, List[List[Any]]] = {}
        self._headers: Dict[str, List[str]] = {}
        self.errors: List[str] = []

    def load_from_file(self, filepath: str) -> bool:
        self.errors = []
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as exc:
            self.errors.append(f"Failed to open workbook: {exc}")
            return False

        try:
            type_id = self._detect_type_from_workbook(wb) or ID_COURSE_SETUP

            self.bp = self.bp_registry.get(type_id)
            if not self.bp:
                self.errors.append(f"Unsupported workbook type: {type_id}")
                return False

            if type_id == ID_COURSE_SETUP:
                ok = self._load_setup_workbook(wb)
                if not ok:
                    return False
                return self._run_business_rules()

            self.errors.append(f"No loader implemented for workbook type: {type_id}")
            return False
        finally:
            wb.close()

    def load_with_external(self, primary_path: str, external_path: str) -> bool:
        ext = UniversalEngine(self.bp_registry)
        if not ext.load_from_file(external_path):
            self.errors = ext.errors.copy()
            return False
        return self.load_with_external_engine(primary_path, ext)

    def load_with_external_engine(self, primary_path: str, external_engine: "UniversalEngine") -> bool:
        self.errors = []
        try:
            wb = openpyxl.load_workbook(primary_path, read_only=True, data_only=True)
        except Exception as exc:
            self.errors.append(f"Failed to read marks workbook: {exc}")
            return False

        try:
            if not external_engine.data_store:
                self.errors.append("Setup context is empty. Load setup first.")
                return False

            setup = external_engine.data_store
            expected = self._derive_expected_marks_layout(setup)
            expected_students = [
                str(row[0]).strip()
                for row in setup.get("Students", [])
                if row and len(row) >= 1 and row[0] is not None and str(row[0]).strip()
            ]

            actual_name_map = {str(name).strip().lower(): name for name in wb.sheetnames}

            missing = [name for name in expected["required"] if name.lower() not in actual_name_map]
            if missing:
                self.errors.append("Missing required marks sheet(s): " + ", ".join(missing))
                return False

            self.bp = type("DynamicBP", (), {"type_id": "MARKS_ENTRY_V1"})()

            for comp_name in expected["direct_components"]:
                ws_name = actual_name_map.get(comp_name.lower())
                if not ws_name:
                    self.errors.append(f"Missing direct sheet: {comp_name}")
                    return False

                ws = wb[ws_name]
                q_headers = expected["questions_by_component"].get(comp_name, [])
                expected_header = ["RegNo", "Student_Name"] + q_headers + ["Total"]
                actual_header = self._read_header(ws, len(expected_header))

                if actual_header != expected_header:
                    self.errors.append(f"{comp_name}: header mismatch.")
                    return False

                co_row = self._read_row_values(ws, row=2, start_col=3, width=len(q_headers))
                max_row = self._read_row_values(ws, row=3, start_col=3, width=len(q_headers))
                if co_row != expected["co_by_component"].get(comp_name, []):
                    self.errors.append(f"{comp_name}: CO row mismatch.")
                    return False
                if not self._rows_numeric_equal(max_row, expected["max_by_component"].get(comp_name, [])):
                    self.errors.append(f"{comp_name}: Max row mismatch.")
                    return False

                students, _ = self._validate_direct_sheet_entries(ws, comp_name, len(q_headers))
                if students != expected_students:
                    self.errors.append(f"{comp_name}: student list mismatch.")
                    return False

            for tool_name in expected["indirect_tools"]:
                expected_sheet = f"{tool_name}_Indirect"
                ws_name = actual_name_map.get(expected_sheet.lower())
                if not ws_name:
                    self.errors.append(f"Missing indirect sheet: {expected_sheet}")
                    return False

                ws = wb[ws_name]
                expected_header = ["RegNo", "Student_Name"] + [f"CO{i}" for i in expected["co_numbers"]]
                actual_header = self._read_header(ws, len(expected_header))

                if actual_header != expected_header:
                    self.errors.append(f"{expected_sheet}: header mismatch.")
                    return False

                students = self._validate_indirect_sheet_entries(ws, expected_sheet, len(expected["co_numbers"]))
                if students != expected_students:
                    self.errors.append(f"{expected_sheet}: student list mismatch.")
                    return False

            return True
        finally:
            wb.close()

    def load_marks_standalone(self, primary_path: str) -> bool:
        self.errors = []
        try:
            wb = openpyxl.load_workbook(primary_path, read_only=True, data_only=True)
        except Exception as exc:
            self.errors.append(f"Failed to read marks workbook: {exc}")
            return False

        try:
            actual_name_map = {str(name).strip().lower(): name for name in wb.sheetnames}
            assess_sheet_name = actual_name_map.get("assessment_config")
            if not assess_sheet_name:
                self.errors.append("Missing Assessment_Config sheet in marks workbook.")
                return False

            direct_components, indirect_tools = self._read_assessment_config(wb[assess_sheet_name])
            if not direct_components:
                self.errors.append("No direct components found in Assessment_Config.")
                return False

            co_numbers = set()
            baseline_students: List[str] | None = None

            for comp_name in direct_components:
                ws_name = actual_name_map.get(comp_name.lower())
                if not ws_name:
                    self.errors.append(f"Missing direct sheet: {comp_name}")
                    return False

                ws = wb[ws_name]
                header, q_headers = self._read_direct_header(ws)
                if len(header) < 4 or header[:2] != ["RegNo", "Student_Name"] or header[-1] != "Total":
                    self.errors.append(f"{comp_name}: invalid header format.")
                    return False
                if not q_headers:
                    self.errors.append(f"{comp_name}: no question columns found.")
                    return False

                if self._cell_text(ws, 2, 1) != "CO":
                    self.errors.append(f"{comp_name}: row 2, col 1 must be CO.")
                    return False
                if self._cell_text(ws, 3, 1) != "Max":
                    self.errors.append(f"{comp_name}: row 3, col 1 must be Max.")
                    return False

                students, sheet_cos = self._validate_direct_sheet_entries(ws, comp_name, len(q_headers))
                for n in sheet_cos:
                    co_numbers.add(n)
                if baseline_students is None:
                    baseline_students = students
                elif students != baseline_students:
                    self.errors.append(f"{comp_name}: student list mismatch.")
                    return False

            if not co_numbers:
                co_numbers = {1}
            expected_co_header = [f"CO{i}" for i in sorted(co_numbers)]

            for tool_name in indirect_tools:
                expected_sheet = f"{tool_name}_Indirect"
                ws_name = actual_name_map.get(expected_sheet.lower())
                if not ws_name:
                    self.errors.append(f"Missing indirect sheet: {expected_sheet}")
                    return False

                ws = wb[ws_name]
                expected_header = ["RegNo", "Student_Name"] + expected_co_header
                actual_header = self._read_header(ws, len(expected_header))
                if actual_header != expected_header:
                    self.errors.append(f"{expected_sheet}: header mismatch.")
                    return False

                students = self._validate_indirect_sheet_entries(ws, expected_sheet, len(expected_co_header))
                if baseline_students is not None and students != baseline_students:
                    self.errors.append(f"{expected_sheet}: student list mismatch.")
                    return False

            self.bp = type("DynamicBP", (), {"type_id": "MARKS_ENTRY_V1"})()
            return True
        finally:
            wb.close()

    def get_col_idx(self, sheet_name: str, col_name: str) -> int:
        headers = self._headers.get(sheet_name, [])
        target = "" if col_name is None else str(col_name).strip().lower()
        for i, val in enumerate(headers):
            if str(val).strip().lower() == target:
                return i
        return -1

    def _detect_type(self, filepath: str) -> str | None:
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            try:
                return self._detect_type_from_workbook(wb)
            finally:
                wb.close()
        except Exception:
            return None

    def _detect_type_from_workbook(self, wb) -> str | None:
        if SYSTEM_HASH_SHEET_NAME in wb.sheetnames:
            ws = wb[SYSTEM_HASH_SHEET_NAME]
            val = ws.cell(row=1, column=1).value
            if val:
                return str(val).strip()
        return None

    def _load_setup_workbook(self, wb) -> bool:
        self.data_store = {}
        self._headers = {}

        for sheet_schema in self.bp.sheets:
            if sheet_schema.name not in wb.sheetnames:
                self.errors.append(f"Missing required sheet: {sheet_schema.name}")
                return False

            ws = wb[sheet_schema.name]
            expected_header = [str(v).strip() for v in sheet_schema.header_matrix[-1]]
            actual_header = self._read_header(ws, len(expected_header))

            if actual_header != expected_header:
                self.errors.append(
                    f"Header mismatch in '{sheet_schema.name}'. Expected {expected_header}, found {actual_header}"
                )
                return False

            self._headers[sheet_schema.name] = expected_header

            rows: List[List[Any]] = []
            for r in ws.iter_rows(min_row=2, max_col=len(expected_header), values_only=True):
                if all(v is None or str(v).strip() == "" for v in r):
                    continue
                rows.append(list(r))
            self.data_store[sheet_schema.name] = rows

        return True

    def _read_header(self, ws, width: int) -> List[str]:
        values = [ws.cell(row=1, column=c).value for c in range(1, width + 1)]
        return ["" if v is None else str(v).strip() for v in values]

    def _run_business_rules(self) -> bool:
        problems = []
        for rule in self.bp.business_rules:
            errs = rule.run(self)
            problems.extend([f"Rule {rule.rule_id}: {e}" for e in errs])

        if problems:
            self.errors.extend(problems)
            return False
        return True

    def _derive_expected_marks_layout(self, setup_store: Dict[str, List[List[Any]]]) -> Dict[str, Any]:
        assess_rows = setup_store.get("Assessment_Config", [])
        qmap_rows = setup_store.get("Question_Map", [])

        direct_components: List[str] = []
        indirect_tools: List[str] = []
        for row in assess_rows:
            if len(row) < 5:
                continue
            name = "" if row[0] is None else str(row[0]).strip()
            direct_flag = "" if row[4] is None else str(row[4]).strip().upper()
            if not name:
                continue
            if direct_flag == "YES":
                direct_components.append(name)
            elif direct_flag == "NO":
                indirect_tools.append(name)

        q_by_comp: Dict[str, List[str]] = defaultdict(list)
        co_by_comp: Dict[str, List[str]] = defaultdict(list)
        max_by_comp: Dict[str, List[str]] = defaultdict(list)
        co_numbers = set()

        for row in qmap_rows:
            if len(row) < 4:
                continue
            comp = "" if row[0] is None else str(row[0]).strip()
            qid = "" if row[1] is None else str(row[1]).strip()
            co_raw = "" if row[3] is None else str(row[3]).strip()
            max_raw = "" if row[2] is None else str(row[2]).strip()

            if comp and qid:
                q_by_comp[comp].append(qid)
                co_by_comp[comp].append(self._normalize_co_text(co_raw))
                max_by_comp[comp].append(max_raw)

            for token in co_raw.replace("CO", "").replace(" ", "").replace(";", ",").split(","):
                if not token:
                    continue
                try:
                    co_numbers.add(int(float(token)))
                except Exception:
                    continue

        if not co_numbers:
            co_numbers = {1}

        required = list(direct_components) + [f"{name}_Indirect" for name in indirect_tools]

        return {
            "required": required,
            "direct_components": direct_components,
            "indirect_tools": indirect_tools,
            "questions_by_component": dict(q_by_comp),
            "co_by_component": dict(co_by_comp),
            "max_by_component": dict(max_by_comp),
            "co_numbers": sorted(co_numbers),
        }

    def _read_assessment_config(self, ws) -> tuple[List[str], List[str]]:
        direct_components: List[str] = []
        indirect_tools: List[str] = []

        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            name = "" if row[0] is None else str(row[0]).strip()
            flag = self._normalize_direct_flag(row[4] if len(row) >= 5 else "")
            if not name:
                continue
            if flag == "YES":
                direct_components.append(name)
            elif flag == "NO":
                indirect_tools.append(name)

        return direct_components, indirect_tools

    def _read_direct_header(self, ws) -> tuple[List[str], List[str]]:
        vals: List[str] = []
        col = 1
        while col <= 200:
            v = ws.cell(row=1, column=col).value
            text = "" if v is None else str(v).strip()
            if text == "":
                break
            vals.append(text)
            col += 1

        q_headers: List[str] = vals[2:-1] if len(vals) >= 4 else []
        return vals, q_headers

    def _read_regnos(self, ws, min_row: int) -> List[str]:
        out: List[str] = []
        row_idx = min_row
        while row_idx <= 50000:
            reg = self._cell_text(ws, row_idx, 1)
            if not reg:
                break
            out.append(reg)
            row_idx += 1
        return out

    def _cell_text(self, ws, row: int, col: int) -> str:
        val = ws.cell(row=row, column=col).value
        return "" if val is None else str(val).strip()

    def _read_row_values(self, ws, row: int, start_col: int, width: int) -> List[str]:
        return [self._cell_text(ws, row, start_col + i) for i in range(width)]

    def _validate_direct_sheet_entries(self, ws, comp_name: str, q_count: int) -> tuple[List[str], set[int]]:
        co_numbers = set()
        max_marks: List[float] = []
        for i in range(q_count):
            co_text = self._cell_text(ws, 2, 3 + i)
            for n in self._parse_co_numbers(co_text):
                co_numbers.add(n)

            ok_max, max_val = self._to_float(ws.cell(row=3, column=3 + i).value)
            if not ok_max:
                self.errors.append(f"{comp_name}: invalid Max value at question {i + 1}.")
                return [], set()
            if max_val < 0:
                self.errors.append(f"{comp_name}: Max cannot be negative at question {i + 1}.")
                return [], set()
            max_marks.append(max_val)

        students: List[str] = []
        row_idx = 4
        while row_idx <= 50000:
            reg = self._cell_text(ws, row_idx, 1)
            if not reg:
                break
            students.append(reg)

            row_sum = 0.0
            for i in range(q_count):
                val = ws.cell(row=row_idx, column=3 + i).value
                if val is None or str(val).strip() == "":
                    continue
                ok, num = self._to_float(val)
                if not ok:
                    self.errors.append(f"{comp_name}: non-numeric mark at row {row_idx}, question {i + 1}.")
                    return [], set()
                if num < 0 or num > max_marks[i]:
                    self.errors.append(
                        f"{comp_name}: mark out of range at row {row_idx}, question {i + 1} (0 to {max_marks[i]})."
                    )
                    return [], set()
                row_sum += num

            total_val = ws.cell(row=row_idx, column=3 + q_count).value
            if total_val is not None and str(total_val).strip() != "":
                ok_total, total_num = self._to_float(total_val)
                if not ok_total:
                    self.errors.append(f"{comp_name}: non-numeric Total at row {row_idx}.")
                    return [], set()
                if abs(total_num - row_sum) > 0.01:
                    self.errors.append(f"{comp_name}: Total mismatch at row {row_idx}.")
                    return [], set()

            row_idx += 1

        return students, co_numbers

    def _validate_indirect_sheet_entries(self, ws, sheet_name: str, co_count: int) -> List[str]:
        students: List[str] = []
        row_idx = 2
        while row_idx <= 50000:
            reg = self._cell_text(ws, row_idx, 1)
            if not reg:
                break
            students.append(reg)

            for i in range(co_count):
                val = ws.cell(row=row_idx, column=3 + i).value
                if val is None or str(val).strip() == "":
                    continue
                ok, num = self._to_float(val)
                if not ok:
                    self.errors.append(f"{sheet_name}: non-numeric CO value at row {row_idx}, col {i + 1}.")
                    return []
                if num < 0 or num > 100:
                    self.errors.append(f"{sheet_name}: CO value out of range at row {row_idx}, col {i + 1} (0 to 100).")
                    return []
            row_idx += 1
        return students

    def _to_float(self, value: Any) -> tuple[bool, float]:
        if value is None:
            return False, 0.0
        text = str(value).strip()
        if text == "":
            return False, 0.0
        try:
            return True, float(text)
        except Exception:
            return False, 0.0

    def _rows_numeric_equal(self, actual: List[str], expected: List[str]) -> bool:
        if len(actual) != len(expected):
            return False
        for a, e in zip(actual, expected):
            ok_a, n_a = self._to_float(a)
            ok_e, n_e = self._to_float(e)
            if not ok_a or not ok_e:
                return False
            if abs(n_a - n_e) > 0.0001:
                return False
        return True

    def _normalize_direct_flag(self, value: Any) -> str:
        token = "" if value is None else str(value).strip().lower()
        if token in {"yes", "y", "true", "1", "direct", "d"}:
            return "YES"
        if token in {"no", "n", "false", "0", "indirect", "i"}:
            return "NO"
        return ""

    def _parse_co_numbers(self, raw: Any) -> List[int]:
        s = "" if raw is None else str(raw).strip()
        if not s:
            return []

        out: List[int] = []
        for token in re.split(r"[;,|]+", s.replace(" ", "").upper().replace("CO", "")):
            if not token:
                continue
            try:
                out.append(int(float(token)))
            except Exception:
                continue
        return out

    def _normalize_co_text(self, raw: Any) -> str:
        nums = self._parse_co_numbers(raw)
        if nums:
            return ",".join(str(x) for x in nums)
        return "" if raw is None else str(raw).strip()
