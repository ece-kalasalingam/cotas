from __future__ import annotations

import argparse
import json
import statistics
import sys
import tempfile
import time
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from common.constants import ID_COURSE_SETUP
from common.jobs import CancellationToken
from domain.template_strategy_router import generate_workbook, validate_workbooks
from domain.template_versions.course_setup_v2_impl.co_attainment import (
    generate_final_report_workbook,
)
from services import InstructorWorkflowService


def _time_call(fn):
    """Time call.
    
    Args:
        fn: Parameter value.
    
    Returns:
        None.
    
    Raises:
        None.
    """
    started = time.perf_counter()
    fn()
    return (time.perf_counter() - started) * 1000.0


def _fill_marks_workbook(marks_path: Path, mark_value: float = 1.0) -> None:
    """Populate generated marks workbook with sample values for perf flow."""
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for perf soak marks fill.") from exc

    workbook = openpyxl.load_workbook(marks_path)
    try:
        manifest_text = workbook["__SYSTEM_LAYOUT__"]["A2"].value
        if not isinstance(manifest_text, str):
            return
        manifest = json.loads(manifest_text)
        for spec in manifest.get("sheets", []):
            kind = str(spec.get("kind") or "")
            if kind not in {"direct_co_wise", "direct_non_co_wise", "indirect"}:
                continue
            sheet = workbook[str(spec["name"])]
            header_row = int(spec["header_row"])
            header_count = len(spec["headers"])
            if kind == "indirect":
                first_data_row = header_row + 1
                mark_cols = range(4, header_count + 1)
            elif kind == "direct_non_co_wise":
                first_data_row = header_row + 3
                mark_cols = range(4, 5)
            else:
                first_data_row = header_row + 3
                mark_cols = range(4, header_count)
            row = first_data_row
            while True:
                reg_no = sheet.cell(row=row, column=2).value
                student_name = sheet.cell(row=row, column=3).value
                if reg_no is None and student_name is None:
                    break
                for col in mark_cols:
                    sheet.cell(row=row, column=col, value=mark_value)
                row += 1
        workbook.save(marks_path)
    finally:
        workbook.close()


def main() -> int:
    """Main.
    
    Args:
        None.
    
    Returns:
        int: Return value.
    
    Raises:
        None.
    """
    parser = argparse.ArgumentParser(
        description="Run instructor workflow performance/soak checks."
    )
    parser.add_argument(
        "--iterations",
        type=int,
        default=5,
        help="How many full workflow iterations to run.",
    )
    parser.add_argument(
        "--enforce",
        action="store_true",
        help="Fail with non-zero exit code when thresholds are exceeded.",
    )
    parser.add_argument(
        "--max-step-ms",
        type=float,
        default=8000.0,
        help="Per-step p95 threshold in milliseconds.",
    )
    args = parser.parse_args()

    service: Any = InstructorWorkflowService()
    timings: dict[str, list[float]] = {
        "generate_course_details_template": [],
        "validate_course_details_workbooks": [],
        "generate_marks_template": [],
        "generate_final_report": [],
    }

    with tempfile.TemporaryDirectory(prefix="instructor_perf_") as temp_dir:
        root = Path(temp_dir)
        for index in range(args.iterations):
            course_details = root / f"course_details_{index}.xlsx"
            marks_template = root / f"marks_template_{index}.xlsx"
            final_report = root / f"final_report_{index}.xlsx"

            timings["generate_course_details_template"].append(
                _time_call(
                    lambda: generate_workbook(
                        template_id=ID_COURSE_SETUP,
                        output_path=course_details,
                        workbook_name=course_details.name,
                        workbook_kind="course_details_template",
                        cancel_token=CancellationToken(),
                    )
                )
            )

            timings["validate_course_details_workbooks"].append(
                _time_call(
                    lambda: validate_workbooks(
                        template_id=ID_COURSE_SETUP,
                        workbook_paths=[course_details],
                        workbook_kind="course_details",
                        cancel_token=CancellationToken(),
                    )
                )
            )

            ctx3 = service.create_job_context(step_id="step2", payload={"i": index})
            timings["generate_marks_template"].append(
                _time_call(
                    lambda: service.generate_marks_template(
                        course_details,
                        marks_template,
                        context=ctx3,
                        cancel_token=CancellationToken(),
                    )
                )
            )
            _fill_marks_workbook(marks_template, mark_value=1.0)

            timings["generate_final_report"].append(
                _time_call(
                    lambda: generate_final_report_workbook(
                        filled_marks_path=marks_template,
                        output_path=final_report,
                        cancel_token=CancellationToken(),
                    )
                )
            )

    summary: dict[str, dict[str, float]] = {}
    threshold_breaches: list[str] = []
    for step_name, values in timings.items():
        ordered = sorted(values)
        p95_index = max(0, int(round(0.95 * (len(ordered) - 1))))
        p95 = ordered[p95_index]
        step_summary = {
            "min_ms": round(min(values), 2),
            "mean_ms": round(statistics.mean(values), 2),
            "p95_ms": round(p95, 2),
            "max_ms": round(max(values), 2),
        }
        summary[step_name] = step_summary
        if step_summary["p95_ms"] > args.max_step_ms:
            threshold_breaches.append(step_name)

    report = {
        "iterations": args.iterations,
        "threshold_ms": args.max_step_ms,
        "summary": summary,
        "threshold_breaches": threshold_breaches,
    }
    print(json.dumps(report, indent=2))

    if args.enforce and threshold_breaches:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
